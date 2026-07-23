# =========================================================
# app.py — MARWHATS (checkpoint) / Poliutech
# Limpio + Roles (ADMIN / USER) + Filtro por Responsable
# =========================================================
from __future__ import annotations

import os, io, csv, sys, math, re, json, traceback, unicodedata, smtplib, zipfile, logging, base64
import mimetypes
import requests
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from typing import Iterable, Optional, List
from urllib.parse import urlparse
from pathlib import Path
from functools import wraps
from email.message import EmailMessage
from email.utils import getaddresses
from html import escape
import xml.etree.ElementTree as ET
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from werkzeug.utils import secure_filename
try:
    import firebase_admin
    from firebase_admin import credentials as firebase_credentials
    from firebase_admin import messaging as firebase_messaging
except Exception:
    firebase_admin = None
    firebase_credentials = None
    firebase_messaging = None


# -------------------------------
# Condiciones comerciales
# -------------------------------
# Ya no se agregan condiciones por defecto. Solo se exporta lo capturado
# por el usuario y, cuando aplique, la trazabilidad de la zona.
DEFAULT_CONDICIONES: list[str] = []
VALID_ESTATUS_SEGUIMIENTO = [
    "ENVIADA",
    "PENDIENTE",
    "EN CURSO",
    "O. TERMINADA",
    "FINALIZADA",
    "GANADA",
    "PERDIDA",
]
VALID_ESTATUS_APROBACION = [
    "APROBADA",
    "RECHAZADA",
    "EN REVISIÓN",
]
VALID_ESTATUS = VALID_ESTATUS_SEGUIMIENTO
PROSPECT_STATUS_OPTIONS = [
    "PENDIENTE",
    "CONTACTADO",
    "COTIZADO",
    "FINALIZADO",
    "RECHAZADO",
]
TICKET_STATUS_OPTIONS = [
    "NUEVO",
    "EN REVISION",
    "EN PROCESO",
    "ESPERANDO RESPUESTA",
    "RESUELTO",
    "CERRADO",
]
TICKET_PRIORITY_OPTIONS = ["BAJA", "MEDIA", "ALTA", "URGENTE"]
TICKET_CATEGORY_OPTIONS = ["GENERAL", "SISTEMA", "COTIZACIONES", "COMPRAS", "FACTURACION", "APP MOVIL"]
TICKET_ALLOWED_EXTENSIONS = {
    ".png", ".jpg", ".jpeg", ".webp", ".gif",
    ".pdf", ".txt", ".log", ".csv", ".xlsx", ".xls", ".doc", ".docx",
}
TICKET_MAX_ATTACHMENTS = 6
PROVIDER_NUMBERS_JSON = Path(__file__).resolve().parent / "provider_numbers.json"
PROVIDER_NUMBERS_XLSX = Path.home() / "Downloads" / "NUMEROS DE PROVEEDOR POLIUTECH.xlsx"
REGISTRO_OBRAS_JSON = Path(__file__).resolve().parent / "registro_obras.json"
OPENAI_API_KEY = (os.getenv("OPENAI_API_KEY") or "").strip()
OPENAI_TRANSCRIBE_MODEL = (os.getenv("OPENAI_TRANSCRIBE_MODEL") or "gpt-4o-mini-transcribe").strip()
XLSX_NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "p": "http://schemas.openxmlformats.org/package/2006/relationships",
}

VOICE_TRANSCRIPTION_PROMPT = (
    "Transcribe audio en espanol de Mexico para captura de cotizaciones. "
    "Conserva y favorece estas etiquetas cuando se escuchen: CLIENTE, EMPRESA, CORREO, TELEFONO, CIUDAD, "
    "CONCEPTO, OTRO CONCEPTO, UNIDAD, CANTIDAD, PRECIO, SISTEMA. "
    "No inventes datos. Si el hablante dice arroba, punto com, metro cuadrado o metros cuadrados, "
    "transcribelos de forma util y legible."
)



def _split_notas_y_zona(notas_raw: str) -> tuple[str, str]:
    notas_raw = (notas_raw or "").strip()
    extras = []
    zona_line = ""
    for ln in notas_raw.splitlines():
        s = ln.strip()
        if s.lower().startswith("zona:"):
            zona_line = s
        else:
            # Conserva los renglones vacios intencionales entre condiciones.
            extras.append(s)
    return "\n".join(extras).strip(), zona_line

def _condiciones_comerciales_finales(notas_raw: str) -> list[str]:
    extras_txt, zona_line = _split_notas_y_zona(notas_raw)
    items = list(DEFAULT_CONDICIONES)
    if zona_line:
        items.append(zona_line)
    if extras_txt:
        for ln in extras_txt.splitlines():
            s = ln.strip()
            items.append(s)
    return items


def _excel_col_to_index(ref: str) -> int:
    letters = "".join(ch for ch in (ref or "") if ch.isalpha()).upper()
    idx = 0
    for ch in letters:
        idx = (idx * 26) + (ord(ch) - 64)
    return max(idx - 1, 0)


def _xlsx_cell_text(cell, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join((node.text or "") for node in cell.findall(".//a:t", XLSX_NS)).strip()

    value_node = cell.find("a:v", XLSX_NS)
    raw_value = "" if value_node is None or value_node.text is None else value_node.text
    if cell_type == "s" and raw_value != "":
        try:
            return str(shared_strings[int(raw_value)]).strip()
        except Exception:
            return ""
    return str(raw_value).strip()


def _load_provider_numbers_from_xlsx() -> list[dict]:
    if not PROVIDER_NUMBERS_XLSX.exists():
        return []

    with zipfile.ZipFile(PROVIDER_NUMBERS_XLSX) as workbook_zip:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in workbook_zip.namelist():
            shared_root = ET.fromstring(workbook_zip.read("xl/sharedStrings.xml"))
            for item in shared_root.findall("a:si", XLSX_NS):
                shared_strings.append(
                    "".join((node.text or "") for node in item.findall(".//a:t", XLSX_NS)).strip()
                )

        workbook_root = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
        rels_root = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))
        rel_map = {
            rel.attrib.get("Id"): rel.attrib.get("Target", "")
            for rel in rels_root.findall("p:Relationship", XLSX_NS)
        }

        target_sheet = None
        for sheet in workbook_root.findall("a:sheets/a:sheet", XLSX_NS):
            sheet_name = (sheet.attrib.get("name") or "").strip().lower()
            rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if rel_id and sheet_name == "table 2":
                target = rel_map.get(rel_id, "")
                if target:
                    target_sheet = f"xl/{target.lstrip('/')}"
                    break

        if not target_sheet or target_sheet not in workbook_zip.namelist():
            return []

        sheet_root = ET.fromstring(workbook_zip.read(target_sheet))
        rows = sheet_root.findall("a:sheetData/a:row", XLSX_NS)
        parsed_rows: list[list[str]] = []

        for row in rows:
            values_by_col: dict[int, str] = {}
            for cell in row.findall("a:c", XLSX_NS):
                ref = cell.attrib.get("r", "")
                values_by_col[_excel_col_to_index(ref)] = _xlsx_cell_text(cell, shared_strings)

            if not values_by_col:
                continue

            max_col = max(values_by_col)
            parsed_rows.append([values_by_col.get(col, "").strip() for col in range(max_col + 1)])

        if not parsed_rows:
            return []

        data_rows = parsed_rows[1:]
        records: list[dict] = []
        for idx, row in enumerate(data_rows, start=1):
            numero = row[0].strip() if len(row) > 0 else ""
            empresa = row[1].strip() if len(row) > 1 else ""
            razon_social = row[2].strip() if len(row) > 2 else ""
            if not any([numero, empresa, razon_social]):
                continue
            records.append({
                "id": idx,
                "numero": numero,
                "empresa": empresa,
                "razon_social_poliutech": razon_social,
                "relacion": "PROVEEDOR",
                "contacto": "",
                "telefono": "",
                "correo": "",
            })
        return records


def _save_provider_numbers(rows: list[dict]) -> None:
    PROVIDER_NUMBERS_JSON.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _normalize_provider_row(row: Optional[dict], idx: int) -> dict:
    row = row or {}
    relacion = str(row.get("relacion", "")).strip().upper()
    if relacion not in {"CLIENTE", "PROVEEDOR"}:
        relacion = "PROVEEDOR"
    return {
        "id": idx,
        "numero": str(row.get("numero", "")).strip(),
        "empresa": str(row.get("empresa", "")).strip(),
        "razon_social_poliutech": str(row.get("razon_social_poliutech", "")).strip(),
        "relacion": relacion,
        "contacto": str(row.get("contacto", "")).strip(),
        "telefono": str(row.get("telefono", "")).strip(),
        "correo": str(row.get("correo", "")).strip(),
    }


def _load_provider_numbers() -> list[dict]:
    if PROVIDER_NUMBERS_JSON.exists():
        try:
            data = json.loads(PROVIDER_NUMBERS_JSON.read_text(encoding="utf-8"))
            if isinstance(data, list):
                return [_normalize_provider_row(row, idx) for idx, row in enumerate(data, start=1)]
        except Exception:
            pass

    seeded = _load_provider_numbers_from_xlsx()
    _save_provider_numbers(seeded)
    return seeded


def _normalize_registro_obra_row(row: Optional[dict], idx: int) -> dict:
    row = row or {}
    try:
        row_id = int(row.get("id") or 0)
    except Exception:
        row_id = 0
    return {
        "id": row_id,
        "numero": str(row.get("numero", "")).strip(),
        "obra": str(row.get("obra", "")).strip(),
        "ubicacion": str(row.get("ubicacion", "")).strip(),
        "encargado": str(row.get("encargado", "")).strip(),
        "puesto": str(row.get("puesto", "")).strip(),
        "telefono": str(row.get("telefono", "")).strip(),
        "correo": str(row.get("correo", "")).strip(),
        "responsable": str(row.get("responsable", "")).strip(),
    }


def _clean_registro_obra_excel_value(value: object) -> str:
    text = str(value or "").strip()
    if text in {'"', "-", "—", "N/A", "n/a"}:
        return ""
    return text


def _normalize_registro_obra_phone(value: object) -> str:
    text = _clean_registro_obra_excel_value(value)
    if not text:
        return ""
    try:
        if re.fullmatch(r"\d+(?:\.\d+)?E\d+", text, re.IGNORECASE):
            text = format(int(float(text)), "d")
    except Exception:
        pass
    return text


def _registro_obra_duplicate_key(row: dict) -> tuple[str, str, str, str, str]:
    def norm(value: object) -> str:
        raw = str(value or "").strip()
        normalized = unicodedata.normalize("NFKD", raw)
        return "".join(ch for ch in normalized if not unicodedata.combining(ch)).lower().strip()

    return (
        norm(row.get("obra")),
        norm(row.get("ubicacion")),
        norm(row.get("encargado")),
        norm(row.get("telefono")),
        norm(row.get("correo")),
    )


def _registro_obra_to_row(item: "RegistroObra", idx: Optional[int] = None) -> dict:
    position = idx if idx is not None else (item.id or 0)
    row = _normalize_registro_obra_row({
        "numero": item.numero,
        "obra": item.obra,
        "ubicacion": item.ubicacion,
        "encargado": item.encargado,
        "puesto": item.puesto,
        "telefono": item.telefono,
        "correo": item.correo,
        "responsable": item.responsable,
    }, position)
    row["id"] = item.id
    row["numero"] = str(position)
    row["seguimiento_count"] = len(item.seguimientos or [])
    return row


def _load_registro_obras() -> list[dict]:
    items = RegistroObra.query.order_by(RegistroObra.numero.asc(), RegistroObra.id.asc()).all()
    return [_registro_obra_to_row(item, idx) for idx, item in enumerate(items, start=1)]


def _registro_obras_filters_from_request() -> dict[str, str]:
    return {
        "obra": (request.args.get("obra") or "").strip().lower(),
        "responsable": (request.args.get("responsable") or "").strip().lower(),
    }


def _registro_obra_matches_filters(row: dict, filters: dict[str, str]) -> bool:
    for field, needle in filters.items():
        if needle and needle not in str(row.get(field, "")).strip().lower():
            return False
    return True


def _filter_registro_obras(rows: list[dict], filters: dict[str, str]) -> list[dict]:
    filtered = [row for row in rows if _registro_obra_matches_filters(row, filters)]
    if is_admin():
        return filtered
    ra = (responsable_actual() or "").strip().lower()
    if not ra:
        return []
    return [row for row in filtered if (row.get("responsable") or "").strip().lower() == ra]


def _sync_cliente_from_registro_obra(row: dict) -> None:
    nombre_cliente = (row.get("encargado") or "").strip()
    empresa = (row.get("obra") or "").strip()
    if not nombre_cliente:
        return

    query = Cliente.query.filter(db.func.lower(Cliente.nombre_cliente) == nombre_cliente.lower())
    if empresa:
        query = query.filter(db.func.lower(Cliente.empresa) == empresa.lower())
    cliente = query.first()
    if not cliente:
        cliente = Cliente(nombre_cliente=nombre_cliente, empresa=empresa or None)
        db.session.add(cliente)

    responsable = (row.get("responsable") or "").strip()
    cliente.responsable = responsable or cliente.responsable
    cliente.correo = (row.get("correo") or "").strip() or cliente.correo
    cliente.telefono = (row.get("telefono") or "").strip() or cliente.telefono
    cliente.direccion = (row.get("ubicacion") or "").strip() or cliente.direccion


def _registro_obra_email_body() -> str:
    return (
        "Buen día,\n\n"
        "Con gusto de saludarlo y de acuerdo a la plática que pudimos sostener con usted o un representante de su empresa, por medio del presente, nos permitimos presentar a Corporativo Poliutech, una empresa especializada en la aplicación de recubrimientos para la construcción. Contamos con certificaciones como aplicadores en pisos epóxicos, impermeabilizantes, poliureas, pinturas y diversos recubrimientos especializados para proyectos en los sectores industrial, comercial, público y privado.\n\n"
        "Nos distinguimos por adaptarnos a los requerimientos de nuestros clientes, optimizando al máximo los recursos y espacios disponibles para garantizar soluciones eficientes y de alta calidad.\n\n"
        "Adjunto a este correo encontrará nuestro CV empresarial, donde podrá conocer más sobre nuestros servicios y proyectos.\n\n"
        "Quedamos a sus órdenes para cualquier necesidad o consulta.\n\n"
        "Atentamente Poliutech Recubrimientos Especializados"
    )


def _send_registro_obra_email(row: dict) -> None:
    recipients = _parse_email_list(row.get("correo"))
    if not recipients:
        raise ValueError("El registro no tiene correo destino.")
    if not REGISTRO_MAIL_ATTACHMENT.exists():
        raise FileNotFoundError(f"No se encontró el adjunto requerido: {REGISTRO_MAIL_ATTACHMENT.name}")

    msg = EmailMessage()
    msg["Subject"] = "Te visitamos recientemente"
    msg["From"] = f"Poliutech <{REGISTRO_MAIL_FROM}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content(_registro_obra_email_body())

    attachment_bytes = REGISTRO_MAIL_ATTACHMENT.read_bytes()
    msg.add_attachment(
        attachment_bytes,
        maintype="application",
        subtype="pdf",
        filename=REGISTRO_MAIL_ATTACHMENT.name,
    )

    with smtplib.SMTP(REGISTRO_MAIL_HOST, REGISTRO_MAIL_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(REGISTRO_MAIL_USERNAME, REGISTRO_MAIL_PASSWORD)
        smtp.send_message(msg, to_addrs=recipients)


def _save_registro_obras(rows: list[dict]) -> None:
    existing = {item.id: item for item in RegistroObra.query.all()}
    seen_ids: set[int] = set()
    for idx, raw_row in enumerate(rows, start=1):
        row = _normalize_registro_obra_row(raw_row, idx)
        row_id = row.get("id")
        item = existing.get(row_id) if isinstance(row_id, int) and row_id > 0 else None
        if item is None:
            item = RegistroObra()
            db.session.add(item)
            db.session.flush()
        seen_ids.add(item.id)
        item.numero = idx
        item.obra = row.get("obra", "")
        item.ubicacion = row.get("ubicacion", "")
        item.encargado = row.get("encargado", "")
        item.puesto = row.get("puesto", "")
        item.telefono = row.get("telefono", "")
        item.correo = row.get("correo", "")
        item.responsable = row.get("responsable", "")
        raw_row["id"] = item.id
        raw_row["numero"] = str(idx)

    for item_id, item in existing.items():
        if item_id not in seen_ids:
            db.session.delete(item)


def _migrate_registro_obras_from_json() -> None:
    if RegistroObra.query.first() or not REGISTRO_OBRAS_JSON.exists():
        return

    try:
        data = json.loads(REGISTRO_OBRAS_JSON.read_text(encoding="utf-8"))
        if not isinstance(data, list):
            return

        rows = [_normalize_registro_obra_row(row, idx) for idx, row in enumerate(data, start=1)]
        _save_registro_obras(rows)
        for row in rows:
            _sync_cliente_from_registro_obra(row)
        db.session.commit()
        print(f"✅ Migrados {len(rows)} registros de obras desde JSON a base de datos.")
    except Exception as e:
        db.session.rollback()
        print("⚠️ ensure_schema(registro_obra.migracion_json):", e)


def _load_registro_obras_from_xlsx(file_bytes: bytes, default_responsable: str = "") -> list[dict]:
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as workbook_zip:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in workbook_zip.namelist():
            shared_root = ET.fromstring(workbook_zip.read("xl/sharedStrings.xml"))
            for item in shared_root.findall("a:si", XLSX_NS):
                shared_strings.append(
                    "".join((node.text or "") for node in item.findall(".//a:t", XLSX_NS)).strip()
                )

        workbook_root = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
        rels_root = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))
        rel_map = {
            rel.attrib.get("Id"): rel.attrib.get("Target", "")
            for rel in rels_root.findall("p:Relationship", XLSX_NS)
        }

        first_sheet_path = None
        for sheet in workbook_root.findall("a:sheets/a:sheet", XLSX_NS):
            rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            target = rel_map.get(rel_id or "", "")
            if target:
                first_sheet_path = f"xl/{target.lstrip('/')}"
                break

        if not first_sheet_path or first_sheet_path not in workbook_zip.namelist():
            return []

        sheet_root = ET.fromstring(workbook_zip.read(first_sheet_path))
        parsed_rows: list[list[str]] = []
        for row in sheet_root.findall("a:sheetData/a:row", XLSX_NS):
            values_by_col: dict[int, str] = {}
            for cell in row.findall("a:c", XLSX_NS):
                ref = cell.attrib.get("r", "")
                values_by_col[_excel_col_to_index(ref)] = _xlsx_cell_text(cell, shared_strings)
            if not values_by_col:
                continue
            max_col = max(values_by_col)
            parsed_rows.append([values_by_col.get(col, "").strip() for col in range(max_col + 1)])

    if not parsed_rows:
        return []

    header_aliases = {
        "numero": "numero",
        "n°": "numero",
        "no": "numero",
        "obra": "obra",
        "ubicacion": "ubicacion",
        "encargado": "encargado",
        "puesto": "puesto",
        "telefono": "telefono",
        "correo": "correo",
        "responsable": "responsable",
    }
    column_map: dict[str, int] = {}
    header_row_index = -1
    for idx, row in enumerate(parsed_rows):
        current_map: dict[str, int] = {}
        for col_idx, value in enumerate(row):
            key = _normalize_text_for_match(value).replace(".", "")
            if key in header_aliases:
                current_map[header_aliases[key]] = col_idx
        if "obra" in current_map:
            column_map = current_map
            header_row_index = idx
            break

    if header_row_index < 0:
        return []

    imported_rows: list[dict] = []
    for row in parsed_rows[header_row_index + 1:]:
        obra = _clean_registro_obra_excel_value(row[column_map["obra"]]) if "obra" in column_map and len(row) > column_map["obra"] else ""
        ubicacion = _clean_registro_obra_excel_value(row[column_map["ubicacion"]]) if "ubicacion" in column_map and len(row) > column_map["ubicacion"] else ""
        encargado = _clean_registro_obra_excel_value(row[column_map["encargado"]]) if "encargado" in column_map and len(row) > column_map["encargado"] else ""
        puesto = _clean_registro_obra_excel_value(row[column_map["puesto"]]) if "puesto" in column_map and len(row) > column_map["puesto"] else ""
        telefono = _normalize_registro_obra_phone(row[column_map["telefono"]]) if "telefono" in column_map and len(row) > column_map["telefono"] else ""
        correo = _clean_registro_obra_excel_value(row[column_map["correo"]]) if "correo" in column_map and len(row) > column_map["correo"] else ""
        responsable = _clean_registro_obra_excel_value(row[column_map["responsable"]]) if "responsable" in column_map and len(row) > column_map["responsable"] else ""
        if not any([obra, ubicacion, encargado, puesto, telefono, correo, responsable]):
            continue
        imported_rows.append(_normalize_registro_obra_row({
            "numero": "",
            "obra": obra,
            "ubicacion": ubicacion,
            "encargado": encargado,
            "puesto": puesto,
            "telefono": telefono,
            "correo": correo,
            "responsable": responsable or default_responsable,
        }, len(imported_rows) + 1))

    return imported_rows


def _mobile_json_error(message: str, status: int = 400):
    return jsonify({"ok": False, "error": message}), status


def _mobile_token_serializer() -> URLSafeTimedSerializer:
    return URLSafeTimedSerializer(current_app.secret_key, salt="registro-obras-mobile")


def _issue_mobile_token(user: Usuario) -> str:
    return _mobile_token_serializer().dumps({"user_id": user.id})


def _mobile_user_from_token() -> Optional[Usuario]:
    auth_header = (request.headers.get("Authorization") or "").strip()
    if not auth_header.lower().startswith("bearer "):
        return None
    token = auth_header[7:].strip()
    if not token:
        return None
    try:
        payload = _mobile_token_serializer().loads(token, max_age=60 * 60 * 24 * 30)
    except (BadSignature, SignatureExpired):
        return None
    user_id = payload.get("user_id")
    if not user_id:
        return None
    return Usuario.query.get(int(user_id))


def _mobile_user_is_admin(user: Usuario) -> bool:
    return ((getattr(user, "rol", "") or "").upper() == "ADMIN")


def _mobile_user_responsable(user: Usuario) -> str:
    return _usuario_nombre_representante(user)


def require_mobile_auth(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = _mobile_user_from_token()
        if not user:
            return _mobile_json_error("No autorizado.", 401)
        g.mobile_user = user
        return fn(*args, **kwargs)
    return wrapper


def _mobile_user_can_access_quote(user: Usuario, cot: Cotizacion) -> bool:
    if _mobile_user_is_admin(user):
        return True
    return (cot.responsable or "").strip().lower() == _mobile_user_responsable(user).lower()


def _mobile_quote_pdf_url(cot_id: int) -> str:
    url = url_for("api_mobile_quote_pdf", cot_id=cot_id, _external=True)
    if url.startswith("http://"):
        return "https://" + url[len("http://"):]
    return url


VOICE_NUMBER_WORDS = {
    "un": 1,
    "uno": 1,
    "una": 1,
    "primer": 1,
    "primero": 1,
    "dos": 2,
    "tres": 3,
    "cuatro": 4,
    "cinco": 5,
    "seis": 6,
    "siete": 7,
    "ocho": 8,
    "nueve": 9,
    "diez": 10,
    "once": 11,
    "doce": 12,
    "trece": 13,
    "catorce": 14,
    "quince": 15,
    "veinte": 20,
}
VOICE_STOPWORDS = {
    "cotiza", "cotizar", "cotizacion", "cotízame", "cotizame", "quiero", "necesito",
    "agrega", "agregar", "para", "del", "de", "la", "el", "los", "las", "un", "una",
    "por", "con", "color", "acabado", "cliente", "nombre", "favor", "favor,", "metros",
    "metro", "piezas", "pieza", "pza", "pz", "m2", "mt2", "x", "mas", "ademas",
    "cotiza:", "cotizar:", "precio", "unitario", "medida", "medidas", "ancho", "alto",
}


def _voice_normalize_text(value: str) -> str:
    raw = str(value or "").strip().lower()
    raw = unicodedata.normalize("NFKD", raw)
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
    raw = raw.replace("\n", " ")
    raw = re.sub(r"[^\w\s\.,x/-]", " ", raw)
    raw = re.sub(r"\s+", " ", raw)
    return raw.strip()


def _voice_parse_number(value, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        raw = str(value).strip()
        if not raw:
            return default
        raw = re.sub(r"[^\d,.\-]", "", raw)
        if not raw:
            return default
        if "," in raw and "." in raw:
            if raw.rfind(",") > raw.rfind("."):
                raw = raw.replace(".", "").replace(",", ".")
            else:
                raw = raw.replace(",", "")
        elif "," in raw:
            parts = raw.split(",")
            if len(parts) == 2 and len(parts[1]) == 3 and len(parts[0]) >= 1:
                raw = "".join(parts)
            else:
                raw = raw.replace(",", ".")
        elif "." in raw:
            parts = raw.split(".")
            if len(parts) == 2 and len(parts[1]) == 3 and len(parts[0]) >= 1:
                raw = "".join(parts)
        return float(raw)
    except Exception:
        return default


def _voice_parse_number_word(value: str) -> Optional[float]:
    token = _voice_normalize_text(value).strip()
    if token in VOICE_NUMBER_WORDS:
        return float(VOICE_NUMBER_WORDS[token])
    return None


def _voice_extract_client(command_text: str) -> str:
    patterns = [
        r"(?:cliente|para cliente|a nombre de)\s+([a-z0-9áéíóúñ .-]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, command_text, flags=re.IGNORECASE)
        if match:
            client = re.split(
                r"\b(?:"
                r"concepto|descripcion|descripción|cantidad|cant(?:idad)?|precio(?:\s+unitario)?|"
                r"empresa|razon\s+social|razón\s+social|correo|telefono|teléfono|celular|movil|móvil|"
                r"responsable|contacto|atencion|atención|ciudad|municipio|"
                r"sistema|color|acabado|condicion|condición"
                r")\b",
                match.group(1),
                maxsplit=1,
                flags=re.IGNORECASE,
            )[0]
            client = re.split(r"\b(?:con|medida|medidas)\b", client, maxsplit=1, flags=re.IGNORECASE)[0]
            return _voice_clean_field(client).title()
    return ""


def _voice_clean_field(value: str) -> str:
    cleaned = re.sub(r"\s+", " ", str(value or "").strip())
    return cleaned.strip(" ,.-")


def _voice_extract_company(command_text: str) -> str:
    patterns = [
        r"(?:empresa|razon social|razón social)\s+([a-z0-9áéíóúñ.&,\-/ ]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, command_text, flags=re.IGNORECASE)
        if match:
            value = re.split(
                r"\b(?:correo|telefono|teléfono|ciudad|responsable|concepto|cantidad|precio)\b",
                match.group(1),
                maxsplit=1,
                flags=re.IGNORECASE,
            )[0]
            return _voice_clean_field(value).title()
    return ""


def _voice_extract_email(command_text: str) -> str:
    match = re.search(r"\b([A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,})\b", command_text, flags=re.IGNORECASE)
    return match.group(1).strip() if match else ""


def _voice_extract_phone(command_text: str) -> str:
    explicit = re.search(
        r"(?:telefono|teléfono|celular|movil|móvil|whatsapp)\s+([0-9\-\+\(\)\s]{8,})",
        command_text,
        flags=re.IGNORECASE,
    )
    candidate = explicit.group(1) if explicit else ""
    if not candidate:
        any_phone = re.search(r"\b(?:\+?\d[\d\-\(\)\s]{8,}\d)\b", command_text)
        candidate = any_phone.group(0) if any_phone else ""
    digits = re.sub(r"\D", "", candidate)
    if len(digits) < 8:
        return ""
    return digits


def _voice_extract_city(command_text: str) -> str:
    patterns = [
        r"(?:ciudad|municipio)\s+([a-z0-9áéíóúñ.\- ]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, command_text, flags=re.IGNORECASE)
        if match:
            value = re.split(
                r"\b(?:correo|telefono|teléfono|responsable|concepto|cantidad|precio)\b",
                match.group(1),
                maxsplit=1,
                flags=re.IGNORECASE,
            )[0]
            return _voice_clean_field(value).title()
    return ""


def _voice_extract_contact_responsible(command_text: str) -> str:
    patterns = [
        r"(?:responsable|contacto|atencion|atención)\s+([a-z0-9áéíóúñ.\- ]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, command_text, flags=re.IGNORECASE)
        if match:
            value = re.split(
                r"\b(?:correo|telefono|teléfono|ciudad|concepto|cantidad|precio)\b",
                match.group(1),
                maxsplit=1,
                flags=re.IGNORECASE,
            )[0]
            return _voice_clean_field(value).title()
    return ""


def _voice_strip_client_phrase(command_text: str) -> str:
    cleaned = str(command_text or "")
    cleaned = re.sub(
        r"(?i)\b(?:para cliente|cliente|a nombre de)\s+([a-z0-9áéíóúñ .-]+?)"
        r"(?=\b(?:concepto|descripcion|descripción|cantidad|precio|empresa|correo|telefono|teléfono|"
        r"responsable|ciudad|sistema|color|acabado|condicion|condición)\b|$)",
        " ",
        cleaned,
    )
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip(" ,.-")


def _voice_extract_dimensions(command_text: str) -> tuple[Optional[float], Optional[float]]:
    match = re.search(
        r"(\d+(?:[\.,]\d+)?)\s*(?:m|mt|mts|metros)?\s*(?:x|por)\s*(\d+(?:[\.,]\d+)?)\s*(?:m|mt|mts|metros)?",
        command_text,
        flags=re.IGNORECASE,
    )
    if not match:
        return None, None
    width = _voice_parse_number(match.group(1))
    height = _voice_parse_number(match.group(2))
    if width <= 0 or height <= 0:
        return None, None
    return width, height


def _voice_extract_quantity(command_text: str) -> Optional[float]:
    match = re.search(r"^\s*(\d+(?:[\.,]\d+)?)\s*(?:piezas?|pieza|pzas?|pz)?\b", command_text)
    if match:
        quantity = _voice_parse_number(match.group(1), 1.0)
        return quantity if quantity > 0 else None
    first_word = command_text.split(" ", 1)[0].strip()
    word_quantity = _voice_parse_number_word(first_word)
    if word_quantity is not None:
        return word_quantity
    match = re.search(r"\b(\d+(?:[\.,]\d+)?)\s*(?:piezas?|pieza|pzas?|pz)\b", command_text)
    if match:
        quantity = _voice_parse_number(match.group(1), 1.0)
        return quantity if quantity > 0 else None
    match = re.search(
        r"\bcantidad(?:\s+de)?\s+(\d+(?:[\.,]\d+)?)\b",
        command_text,
        flags=re.IGNORECASE,
    )
    if match:
        quantity = _voice_parse_number(match.group(1), 1.0)
        return quantity if quantity > 0 else None
    match = re.search(
        r"\bcantidad(?:\s+de)?\s+(un|uno|una|dos|tres|cuatro|cinco|seis|siete|ocho|nueve|diez|once|doce|trece|catorce|quince|veinte)\b",
        command_text,
        flags=re.IGNORECASE,
    )
    if match:
        quantity = _voice_parse_number_word(match.group(1))
        return quantity if quantity and quantity > 0 else None
    match = re.search(
        r"\b(\d+(?:[\.,]\d+)?)\s*(?:hectareas?|hectáreas?|ha|m2|mt2|metros?\s+cuadrados?|ml|metros?\s+lineales?)\b",
        command_text,
        flags=re.IGNORECASE,
    )
    if match:
        quantity = _voice_parse_number(match.group(1), 1.0)
        return quantity if quantity > 0 else None
    match = re.search(
        r"\b(un|uno|una|dos|tres|cuatro|cinco|seis|siete|ocho|nueve|diez|once|doce|trece|catorce|quince|veinte)\s*"
        r"(?:hectareas?|hectáreas?|ha|m2|mt2|metros?\s+cuadrados?|ml|metros?\s+lineales?)\b",
        command_text,
        flags=re.IGNORECASE,
    )
    if match:
        quantity = _voice_parse_number_word(match.group(1))
        return quantity if quantity and quantity > 0 else None
    return None


def _voice_extract_price(command_text: str) -> Optional[float]:
    patterns = [
        r"(?:precio(?: unitario)?|a|en)\s+\$?\s*(\d+(?:[\.,]\d+)?)",
        r"(?:cada\s+uno|cada\s+una|c/u|cu)\s+\$?\s*(\d+(?:[\.,]\d+)?)",
        r"(\d+(?:[\.,]\d+)?)\s*(?:pesos|mxn)\b",
        r"\$+\s*(\d+(?:[\.,]\d+)?)",
    ]
    for pattern in patterns:
        match = re.search(pattern, command_text, flags=re.IGNORECASE)
        if match:
            price = _voice_parse_number(match.group(1))
            if price > 0:
                return price
    return None


def _voice_extract_unit(command_text: str) -> str:
    if re.search(r"\bhectareas?\b|\bhectáreas?\b|\bha\b", command_text):
        return "ha"
    if re.search(r"\bm2\b|\bmetro(?:s)? cuadrados?\b|\bmt2\b", command_text):
        return "m2"
    if re.search(r"\bmetro(?:s)? lineales?\b|\bml\b", command_text):
        return "ml"
    if re.search(r"\bpiezas?\b|\bpzas?\b|\bpz\b", command_text):
        return "pza"
    return ""


def _voice_extract_system(command_text: str) -> str:
    patterns = [
        r"(?:sistema)\s+([a-z0-9áéíóúñ.\-/ ]+)",
        r"\b(sfrm|cementicio|intumescente|monokote|cafco|promat|vermiculita)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, command_text, flags=re.IGNORECASE)
        if match:
            value = match.group(1)
            value = re.split(r"\b(?:precio|cantidad|descripcion|descripción|correo|telefono|teléfono|direccion|dirección|ciudad)\b", value, maxsplit=1)[0]
            return _voice_clean_field(value).title()
    return ""


def _voice_extract_color_or_finish(command_text: str) -> str:
    patterns = [
        r"\bcolor\s+([a-z0-9\s-]+)",
        r"\bacabado\s+([a-z0-9\s-]+)",
        r"\ben\s+(blanco|negro|gris|natural|bronce|mate|brillante)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, command_text, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip(" ,.-").title()
    return ""


def _voice_build_search_text(command_text: str, client_name: str) -> str:
    text = command_text
    for piece in [client_name, "cliente", "para cliente", "a nombre de"]:
        if piece:
            text = re.sub(re.escape(piece), " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\d+(?:[\.,]\d+)?", " ", text)
    text = re.sub(
        r"\b(?:x|por|color|acabado|precio|medida|medidas|de|del|la|el|los|las|con|en|incluye|incluyan|incluido|incluida|pesos|mxn|cada|uno|una)\b",
        " ",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _voice_match_concept(search_text: str) -> Optional[Concepto]:
    tokens = [
        token for token in _voice_normalize_text(search_text).split()
        if len(token) > 2 and token not in VOICE_STOPWORDS and not token.isdigit()
    ]
    if not tokens:
        return None

    best_score = 0
    best_concept = None
    for concept in Concepto.query.all():
        name = _voice_normalize_text(concept.nombre_concepto or "")
        if not name:
            continue
        score = sum(4 for token in tokens if token in name)
        if tokens and all(token in name for token in tokens):
            score += 3
        if score > best_score:
            best_score = score
            best_concept = concept
    return best_concept if best_score >= 4 else None


def _voice_split_segments(command_text: str) -> list[str]:
    base = _voice_strip_client_phrase(command_text)
    base = re.sub(r"\b(?:en otro concepto|otro concepto|siguiente concepto)\b", " | ", base, flags=re.IGNORECASE)
    base = re.sub(r"\s+(?:ademas|más|mas)\s+", " | ", base, flags=re.IGNORECASE)
    base = re.sub(
        r"\s+y\s+(?=(?:\d+(?:[\.,]\d+)?|un|uno|una|dos|tres|cuatro|cinco|seis|siete|ocho|nueve|diez)\b)",
        " | ",
        base,
        flags=re.IGNORECASE,
    )
    base = base.replace(";", " | ").replace(",", " | ")
    items = [part.strip(" ,.-") for part in base.split("|") if part.strip(" ,.-")]
    return items or [command_text.strip()]


def _voice_build_item_payload(segment_raw: str, client_name: str, index: int) -> dict:
    segment_text = _voice_normalize_text(segment_raw)
    width, height = _voice_extract_dimensions(segment_text)
    quantity = _voice_extract_quantity(segment_text)
    explicit_price = _voice_extract_price(segment_text)
    explicit_unit = _voice_extract_unit(segment_text)
    explicit_system = _voice_extract_system(segment_text)
    finish = _voice_extract_color_or_finish(segment_text)
    search_text = _voice_build_search_text(segment_text, client_name)
    concept = _voice_match_concept(search_text)

    concept_name = (segment_raw or "").strip()
    if not concept_name:
        concept_name = (search_text or "").strip() or f"Concepto por voz {index}"
    unit = explicit_unit or (getattr(concept, "unidad", "") or "").strip() or ""
    unit_price = explicit_price if explicit_price is not None else float(getattr(concept, "precio_unitario", 0) or 0)
    if explicit_price is None and unit_price <= 0:
        unit_price = 0.0
    system = explicit_system or (getattr(concept, "sistema", "") or "").strip()
    area = fmt(width * height) if width and height else 0.0
    effective_quantity = quantity
    if area > 0 and quantity is not None and unit.lower() in {"m2", "mt2", "metro cuadrado", "metros cuadrados"}:
        effective_quantity = fmt(quantity * area)
    subtotal = fmt((effective_quantity or 0) * unit_price) if effective_quantity is not None and unit_price > 0 else 0.0

    warnings = []
    if not concept:
        warnings.append(f"Partida {index}: no encontré un concepto exacto en catálogo; usaré el texto dictado.")
    if quantity is None:
        warnings.append(f"Partida {index}: la cantidad no se detectó y quedó en blanco.")
    if not unit:
        warnings.append(f"Partida {index}: la unidad no se detectó y quedó en blanco.")
    if unit_price <= 0:
        warnings.append(f"Partida {index}: el precio unitario quedó en 0.")

    description_parts = []
    if area > 0:
        description_parts.append(
            f"Medidas detectadas: {fmt(width)} x {fmt(height)} m ({area} m2 por pieza)."
        )
    if finish:
        description_parts.append(f"Acabado/color: {finish}.")

    return {
        "id": concept.id if concept else None,
        "nombre": concept_name,
        "unidad": unit,
        "cantidad": fmt(effective_quantity) if effective_quantity is not None else "",
        "cantidad_capturada": fmt(quantity) if quantity is not None else "",
        "precio_unitario": fmt(unit_price) if unit_price > 0 else "",
        "sistema": system,
        "subtotal": subtotal if subtotal > 0 else "",
        "ancho": fmt(width) if width else 0.0,
        "alto": fmt(height) if height else 0.0,
        "area_por_pieza": area,
        "acabado": finish,
        "descripcion": "\n".join(part for part in description_parts if part).strip(),
        "origen_segmento": segment_raw.strip(),
        "warnings": warnings,
    }


def _voice_log_command(user: Usuario, preview: dict, status: str, cotizacion: Optional[Cotizacion] = None) -> None:
    try:
        row = VoiceCommandLog(
            usuario_id=getattr(user, "id", None),
            cotizacion_id=getattr(cotizacion, "id", None) if cotizacion else None,
            cliente=(preview.get("cliente") or "").strip() or None,
            comando_raw=(preview.get("comando_original") or "").strip(),
            comando_normalizado=_voice_normalize_text(preview.get("comando_original") or ""),
            preview_json=json.dumps(preview, ensure_ascii=False),
            status=status,
        )
        db.session.add(row)
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass


def _voice_parse_conditions(conditions_raw: str) -> list[str]:
    raw = str(conditions_raw or "").strip()
    if not raw:
        return []
    normalized = re.sub(
        r"\b(?:otra condicion es que|otra condición es que|condicion es que|condición es que)\b",
        "|",
        raw,
        flags=re.IGNORECASE,
    )
    parts = [part.strip(" ,.-") for part in normalized.split("|") if part.strip(" ,.-")]
    return parts


def _voice_transcribe_audio_bytes(audio_bytes: bytes, filename: str = "voz.m4a", mime_type: str = "audio/mp4") -> str:
    if not OPENAI_API_KEY:
        raise RuntimeError("Falta configurar OPENAI_API_KEY en el servidor.")
    if not audio_bytes:
        raise ValueError("El audio llegó vacío.")
    if len(audio_bytes) > 25 * 1024 * 1024:
        raise ValueError("El audio supera el límite de 25 MB.")

    safe_name = Path(filename or "voz.m4a").name or "voz.m4a"
    guessed_type = mime_type or mimetypes.guess_type(safe_name)[0] or "application/octet-stream"
    response = requests.post(
        "https://api.openai.com/v1/audio/transcriptions",
        headers={
            "Authorization": f"Bearer {OPENAI_API_KEY}",
        },
        data={
            "model": OPENAI_TRANSCRIBE_MODEL,
            "language": "es",
            "prompt": VOICE_TRANSCRIPTION_PROMPT,
        },
        files={
            "file": (safe_name, audio_bytes, guessed_type),
        },
        timeout=120,
    )
    if response.status_code >= 400:
        detail = ""
        try:
            detail = response.json().get("error", {}).get("message", "")
        except Exception:
            detail = response.text[:300]
        raise RuntimeError(detail or f"OpenAI devolvió HTTP {response.status_code}.")
    payload = response.json()
    transcript = str(payload.get("text") or "").strip()
    if not transcript:
        raise RuntimeError("La transcripción llegó vacía.")
    return transcript


VOICE_GUIDED_HEADER_LABELS = [
    ("cliente", "cliente"),
    ("empresa", "empresa"),
    ("esa", "empresa"),
    ("correo", "correo"),
    ("telefono", "telefono"),
    ("teléfono", "telefono"),
    ("ciudad", "ciudad"),
]

VOICE_GUIDED_ITEM_LABELS = [
    ("concepto", "concepto"),
    ("otro concepto", "concepto"),
    ("unidad", "unidad"),
    ("cantidad", "cantidad"),
    ("precio", "precio"),
    ("sistema", "sistema"),
    ("tema", "sistema"),
]


def _voice_is_guided_script(command_raw: str) -> bool:
    text = str(command_raw or "")
    labels = ["CLIENTE", "EMPRESA", "CORREO", "CONCEPTO", "OTRO CONCEPTO", "CANTIDAD", "PRECIO"]
    matches = sum(1 for label in labels if re.search(rf"\b{re.escape(label)}\s*:?", text, flags=re.IGNORECASE))
    return matches >= 4


def _voice_extract_labeled_sections(
    text: str,
    labels: list[tuple[str, str]],
    stop_labels: Optional[list[tuple[str, str]]] = None,
) -> list[tuple[str, str]]:
    if not text:
        return []
    combined = []
    seen = set()
    for label, key in (labels + (stop_labels or [])):
        norm = (label.lower(), key)
        if norm in seen:
            continue
        seen.add(norm)
        combined.append((label, key))
    combined.sort(key=lambda item: len(item[0]), reverse=True)
    pattern = "|".join(
        rf"(?P<label_{idx}>{re.escape(label)})\s*:?"
        for idx, (label, _) in enumerate(combined)
    )
    matches = list(re.finditer(pattern, text, flags=re.IGNORECASE))
    sections: list[tuple[str, str]] = []
    for idx, match in enumerate(matches):
        canonical = None
        source_key = None
        for group_idx, (_, key) in enumerate(combined):
            if match.group(f"label_{group_idx}"):
                canonical = key
                source_key = group_idx
                break
        if canonical is None or source_key is None:
            continue
        matched_label = combined[source_key][0].lower()
        allowed = any(matched_label == label.lower() for label, _ in labels)
        if not allowed:
            continue
        start = match.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(text)
        value = text[start:end].strip(" \r\n\t:-")
        sections.append((canonical, value))
    return sections


def _voice_parse_guided_quantity(value: str) -> Optional[float]:
    raw = _voice_clean_field(value)
    if not raw:
        return None
    number_match = re.search(r"\d+(?:[\.,]\d+)?", raw)
    if number_match:
        quantity = _voice_parse_number(number_match.group(0), 0.0)
        return quantity if quantity > 0 else None
    for token in _voice_normalize_text(raw).split():
        parsed = _voice_parse_number_word(token)
        if parsed is not None and parsed > 0:
            return parsed
    return None


def _voice_parse_guided_price(value: str) -> Optional[float]:
    raw = _voice_clean_field(value)
    if not raw:
        return None
    number_match = re.search(r"\d+(?:[\.,]\d+)?", raw)
    if not number_match:
        return None
    price = _voice_parse_number(number_match.group(0), 0.0)
    return price if price > 0 else None


def _voice_parse_guided_email(value: str) -> str:
    raw = _voice_clean_field(value).lower()
    if not raw:
        return ""
    raw = re.split(
        r"\b(?:telefono|teléfono|responsable|ciudad|concepto|otro concepto|unidad|cantidad|precio|sistema|descripcion|descripción)\b",
        raw,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0]
    normalized = f" {raw} "
    replacements = {
        " arroba ": "@",
        " arrova ": "@",
        " punto com ": ".com",
        " punto mx ": ".mx",
        " punto net ": ".net",
        " punto org ": ".org",
        " punto ": ".",
        " guion bajo ": "_",
        " guion medio ": "-",
    }
    for source, target in replacements.items():
        normalized = normalized.replace(source, target)
    normalized = normalized.replace(" ", "").strip()
    normalized = normalized.strip(".,;:-")
    match = re.search(r"[^@\s]+@[^@\s]+\.[^@\s]+", normalized)
    if match:
        return match.group(0)
    return ""


def _voice_strip_guided_label_echo(value: str, field_name: str) -> str:
    raw = _voice_clean_field(value)
    if not raw:
        return ""
    aliases = {
        "empresa": ["empresa", "esa"],
        "correo": ["correo"],
        "telefono": ["telefono", "teléfono"],
        "ciudad": ["ciudad"],
        "concepto": ["concepto", "otro concepto"],
        "unidad": ["unidad"],
        "cantidad": ["cantidad"],
        "precio": ["precio"],
        "sistema": ["sistema", "tema"],
    }
    changed = True
    while changed:
        changed = False
        for alias in aliases.get(field_name, [field_name]):
            updated = re.sub(rf"^\s*{re.escape(alias)}\s*:?\s*", "", raw, flags=re.IGNORECASE)
            updated = _voice_clean_field(updated)
            if updated != raw:
                raw = updated
                changed = True
    return raw


def _voice_split_guided_unit_and_quantity(unit_raw: str) -> tuple[str, Optional[float]]:
    raw = _voice_strip_guided_label_echo(unit_raw, "unidad")
    if not raw:
        return "", None
    quantity = None
    number_match = re.search(r"(\d+(?:[\.,]\d+)?)", raw)
    if number_match:
        quantity = _voice_parse_number(number_match.group(1), 0.0)
        quantity = quantity if quantity > 0 else None
        raw = re.sub(r"\d+(?:[\.,]\d+)?", " ", raw)
        raw = re.sub(r"\s+", " ", raw).strip(" ,.-")
    normalized = _voice_normalize_text(raw)
    if normalized in {"metro lineal", "metros lineales", "lineal", "lineales", "ml"}:
        raw = "ml"
    elif normalized in {"metro cuadrado", "metros cuadrados", "m2", "mt2"}:
        raw = "m2"
    elif normalized in {"hectarea", "hectareas", "hectárea", "hectáreas", "ha"}:
        raw = "ha"
    elif normalized in {"pieza", "piezas", "pza", "pz"}:
        raw = "pza"
    return raw, quantity


def _voice_split_guided_system_and_tail(system_raw: str) -> tuple[str, str]:
    raw = _voice_clean_field(system_raw)
    if not raw:
        return "", ""
    brand_aliases = {
        "comex": "Comex",
        "ppg": "PPG",
        "sherwin": "Sherwin",
        "sika": "Sika",
        "promat": "Promat",
        "cafco": "Cafco",
        "monokote": "Monokote",
        "nullifire": "Nullifire",
        "international": "International",
    }
    for alias, canonical in brand_aliases.items():
        match = re.match(rf"^{re.escape(alias)}\b[\s:,-]*(.*)$", raw, flags=re.IGNORECASE)
        if match:
            return canonical, _voice_clean_field(match.group(1))
    return raw, ""


def _voice_rescue_unlabeled_first_item(command_raw: str, city_value: str) -> tuple[str, dict[str, str]]:
    text = str(command_raw or "")
    if not text:
        return city_value, {}
    match = re.search(r"\bciudad\b\s*:?\s*(.+)$", text, flags=re.IGNORECASE)
    if not match:
        return city_value, {}
    tail = _voice_clean_field(match.group(1))
    if not tail:
        return city_value, {}

    next_field = re.search(
        r"\b(?:otro concepto|concepto|cantidad|precio|sistema|tema|descripcion|descripción)\b\s*:?",
        tail,
        flags=re.IGNORECASE,
    )
    prelude = tail[:next_field.start()] if next_field else tail
    prelude = _voice_strip_guided_label_echo(prelude, "ciudad")
    current_city = _voice_strip_guided_label_echo(city_value, "ciudad")
    if current_city:
        city_pattern = rf"^\s*{re.escape(current_city)}\b"
        stripped = re.sub(city_pattern, "", prelude, count=1, flags=re.IGNORECASE)
        if stripped != prelude:
            prelude = _voice_clean_field(stripped)
    tokens = prelude.split()
    rescued_city = current_city
    if not rescued_city and tokens:
        rescued_city = tokens[0].title()
        prelude = _voice_clean_field(" ".join(tokens[1:]))
    elif rescued_city and len(tokens) > 1 and _voice_normalize_text(prelude).startswith(_voice_normalize_text(rescued_city)):
        stripped = prelude[len(rescued_city):]
        prelude = _voice_clean_field(stripped)

    if not prelude:
        return rescued_city, {}

    concept_raw = prelude
    unit_raw = ""
    unit_match = re.search(
        r"\b(metro cuadrado|metros cuadrados|m2|mt2|metro lineal|metros lineales|lineal|lineales|ml|pieza|piezas|pza|pz|hectarea|hectareas|hectárea|hectáreas|ha)\b(?:\s+\1\b)?\s*$",
        prelude,
        flags=re.IGNORECASE,
    )
    if unit_match:
        unit_raw = _voice_clean_field(unit_match.group(0))
        concept_raw = _voice_clean_field(prelude[:unit_match.start()])

    item = {}
    if concept_raw:
        item["concepto"] = concept_raw
    if unit_raw:
        item["unidad"] = unit_raw
    return rescued_city, item


def _voice_build_guided_item_payload(item_fields: dict, index: int) -> dict:
    concept_name = _voice_strip_guided_label_echo(item_fields.get("concepto") or "", "concepto")
    unit, quantity_from_unit = _voice_split_guided_unit_and_quantity(item_fields.get("unidad") or "")
    quantity_value = _voice_parse_guided_quantity(_voice_strip_guided_label_echo(item_fields.get("cantidad") or "", "cantidad"))
    price_value = _voice_parse_guided_price(_voice_strip_guided_label_echo(item_fields.get("precio") or "", "precio"))
    system, system_tail = _voice_split_guided_system_and_tail(_voice_strip_guided_label_echo(item_fields.get("sistema") or "", "sistema"))
    description = _voice_strip_guided_label_echo(item_fields.get("descripcion") or "", "descripcion")
    if quantity_value is None and quantity_from_unit is not None:
        quantity_value = quantity_from_unit
    if system_tail:
        if description:
            description = f"{system_tail}. {description}"
        else:
            description = system_tail
    if not concept_name and description:
        concept_name = description
        description = ""
    if not concept_name and system:
        concept_name = system
        system = ""
    subtotal = fmt((quantity_value or 0) * (price_value or 0)) if quantity_value and price_value else 0.0

    warnings = []
    if not concept_name:
        warnings.append(f"Partida {index}: el concepto no se detectó y quedó en blanco.")
    if not unit:
        warnings.append(f"Partida {index}: la unidad no se detectó y quedó en blanco.")
    if quantity_value is None:
        warnings.append(f"Partida {index}: la cantidad no se detectó y quedó en blanco.")
    if price_value is None:
        warnings.append(f"Partida {index}: el precio unitario no se detectó y quedó en blanco.")

    return {
        "id": None,
        "nombre": concept_name,
        "unidad": unit,
        "cantidad": fmt(quantity_value) if quantity_value is not None else "",
        "cantidad_capturada": fmt(quantity_value) if quantity_value is not None else "",
        "precio_unitario": fmt(price_value) if price_value is not None else "",
        "sistema": system,
        "subtotal": subtotal if subtotal > 0 else "",
        "ancho": 0.0,
        "alto": 0.0,
        "area_por_pieza": 0.0,
        "acabado": "",
        "descripcion": description,
        "origen_segmento": concept_name,
        "warnings": warnings,
    }


def _voice_parse_guided_script(command_raw: str) -> dict:
    text = str(command_raw or "").replace("\r", "\n")
    header_sections = _voice_extract_labeled_sections(
        text,
        VOICE_GUIDED_HEADER_LABELS,
        stop_labels=VOICE_GUIDED_ITEM_LABELS,
    )
    header_data = {key: _voice_clean_field(value) for key, value in header_sections}
    combined_labels = sorted(
        {label for label, _ in (VOICE_GUIDED_HEADER_LABELS + VOICE_GUIDED_ITEM_LABELS)},
        key=len,
        reverse=True,
    )
    if combined_labels:
        first_match = re.search(
            "|".join(rf"\b{re.escape(label)}\b\s*:?" for label in combined_labels),
            text,
            flags=re.IGNORECASE,
        )
        if first_match:
            prefix = _voice_clean_field(text[:first_match.start()])
            if prefix and not header_data.get("cliente"):
                header_data["cliente"] = prefix

    item_sections = _voice_extract_labeled_sections(
        text,
        VOICE_GUIDED_ITEM_LABELS,
        stop_labels=VOICE_GUIDED_HEADER_LABELS,
    )
    items = []
    current_item: dict[str, str] = {}
    rescued_city, rescued_item = _voice_rescue_unlabeled_first_item(text, header_data.get("ciudad", ""))
    if rescued_city:
        header_data["ciudad"] = rescued_city
    if rescued_item and (not item_sections or item_sections[0][0] != "concepto"):
        current_item.update(rescued_item)
    for key, value in item_sections:
        if key == "concepto":
            if current_item.get("concepto"):
                items.append(_voice_build_guided_item_payload(current_item, len(items) + 1))
            current_item = {"concepto": value}
        else:
            if not current_item:
                continue
            current_item[key] = value
    if current_item.get("concepto"):
        items.append(_voice_build_guided_item_payload(current_item, len(items) + 1))

    return {
        "cliente": _voice_strip_guided_label_echo(header_data.get("cliente", ""), "cliente"),
        "empresa": _voice_strip_guided_label_echo(header_data.get("empresa", ""), "empresa"),
        "correo": _voice_parse_guided_email(header_data.get("correo", "")),
        "telefono": _voice_strip_guided_label_echo(header_data.get("telefono", ""), "telefono"),
        "responsable_contacto": "",
        "ciudad": _voice_strip_guided_label_echo(header_data.get("ciudad", ""), "ciudad"),
        "items": items,
    }


def _voice_preview_payload_for_mobile(
    command_raw: str,
    user: Usuario,
    client_override: str = "",
    notes: str = "",
    conditions_raw: str = "",
) -> dict:
    command_text = _voice_normalize_text(command_raw)
    if not command_text:
        raise ValueError("No se recibió ningún comando de voz.")

    if _voice_is_guided_script(command_raw):
        guided = _voice_parse_guided_script(command_raw)
        client_name = (client_override or "").strip() or guided.get("cliente", "")
        company = guided.get("empresa", "")
        email = guided.get("correo", "")
        phone = guided.get("telefono", "")
        address = ""
        city = guided.get("ciudad", "")
        contact_responsible = guided.get("responsable_contacto", "")
        items = guided.get("items", [])
    else:
        client_name = (client_override or "").strip() or _voice_extract_client(command_text)
        company = _voice_extract_company(command_text)
        email = _voice_extract_email(command_text)
        phone = _voice_extract_phone(command_text)
        address = ""
        city = _voice_extract_city(command_text)
        contact_responsible = _voice_extract_contact_responsible(command_text)
        segments = _voice_split_segments(command_raw)
        items = [_voice_build_item_payload(segment, client_name, idx) for idx, segment in enumerate(segments, start=1)]
    conditions = _voice_parse_conditions(conditions_raw)
    subtotal = fmt(sum(float(item.get("subtotal") or 0) for item in items))
    iva = fmt(subtotal * 0.16)
    total = fmt(subtotal + iva)
    warnings = []
    if not client_name:
        warnings.append("No se detectó el cliente. Puedes escribirlo antes de guardar.")
    if len(items) > 1:
        warnings.append(f"Se detectaron {len(items)} partidas dentro del mismo comando.")
    for item in items:
        warnings.extend(item.get("warnings") or [])

    preview = {
        "cliente": client_name,
        "responsable": _mobile_user_responsable(user),
        "datos_encabezado": {
            "cliente": client_name,
            "empresa": company,
            "correo": email,
            "telefono": phone,
            "proyecto": "",
            "responsable_contacto": contact_responsible,
            "direccion": address,
            "ciudad": city,
        },
        "items": items,
        "concepto_detectado": items[0] if items else {},
        "resumen": {
            "partidas": len(items),
            "subtotal": subtotal,
            "iva": iva,
            "total": total,
        },
        "condiciones": conditions,
        "condiciones_raw": (conditions_raw or "").strip(),
        "notas": (notes or "").strip(),
        "comando_original": command_raw.strip(),
        "warnings": warnings,
    }
    _voice_log_command(user, preview, status="PREVIEW")
    return preview


def _create_mobile_voice_quote(preview: dict, user: Usuario) -> Cotizacion:
    cliente_nombre = (preview.get("cliente") or "").strip()
    responsible = _mobile_user_responsable(user)
    header_data = preview.get("datos_encabezado") or {}
    cliente = None
    if cliente_nombre:
        query = Cliente.query.filter(db.func.lower(Cliente.nombre_cliente) == cliente_nombre.lower())
        if not _mobile_user_is_admin(user):
            query = query.filter(Cliente.responsable == responsible)
        cliente = query.first()
        if not cliente:
            cliente = Cliente(
                nombre_cliente=cliente_nombre,
                empresa=(header_data.get("empresa") or "").strip() or None,
                responsable=responsible,
                correo=(header_data.get("correo") or "").strip() or None,
                telefono=(header_data.get("telefono") or "").strip() or None,
                direccion=(header_data.get("direccion") or "").strip() or None,
            )
            db.session.add(cliente)
            db.session.flush()
        else:
            if (header_data.get("empresa") or "").strip():
                cliente.empresa = (header_data.get("empresa") or "").strip()
            if (header_data.get("correo") or "").strip():
                cliente.correo = (header_data.get("correo") or "").strip()
            if (header_data.get("telefono") or "").strip():
                cliente.telefono = (header_data.get("telefono") or "").strip()
            if (header_data.get("direccion") or "").strip():
                cliente.direccion = (header_data.get("direccion") or "").strip()

    notes_parts = []
    if preview.get("notas"):
        notes_parts.append(str(preview["notas"]).strip())
    for condition in preview.get("condiciones") or []:
        notes_parts.append(str(condition).strip())
    if (header_data.get("responsable_contacto") or "").strip():
        notes_parts.append(f"Responsable contacto: {(header_data.get('responsable_contacto') or '').strip()}")

    subtotal = fmt(sum(fmt(item.get("subtotal")) for item in (preview.get("items") or [])))
    iva = fmt(subtotal * 0.16)
    total = fmt(subtotal + iva)

    cot = Cotizacion(
        folio=generar_folio(),
        fecha=now_cdmx_naive(),
        cliente_id=cliente.id if cliente else None,
        estatus="PENDIENTE",
        estatus_aprobacion="EN REVISIÓN",
        notas="\n".join(part for part in notes_parts if part).strip() or None,
        responsable=responsible,
        proyecto=(header_data.get("proyecto") or preview.get("proyecto") or "").strip() or None,
        ciudad_trabajo=(header_data.get("ciudad") or "").strip().upper() or None,
    )
    cot.subtotal = subtotal
    cot.descuento_total = 0.0
    cot.iva_porc = 16.0
    cot.iva_monto = iva
    cot.total = total
    db.session.add(cot)
    db.session.flush()

    for item in preview.get("items") or []:
        unit_price = fmt(item.get("precio_unitario"))
        quantity = fmt(item.get("cantidad"))
        item_subtotal = fmt(item.get("subtotal"))
        concept_name = (item.get("nombre") or "Concepto por voz").strip()
        concept = Concepto.query.filter(db.func.lower(Concepto.nombre_concepto) == concept_name.lower()).first()
        if not concept:
            concept = Concepto(
                nombre_concepto=concept_name,
                unidad=(item.get("unidad") or "").strip() or None,
                precio_unitario=unit_price,
                sistema=(item.get("sistema") or "").strip() or None,
                descripcion=(item.get("descripcion") or "").strip() or None,
            )
            db.session.add(concept)
            db.session.flush()

        det = CotizacionDetalle(**_safe_detalle_kwargs(
            cotizacion_id=cot.id,
            concepto_id=concept.id if concept else None,
            nombre_concepto=concept_name,
            unidad=(item.get("unidad") or "").strip(),
            cantidad=quantity,
            precio_unitario=unit_price,
            sistema=(item.get("sistema") or "").strip() or None,
            descripcion=(item.get("descripcion") or "").strip(),
            subtotal=item_subtotal,
            origen="voz",
        ))
        db.session.add(det)
    db.session.commit()
    _voice_log_command(user, preview, status="CREATED", cotizacion=cot)
    _send_quote_created_notification(cot)
    _send_quote_review_email_safely(cot)
    return cot


def _firebase_is_configured() -> bool:
    return PUSH_NOTIFICATIONS_ENABLED and firebase_admin is not None and bool(FIREBASE_CREDENTIALS_FILE or FIREBASE_CREDENTIALS_JSON)


def _get_firebase_app():
    if not _firebase_is_configured():
        return None
    try:
        return firebase_admin.get_app()
    except Exception:
        pass

    try:
        if FIREBASE_CREDENTIALS_JSON:
            cred = firebase_credentials.Certificate(json.loads(FIREBASE_CREDENTIALS_JSON))
        elif FIREBASE_CREDENTIALS_FILE:
            cred = firebase_credentials.Certificate(FIREBASE_CREDENTIALS_FILE)
        else:
            return None
        return firebase_admin.initialize_app(cred)
    except Exception as exc:
        logger.warning("Firebase no se pudo inicializar: %s", exc)
        return None


def _upsert_mobile_device(user: Usuario, token: str, plataforma: str = "android", device_name: str = "", app_version: str = "") -> MobileDevice:
    existing = MobileDevice.query.filter_by(token=token).first()
    if existing:
        existing.usuario_id = user.id
        existing.plataforma = (plataforma or "android").strip().lower()[:30]
        existing.device_name = (device_name or "").strip()[:120]
        existing.app_version = (app_version or "").strip()[:40]
        existing.is_active = True
        existing.last_seen_at = now_cdmx_naive()
        db.session.add(existing)
        db.session.commit()
        return existing

    device = MobileDevice(
        usuario_id=user.id,
        token=token,
        plataforma=(plataforma or "android").strip().lower()[:30] or "android",
        device_name=(device_name or "").strip()[:120],
        app_version=(app_version or "").strip()[:40],
        is_active=True,
        last_seen_at=now_cdmx_naive(),
    )
    db.session.add(device)
    db.session.commit()
    return device


def _deactivate_mobile_device(token: str) -> None:
    if not token:
        return
    device = MobileDevice.query.filter_by(token=token).first()
    if not device:
        return
    device.is_active = False
    device.updated_at = now_cdmx_naive()
    db.session.add(device)
    db.session.commit()


def _mobile_push_tokens_for_users(user_ids: list[int]) -> list[str]:
    if not user_ids:
        return []
    rows = (
        MobileDevice.query
        .filter(MobileDevice.usuario_id.in_(user_ids), MobileDevice.is_active.is_(True))
        .all()
    )
    unique_tokens: list[str] = []
    seen: set[str] = set()
    for row in rows:
        token = (row.token or "").strip()
        if not token or token in seen:
            continue
        seen.add(token)
        unique_tokens.append(token)
    return unique_tokens


def _mobile_all_active_push_tokens() -> list[str]:
    rows = MobileDevice.query.filter(MobileDevice.is_active.is_(True)).all()
    unique_tokens: list[str] = []
    seen: set[str] = set()
    for row in rows:
        token = (row.token or "").strip()
        if not token or token in seen:
            continue
        seen.add(token)
        unique_tokens.append(token)
    return unique_tokens


def _mobile_push_user_ids_for_approval_reviewer() -> list[int]:
    review_emails = {email.lower() for email in _parse_email_list(COTIZACION_REVIEW_EMAIL)}
    review_emails.add("hjaramillo@poliutech.com")
    review_emails.add("mescalera@poliutech.com")
    hansel_aliases = {"hansel", "hansel alejandro", "hansel angel", "hansel ángel"}
    mescalera_aliases = {"mescalera", "mesacalera"}
    fixed_reviewer_ids = {18}
    users = Usuario.query.all()
    user_ids: set[int] = set()
    for user in users:
        user_name = (getattr(user, "nombre", "") or "").strip().lower()
        visible_name = (_mobile_user_responsable(user) or "").strip().lower()
        raw_visible_name = (getattr(user, "nombre_visible", "") or "").strip().lower()
        user_email = (getattr(user, "correo", "") or "").strip().lower()
        identity_parts = {user_name, visible_name, raw_visible_name, user_email}
        if (
            user.id in fixed_reviewer_ids
            or any(part in hansel_aliases or part.startswith("hansel ") for part in identity_parts if part)
            or any(part in mescalera_aliases or part.startswith("mescalera ") or part.startswith("mesacalera ") for part in identity_parts if part)
            or user_email in review_emails
        ):
            if user.id:
                user_ids.add(user.id)
    result = list(user_ids)
    if not result:
        logger.warning("Push aprobación: no se encontró usuario revisor Hansel/Mescalera ni correo %s.", sorted(review_emails))
    return result


def _mobile_push_user_ids_for_quote_owner(cot: Cotizacion) -> list[int]:
    user_ids: set[int] = set()
    responsable = (cot.responsable or "").strip().lower()
    if responsable:
        owner = Usuario.query.filter(
            or_(
                db.func.lower(Usuario.nombre) == responsable,
                db.func.lower(db.func.coalesce(Usuario.nombre_visible, "")) == responsable,
            )
        ).first()
        if owner and owner.id:
            user_ids.add(owner.id)
        else:
            users = Usuario.query.all()
            for user in users:
                first_name = _mobile_user_responsable(user).strip().lower()
                if first_name and first_name == responsable and user.id:
                    user_ids.add(user.id)
    return list(user_ids)


def _mobile_push_user_ids_for_aazcona() -> list[int]:
    aliases = {"aazcona", "azcona"}
    target_email = COTIZACION_REVIEW_RESULT_AAZCONA_EMAIL.lower()
    user_ids: set[int] = set()
    for user in Usuario.query.all():
        user_name = (getattr(user, "nombre", "") or "").strip().lower()
        visible_name = (_mobile_user_responsable(user) or "").strip().lower()
        raw_visible_name = (getattr(user, "nombre_visible", "") or "").strip().lower()
        user_email = (getattr(user, "correo", "") or "").strip().lower()
        identity_parts = {user_name, visible_name, raw_visible_name}
        if (
            user_email == target_email
            or any(part in aliases or part.startswith("aazcona ") or part.startswith("azcona ") for part in identity_parts if part)
        ):
            if user.id:
                user_ids.add(user.id)
    if not user_ids:
        logger.warning("Push resultado cotizacion: no se encontro usuario Aazcona con correo %s.", target_email)
    return list(user_ids)


def _send_quote_status_push(cot: Cotizacion, previous_status: str, new_status: str) -> dict[str, int]:
    if (new_status or "").strip().upper() == "FINALIZADA":
        return {"sent": 0, "failed": 0}
    pdf_url = _mobile_quote_pdf_url(cot.id)
    owner_ids = _mobile_push_user_ids_for_quote_owner(cot)
    tokens = _mobile_push_tokens_for_users(owner_ids)
    return _send_push_notification(
        tokens,
        title=f"Cotización {new_status}",
        body=f"{cot.folio or 'Sin folio'} · {money(cot.total)}",
        data={
            "type": "quote_status",
            "cotizacion_id": str(cot.id or ""),
            "folio": str(cot.folio or ""),
            "previous_status": str(previous_status or ""),
            "estatus": str(new_status or ""),
            "pdf_url": pdf_url,
            "target_user_id": str(owner_ids[0]) if len(owner_ids) == 1 else "",
        },
    )


def _send_quote_review_result_push(cot: Cotizacion, selected_status: str, reason: str = "") -> dict[str, int]:
    normalized = (selected_status or "").strip().upper()
    target_ids = list(dict.fromkeys([
        *_mobile_push_user_ids_for_quote_owner(cot),
        *_mobile_push_user_ids_for_aazcona(),
    ]))
    tokens = _mobile_push_tokens_for_users(target_ids)
    if normalized in {"APROBADO", "APROBADA", "AUTORIZADO"}:
        title = "Cotización aprobada"
        body = f"{cot.folio or 'Sin folio'} fue aprobada."
    elif normalized in {"RECHAZADO", "RECHAZADA"}:
        title = "Cotización rechazada"
        reason_text = " ".join((reason or "").split())
        body = f"{cot.folio or 'Sin folio'} fue rechazada."
        if reason_text:
            body = f"{body} Motivo: {reason_text}"
    else:
        title = "Cotización en revisión"
        body = f"{cot.folio or 'Sin folio'} quedó en revisión."
    return _send_push_notification(
        tokens,
        title=title,
        body=body,
        data={
            "type": "quote_review_result",
            "cotizacion_id": str(cot.id or ""),
            "folio": str(cot.folio or ""),
            "estatus": normalized,
            "reason": str(reason or ""),
            "pdf_url": _mobile_quote_pdf_url(cot.id),
            "target_user_id": str(target_ids[0]) if len(target_ids) == 1 else "",
        },
    )


def _send_quote_updated_push(cot: Cotizacion) -> dict[str, int]:
    target_ids = list(dict.fromkeys([
        *_mobile_push_user_ids_for_quote_owner(cot),
        *_mobile_push_user_ids_for_approval_reviewer(),
        *_mobile_push_user_ids_for_aazcona(),
    ]))
    tokens = _mobile_push_tokens_for_users(target_ids)
    if not tokens:
        logger.warning("Push edición %s: destinatarios configurados sin token móvil activo.", cot.folio or cot.id)
        tokens = _mobile_all_active_push_tokens()
    approve_url = url_for("cotizacion_revision_decidir", cot_id=cot.id, action="approve", token=_quote_review_token(cot, "approve"), _external=True)
    reject_url = url_for("cotizacion_revision_decidir", cot_id=cot.id, action="reject", token=_quote_review_token(cot, "reject"), _external=True)
    return _send_push_notification(
        tokens,
        title="Cotización editada pendiente de aprobación",
        body=f"{cot.folio or 'Sin folio'} · {money(cot.total)} · Aprobar o rechazar",
        data={
            "type": "quote_pending_approval",
            "cotizacion_id": str(cot.id or ""),
            "folio": str(cot.folio or ""),
            "estatus": str(cot.estatus or ""),
            "estatus_aprobacion": str(cot.estatus_aprobacion or ""),
            "pdf_url": _mobile_quote_pdf_url(cot.id),
            "approve_url": approve_url,
            "reject_url": reject_url,
            "target_user": "Hansel",
            "target_user_name": "Hansel",
            "recipient_user_name": "Hansel",
            "approval_reviewer": "Hansel",
            "requires_decision": "true",
            "source": "quote_updated",
            "target_user_id": str(target_ids[0]) if len(target_ids) == 1 else "",
        },
    )


def _send_quote_approval_request_push(cot: Cotizacion) -> dict[str, int]:
    reviewer_ids = _mobile_push_user_ids_for_approval_reviewer()
    hansel_ids = [18]
    tokens = _mobile_push_tokens_for_users(reviewer_ids)
    using_active_device_fallback = False
    if not reviewer_ids:
        logger.warning("Push aprobación %s: no hay usuario Hansel/Mescalera/revisor configurado.", cot.folio or cot.id)
    if not tokens:
        logger.warning("Push aprobación %s: Hansel/Mescalera/revisor %s no tiene token móvil activo.", cot.folio or cot.id, reviewer_ids)
        tokens = _mobile_all_active_push_tokens()
        using_active_device_fallback = bool(tokens)
        if using_active_device_fallback:
            logger.warning(
                "Push aprobación %s: usando respaldo a todos los dispositivos activos (%s tokens).",
                cot.folio or cot.id,
                len(tokens),
            )
    result = _send_push_notification(
        tokens,
        title="Cotización pendiente de aprobación",
        body=f"{cot.folio or 'Sin folio'} · {money(cot.total)}",
        data={
            "type": "quote_pending_approval",
            "cotizacion_id": str(cot.id or ""),
            "folio": str(cot.folio or ""),
            "estatus": str(cot.estatus_aprobacion or "EN REVISIÓN"),
            "pdf_url": _mobile_quote_pdf_url(cot.id),
            "target_user": "Hansel/Mescalera",
            "target_user_name": "Hansel/Mescalera",
            "recipient_user_name": "Hansel/Mescalera",
            "approval_reviewer": "Hansel/Mescalera",
            "requires_decision": "true",
        },
    )
    logger.info(
        "Push aprobación %s: hansel_ids=%s reviewers=%s tokens=%s fallback=%s sent=%s failed=%s",
        cot.folio or cot.id,
        hansel_ids,
        reviewer_ids,
        len(tokens),
        using_active_device_fallback,
        result.get("sent", 0),
        result.get("failed", 0),
    )
    return result


def _send_quote_created_notification(cot: Cotizacion) -> None:
    estatus_actual = (cot.estatus or "").strip().upper()
    aprobacion_actual = (cot.estatus_aprobacion or "EN REVISIÓN").strip().upper()
    try:
        msg = (
            "🧾 *Nueva Cotización Creada*\\n"
            f"Folio: *{cot.folio or 'Sin folio'}*\\n"
            f"Estatus seguimiento: *{estatus_actual or 'SIN ESTATUS'}*\\n"
            f"Estatus aprobación: *{aprobacion_actual}*\\n"
            f"Fecha (CDMX): {cot.fecha.strftime('%d/%m/%Y %H:%M') if cot.fecha else ''}\\n"
            f"Total: {money(cot.total)}"
        )
        send_whatsapp_multi(ADMIN_LIST, msg)
    except Exception as exc:
        logger.warning("WhatsApp de creación falló: %s", exc)

    try:
        _send_quote_approval_request_push(cot)
    except Exception as exc:
        logger.warning("Push de aprobación pendiente falló: %s", exc)


def _send_quote_followup_push(cot: Cotizacion, seg: CotizacionSeguimiento) -> dict[str, int]:
    tokens = _mobile_push_tokens_for_users(_mobile_push_user_ids_for_quote_owner(cot))
    preview = " ".join((seg.comentario or "").split())
    if len(preview) > 120:
        preview = preview[:117] + "..."
    return _send_push_notification(
        tokens,
        title=f"Nuevo seguimiento · {cot.folio or 'Sin folio'}",
        body=preview or f"{seg.autor} agregó un seguimiento.",
        data={
            "type": "quote_followup",
            "cotizacion_id": str(cot.id or ""),
            "seguimiento_id": str(seg.id or ""),
            "folio": str(cot.folio or ""),
            "estatus": str(cot.estatus or ""),
            "autor": str(seg.autor or ""),
            "pdf_url": _mobile_quote_pdf_url(cot.id),
        },
    )


def _send_daily_status_reminder(cot: Cotizacion, ahora: datetime) -> None:
    estatus_actual = (cot.estatus or "").strip().upper()
    if not estatus_actual or estatus_actual == "FINALIZADA":
        return

    body = (
        "🔔 *Recordatorio diario de cotización*\\n"
        f"Folio: *{cot.folio or 'Sin folio'}*\\n"
        f"Estatus: *{estatus_actual}*\\n"
        f"Fecha (CDMX): {cot.fecha.strftime('%d/%m/%Y %H:%M') if cot.fecha else ''}\\n"
        f"Total: {money(cot.total)}"
    )
    send_whatsapp_multi(ADMIN_LIST, body)
    _send_quote_status_push(cot, estatus_actual, estatus_actual)
    cot.last_whatsapp_at = ahora
    db.session.commit()


def _send_push_notification(tokens: list[str], title: str, body: str, data: Optional[dict[str, str]] = None) -> dict[str, int]:
    if not tokens:
        return {"sent": 0, "failed": 0}
    app_instance = _get_firebase_app()
    if app_instance is None or firebase_messaging is None:
        logger.warning("Push no enviado: Firebase no está configurado o firebase_messaging no está disponible.")
        return {"sent": 0, "failed": len(tokens)}

    sent = 0
    failed = 0
    payload_data = {str(k): str(v) for k, v in (data or {}).items()}
    payload_data["title"] = str(title)
    payload_data["body"] = str(body)
    for token in tokens:
        try:
            message = firebase_messaging.Message(
                token=token,
                data=payload_data,
                android=firebase_messaging.AndroidConfig(priority="high"),
            )
            firebase_messaging.send(message, app=app_instance)
            sent += 1
        except Exception as exc:
            failed += 1
            logger.warning("Push fallido para token móvil: %s", exc)
            err = str(exc).lower()
            if any(fragment in err for fragment in ["registration-token", "not registered", "invalid argument", "requested entity was not found"]):
                _deactivate_mobile_device(token)
    return {"sent": sent, "failed": failed}


def _send_push_notification_debug(tokens: list[str], title: str, body: str, data: Optional[dict[str, str]] = None) -> dict:
    if not tokens:
        return {"sent": 0, "failed": 0, "errors": []}
    app_instance = _get_firebase_app()
    if app_instance is None or firebase_messaging is None:
        return {
            "sent": 0,
            "failed": len(tokens),
            "errors": [{"error": "Firebase no está configurado o firebase_messaging no está disponible."}],
        }

    sent = 0
    failed = 0
    errors = []
    payload_data = {str(k): str(v) for k, v in (data or {}).items()}
    payload_data["title"] = str(title)
    payload_data["body"] = str(body)
    for token in tokens:
        try:
            message = firebase_messaging.Message(
                token=token,
                data=payload_data,
                android=firebase_messaging.AndroidConfig(priority="high"),
            )
            firebase_messaging.send(message, app=app_instance)
            sent += 1
        except Exception as exc:
            failed += 1
            err = str(exc)
            errors.append({
                "token_prefix": token[:18],
                "error_type": exc.__class__.__name__,
                "error": err,
            })
            if any(fragment in err.lower() for fragment in ["registration-token", "not registered", "invalid argument", "requested entity was not found"]):
                _deactivate_mobile_device(token)
    return {"sent": sent, "failed": failed, "errors": errors}


def _filter_registro_obras_for_mobile(rows: list[dict], user: Usuario, obra: str = "", responsable: str = "") -> list[dict]:
    obra_filter = (obra or "").strip().lower()
    responsable_filter = (responsable or "").strip().lower()
    out = []
    for row in rows:
        obra_value = (row.get("obra") or "").strip().lower()
        responsable_value = (row.get("responsable") or "").strip().lower()
        if obra_filter and obra_filter not in obra_value:
            continue
        if responsable_filter and responsable_filter not in responsable_value:
            continue
        if not _mobile_user_is_admin(user):
            if responsable_value != _mobile_user_responsable(user).lower():
                continue
        out.append(row)
    return out


def _provider_filters_from_request() -> dict[str, str]:
    return {
        "razon_social_poliutech": (request.args.get("razon_social_poliutech") or "").strip().lower(),
        "relacion": (request.args.get("relacion") or "").strip().upper(),
    }


def _provider_row_matches_filters(row: dict, filters: dict[str, str]) -> bool:
    for field, needle in filters.items():
        value = str(row.get(field, "")).strip()
        if field == "relacion":
            if needle and value.upper() != needle:
                return False
        elif needle and needle not in value.lower():
            return False
    return True


def _filter_provider_rows(rows: list[dict], filters: dict[str, str]) -> list[dict]:
    return [row for row in rows if _provider_row_matches_filters(row, filters)]


def _normalize_prospecto_status(value: object) -> str:
    status = str(value or "").strip().upper()
    return status if status in PROSPECT_STATUS_OPTIONS else "PENDIENTE"


def _prospecto_to_row(item: "Prospecto", idx: Optional[int] = None) -> dict:
    position = idx if idx is not None else (item.id or 0)
    return {
        "id": item.id,
        "numero": position,
        "titulo": (item.titulo or "").strip(),
        "descripcion": (item.descripcion or "").strip(),
        "contacto": (item.contacto or "").strip(),
        "telefono": (item.telefono or "").strip(),
        "correo": (item.correo or "").strip(),
        "status": _normalize_prospecto_status(item.status),
        "responsable": (item.responsable or "").strip(),
        "seguimiento_count": len(item.seguimientos or []),
    }


def _load_prospectos() -> list[dict]:
    query = Prospecto.query.order_by(Prospecto.id.desc())
    items = query.all()
    return [_prospecto_to_row(item, idx) for idx, item in enumerate(items, start=1)]


def _prospectos_filters_from_request() -> dict[str, str]:
    status_raw = (request.args.get("status") or "").strip()
    return {
        "titulo": (request.args.get("titulo") or "").strip().lower(),
        "status": _normalize_prospecto_status(status_raw) if status_raw else "",
        "contacto": (request.args.get("contacto") or "").strip().lower(),
    }


def _prospecto_matches_filters(row: dict, filters: dict[str, str]) -> bool:
    titulo = filters.get("titulo") or ""
    status = filters.get("status") or ""
    contacto = filters.get("contacto") or ""

    if titulo and titulo not in str(row.get("titulo", "")).strip().lower():
        return False
    if status and status != _normalize_prospecto_status(row.get("status")):
        return False
    if contacto and contacto not in str(row.get("contacto", "")).strip().lower():
        return False
    return True


def _filter_prospectos(rows: list[dict], filters: dict[str, str]) -> list[dict]:
    return [row for row in rows if _prospecto_matches_filters(row, filters)]


def _normalize_ticket_status(value: object) -> str:
    status = str(value or "").strip().upper()
    return status if status in TICKET_STATUS_OPTIONS else "NUEVO"


def _normalize_ticket_priority(value: object) -> str:
    priority = str(value or "").strip().upper()
    return priority if priority in TICKET_PRIORITY_OPTIONS else "MEDIA"


def _normalize_ticket_category(value: object) -> str:
    category = str(value or "").strip().upper()
    return category if category in TICKET_CATEGORY_OPTIONS else "GENERAL"


def _ticket_is_closed(status: str) -> bool:
    return _normalize_ticket_status(status) in {"RESUELTO", "CERRADO"}


def _ticket_upload_root(ticket_id: int) -> Path:
    base = Path(app.static_folder or "static") / "uploads" / "tickets" / str(ticket_id)
    base.mkdir(parents=True, exist_ok=True)
    return base


def _ticket_file_is_allowed(filename: str) -> bool:
    ext = Path(filename or "").suffix.lower()
    return bool(ext and ext in TICKET_ALLOWED_EXTENSIONS)


def _ticket_public_url(adjunto: "TicketAdjunto") -> str:
    return url_for("static", filename=(adjunto.ruta_relativa or "").replace("\\", "/"))


def _save_ticket_attachments(ticket: "TicketSoporte", files, comentario: Optional["TicketComentario"] = None) -> int:
    saved = 0
    for uploaded in list(files or [])[:TICKET_MAX_ATTACHMENTS]:
        original = (uploaded.filename or "").strip()
        if not original:
            continue
        if not _ticket_file_is_allowed(original):
            raise ValueError(f"El archivo '{original}' no tiene un formato permitido.")

        safe_original = secure_filename(original) or "adjunto"
        ext = Path(safe_original).suffix.lower()
        unique_name = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}_{saved + 1}{ext}"
        dest_dir = _ticket_upload_root(ticket.id)
        dest_path = dest_dir / unique_name
        uploaded.save(dest_path)

        rel_path = Path("uploads") / "tickets" / str(ticket.id) / unique_name
        size = dest_path.stat().st_size if dest_path.exists() else 0
        db.session.add(TicketAdjunto(
            ticket_id=ticket.id,
            comentario_id=getattr(comentario, "id", None),
            usuario_id=getattr(current_user, "id", None),
            nombre_original=original[:260],
            nombre_archivo=unique_name,
            ruta_relativa=str(rel_path).replace("\\", "/"),
            mime_type=uploaded.mimetype or mimetypes.guess_type(original)[0],
            tamano_bytes=int(size or 0),
        ))
        saved += 1
    return saved


def _ticket_to_row(ticket: "TicketSoporte") -> dict:
    return {
        "id": ticket.id,
        "folio": ticket.folio or f"TCK-{ticket.id:06d}",
        "asunto": ticket.asunto or "",
        "solicitante": ticket.solicitante or "",
        "empresa": ticket.empresa or "",
        "categoria": _normalize_ticket_category(ticket.categoria),
        "prioridad": _normalize_ticket_priority(ticket.prioridad),
        "estado": _normalize_ticket_status(ticket.estado),
        "responsable": ticket.responsable or "",
        "creado_en": ticket.creado_en,
        "actualizado_en": ticket.actualizado_en,
        "comentarios_count": len(ticket.comentarios or []),
        "adjuntos_count": len(ticket.adjuntos or []),
    }


def _ticket_filters_from_request() -> dict[str, str]:
    estado_raw = (request.args.get("estado") or "").strip()
    prioridad_raw = (request.args.get("prioridad") or "").strip()
    return {
        "q": (request.args.get("q") or "").strip().lower(),
        "estado": _normalize_ticket_status(estado_raw) if estado_raw else "",
        "prioridad": _normalize_ticket_priority(prioridad_raw) if prioridad_raw else "",
        "responsable": (request.args.get("responsable") or "").strip().lower(),
    }


def _tickets_base_query():
    query = TicketSoporte.query
    if not is_admin():
        ra = responsable_actual()
        query = query.filter(or_(TicketSoporte.responsable == ra, TicketSoporte.creado_por_id == getattr(current_user, "id", None)))
    return query


def _load_ticket_rows(filters: dict[str, str]) -> list[dict]:
    query = _tickets_base_query().order_by(TicketSoporte.creado_en.desc())
    if filters.get("estado"):
        query = query.filter(TicketSoporte.estado == filters["estado"])
    if filters.get("prioridad"):
        query = query.filter(TicketSoporte.prioridad == filters["prioridad"])
    if filters.get("responsable"):
        query = query.filter(db.func.lower(db.func.coalesce(TicketSoporte.responsable, "")).contains(filters["responsable"]))
    if filters.get("q"):
        q = f"%{filters['q']}%"
        query = query.filter(or_(
            db.func.lower(db.func.coalesce(TicketSoporte.folio, "")).like(q),
            db.func.lower(db.func.coalesce(TicketSoporte.asunto, "")).like(q),
            db.func.lower(db.func.coalesce(TicketSoporte.solicitante, "")).like(q),
        ))
    return [_ticket_to_row(ticket) for ticket in query.limit(300).all()]


def require_ticket_owner_or_admin(ticket: "TicketSoporte") -> None:
    if is_admin():
        return
    ra = responsable_actual()
    current_user_id = getattr(current_user, "id", None)
    if current_user_id and ticket.creado_por_id == current_user_id:
        return
    if ra and (ticket.responsable or "").strip().lower() == ra.strip().lower():
        return
    abort(403)


def _support_ticket_email_html(ticket: "TicketSoporte", detail_url: str) -> str:
    folio = escape(ticket.folio or f"TCK-{ticket.id:06d}")
    asunto = escape(ticket.asunto or "")
    descripcion = escape(ticket.descripcion or "")
    solicitante = escape(ticket.solicitante or "")
    correo = escape(ticket.correo or "Sin correo")
    categoria = escape(ticket.categoria or "GENERAL")
    prioridad = escape(ticket.prioridad or "MEDIA")
    responsable = escape(ticket.responsable or "Sin asignar")
    button_style = (
        "display:inline-block;min-width:150px;text-align:center;padding:14px 22px;"
        "border-radius:8px;text-decoration:none;font-weight:700;font-size:15px;"
        "background:#0C3C78;color:#ffffff;border:1px solid #0C3C78;"
    )
    return f"""
    <html>
      <body style="margin:0;padding:0;background:#eef2f7;font-family:Arial,Helvetica,sans-serif;color:#1f2937;">
        <div style="max-width:760px;margin:0 auto;padding:30px 16px;">
          <div style="background:#ffffff;border:1px solid #d9e2ec;border-radius:10px;overflow:hidden;box-shadow:0 8px 24px rgba(15,45,80,.08);">
            <div style="background:#0C3C78;color:#ffffff;padding:22px 26px;">
              <div style="font-size:12px;font-weight:700;letter-spacing:.9px;text-transform:uppercase;opacity:.9;">MAR · Soporte</div>
              <div style="font-size:23px;font-weight:800;margin-top:5px;">Nuevo ticket de soporte</div>
              <div style="font-size:14px;opacity:.92;margin-top:6px;">{folio}</div>
            </div>
            <div style="padding:26px;">
              <table style="border-collapse:collapse;width:100%;background:#ffffff;margin-bottom:20px;">
                <tr><td style="padding:10px 12px;border:1px solid #dde3ea;background:#f8fafc;width:34%;font-weight:700;color:#64748b;">Asunto</td><td style="padding:10px 12px;border:1px solid #dde3ea;font-weight:700;">{asunto}</td></tr>
                <tr><td style="padding:10px 12px;border:1px solid #dde3ea;background:#f8fafc;font-weight:700;color:#64748b;">Solicitante</td><td style="padding:10px 12px;border:1px solid #dde3ea;">{solicitante}</td></tr>
                <tr><td style="padding:10px 12px;border:1px solid #dde3ea;background:#f8fafc;font-weight:700;color:#64748b;">Correo</td><td style="padding:10px 12px;border:1px solid #dde3ea;">{correo}</td></tr>
                <tr><td style="padding:10px 12px;border:1px solid #dde3ea;background:#f8fafc;font-weight:700;color:#64748b;">Categoria</td><td style="padding:10px 12px;border:1px solid #dde3ea;">{categoria}</td></tr>
                <tr><td style="padding:10px 12px;border:1px solid #dde3ea;background:#f8fafc;font-weight:700;color:#64748b;">Prioridad</td><td style="padding:10px 12px;border:1px solid #dde3ea;">{prioridad}</td></tr>
                <tr><td style="padding:10px 12px;border:1px solid #dde3ea;background:#f8fafc;font-weight:700;color:#64748b;">Responsable</td><td style="padding:10px 12px;border:1px solid #dde3ea;">{responsable}</td></tr>
              </table>
              <div style="background:#f8fafc;border:1px solid #dbe4ef;border-radius:10px;padding:16px 18px;margin-bottom:22px;">
                <div style="font-size:12px;text-transform:uppercase;letter-spacing:.7px;color:#64748b;font-weight:800;margin-bottom:8px;">Descripcion del problema</div>
                <div style="white-space:pre-wrap;line-height:1.5;">{descripcion}</div>
              </div>
              <a href="{detail_url}" style="{button_style}">Ver Ticket</a>
            </div>
          </div>
        </div>
      </body>
    </html>
    """.strip()


def _send_support_ticket_email(ticket: "TicketSoporte") -> None:
    recipients = _parse_email_list(SUPPORT_TICKET_EMAIL)
    if not recipients:
        raise ValueError("No hay correo configurado para soporte.")
    detail_url = url_for("soporte_ticket_detalle", ticket_id=ticket.id, _external=True)
    msg = EmailMessage()
    msg["Subject"] = f"Nuevo ticket de soporte {ticket.folio or ticket.id}"
    msg["From"] = f"SISTEMA MAR DE TICKETS <{SMTP_FROM or SMTP_USERNAME}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content(
        f"Nuevo ticket de soporte {ticket.folio or ticket.id}\n"
        f"Asunto: {ticket.asunto or ''}\n"
        f"Solicitante: {ticket.solicitante or ''}\n"
        f"Correo: {ticket.correo or ''}\n"
        f"Categoria: {ticket.categoria or ''}\n"
        f"Prioridad: {ticket.prioridad or ''}\n"
        f"Responsable: {ticket.responsable or ''}\n\n"
        f"Descripcion:\n{ticket.descripcion or ''}\n\n"
        f"Ver ticket: {detail_url}\n"
    )
    msg.add_alternative(_support_ticket_email_html(ticket, detail_url), subtype="html")

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg, to_addrs=recipients)


def _build_simple_xls(sheet_name: str, headers: list[str], rows: list[list[str]]) -> bytes:
    def html_cell(value: object) -> str:
        text = "" if value is None else str(value)
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        return escape(text).replace("\n", "<br>")

    parts = [
        "<html>",
        "<head>",
        '<meta http-equiv="Content-Type" content="text/html; charset=utf-8" />',
        f"<title>{escape(sheet_name)}</title>",
        "</head>",
        "<body>",
        "<table border='1'>",
        "<thead><tr>",
    ]
    for header in headers:
        parts.append(f"<th>{html_cell(header)}</th>")
    parts.append("</tr></thead><tbody>")
    for row in rows:
        parts.append("<tr>")
        for cell in row:
            parts.append(f"<td>{html_cell(cell)}</td>")
        parts.append("</tr>")
    parts.append("</tbody></table></body></html>")
    return "".join(parts).encode("utf-8")


def _normalize_import_header(value: object) -> str:
    raw = str(value or "").strip().lower()
    normalized = unicodedata.normalize("NFKD", raw)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return normalized


def _load_prospectos_from_xlsx(file_bytes: bytes) -> list[dict]:
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as workbook_zip:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in workbook_zip.namelist():
            shared_root = ET.fromstring(workbook_zip.read("xl/sharedStrings.xml"))
            for item in shared_root.findall("a:si", XLSX_NS):
                shared_strings.append(
                    "".join((node.text or "") for node in item.findall(".//a:t", XLSX_NS)).strip()
                )

        workbook_root = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
        rels_root = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))
        rel_map = {
            rel.attrib.get("Id"): rel.attrib.get("Target", "")
            for rel in rels_root.findall("p:Relationship", XLSX_NS)
        }

        target_sheet = None
        for sheet in workbook_root.findall("a:sheets/a:sheet", XLSX_NS):
            rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if rel_id:
                target = rel_map.get(rel_id, "")
                if target:
                    target_sheet = f"xl/{target.lstrip('/')}"
                    break

        if not target_sheet or target_sheet not in workbook_zip.namelist():
            return []

        sheet_root = ET.fromstring(workbook_zip.read(target_sheet))
        parsed_rows: list[list[str]] = []
        for row in sheet_root.findall("a:sheetData/a:row", XLSX_NS):
            values_by_col: dict[int, str] = {}
            for cell in row.findall("a:c", XLSX_NS):
                ref = cell.attrib.get("r", "")
                values_by_col[_excel_col_to_index(ref)] = _xlsx_cell_text(cell, shared_strings)
            if not values_by_col:
                continue
            max_col = max(values_by_col)
            parsed_rows.append([values_by_col.get(col, "").strip() for col in range(max_col + 1)])

        if len(parsed_rows) < 2:
            return []

        header_map = {
            _normalize_import_header(name): idx
            for idx, name in enumerate(parsed_rows[0])
            if str(name or "").strip()
        }

        required = ["titulo", "descripcion", "contacto", "telefono", "correo"]
        if any(col not in header_map for col in required):
            raise ValueError("El Excel debe incluir las columnas: Título, Descripción, Contacto, Teléfono y Correo.")

        rows: list[dict] = []
        for row in parsed_rows[1:]:
            titulo = row[header_map["titulo"]].strip() if len(row) > header_map["titulo"] else ""
            descripcion = row[header_map["descripcion"]].strip() if len(row) > header_map["descripcion"] else ""
            contacto = row[header_map["contacto"]].strip() if len(row) > header_map["contacto"] else ""
            telefono = row[header_map["telefono"]].strip() if len(row) > header_map["telefono"] else ""
            correo = row[header_map["correo"]].strip() if len(row) > header_map["correo"] else ""
            if not any([titulo, descripcion, contacto, telefono, correo]):
                continue
            rows.append({
                "titulo": titulo,
                "descripcion": descripcion,
                "contacto": contacto,
                "telefono": telefono,
                "correo": correo,
                "status": "PENDIENTE",
            })
        return rows


def _build_simple_xlsx(sheet_name: str, headers: list[str], rows: list[list[str]], column_widths: Optional[list[int]] = None) -> bytes:
    def cell_ref(row_idx: int, col_idx: int) -> str:
        label = ""
        num = col_idx
        while num > 0:
            num, rem = divmod(num - 1, 26)
            label = chr(65 + rem) + label
        return f"{label}{row_idx}"

    def inline_cell(row_idx: int, col_idx: int, value: object) -> str:
        text = escape("" if value is None else str(value))
        return (
            f'<c r="{cell_ref(row_idx, col_idx)}" t="inlineStr">'
            f"<is><t>{text}</t></is></c>"
        )

    cols_xml = ""
    if column_widths:
        cols_parts = []
        for idx, width in enumerate(column_widths, start=1):
            cols_parts.append(
                f'<col min="{idx}" max="{idx}" width="{width}" customWidth="1"/>'
            )
        cols_xml = f"<cols>{''.join(cols_parts)}</cols>"

    sheet_rows = []
    header_cells = "".join(inline_cell(1, idx, value) for idx, value in enumerate(headers, start=1))
    sheet_rows.append(f'<row r="1" s="1">{header_cells}</row>')

    for row_idx, row in enumerate(rows, start=2):
        cells = "".join(inline_cell(row_idx, col_idx, value) for col_idx, value in enumerate(row, start=1))
        sheet_rows.append(f'<row r="{row_idx}">{cells}</row>')

    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"{cols_xml}"
        f"<sheetData>{''.join(sheet_rows)}</sheetData>"
        "</worksheet>"
    )

    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheets><sheet name="{escape(sheet_name)}" sheetId="1" r:id="rId1"/></sheets>'
        "</workbook>"
    )

    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
        'Target="styles.xml"/>'
        "</Relationships>"
    )

    root_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/>'
        "</Relationships>"
    )

    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/styles.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        "</Types>"
    )

    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="2">'
        '<font><sz val="11"/><name val="Calibri"/></font>'
        '<font><b/><sz val="11"/><name val="Calibri"/></font>'
        '</fonts>'
        '<fills count="2">'
        '<fill><patternFill patternType="none"/></fill>'
        '<fill><patternFill patternType="gray125"/></fill>'
        '</fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="2">'
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
        '<xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyFont="1"/>'
        '</cellXfs>'
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        "</styleSheet>"
    )

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types_xml)
        zf.writestr("_rels/.rels", root_rels_xml)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)
        zf.writestr("xl/styles.xml", styles_xml)
    return output.getvalue()


def _build_matrix_xlsx(sheet_name: str, rows: list[list[str]], column_widths: Optional[list[int]] = None) -> bytes:
    def cell_ref(row_idx: int, col_idx: int) -> str:
        label = ""
        num = col_idx
        while num > 0:
            num, rem = divmod(num - 1, 26)
            label = chr(65 + rem) + label
        return f"{label}{row_idx}"

    def inline_cell(row_idx: int, col_idx: int, value: object) -> str:
        text = escape("" if value is None else str(value))
        return f'<c r="{cell_ref(row_idx, col_idx)}" t="inlineStr"><is><t>{text}</t></is></c>'

    cols_xml = ""
    if column_widths:
        cols_xml = "<cols>" + "".join(
            f'<col min="{idx}" max="{idx}" width="{width}" customWidth="1"/>'
            for idx, width in enumerate(column_widths, start=1)
        ) + "</cols>"

    sheet_rows = []
    for row_idx, row in enumerate(rows, start=1):
        cells = "".join(inline_cell(row_idx, col_idx, value) for col_idx, value in enumerate(row, start=1))
        sheet_rows.append(f'<row r="{row_idx}">{cells}</row>')

    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"{cols_xml}<sheetData>{''.join(sheet_rows)}</sheetData></worksheet>"
    )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheets><sheet name="{escape(sheet_name)}" sheetId="1" r:id="rId1"/></sheets></workbook>'
    )
    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        "</Relationships>"
    )
    root_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        "</Types>"
    )
    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
        '<fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        "</styleSheet>"
    )
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types_xml)
        zf.writestr("_rels/.rels", root_rels_xml)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)
        zf.writestr("xl/styles.xml", styles_xml)
    return output.getvalue()

from flask import (
    Flask, render_template, render_template_string, request, redirect, url_for,
    flash, jsonify, Response, abort, g, current_app
)

from sqlalchemy import text, or_, and_, case

# ReportLab (PDF)
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Table, TableStyle, Paragraph, SimpleDocTemplate, Spacer, KeepTogether
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.utils import ImageReader
from reportlab.lib.enums import TA_JUSTIFY

# Excel
try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except Exception:
    Workbook = None  # la app sigue arrancando aunque falte openpyxl
    get_column_letter = None

# Twilio + Scheduler
from twilio.rest import Client as TwilioClient
from apscheduler.schedulers.background import BackgroundScheduler

# Auth (Flask-Login)
from flask_login import LoginManager, login_user, login_required, logout_user, current_user

# ---------------------------------------------------------
# Config
# ---------------------------------------------------------
TZ_CDMX = ZoneInfo("America/Mexico_City")

def now_cdmx_naive() -> datetime:
    """Hora CDMX (naive). Úsala para timestamps en DB/UX sin desfases."""
    return datetime.now(TZ_CDMX).replace(tzinfo=None)

DEFAULT_SECRET_KEY = "poliutech_mar_checkpoint_superseguro"
DEFAULT_DATABASE_URL = "sqlite:///mar3.db"
MAR_BLUE = "#0C3C78"
MAR_BLUE_XLSX = "0C3C78"

TWILIO_ACCOUNT_SID = os.getenv("TWILIO_ACCOUNT_SID", "").strip()
TWILIO_AUTH_TOKEN  = os.getenv("TWILIO_AUTH_TOKEN", "").strip()
TWILIO_WHATSAPP    = os.getenv("TWILIO_WHATSAPP", "whatsapp:+14155238886").strip()

DEFAULT_ADMIN_WHATSAPP_RECIPIENTS = (
    "whatsapp:+5215521323076,whatsapp:+5215610035643,whatsapp:+14055619808"
)
ADMIN_WHATSAPP_RECIPIENTS = os.getenv(
    "ADMIN_WHATSAPP_RECIPIENTS",
    DEFAULT_ADMIN_WHATSAPP_RECIPIENTS
).strip()
ADMIN_LIST: List[str] = [x.strip() for x in ADMIN_WHATSAPP_RECIPIENTS.split(",") if x.strip()]

SMTP_HOST = os.getenv("SMTP_HOST", "servidor15.escala.net.mx").strip()
SMTP_PORT = int(os.getenv("SMTP_PORT", "26"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME", "cotizaciones@poliutech.com").strip()
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "Cotizaciones2025@").strip()
SMTP_FROM = os.getenv("SMTP_FROM", SMTP_USERNAME).strip()
COTIZACION_REVIEW_EMAIL = "hjaramillo@poliutech.com,mescalera@poliutech.com"
COTIZACION_REVIEW_BCC_EMAIL = "sistemas@poliutech.com"
COTIZACION_RESPONSE_EMAIL = (os.getenv("COTIZACION_RESPONSE_EMAIL") or "umorales@poliutech.com").strip()
COTIZACION_APPROVALS_EMAIL = "aprobaciones@poliutech.com,mescalera@poliutech.com"
COTIZACION_REVIEW_RESULT_AAZCONA_EMAIL = "aazcona@poliutech.com"
GASTOS_REVIEW_EMAIL = "hjaramillo@poliutech.com,sistemas@poliutech.com"
GASTOS_REVIEW_BCC_EMAIL = ""
FINANZAS_AUTH_NOTIFY_EMAILS = "mescalera@poliutech.com,miguele@poliutech.com"
SUPPORT_TICKET_EMAIL = (os.getenv("SUPPORT_TICKET_EMAIL") or "sistemas@poliutech.com").strip()
USER_CREATION_EMAIL = "sistemas@poliutech.com"
COTIZACION_TRASH_RETENTION_DAYS = 30
REGISTRO_MAIL_HOST = os.getenv("REGISTRO_MAIL_HOST", "servidor15.escala.net.mx").strip()
REGISTRO_MAIL_PORT = int(os.getenv("REGISTRO_MAIL_PORT", "26"))
REGISTRO_MAIL_USERNAME = os.getenv("REGISTRO_MAIL_USERNAME", "info@poliutech.com").strip()
REGISTRO_MAIL_PASSWORD = os.getenv("REGISTRO_MAIL_PASSWORD", "Info@2025?").strip()
REGISTRO_MAIL_FROM = os.getenv("REGISTRO_MAIL_FROM", REGISTRO_MAIL_USERNAME).strip()
REGISTRO_MAIL_ATTACHMENT = Path(__file__).resolve().parent / "presentacion2026OK.pdf"
FIREBASE_CREDENTIALS_FILE = os.getenv("FIREBASE_CREDENTIALS_FILE", "").strip()
FIREBASE_CREDENTIALS_JSON = os.getenv("FIREBASE_CREDENTIALS_JSON", "").strip()
PUSH_NOTIFICATIONS_ENABLED = os.getenv("PUSH_NOTIFICATIONS_ENABLED", "1").strip().lower() not in {"0", "false", "no"}

# Usa SIEMPRE los modelos desde models.py para evitar duplicados
from models import (
    db,
    Cliente,
    Concepto,
    Cotizacion,
    CotizacionDetalle,
    CotizacionSeguimiento,
    VoiceCommandLog,
    Usuario,
    MobileDevice,
    RegistroObra,
    RegistroObraSeguimiento,
    Prospecto,
    ProspectoSeguimiento,
    TicketSoporte,
    TicketComentario,
    TicketAdjunto,
    ActivityLog,
    InventarioProducto,
    InventarioMovimiento,
    OrdenCompra,
    OrdenCompraPartida,
    SolicitudRecurso,
    SolicitudRecursoPartida,
    ReporteDiario,
    MovimientoFinanciero,
    ComprobacionGasto,
    ComprobacionAdjunto,
)

try:
    from models import MovimientoFinancieroPago
except ImportError:
    class MovimientoFinancieroPago(db.Model):
        __tablename__ = "movimiento_financiero_pago"

        id = db.Column(db.Integer, primary_key=True)
        movimiento_id = db.Column(db.Integer, db.ForeignKey("movimiento_financiero.id"), nullable=False, index=True)
        fecha = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
        monto = db.Column(db.Float, default=0.0, nullable=False)
        referencia = db.Column(db.String(120))
        notas = db.Column(db.Text)
        responsable = db.Column(db.String(120))
        usuario_id = db.Column(db.Integer, db.ForeignKey("usuario.id"), nullable=True)
        creado_en = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

        movimiento = db.relationship("MovimientoFinanciero")
        usuario = db.relationship("Usuario", backref=db.backref("pagos_creditos_financieros_fallback", lazy=True))

# ---------------------------------------------------------
# Flask + DB + Login
# ---------------------------------------------------------
app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", DEFAULT_SECRET_KEY)
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("DATABASE_URL", DEFAULT_DATABASE_URL)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
logger = logging.getLogger(__name__)

db.init_app(app)


@app.context_processor
def inject_endpoint_helpers():
    def endpoint_exists(endpoint: str) -> bool:
        return endpoint in current_app.view_functions

    return {
        "endpoint_exists": endpoint_exists,
        "gastos_admin_can_view": lambda: _gastos_admin_can_view(),
        "estado_cuenta_recursos_can_view": lambda: _estado_cuenta_recursos_can_view(),
        "evaluacion_departamental_can_view": lambda: _evaluacion_departamental_can_view(),
    }


login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

@login_manager.user_loader
def load_user(user_id):
    try:
        return Usuario.query.get(int(user_id))
    except Exception:
        return None

# ---------------------------------------------------------
# 🔒 Enforce login for ALL pages (except /login + static)
# ---------------------------------------------------------
@app.before_request
def _require_login_everywhere():
    """Protege TODAS las páginas del sistema.

    Si el usuario NO está autenticado, lo mandamos a /login.
    Esto cubre cualquier ruta/HTML aunque olvides poner @login_required.
    """
    # Permitir estáticos
    if request.path.startswith("/static/") or request.endpoint == "static":
        return

    # Permitir autenticación y recuperación de acceso
    if request.endpoint in {"login", "forgot_password", "reset_password"}:
        return
    if request.path in ("/health", "/ping"):
        return
    if request.path.startswith("/gastos-viaticos/revision/"):
        return
    if request.path.startswith("/cotizaciones/revision/"):
        return
    if request.path.startswith("/api/mobile/"):
        return
    if request.path.startswith("/cotizaciones/") and request.path.endswith("/export.pdf"):
        auth_header = (request.headers.get("Authorization") or "").strip()
        if auth_header.lower().startswith("bearer "):
            return

    # Si ya está logueado, ok
    if current_user.is_authenticated:
        return

    # Redirigir a login, preservando a dónde quería ir
    nxt = request.full_path
    if nxt.endswith("?"):
        nxt = nxt[:-1]
    return redirect(url_for("login", next=nxt))

# ---------------------------------------------------------
# Bitácora de actividad (Audit Log)
# ---------------------------------------------------------
def _safe_join_keys(keys, limit=60):
    try:
        if not keys:
            return None
        out = []
        for k in list(keys)[:limit]:
            # evitamos guardar cosas sensibles por nombre
            lk = str(k).lower()
            if any(x in lk for x in ["pass", "password", "clave", "token", "secret"]):
                out.append(f"{k}=<hidden>")
            else:
                out.append(str(k))
        s = ", ".join(out)
        return s[:780]
    except Exception:
        return None

def _get_client_ip():
    # Render / proxies: X-Forwarded-For suele venir
    xf = request.headers.get("X-Forwarded-For", "")
    if xf:
        return xf.split(",")[0].strip()[:60]
    return (request.remote_addr or "")[:60]



# ---------------------------------------------------------
# Audit retention / cleanup (keep DB from growing forever)
# ---------------------------------------------------------
AUDIT_LOG_RETENTION_DAYS = int(os.getenv("AUDIT_LOG_RETENTION_DAYS", "90"))
AUDIT_CLEANUP_EVERY_HOURS = int(os.getenv("AUDIT_CLEANUP_EVERY_HOURS", "24"))

def _audit_cleanup_stamp_path() -> str:
    try:
        os.makedirs(app.instance_path, exist_ok=True)
    except Exception:
        pass
    return os.path.join(app.instance_path, "audit_cleanup_stamp.txt")

def _should_run_audit_cleanup(now: datetime) -> bool:
    # Run cleanup at most once per AUDIT_CLEANUP_EVERY_HOURS (best-effort).
    try:
        stamp_path = _audit_cleanup_stamp_path()
        if not os.path.exists(stamp_path):
            return True
        raw = Path(stamp_path).read_text(encoding="utf-8", errors="ignore").strip()
        if not raw:
            return True
        try:
            last = datetime.fromisoformat(raw)
        except Exception:
            return True
        delta = now - last
        return delta.total_seconds() >= AUDIT_CLEANUP_EVERY_HOURS * 3600
    except Exception:
        # If anything weird, skip (never break requests)
        return False

def _mark_audit_cleanup(now: datetime) -> None:
    try:
        Path(_audit_cleanup_stamp_path()).write_text(now.isoformat(), encoding="utf-8")
    except Exception:
        pass

def cleanup_audit_logs(retention_days: int | None = None) -> int:
    # Delete ActivityLog rows older than retention_days. Returns deleted count (best-effort).
    try:
        days = int(retention_days if retention_days is not None else AUDIT_LOG_RETENTION_DAYS)
        cutoff = now_cdmx_naive() - timedelta(days=days)
        deleted = ActivityLog.query.filter(ActivityLog.fecha < cutoff).delete(synchronize_session=False)
        db.session.commit()
        return int(deleted or 0)
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        return 0

def maybe_cleanup_audit_logs() -> None:
    # Best-effort periodic cleanup.
    try:
        now = now_cdmx_naive()
        if _should_run_audit_cleanup(now):
            cleanup_audit_logs()
            _mark_audit_cleanup(now)
    except Exception:
        pass
def _describe_action():
    # Acción legible sin datos sensibles
    try:
        ep = (request.endpoint or "").strip()
        m = request.method
        p = request.path

        # Login explícito
        if ep == "login" and m == "POST":
            nombre = (
                request.form.get("nombre")
                or request.form.get("username")
                or request.form.get("usuario")
                or request.form.get("user")
                or ""
            ).strip()[:60]
            return f"LOGIN intento usuario={nombre}"

        if ep == "logout":
            return "LOGOUT"

        # Cotizaciones (patrones comunes)
        if "cotizacion" in p.lower():
            return f"{m} {p}"

        if "cliente" in p.lower():
            return f"{m} {p}"

        if "catalog" in p.lower() or "catalogo" in p.lower():
            return f"{m} {p}"

        # Default
        return f"{m} {p}"
    except Exception:
        return f"{request.method} {request.path}"

@app.before_request
def _audit_before_request():
    try:
        # Ignorar estáticos y healthchecks
        if request.path.startswith("/static/") or request.path == "/favicon.ico":
            g._skip_audit = True
            return
        g._skip_audit = False

        g._audit_started_at = now_cdmx_naive()

        # Captura keys sin valores
        form_keys = None
        json_keys = None
        if request.method in ("POST", "PUT", "PATCH", "DELETE"):
            if request.form:
                form_keys = _safe_join_keys(request.form.keys())
            j = request.get_json(silent=True)
            if isinstance(j, dict):
                json_keys = _safe_join_keys(j.keys())

        g._audit_payload = {
            "form_keys": form_keys,
            "json_keys": json_keys,
            "query_string": (request.query_string.decode("utf-8", "ignore")[:780] if request.query_string else None),
        }
    except Exception:
        # no rompemos request por falla de bitácora
        g._skip_audit = True

@app.after_request
def _audit_after_request(response):
    try:
        if getattr(g, "_skip_audit", False):
            return response

        # Usuario
        usuario = "ANON"
        usuario_id = None
        rol = None
        try:
            if current_user and getattr(current_user, "is_authenticated", False):
                usuario = (_usuario_nombre_representante(current_user) or "ANON")[:60]
                usuario_id = getattr(current_user, "id", None)
                rol = getattr(current_user, "rol", None)
        except Exception:
            pass

        # acción
        accion = _describe_action()

        log = ActivityLog(
            fecha=now_cdmx_naive(),
            usuario_id=usuario_id,
            usuario=usuario,
            rol=rol,
            metodo=request.method,
            ruta=(request.path or "")[:300],
            endpoint=(request.endpoint or "")[:120] if request.endpoint else None,
            status_code=int(getattr(response, "status_code", 0) or 0),
            ip=_get_client_ip(),
            user_agent=(request.headers.get("User-Agent", "")[:300] if request.headers else None),
            query_string=(g._audit_payload.get("query_string") if hasattr(g, "_audit_payload") else None),
            form_keys=(g._audit_payload.get("form_keys") if hasattr(g, "_audit_payload") else None),
            json_keys=(g._audit_payload.get("json_keys") if hasattr(g, "_audit_payload") else None),
            accion=accion[:500],
        )
        db.session.add(log)
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
    return response

# ---------------------------------------------------------
# Twilio init (Render)
# ---------------------------------------------------------
twilio_client: Optional[TwilioClient] = None

def init_twilio_client():
    global twilio_client, TWILIO_WHATSAPP
    try:
        sid = os.getenv("TWILIO_ACCOUNT_SID", "").strip()
        token = os.getenv("TWILIO_AUTH_TOKEN", "").strip()
        wsp_from = os.getenv("TWILIO_WHATSAPP", "").strip()

        if sid and token and wsp_from:
            twilio_client = TwilioClient(sid, token)
            TWILIO_WHATSAPP = wsp_from
            print("[Twilio] Cliente inicializado correctamente.")
            print(f"[Twilio] Remitente WhatsApp: {TWILIO_WHATSAPP}")
        else:
            twilio_client = None
            print("[Twilio] Configuración incompleta. WhatsApp deshabilitado.")
    except Exception as e:
        twilio_client = None
        print(f"[Twilio] Error al inicializar cliente: {e}", file=sys.stderr)

with app.app_context():
    init_twilio_client()

# ---------------------------------------------------------
# Migraciones mínimas (SQLite)
# ---------------------------------------------------------
def _table_columns(table_name: str) -> set[str]:
    rows = db.session.execute(text(f"PRAGMA table_info('{table_name}')")).mappings().all()
    return {r["name"] for r in rows}

def ensure_schema():
    """Crea tablas si no existen y agrega/normaliza columnas clave."""
    print("🔍 Verificando estructura de la base de datos...")
    db.create_all()

    # --- CLIENTE.responsable ---
    try:
        cols_cli = _table_columns("cliente")
        if "responsable" not in cols_cli:
            db.session.execute(text("ALTER TABLE cliente ADD COLUMN responsable VARCHAR(120)"))
            db.session.commit()
            print("✅ Campo 'responsable' agregado en 'cliente'.")
    except Exception as e:
        print("⚠️ ensure_schema(cliente.responsable):", e)

    # --- CLIENTE.sistema (si existía en tu proyecto) ---
    try:
        cols_cli = _table_columns("cliente")
        if "sistema" not in cols_cli:
            db.session.execute(text("ALTER TABLE cliente ADD COLUMN sistema VARCHAR(120)"))
            db.session.commit()
            print("✅ Campo 'sistema' agregado en 'cliente'.")
    except Exception as e:
        print("⚠️ ensure_schema(cliente.sistema):", e)

    # --- USUARIO.nombre_visible ---
    try:
        cols_user = _table_columns("usuario")
        if "nombre_visible" not in cols_user:
            db.session.execute(text("ALTER TABLE usuario ADD COLUMN nombre_visible VARCHAR(120)"))
            db.session.execute(text("UPDATE usuario SET nombre_visible = nombre WHERE nombre_visible IS NULL OR TRIM(nombre_visible) = ''"))
            db.session.commit()
            print("✅ Campo 'nombre_visible' agregado en 'usuario'.")
    except Exception as e:
        print("⚠️ ensure_schema(usuario.nombre_visible):", e)

    # --- COTIZACION.responsable ---
    try:
        cols_cot = _table_columns("cotizacion")
        if "responsable" not in cols_cot:
            if "representante" in cols_cot:
                db.session.execute(text("ALTER TABLE cotizacion ADD COLUMN responsable VARCHAR(120)"))
                try:
                    db.session.execute(text("UPDATE cotizacion SET responsable = representante WHERE responsable IS NULL"))
                except Exception:
                    pass
                db.session.commit()
                print("✅ Campo 'responsable' creado y poblado desde 'representante'.")
            else:
                db.session.execute(text("ALTER TABLE cotizacion ADD COLUMN responsable VARCHAR(120)"))
                db.session.commit()
                print("✅ Campo 'responsable' agregado en 'cotizacion'.")
    except Exception as e:
        print("⚠️ ensure_schema(cotizacion.responsable):", e)

    # --- Otros mínimos para estabilidad ---
    try:
        user_cols = _table_columns("usuario")
        if "correo" not in user_cols:
            db.session.execute(text("ALTER TABLE usuario ADD COLUMN correo VARCHAR(160)"))
            db.session.commit()
            print("✅ Campo 'correo' agregado en 'usuario'.")
    except Exception as e:
        print("⚠️ ensure_schema(usuario.correo):", e)

    # --- Otros mínimos para estabilidad ---
    try:
        cols = _table_columns("cotizacion")
        for col, stmt in [
            ("subtotal", "ALTER TABLE cotizacion ADD COLUMN subtotal FLOAT DEFAULT 0.0"),
            ("descuento_total", "ALTER TABLE cotizacion ADD COLUMN descuento_total FLOAT DEFAULT 0.0"),
            ("iva_porc", "ALTER TABLE cotizacion ADD COLUMN iva_porc FLOAT DEFAULT 16.0"),
            ("iva_monto", "ALTER TABLE cotizacion ADD COLUMN iva_monto FLOAT DEFAULT 0.0"),
            ("total", "ALTER TABLE cotizacion ADD COLUMN total FLOAT DEFAULT 0.0"),
            ("moneda", "ALTER TABLE cotizacion ADD COLUMN moneda VARCHAR(10) DEFAULT 'MXN'"),
            ("estatus_aprobacion", "ALTER TABLE cotizacion ADD COLUMN estatus_aprobacion VARCHAR(20) DEFAULT 'EN REVISIÓN'"),
            ("especialidad", "ALTER TABLE cotizacion ADD COLUMN especialidad VARCHAR(160)"),
            ("notas", "ALTER TABLE cotizacion ADD COLUMN notas VARCHAR(3000)"),
            ("last_whatsapp_at", "ALTER TABLE cotizacion ADD COLUMN last_whatsapp_at TIMESTAMP NULL"),
            ("proyecto", "ALTER TABLE cotizacion ADD COLUMN proyecto VARCHAR(200)"),
            ("ciudad_trabajo", "ALTER TABLE cotizacion ADD COLUMN ciudad_trabajo VARCHAR(120)"),
            ("eliminada_en", "ALTER TABLE cotizacion ADD COLUMN eliminada_en TIMESTAMP NULL"),
            ("eliminada_por", "ALTER TABLE cotizacion ADD COLUMN eliminada_por VARCHAR(120)"),
            ("eliminacion_definitiva_en", "ALTER TABLE cotizacion ADD COLUMN eliminacion_definitiva_en TIMESTAMP NULL"),
        ]:
            if col not in cols:
                try:
                    db.session.execute(text(stmt))
                except Exception:
                    pass
        db.session.commit()
        cols = _table_columns("cotizacion")
        if "estatus_aprobacion" in cols and "estatus" in cols:
            db.session.execute(text("""
                UPDATE cotizacion
                SET estatus_aprobacion = CASE
                    WHEN UPPER(COALESCE(estatus, '')) IN ('APROBADO', 'APROBADA', 'AUTORIZADO') THEN 'APROBADA'
                    WHEN UPPER(COALESCE(estatus, '')) IN ('RECHAZADO', 'RECHAZADA') THEN 'RECHAZADA'
                    WHEN UPPER(COALESCE(estatus, '')) IN ('EN REVISIÓN', 'EN REVISION') THEN 'EN REVISIÓN'
                    ELSE COALESCE(NULLIF(TRIM(estatus_aprobacion), ''), 'EN REVISIÓN')
                END
                WHERE estatus_aprobacion IS NULL OR TRIM(estatus_aprobacion) = ''
                   OR UPPER(COALESCE(estatus, '')) IN ('APROBADO', 'APROBADA', 'AUTORIZADO', 'RECHAZADO', 'RECHAZADA', 'EN REVISIÓN', 'EN REVISION')
            """))
            db.session.execute(text("""
                UPDATE cotizacion
                SET estatus = 'PENDIENTE'
                WHERE UPPER(COALESCE(estatus, '')) IN ('APROBADO', 'APROBADA', 'AUTORIZADO', 'RECHAZADO', 'RECHAZADA', 'EN REVISIÓN', 'EN REVISION')
                   OR estatus IS NULL OR TRIM(estatus) = ''
            """))
            db.session.commit()
    except Exception as e:
        print("⚠️ ensure_schema(cotizacion extras):", e)

    try:
        inv_cols = _table_columns("inventario_producto")
        for col, stmt in [
            ("stock_maximo", "ALTER TABLE inventario_producto ADD COLUMN stock_maximo FLOAT DEFAULT 0.0"),
        ]:
            if col not in inv_cols:
                db.session.execute(text(stmt))
        db.session.commit()
    except Exception as e:
        print("⚠️ ensure_schema(inventario_producto):", e)

    try:
        oc_cols = _table_columns("orden_compra")
        for col, stmt in [
            ("numero_cliente_proveedor", "ALTER TABLE orden_compra ADD COLUMN numero_cliente_proveedor VARCHAR(80)"),
            ("forma_pago", "ALTER TABLE orden_compra ADD COLUMN forma_pago VARCHAR(20) DEFAULT 'CONTADO'"),
            ("descuento_total", "ALTER TABLE orden_compra ADD COLUMN descuento_total FLOAT DEFAULT 0.0"),
            ("factura_folio", "ALTER TABLE orden_compra ADD COLUMN factura_folio VARCHAR(80)"),
            ("factura_monto", "ALTER TABLE orden_compra ADD COLUMN factura_monto FLOAT DEFAULT 0.0"),
            ("factura_archivo", "ALTER TABLE orden_compra ADD COLUMN factura_archivo VARCHAR(260)"),
            ("pago_referencia", "ALTER TABLE orden_compra ADD COLUMN pago_referencia VARCHAR(120)"),
            ("pago_monto", "ALTER TABLE orden_compra ADD COLUMN pago_monto FLOAT DEFAULT 0.0"),
            ("pago_archivo", "ALTER TABLE orden_compra ADD COLUMN pago_archivo VARCHAR(260)"),
            ("condiciones", "ALTER TABLE orden_compra ADD COLUMN condiciones TEXT"),
        ]:
            if col not in oc_cols:
                db.session.execute(text(stmt))
        db.session.commit()
    except Exception as e:
        print("⚠️ ensure_schema(orden_compra):", e)

    try:
        sr_cols = _table_columns("solicitud_recurso")
        for col, stmt in [
            ("gasto_generado_id", "ALTER TABLE solicitud_recurso ADD COLUMN gasto_generado_id INTEGER"),
            ("gasto_generado_en", "ALTER TABLE solicitud_recurso ADD COLUMN gasto_generado_en TIMESTAMP"),
        ]:
            if col not in sr_cols:
                db.session.execute(text(stmt))
        db.session.commit()
    except Exception as e:
        print("⚠️ ensure_schema(solicitud_recurso):", e)

    try:
        sr_partida_cols = _table_columns("solicitud_recurso_partida")
        if "total" not in sr_partida_cols:
            db.session.execute(text("ALTER TABLE solicitud_recurso_partida ADD COLUMN total FLOAT DEFAULT 0.0"))
            db.session.execute(text("""
                UPDATE solicitud_recurso_partida
                SET total = COALESCE(importe, 0.0)
                WHERE total IS NULL OR total = 0
            """))
            db.session.commit()
            print("✅ Campo 'total' agregado en 'solicitud_recurso_partida'.")
    except Exception as e:
        print("⚠️ ensure_schema(solicitud_recurso_partida):", e)

    try:
        comprobacion_cols = _table_columns("comprobacion_gasto")
        if "solicitud_recurso_id" not in comprobacion_cols:
            db.session.execute(text("ALTER TABLE comprobacion_gasto ADD COLUMN solicitud_recurso_id INTEGER"))
            db.session.execute(text("""
                UPDATE comprobacion_gasto
                SET solicitud_recurso_id = (
                    SELECT solicitud_recurso.id
                    FROM solicitud_recurso
                    WHERE comprobacion_gasto.referencia = 'SR:' || solicitud_recurso.folio
                    LIMIT 1
                )
                WHERE solicitud_recurso_id IS NULL
                  AND referencia LIKE 'SR:%'
            """))
            db.session.commit()
            print("✅ Campo 'solicitud_recurso_id' agregado en 'comprobacion_gasto'.")
    except Exception as e:
        print("⚠️ ensure_schema(comprobacion_gasto.solicitud_recurso_id):", e)

    try:
        dcols = _table_columns("cotizacion_detalle")
        if "sistema" not in dcols:
            db.session.execute(text("ALTER TABLE cotizacion_detalle ADD COLUMN sistema VARCHAR(200)"))
        if "descripcion" not in dcols:
            db.session.execute(text("ALTER TABLE cotizacion_detalle ADD COLUMN descripcion VARCHAR(1000)"))
        for col, stmt in [
            ("capitulo", "ALTER TABLE cotizacion_detalle ADD COLUMN capitulo VARCHAR(120)"),
            ("origen", "ALTER TABLE cotizacion_detalle ADD COLUMN origen VARCHAR(50)"),
        ]:
            if col not in dcols:
                db.session.execute(text(stmt))
        db.session.commit()
    except Exception as e:
        print("[WARN] ensure_schema(detalle extras):", e)

    # --- PRECIOS UNITARIOS: columnas nativas ---
    try:
        pu_obra_cols = _table_columns("pu_obra")
        for col, stmt in [
            ("cliente", "ALTER TABLE pu_obra ADD COLUMN cliente VARCHAR(180)"),
            ("ubicacion", "ALTER TABLE pu_obra ADD COLUMN ubicacion VARCHAR(220)"),
            ("descripcion", "ALTER TABLE pu_obra ADD COLUMN descripcion TEXT"),
            ("moneda", "ALTER TABLE pu_obra ADD COLUMN moneda VARCHAR(20) DEFAULT 'MXN'"),
            ("m2_proyecto", "ALTER TABLE pu_obra ADD COLUMN m2_proyecto FLOAT DEFAULT 0.0"),
            ("creado_en", "ALTER TABLE pu_obra ADD COLUMN creado_en TIMESTAMP"),
            ("actualizado_en", "ALTER TABLE pu_obra ADD COLUMN actualizado_en TIMESTAMP"),
        ]:
            if col not in pu_obra_cols:
                db.session.execute(text(stmt))
        db.session.commit()
    except Exception as e:
        print("[WARN] ensure_schema(pu_obra):", e)

    try:
        pu_sob_cols = _table_columns("pu_sobrecosto")
        for col, stmt in [
            ("indirecto_campo_pct", "ALTER TABLE pu_sobrecosto ADD COLUMN indirecto_campo_pct FLOAT DEFAULT 0.0"),
            ("indirecto_oficina_pct", "ALTER TABLE pu_sobrecosto ADD COLUMN indirecto_oficina_pct FLOAT DEFAULT 0.0"),
            ("financiamiento_pct", "ALTER TABLE pu_sobrecosto ADD COLUMN financiamiento_pct FLOAT DEFAULT 0.0"),
            ("utilidad_pct", "ALTER TABLE pu_sobrecosto ADD COLUMN utilidad_pct FLOAT DEFAULT 10.0"),
            ("cargos_adicionales_pct", "ALTER TABLE pu_sobrecosto ADD COLUMN cargos_adicionales_pct FLOAT DEFAULT 0.0"),
            ("creado_en", "ALTER TABLE pu_sobrecosto ADD COLUMN creado_en TIMESTAMP"),
            ("actualizado_en", "ALTER TABLE pu_sobrecosto ADD COLUMN actualizado_en TIMESTAMP"),
        ]:
            if col not in pu_sob_cols:
                db.session.execute(text(stmt))
        db.session.commit()
    except Exception as e:
        print("[WARN] ensure_schema(pu_sobrecosto):", e)

    try:
        pu_recurso_cols = _table_columns("pu_recurso")
        for col, stmt in [
            ("tipo", "ALTER TABLE pu_recurso ADD COLUMN tipo VARCHAR(30) DEFAULT 'material'"),
            ("codigo", "ALTER TABLE pu_recurso ADD COLUMN codigo VARCHAR(60)"),
            ("descripcion", "ALTER TABLE pu_recurso ADD COLUMN descripcion VARCHAR(300)"),
            ("unidad", "ALTER TABLE pu_recurso ADD COLUMN unidad VARCHAR(50)"),
            ("costo_unitario", "ALTER TABLE pu_recurso ADD COLUMN costo_unitario FLOAT DEFAULT 0.0"),
            ("creado_en", "ALTER TABLE pu_recurso ADD COLUMN creado_en TIMESTAMP"),
            ("actualizado_en", "ALTER TABLE pu_recurso ADD COLUMN actualizado_en TIMESTAMP"),
        ]:
            if col not in pu_recurso_cols:
                db.session.execute(text(stmt))
        db.session.commit()
    except Exception as e:
        print("[WARN] ensure_schema(pu_recurso):", e)

    try:
        pu_partida_cols = _table_columns("pu_partida")
        for col, stmt in [
            ("capitulo", "ALTER TABLE pu_partida ADD COLUMN capitulo VARCHAR(160) DEFAULT 'General'"),
            ("clave", "ALTER TABLE pu_partida ADD COLUMN clave VARCHAR(80)"),
            ("descripcion", "ALTER TABLE pu_partida ADD COLUMN descripcion VARCHAR(600)"),
            ("unidad", "ALTER TABLE pu_partida ADD COLUMN unidad VARCHAR(50) DEFAULT 'pza'"),
            ("cantidad", "ALTER TABLE pu_partida ADD COLUMN cantidad FLOAT DEFAULT 1.0"),
            ("costo_directo", "ALTER TABLE pu_partida ADD COLUMN costo_directo FLOAT DEFAULT 0.0"),
            ("precio_unitario", "ALTER TABLE pu_partida ADD COLUMN precio_unitario FLOAT DEFAULT 0.0"),
            ("importe", "ALTER TABLE pu_partida ADD COLUMN importe FLOAT DEFAULT 0.0"),
            ("creado_en", "ALTER TABLE pu_partida ADD COLUMN creado_en TIMESTAMP"),
            ("actualizado_en", "ALTER TABLE pu_partida ADD COLUMN actualizado_en TIMESTAMP"),
        ]:
            if col not in pu_partida_cols:
                db.session.execute(text(stmt))
        db.session.commit()
    except Exception as e:
        print("[WARN] ensure_schema(pu_partida):", e)

    try:
        pu_insumo_cols = _table_columns("pu_partida_insumo")
        for col, stmt in [
            ("recurso_id", "ALTER TABLE pu_partida_insumo ADD COLUMN recurso_id INTEGER"),
            ("tipo", "ALTER TABLE pu_partida_insumo ADD COLUMN tipo VARCHAR(30) DEFAULT 'material'"),
            ("codigo", "ALTER TABLE pu_partida_insumo ADD COLUMN codigo VARCHAR(60)"),
            ("descripcion", "ALTER TABLE pu_partida_insumo ADD COLUMN descripcion VARCHAR(400)"),
            ("unidad", "ALTER TABLE pu_partida_insumo ADD COLUMN unidad VARCHAR(50)"),
            ("presentacion", "ALTER TABLE pu_partida_insumo ADD COLUMN presentacion VARCHAR(120)"),
            ("rendimiento_m2", "ALTER TABLE pu_partida_insumo ADD COLUMN rendimiento_m2 FLOAT DEFAULT 0.0"),
            ("cantidad_exacta", "ALTER TABLE pu_partida_insumo ADD COLUMN cantidad_exacta FLOAT DEFAULT 0.0"),
            ("cantidad", "ALTER TABLE pu_partida_insumo ADD COLUMN cantidad FLOAT DEFAULT 0.0"),
            ("costo_unitario", "ALTER TABLE pu_partida_insumo ADD COLUMN costo_unitario FLOAT DEFAULT 0.0"),
            ("importe", "ALTER TABLE pu_partida_insumo ADD COLUMN importe FLOAT DEFAULT 0.0"),
            ("gravable", "ALTER TABLE pu_partida_insumo ADD COLUMN gravable BOOLEAN DEFAULT 1 NOT NULL"),
            ("creado_en", "ALTER TABLE pu_partida_insumo ADD COLUMN creado_en TIMESTAMP"),
            ("actualizado_en", "ALTER TABLE pu_partida_insumo ADD COLUMN actualizado_en TIMESTAMP"),
        ]:
            if col not in pu_insumo_cols:
                db.session.execute(text(stmt))
        db.session.commit()
    except Exception as e:
        print("[WARN] ensure_schema(pu_partida_insumo):", e)

    _migrate_registro_obras_from_json()

# ---------------------------------------------------------
# Seed: usuarios base (idempotente)
# ---------------------------------------------------------
def seed_default_users():
    """Crea usuarios base si no existen (no duplica)."""
    defaults = [
        ("Ing. Antonio Azcona", "Azcona123!", "USER"),
        ("Joandlc", "Joan123!", "USER"),
        ("JSolis", "Solis123!", "ADMIN"),
    ]
    created = 0
    for nombre, password, rol in defaults:
        try:
            exists = Usuario.query.filter(db.func.lower(Usuario.nombre) == nombre.lower()).first()
            if exists:
                continue
            u = Usuario(nombre=nombre, nombre_visible=nombre, rol=rol)
            # Usa el helper del modelo para hashear
            try:
                u.set_password(password)
            except Exception:
                from werkzeug.security import generate_password_hash
                u.password = generate_password_hash(password)
            db.session.add(u)
            created += 1
        except Exception:
            continue
    try:
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
    if created:
        print(f"✅ Seed users: creados {created} usuario(s).")

with app.app_context():
    ensure_schema()

with app.app_context():
    seed_default_users()


# ==============================
# SETUP TEMPORAL ADMIN
# ==============================
@app.route("/setup_admin")
def setup_admin():
    nombre = "Rafa"       # ← cámbialo si quieres
    password = "1234"     # ← cámbialo si quieres
    rol = "ADMIN"         # ADMIN o USER

    u = Usuario.query.filter_by(nombre=nombre).first()
    if u:
        return f"Ya existe el usuario {nombre}"

    u = Usuario(nombre=nombre, nombre_visible=nombre, rol=rol)
    u.set_password(password)
    db.session.add(u)
    db.session.commit()

    return f"✅ Usuario creado: {nombre} / {password} ({rol})"

# ---------------------------------------------------------
# Helpers (roles + formatting)
# ---------------------------------------------------------
def is_admin() -> bool:
    return bool(getattr(current_user, "is_authenticated", False) and (getattr(current_user, "rol", "") or "").upper() == "ADMIN")

def is_admin_account() -> bool:
    nombre = (getattr(current_user, "nombre", "") or "").strip().lower()
    return bool(getattr(current_user, "is_authenticated", False) and nombre == "admin")

def _gastos_admin_can_view() -> bool:
    if not getattr(current_user, "is_authenticated", False):
        return False
    nombre = (getattr(current_user, "nombre", "") or "").strip().lower()
    visible = (getattr(current_user, "nombre_visible", "") or "").strip().lower()
    correo = (getattr(current_user, "correo", "") or "").strip().lower()
    return (
        is_admin()
        or getattr(current_user, "id", None) == 18
        or correo == "hjaramillo@poliutech.com"
        or nombre in {"hansel", "hjaramillo"}
        or nombre.startswith("hansel")
        or visible.startswith("hansel")
    )

def _estado_cuenta_recursos_can_view() -> bool:
    if not getattr(current_user, "is_authenticated", False):
        return False
    nombre = (getattr(current_user, "nombre", "") or "").strip().lower()
    visible = (getattr(current_user, "nombre_visible", "") or "").strip().lower()
    correo = (getattr(current_user, "correo", "") or "").strip().lower()
    return (
        is_admin()
        or nombre == "admin"
        or getattr(current_user, "id", None) == 18
        or correo in {"hjaramillo@poliutech.com", "miguele@poliutech.com"}
        or nombre in {"hansel", "hjaramillo", "miguel", "miguele"}
        or nombre.startswith("hansel")
        or nombre.startswith("miguel")
        or visible.startswith("hansel")
        or visible.startswith("miguel")
    )

def normalize_user_role(value: str) -> str:
    rol = (value or "").strip().upper()
    return "ADMIN" if rol == "ADMIN" else "USER"

def admin_users_base_query():
    admin_first = case((db.func.upper(Usuario.rol) == "ADMIN", 0), else_=1)
    return Usuario.query.order_by(admin_first, Usuario.nombre.asc())

def _user_notification_recipients(usuario: Usuario) -> list[str]:
    recipients = _parse_email_list(USER_CREATION_EMAIL)
    user_email = getattr(usuario, "correo", None)
    if user_email:
        recipients.extend(_parse_email_list(user_email))

    unique: list[str] = []
    seen: set[str] = set()
    for email in recipients:
        key = email.lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(email)
    return unique

def _send_user_created_email(usuario: Usuario, created_by: Usuario | None = None, initial_password: str = "") -> None:
    recipients = _user_notification_recipients(usuario)
    if not recipients:
        raise ValueError("No hay correo configurado para altas de usuarios.")

    created_at = now_cdmx_naive().strftime("%d/%m/%Y %H:%M")
    created_by_name = (_usuario_nombre_representante(created_by) or "Sistema").strip() or "Sistema"
    created_by_id = getattr(created_by, "id", None)
    ip = (request.headers.get("X-Forwarded-For") or request.remote_addr or "").split(",")[0].strip() or "No disponible"
    user_agent = (request.headers.get("User-Agent") or "No disponible").strip()

    rows = [
        ("ID", usuario.id),
        ("Usuario", usuario.nombre or ""),
        ("Nombre", _usuario_nombre_representante(usuario) or ""),
        ("Correo", usuario.correo or "No capturado"),
        ("Rol", (usuario.rol or "USER").upper()),
        ("Contraseña", initial_password or "No disponible"),
        ("Creado por", f"{created_by_name}" + (f" (ID {created_by_id})" if created_by_id else "")),
        ("Fecha de alta (CDMX)", created_at),
        ("IP de origen", ip),
        ("Navegador", user_agent),
    ]
    text_body = "Nuevo usuario creado en Sistema MAR\n\n" + "\n".join(f"{label}: {value}" for label, value in rows)
    html_rows = "".join(
        f"<tr><td style='padding:10px 12px;border:1px solid #dde3ea;background:#f8fafc;font-weight:700;color:#64748b;width:34%;'>{escape(str(label))}</td>"
        f"<td style='padding:10px 12px;border:1px solid #dde3ea;color:#111827;'>{escape(str(value))}</td></tr>"
        for label, value in rows
    )
    html_body = f"""
    <html>
      <body style="margin:0;padding:0;background:#eef2f7;font-family:Arial,Helvetica,sans-serif;color:#1f2937;">
        <div style="max-width:720px;margin:0 auto;padding:28px 16px;">
          <div style="background:#ffffff;border:1px solid #dbe4ef;border-radius:10px;overflow:hidden;">
            <div style="background:#0C3C78;color:#ffffff;padding:22px 26px;">
              <div style="font-size:12px;font-weight:700;letter-spacing:.9px;text-transform:uppercase;">MAR · Poliutech</div>
              <div style="font-size:23px;font-weight:800;margin-top:5px;">Nuevo usuario creado</div>
            </div>
            <div style="padding:24px;">
              <table style="border-collapse:collapse;width:100%;background:#ffffff;">{html_rows}</table>
              <p style="margin:16px 0 0 0;color:#64748b;font-size:12px;">Este mensaje fue generado automaticamente por Sistema MAR.</p>
            </div>
          </div>
        </div>
      </body>
    </html>
    """.strip()

    msg = EmailMessage()
    msg["Subject"] = f"Nuevo usuario MAR: {usuario.nombre or usuario.id}"
    msg["From"] = f"SISTEMA MAR <{SMTP_FROM or SMTP_USERNAME}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content(text_body)
    msg.add_alternative(html_body, subtype="html")

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg, to_addrs=recipients)

def _send_user_updated_email(
    usuario: Usuario,
    updated_by: Usuario | None = None,
    previous_nombre: str = "",
    previous_nombre_visible: str = "",
    previous_rol: str = "",
    previous_correo: str = "",
    new_password: str = "",
) -> None:
    recipients = _user_notification_recipients(usuario)
    if not recipients:
        raise ValueError("No hay correo configurado para cambios de usuarios.")

    updated_at = now_cdmx_naive().strftime("%d/%m/%Y %H:%M")
    updated_by_name = (_usuario_nombre_representante(updated_by) or "Sistema").strip() or "Sistema"
    updated_by_id = getattr(updated_by, "id", None)
    ip = (request.headers.get("X-Forwarded-For") or request.remote_addr or "").split(",")[0].strip() or "No disponible"
    user_agent = (request.headers.get("User-Agent") or "No disponible").strip()
    current_nombre = usuario.nombre or ""
    current_nombre_visible = _usuario_nombre_representante(usuario)
    current_correo = usuario.correo or ""
    current_rol = (usuario.rol or "USER").upper()
    previous_rol = (previous_rol or "USER").upper()

    changes: list[tuple[str, str, str]] = []
    if previous_nombre != current_nombre:
        changes.append(("Usuario", previous_nombre or "-", current_nombre or "-"))
    if (previous_nombre_visible or "") != current_nombre_visible:
        changes.append(("Nombre", previous_nombre_visible or "-", current_nombre_visible or "-"))
    if (previous_correo or "") != current_correo:
        changes.append(("Correo", previous_correo or "-", current_correo or "-"))
    if previous_rol != current_rol:
        changes.append(("Rol", previous_rol or "-", current_rol or "-"))
    if new_password:
        changes.append(("Contraseña", "Anterior no disponible", new_password))

    if not changes:
        return

    rows = [
        ("ID", usuario.id),
        ("Usuario actual", current_nombre),
        ("Correo actual", current_correo or "No capturado"),
        ("Rol actual", current_rol),
        ("Actualizado por", f"{updated_by_name}" + (f" (ID {updated_by_id})" if updated_by_id else "")),
        ("Fecha de cambio (CDMX)", updated_at),
        ("IP de origen", ip),
        ("Navegador", user_agent),
    ]
    changes_text = "\n".join(f"- {label}: {before} -> {after}" for label, before, after in changes)
    text_body = (
        "Cambio de usuario en Sistema MAR\n\n"
        + "\n".join(f"{label}: {value}" for label, value in rows)
        + "\n\nCambios realizados:\n"
        + changes_text
    )
    detail_rows = "".join(
        f"<tr><td style='padding:10px 12px;border:1px solid #dde3ea;background:#f8fafc;font-weight:700;color:#64748b;width:34%;'>{escape(str(label))}</td>"
        f"<td style='padding:10px 12px;border:1px solid #dde3ea;color:#111827;'>{escape(str(value))}</td></tr>"
        for label, value in rows
    )
    change_rows = "".join(
        f"<tr><td style='padding:10px 12px;border:1px solid #dde3ea;font-weight:700;'>{escape(label)}</td>"
        f"<td style='padding:10px 12px;border:1px solid #dde3ea;color:#64748b;'>{escape(before)}</td>"
        f"<td style='padding:10px 12px;border:1px solid #dde3ea;color:#111827;font-weight:700;'>{escape(after)}</td></tr>"
        for label, before, after in changes
    )
    html_body = f"""
    <html>
      <body style="margin:0;padding:0;background:#eef2f7;font-family:Arial,Helvetica,sans-serif;color:#1f2937;">
        <div style="max-width:760px;margin:0 auto;padding:28px 16px;">
          <div style="background:#ffffff;border:1px solid #dbe4ef;border-radius:10px;overflow:hidden;">
            <div style="background:#0C3C78;color:#ffffff;padding:22px 26px;">
              <div style="font-size:12px;font-weight:700;letter-spacing:.9px;text-transform:uppercase;">MAR · Poliutech</div>
              <div style="font-size:23px;font-weight:800;margin-top:5px;">Usuario actualizado</div>
            </div>
            <div style="padding:24px;">
              <table style="border-collapse:collapse;width:100%;background:#ffffff;margin-bottom:20px;">{detail_rows}</table>
              <div style="font-size:15px;font-weight:800;color:#0C3C78;margin-bottom:8px;">Cambios realizados</div>
              <table style="border-collapse:collapse;width:100%;background:#ffffff;">
                <thead>
                  <tr>
                    <th style="padding:10px 12px;border:1px solid #dde3ea;background:#f8fafc;text-align:left;">Campo</th>
                    <th style="padding:10px 12px;border:1px solid #dde3ea;background:#f8fafc;text-align:left;">Antes</th>
                    <th style="padding:10px 12px;border:1px solid #dde3ea;background:#f8fafc;text-align:left;">Ahora</th>
                  </tr>
                </thead>
                <tbody>{change_rows}</tbody>
              </table>
              <p style="margin:16px 0 0 0;color:#64748b;font-size:12px;">Este mensaje fue generado automaticamente por Sistema MAR.</p>
            </div>
          </div>
        </div>
      </body>
    </html>
    """.strip()

    msg = EmailMessage()
    msg["Subject"] = f"Usuario MAR actualizado: {current_nombre or usuario.id}"
    msg["From"] = f"SISTEMA MAR <{SMTP_FROM or SMTP_USERNAME}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content(text_body)
    msg.add_alternative(html_body, subtype="html")

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg, to_addrs=recipients)

def _usuario_nombre_representante(user: Usuario | None) -> str:
    if not user:
        return ""
    visible = (getattr(user, "nombre_visible", None) or "").strip()
    if visible:
        return visible
    nombre = (getattr(user, "nombre", "") or "").strip()
    return nombre

def responsable_actual() -> str:
    """
    Nombre visible del usuario autenticado para representante/autor.
    Si aun no se capturo, usa el usuario como respaldo.
    """
    if not getattr(current_user, "is_authenticated", False):
        return ""
    return _usuario_nombre_representante(current_user)

def require_owner_or_admin(cot: Cotizacion) -> None:
    if is_admin():
        return
    ra = responsable_actual()
    if not ra or (cot.responsable or "") != ra:
        abort(403)

def require_followup_author_or_admin(seg: CotizacionSeguimiento) -> None:
    if is_admin():
        return
    current_user_id = getattr(current_user, "id", None)
    if current_user_id and seg.usuario_id == current_user_id:
        return
    abort(403)

def require_cliente_owner_or_admin(cli: Cliente) -> None:
    if is_admin():
        return
    ra = responsable_actual()
    if not ra or (cli.responsable or "") != ra:
        abort(403)


def require_prospecto_owner_or_admin(prospecto: Prospecto) -> None:
    if is_admin():
        return
    ra = responsable_actual()
    if not ra or (prospecto.responsable or "") != ra:
        abort(403)


def require_prospecto_followup_author_or_admin(seg: ProspectoSeguimiento) -> None:
    if is_admin():
        return
    current_user_id = getattr(current_user, "id", None)
    if current_user_id and seg.usuario_id == current_user_id:
        return
    abort(403)


def require_registro_obra_owner_or_admin(registro: RegistroObra) -> None:
    if is_admin():
        return
    ra = responsable_actual()
    if not ra or (registro.responsable or "").strip().lower() != ra.strip().lower():
        abort(403)


def require_registro_obra_followup_author_or_admin(seg: RegistroObraSeguimiento) -> None:
    if is_admin():
        return
    current_user_id = getattr(current_user, "id", None)
    if current_user_id and seg.usuario_id == current_user_id:
        return
    abort(403)


def _clean_cliente_value(value) -> str:
    return str(value or "").strip()


def _find_cliente_for_seguimiento(
    *,
    nombre: str = "",
    empresa: str = "",
    correo: str = "",
    telefono: str = "",
) -> Optional[Cliente]:
    correo = _clean_cliente_value(correo)
    telefono = _clean_cliente_value(telefono)
    nombre = _clean_cliente_value(nombre)
    empresa = _clean_cliente_value(empresa)

    if correo:
        cliente = Cliente.query.filter(db.func.lower(Cliente.correo) == correo.lower()).first()
        if cliente:
            return cliente
    if telefono:
        cliente = Cliente.query.filter(Cliente.telefono == telefono).first()
        if cliente:
            return cliente
    if nombre:
        query = Cliente.query.filter(db.func.lower(Cliente.nombre_cliente) == nombre.lower())
        if empresa:
            query = query.filter(db.func.lower(db.func.coalesce(Cliente.empresa, "")) == empresa.lower())
        cliente = query.first()
        if cliente:
            return cliente
    return None


def _cliente_seguimiento_payload(
    *,
    cliente: Optional[Cliente] = None,
    nombre: str = "",
    empresa: str = "",
    responsable: str = "",
    correo: str = "",
    telefono: str = "",
    direccion: str = "",
    rfc: str = "",
    titulo: str = "Datos del cliente",
    extras: Optional[list[dict]] = None,
) -> dict:
    if cliente is None:
        cliente = _find_cliente_for_seguimiento(
            nombre=nombre,
            empresa=empresa,
            correo=correo,
            telefono=telefono,
        )

    base = {
        "titulo": titulo,
        "nombre_cliente": _clean_cliente_value(getattr(cliente, "nombre_cliente", "")) if cliente else "",
        "empresa": _clean_cliente_value(getattr(cliente, "empresa", "")) if cliente else "",
        "responsable": _clean_cliente_value(getattr(cliente, "responsable", "")) if cliente else "",
        "correo": _clean_cliente_value(getattr(cliente, "correo", "")) if cliente else "",
        "telefono": _clean_cliente_value(getattr(cliente, "telefono", "")) if cliente else "",
        "direccion": _clean_cliente_value(getattr(cliente, "direccion", "")) if cliente else "",
        "rfc": _clean_cliente_value(getattr(cliente, "rfc", "")) if cliente else "",
        "extras": extras or [],
    }
    fallbacks = {
        "nombre_cliente": nombre,
        "empresa": empresa,
        "responsable": responsable,
        "correo": correo,
        "telefono": telefono,
        "direccion": direccion,
        "rfc": rfc,
    }
    for key, value in fallbacks.items():
        value = _clean_cliente_value(value)
        if value:
            base[key] = value
    return base



def _build_dashboard_cotizaciones_query(
    *,
    desde: str = "",
    hasta: str = "",
    estatus: str = "",
    cliente: str = "",
    especialidad: str = "",
):
    q = Cotizacion.query.outerjoin(Cliente, Cotizacion.cliente_id == Cliente.id)
    q = q.filter(Cotizacion.eliminada_en.is_(None))

    if not is_admin():
        q = q.filter(Cotizacion.responsable == responsable_actual())

    if desde:
        try:
            d = datetime.strptime(desde, "%Y-%m-%d")
        except ValueError as exc:
            raise ValueError("Filtro 'Desde' invalido") from exc
        q = q.filter(Cotizacion.fecha >= d)

    if hasta:
        try:
            h = datetime.strptime(hasta, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        except ValueError as exc:
            raise ValueError("Filtro 'Hasta' invalido") from exc
        q = q.filter(Cotizacion.fecha <= h)

    if estatus:
        q = q.filter(Cotizacion.estatus == estatus)

    especialidad = (especialidad or "").strip().lower()
    if especialidad:
        q = q.filter(
            db.func.lower(db.func.coalesce(Cotizacion.especialidad, "")).like(f"%{especialidad}%")
        )

    cliente = (cliente or "").strip().lower()
    if cliente:
        pattern = f"%{cliente}%"
        q = q.filter(or_(
            db.func.lower(db.func.coalesce(Cliente.nombre_cliente, "")).like(pattern),
            db.func.lower(db.func.coalesce(Cliente.empresa, "")).like(pattern),
            db.func.lower(db.func.coalesce(Cotizacion.proyecto, "")).like(pattern),
            db.func.lower(db.func.coalesce(Cotizacion.especialidad, "")).like(pattern),
        ))

    return q

def _cotizaciones_base_query():
    q = Cotizacion.query.outerjoin(Cliente, Cotizacion.cliente_id == Cliente.id)
    q = q.filter(Cotizacion.eliminada_en.is_(None))
    if not is_admin():
        q = q.filter(Cotizacion.responsable == responsable_actual())
    return q

def _cotizaciones_activas_query():
    return Cotizacion.query.filter(Cotizacion.eliminada_en.is_(None))

def _cotizacion_activa_or_404(cot_id: int) -> Cotizacion:
    return _cotizaciones_activas_query().filter(Cotizacion.id == cot_id).first_or_404()

def _soft_delete_cotizacion(cot: Cotizacion) -> None:
    now = now_cdmx_naive()
    cot.eliminada_en = now
    cot.eliminada_por = responsable_actual() or "Sistema"
    cot.eliminacion_definitiva_en = now + timedelta(days=COTIZACION_TRASH_RETENTION_DAYS)

def _restore_cotizacion(cot: Cotizacion) -> None:
    cot.eliminada_en = None
    cot.eliminada_por = None
    cot.eliminacion_definitiva_en = None

def _purge_expired_cotizacion_trash() -> int:
    cutoff = now_cdmx_naive() - timedelta(days=COTIZACION_TRASH_RETENTION_DAYS)
    expired = Cotizacion.query.filter(Cotizacion.eliminada_en.isnot(None), Cotizacion.eliminada_en <= cutoff).all()
    for cot in expired:
        for detalle in list(cot.detalles):
            db.session.delete(detalle)
        db.session.delete(cot)
    if expired:
        db.session.commit()
    return len(expired)

with app.app_context():
    try:
        purged = _purge_expired_cotizacion_trash()
        if purged:
            print(f"🧹 Cotizaciones purgadas de papelera: {purged}")
    except Exception as e:
        print("⚠️ purge cotizaciones papelera:", e)

def _project_label_expr():
    return db.func.coalesce(
        db.func.nullif(db.func.trim(Cotizacion.proyecto), ""),
        "Sin proyecto",
    )

def _project_key_expr():
    return db.func.coalesce(
        db.func.nullif(db.func.lower(db.func.trim(Cotizacion.proyecto)), ""),
        "sin proyecto",
    )

def _project_display_expr():
    return db.func.coalesce(
        db.func.nullif(db.func.min(db.func.nullif(db.func.trim(Cotizacion.proyecto), "")), ""),
        "Sin proyecto",
    )

def _known_project_names(limit: int = 100) -> list[str]:
    names: dict[str, str] = {}

    def add(value: str | None) -> None:
        value = (value or "").strip()
        if not value or value.lower() == "sin proyecto":
            return
        names.setdefault(value.lower(), value)

    key_expr = _project_key_expr()
    name_expr = _project_display_expr()
    for row in (
        _cotizaciones_base_query()
        .with_entities(key_expr.label("key"), name_expr.label("proyecto"))
        .group_by(key_expr)
        .order_by(name_expr.asc())
        .limit(limit)
        .all()
    ):
        add(row.proyecto)

    for model in (ComprobacionGasto, SolicitudRecurso, MovimientoFinanciero):
        try:
            rows = (
                db.session.query(model.proyecto)
                .filter(model.proyecto.isnot(None), db.func.trim(model.proyecto) != "")
                .distinct()
                .limit(limit)
                .all()
            )
            for row in rows:
                add(row[0])
        except Exception:
            continue

    return sorted(names.values(), key=str.lower)[:limit]

def generar_folio() -> str:
    prefix = "PTCH-"
    maxn = 0
    rows = db.session.execute(text("SELECT folio FROM cotizacion WHERE folio LIKE 'PTCH-%'")).fetchall()
    for (folio,) in rows:
        m = re.match(r"PTCH-(\d{4})$", str(folio))
        if m:
            n = int(m.group(1))
            maxn = max(maxn, n)
    for i in range(1, 11):
        cand = f"{prefix}{maxn+i:04d}"
        exists = db.session.execute(text("SELECT 1 FROM cotizacion WHERE folio=:f LIMIT 1"), {"f": cand}).fetchone()
        if not exists:
            return cand
    return f"{prefix}{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

def fmt(n: float) -> float:
    try:
        return round(float(n or 0), 2)
    except Exception:
        return 0.0

def parse_float(v, default=0.0) -> float:
    try:
        if v is None or v == "":
            return default
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace("$", "").replace(",", "").strip()
        return float(s) if s else default
    except Exception:
        return default


def parse_int(v, default=0):
    try:
        if v is None or v == "":
            return default
        if isinstance(v, int):
            return v
        if isinstance(v, float):
            return int(v)
        s = str(v).strip()
        return int(float(s)) if s else default
    except Exception:
        return default


def _safe_detalle_kwargs(**kwargs):
    valid = set(getattr(CotizacionDetalle, "__table__").columns.keys())
    return {k: v for k, v in kwargs.items() if k in valid}


def _truncate_pdf_text(value, limit=90):
    text = str(value or "").strip()
    if len(text) <= limit:
        return text
    return text[: max(0, limit - 3)].rstrip() + "..."


def parse_datetime_flexible(v) -> Optional[datetime]:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v
    raw = str(v).strip()
    if not raw:
        return None
    candidates = [raw, raw.replace("Z", "+00:00"), raw + " 00:00:00"]
    formats = [
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
    ]
    for cand in candidates:
        try:
            return datetime.fromisoformat(cand)
        except Exception:
            pass
        for fmt_s in formats:
            try:
                return datetime.strptime(cand, fmt_s)
            except Exception:
                continue
    return None


def _append_note(base: Optional[str], extra: Optional[str]) -> Optional[str]:
    b = (base or "").strip()
    e = (extra or "").strip()
    if not e:
        return b or None
    return f"{b}\n{e}".strip() if b else e


def sample_import_payload() -> dict:
    return {
        "folio": "COT-2026-02-026-2",
        "fecha": "2026-02-26",
        "estatus": "EN REVISIÓN",
        "responsable": responsable_actual() or "",
        "cliente": {
            "nombre_cliente": "Ing. Adriana Vazquez / Ing. Karla Reyes",
            "empresa": "GIA",
            "correo": "",
            "telefono": "",
            "direccion": "Oracle, Guadalajara",
            "rfc": ""
        },
        "zona": "",
        "iva_porc": 16,
        "notas": "Importada desde cotizacion externa.\nVigencia de la cotizacion: 30 dias.\nAnticipo: 50%.\nEl precio se respeta siempre que se haga el trabajo total en aplicacion continua.\nEl precio no respeta siempre que las areas no sean continuas.\nSe requiere muestreo de tablero de 150 cm a 150 cm por ejecucion que impide la instalacion del sistema.\nEsperando contar con su preferencia me despido y quedo a sus apreciables ordenes.",
        "items": [
            {
                "nombre_concepto": "Suministro y aplicacion de sistema impermeable de curado rapido sobre superficie de concreto",
                "unidad": "m2",
                "cantidad": 880,
                "precio_unitario": 1907.69,
                "sistema": "TREMPROOF JARDIN",
                "descripcion": "Incluye: preparacion de superficie por medios manual mecanicos hasta alcanzar perfil de anclaje; limpieza y sello de juntas con sellador de poliuretano flexible; aplicacion de Tremproof 250 GC; aplicacion de Vapor Barrier; trazo, corte y colocacion de Eucodrain H15P Geotextil; incluye material, equipos, herramienta y personal altamente especializado."
            }
        ]
    }


def _normalize_text_for_match(value: str) -> str:
    raw = str(value or "")
    normalized = unicodedata.normalize("NFKD", raw)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).lower().strip()


def _clean_pdf_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _extract_pdf_text_and_tables(pdf_bytes: bytes) -> tuple[str, list[list[list[str]]]]:
    try:
        import pdfplumber
    except Exception as e:
        raise ValueError("El servidor no tiene habilitada la lectura de PDFs. Instala las dependencias del proyecto.") from e

    text_parts: list[str] = []
    tables: list[list[list[str]]] = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if page_text.strip():
                text_parts.append(page_text)
            for table in page.extract_tables() or []:
                normalized_rows = []
                for row in table or []:
                    cells = [_clean_pdf_text(cell) for cell in (row or [])]
                    if any(cells):
                        normalized_rows.append(cells)
                if normalized_rows:
                    tables.append(normalized_rows)

    full_text = "\n".join(text_parts).strip()
    if not full_text:
        raise ValueError("No se pudo extraer texto legible del PDF.")
    return full_text, tables


def _extract_prefixed_line(text: str, prefix: str) -> str:
    prefix_norm = _normalize_text_for_match(prefix)
    for line in text.splitlines():
        clean = line.strip()
        if not clean:
            continue
        norm = _normalize_text_for_match(clean)
        if norm.startswith(prefix_norm):
            parts = clean.split(":", 1)
            return parts[1].strip() if len(parts) > 1 else clean
    return ""


def _parse_spanish_date_from_pdf(text: str) -> Optional[datetime]:
    match = re.search(r"Ciudad de Mexico a\s+(\d{1,2})\s+de\s+([a-z]+)\s+de\s+(\d{4})", _normalize_text_for_match(text), re.IGNORECASE)
    if not match:
        return None

    months = {
        "enero": 1,
        "febrero": 2,
        "marzo": 3,
        "abril": 4,
        "mayo": 5,
        "junio": 6,
        "julio": 7,
        "agosto": 8,
        "septiembre": 9,
        "setiembre": 9,
        "octubre": 10,
        "noviembre": 11,
        "diciembre": 12,
    }
    day = int(match.group(1))
    month_name = _normalize_text_for_match(match.group(2))
    month = months.get(month_name)
    year = int(match.group(3))
    if not month:
        return None
    return datetime(year, month, day)


def _parse_pdf_currency(value: str) -> float:
    match = re.search(r"([\d,]+\.\d{1,2})", str(value or ""))
    return parse_float(match.group(1), 0.0) if match else 0.0


def _parse_pdf_quantity_and_unit(value: str) -> tuple[float, str]:
    raw = str(value or "").replace(",", "")
    match = re.search(r"([\d.]+)\s*([A-Za-z0-9/]+)?", raw)
    if not match:
        return 0.0, ""
    quantity = parse_float(match.group(1), 0.0)
    unit = (match.group(2) or "").strip()
    unit = unit.replace("?", "2")
    return quantity, unit.lower()


def _build_concept_name(system: str, description: str) -> str:
    base = re.split(r"incluye\s*:", description, maxsplit=1, flags=re.IGNORECASE)[0].strip(" .;:-")
    if len(base) >= 12:
        return base[:220]
    return (system or description or "Concepto importado")[:220]


def _extract_items_from_sistema_descripcion_pdf_text(text: str) -> list[dict]:
    lines = [_clean_pdf_text(line) for line in (text or "").splitlines() if _clean_pdf_text(line)]
    if not lines:
        return []

    def is_header_or_footer(value: str) -> bool:
        norm = _normalize_text_for_match(value)
        return any(
            norm.startswith(prefix)
            for prefix in (
                "folio:",
                "campos eliseos",
                "telefonos",
                "www.poliutech.com",
                "empresa 100% mexicana",
                "ciudad de mexico a",
                "atte.",
                "ing.",
                "director general",
                "sistema descripcion unidad cantidad p. unitario importe",
                "condiciones comerciales",
            )
        )

    def is_unit_line(value: str) -> bool:
        return _normalize_text_for_match(value) in {"m2", "m 2", "m?"}

    def is_numeric_line(value: str) -> bool:
        return bool(re.fullmatch(r"[\d,.]+", value.strip()))

    def is_money_line(value: str) -> bool:
        return bool(re.fullmatch(r"\$\s*[\d,]+\.\d{1,2}", value.strip()))

    def parse_inline_values(value: str):
        match = re.search(r"(?i)(m2|m?|m\s*2)\s+([\d,.]+)\s+\$\s*([\d,]+\.\d{1,2})\s+\$\s*([\d,]+\.\d{2})", value)
        if not match:
            return None
        return {
            "unidad": "m2",
            "cantidad": parse_float(match.group(2), 0.0),
            "precio_unitario": parse_float(match.group(3), 0.0),
            "subtotal_pdf": parse_float(match.group(4), 0.0),
        }

    def is_system_like(value: str) -> bool:
        s = value.strip()
        if len(s) > 40:
            return False
        letters = [ch for ch in s if ch.isalpha()]
        if not letters:
            return False
        upper_ratio = sum(1 for ch in letters if ch.isupper()) / max(len(letters), 1)
        return upper_ratio >= 0.75

    start_idx = 0
    for idx, line in enumerate(lines):
        norm = _normalize_text_for_match(line)
        if norm.startswith("sistema descripcion unidad cantidad"):
            start_idx = idx + 1
            break

    items = []
    chunk = []
    i = start_idx
    while i < len(lines):
        line = lines[i]
        norm = _normalize_text_for_match(line)
        if norm.startswith(("subtotal", "iva", "total", "condiciones comerciales")):
            break
        if is_header_or_footer(line):
            i += 1
            continue

        parsed = parse_inline_values(line)
        if parsed is None and is_unit_line(line):
            unidad = "m2"
            j = i + 1
            while j < len(lines) and is_header_or_footer(lines[j]):
                j += 1
            if j < len(lines) and is_numeric_line(lines[j]):
                cantidad = parse_float(lines[j], 0.0)
                j += 1
                while j < len(lines) and is_header_or_footer(lines[j]):
                    j += 1
                if j < len(lines) and is_money_line(lines[j]):
                    precio = parse_float(lines[j], 0.0)
                    j += 1
                    while j < len(lines) and is_header_or_footer(lines[j]):
                        j += 1
                    if j < len(lines) and is_money_line(lines[j]):
                        subtotal = parse_float(lines[j], 0.0)
                        parsed = {
                            "unidad": unidad,
                            "cantidad": cantidad,
                            "precio_unitario": precio,
                            "subtotal_pdf": subtotal,
                        }
                        i = j
        if parsed is not None:
            parts = [part for part in chunk if not is_header_or_footer(part)]
            system_lines = []
            description_lines = []
            seen_description = False
            for part in parts:
                if not seen_description and is_system_like(part):
                    system_lines.append(part)
                else:
                    seen_description = True
                    description_lines.append(part)
            if not description_lines and system_lines:
                description_lines = system_lines[:]
                system_lines = []
            descripcion = " ".join(description_lines).strip()
            sistema = " ".join(system_lines).strip() or None
            if descripcion and parsed["cantidad"] > 0 and parsed["precio_unitario"] > 0:
                items.append({
                    "nombre_concepto": descripcion,
                    "unidad": parsed["unidad"],
                    "cantidad": parsed["cantidad"],
                    "precio_unitario": parsed["precio_unitario"],
                    "sistema": sistema,
                    "descripcion": descripcion,
                    "subtotal_pdf": parsed["subtotal_pdf"],
                })
            chunk = []
            i += 1
            continue

        chunk.append(line)
        i += 1

    return items


def _extract_items_from_pdf_tables(tables: list[list[list[str]]]) -> list[dict]:
    def find_column_indexes(header_cells: list[str], aliases: dict[str, tuple[str, ...]]) -> dict[str, int]:
        normalized = [_normalize_text_for_match(cell) for cell in header_cells]
        indexes: dict[str, int] = {}
        for field, options in aliases.items():
            for idx, cell in enumerate(normalized):
                if any(option in cell for option in options):
                    indexes[field] = idx
                    break
        return indexes

    def get_cell(cells: list[str], index_map: dict[str, int], field: str) -> str:
        idx = index_map.get(field)
        if idx is None or idx >= len(cells):
            return ""
        return cells[idx]

    def extract_money_values(cells: list[str]) -> list[float]:
        values = []
        for cell in cells:
            cell_text = str(cell or "")
            if "$" not in cell_text:
                continue
            amount = _parse_pdf_currency(cell_text)
            if amount > 0:
                values.append(amount)
        return values

    def append_continuation(target: dict, extra_text: str) -> None:
        extra = _clean_pdf_text(extra_text)
        if not extra:
            return
        current_name = _clean_pdf_text(target.get("nombre_concepto") or "")
        current_desc = _clean_pdf_text(target.get("descripcion") or current_name)
        target["nombre_concepto"] = _clean_pdf_text(f"{current_name} {extra}")
        target["descripcion"] = _clean_pdf_text(f"{current_desc} {extra}")

    aliases_variants = [
        {
            "concepto": ("concepto", "descripcion del trabajo", "descripcion"),
            "unidad": ("unidad", "uni.", "area / unidad", "area", "?rea / unidad"),
            "cantidad": ("cantidad",),
            "precio_unitario": ("p.u.", "p. unitario", "precio unitario", "p unitario"),
            "importe": ("importe", "subtotal"),
            "sistema": ("sistema",),
            "codigo": ("codigo", "c?digo"),
        },
    ]

    items: list[dict] = []
    active_header_map: dict[str, int] | None = None

    for table in tables:
        if not table:
            continue

        header_row = None
        header_map = None
        for row in table[:4]:
            cells = [_clean_pdf_text(cell) for cell in row]
            if len(cells) < 4:
                continue
            for aliases in aliases_variants:
                indexes = find_column_indexes(cells, aliases)
                has_amounts = "precio_unitario" in indexes and "importe" in indexes
                has_shape = (
                    ("concepto" in indexes or "sistema" in indexes)
                    and has_amounts
                    and ("cantidad" in indexes or "unidad" in indexes)
                )
                if has_shape:
                    header_row = row
                    header_map = indexes
                    break
            if header_map:
                break

        if header_map:
            active_header_map = header_map
            start_index = table.index(header_row) + 1
        elif active_header_map:
            header_map = active_header_map
            start_index = 0
        else:
            continue

        for row in table[start_index:]:
            cells = [_clean_pdf_text(cell) for cell in row]
            if not any(cells):
                continue

            row_norm = _normalize_text_for_match(" ".join(cells))
            if row_norm.startswith(("subtotal", "iva", "total", "condiciones comerciales")):
                break

            concepto = get_cell(cells, header_map, "concepto")
            sistema = get_cell(cells, header_map, "sistema")
            unidad_cell = get_cell(cells, header_map, "unidad")
            cantidad_cell = get_cell(cells, header_map, "cantidad")
            precio_cell = get_cell(cells, header_map, "precio_unitario")
            importe_cell = get_cell(cells, header_map, "importe")

            cantidad = parse_float(cantidad_cell, 0.0) if cantidad_cell else 0.0
            unidad = ""
            if unidad_cell:
                parsed_qty, parsed_unit = _parse_pdf_quantity_and_unit(unidad_cell)
                if cantidad <= 0 and parsed_qty > 0:
                    cantidad = parsed_qty
                unidad = parsed_unit or unidad_cell.strip()

            money_values = extract_money_values(cells)
            precio_unitario = _parse_pdf_currency(precio_cell)
            subtotal_pdf = _parse_pdf_currency(importe_cell)
            if precio_unitario <= 0 and money_values:
                precio_unitario = money_values[0]
            if subtotal_pdf <= 0 and len(money_values) >= 2:
                subtotal_pdf = money_values[-1]

            if cantidad <= 0 or precio_unitario <= 0:
                continuation_bits = []
                for idx, cell in enumerate(cells):
                    if not cell:
                        continue
                    if idx == header_map.get("codigo"):
                        continue
                    if idx == header_map.get("unidad"):
                        continue
                    if idx == header_map.get("cantidad"):
                        continue
                    if idx == header_map.get("precio_unitario"):
                        continue
                    if idx == header_map.get("importe"):
                        continue
                    continuation_bits.append(cell)
                continuation_text = " ".join(bit for bit in continuation_bits if bit)
                if items and continuation_text and not row_norm.startswith(("subtotal", "iva", "total")):
                    append_continuation(items[-1], continuation_text)
                continue

            if not concepto and not sistema:
                continue

            descripcion = concepto or sistema
            items.append({
                "nombre_concepto": concepto or _build_concept_name(sistema, descripcion),
                "unidad": unidad or "m2",
                "cantidad": cantidad,
                "precio_unitario": precio_unitario,
                "sistema": sistema or None,
                "descripcion": descripcion,
                "subtotal_pdf": subtotal_pdf if subtotal_pdf > 0 else None,
            })

    return items


def _looks_like_partida_numbers_as_quantity(items: list[dict]) -> bool:
    if not items or len(items) < 2:
        return False
    quantities = []
    for item in items:
        try:
            quantities.append(float(item.get("cantidad") or 0))
        except Exception:
            return False
    expected = [float(i) for i in range(1, len(quantities) + 1)]
    return quantities == expected


def _extract_items_from_pdf_text(text: str) -> list[dict]:
    lines = [_clean_pdf_text(line) for line in (text or "").splitlines() if _clean_pdf_text(line)]
    if not lines:
        return []

    def is_code_line(value: str) -> bool:
        return bool(re.fullmatch(r"\d{2}", value.strip()))

    def is_unit_line(value: str) -> bool:
        return _normalize_text_for_match(value) in {"m2", "m 2", "m?"}

    def is_money_line(value: str) -> bool:
        return bool(re.fullmatch(r"\$\s*[\d,]+\.\d{2}", value.strip()))

    def is_numeric_line(value: str) -> bool:
        return bool(re.fullmatch(r"[\d,.]+", value.strip()))

    def is_header_or_footer(value: str) -> bool:
        norm = _normalize_text_for_match(value)
        return any(
            norm.startswith(prefix)
            for prefix in (
                "campos eliseos",
                "telefonos",
                "www.poliutech.com",
                "empresa 100% mexicana",
                "ciudad de mexico a",
                "atte.",
                "ing.",
                "director general",
                "codigo concepto unidad cantidad p.u. importe",
            )
        )

    start_idx = 0
    for idx, line in enumerate(lines):
        if "codigo" in _normalize_text_for_match(line) and "importe" in _normalize_text_for_match(line):
            start_idx = idx + 1
            break

    items: list[dict] = []
    i = start_idx
    while i < len(lines):
        line = lines[i]
        norm = _normalize_text_for_match(line)
        if norm.startswith("subtotal") or norm.startswith("iva") or norm.startswith("total"):
            break
        if is_header_or_footer(line):
            i += 1
            continue
        if not is_code_line(line):
            i += 1
            continue

        i += 1
        desc_lines: list[str] = []
        while i < len(lines):
            current = lines[i]
            if is_header_or_footer(current):
                i += 1
                continue
            if is_unit_line(current):
                i += 1
                break
            desc_lines.append(current)
            i += 1

        while i < len(lines) and (is_header_or_footer(lines[i]) or not is_numeric_line(lines[i])):
            if is_code_line(lines[i]) or _normalize_text_for_match(lines[i]).startswith(("subtotal", "iva", "total")):
                break
            i += 1
        if i >= len(lines) or not is_numeric_line(lines[i]):
            break
        quantity = parse_float(lines[i], 0.0)
        i += 1

        while i < len(lines) and not is_money_line(lines[i]):
            if is_header_or_footer(lines[i]):
                i += 1
                continue
            break
        if i >= len(lines) or not is_money_line(lines[i]):
            break
        unit_price = _parse_pdf_currency(lines[i])
        i += 1

        while i < len(lines) and not is_money_line(lines[i]):
            if is_header_or_footer(lines[i]):
                i += 1
                continue
            break
        if i >= len(lines) or not is_money_line(lines[i]):
            break
        line_subtotal = _parse_pdf_currency(lines[i])
        i += 1

        continuation: list[str] = []
        while i < len(lines):
            current = lines[i]
            current_norm = _normalize_text_for_match(current)
            if current_norm.startswith(("subtotal", "iva", "total")) or is_code_line(current):
                break
            if is_header_or_footer(current) or is_money_line(current) or is_numeric_line(current) or is_unit_line(current):
                i += 1
                continue
            continuation.append(current)
            i += 1

        description = " ".join(desc_lines + continuation).strip()
        items.append({
            "nombre_concepto": _build_concept_name("", description),
            "unidad": "m2",
            "cantidad": quantity,
            "precio_unitario": unit_price,
            "sistema": None,
            "descripcion": description,
            "subtotal_pdf": line_subtotal,
        })

    return items


def _extract_items_from_pdf_block_regex(text: str) -> list[dict]:
    compact = re.sub(r"\n+", "\n", text or "")
    pattern = re.compile(
        r"(?ms)^\s*(?P<codigo>\d{2})\s+"
        r"(?P<descripcion>.*?)\s+"
        r"(?P<unidad>M2|M\s*2|M?)\s+"
        r"(?P<cantidad>[\d,.]+)\s+"
        r"\$(?P<precio>[\d,]+\.\d{2})\s+"
        r"\$(?P<subtotal>[\d,]+\.\d{2})\s*"
        r"(?=\d{2}\s+|Subtotal\s+\$|IVA\s+\d|Total\s+\$|$)",
    )
    items: list[dict] = []
    for match in pattern.finditer(compact):
        description = _clean_pdf_text(match.group("descripcion"))
        quantity = parse_float(match.group("cantidad"), 0.0)
        unit_price = parse_float(match.group("precio"), 0.0)
        line_subtotal = parse_float(match.group("subtotal"), 0.0)
        if quantity <= 0 or unit_price <= 0 or line_subtotal <= 0:
            continue
        items.append({
            "nombre_concepto": _build_concept_name("", description),
            "unidad": "m2",
            "cantidad": quantity,
            "precio_unitario": unit_price,
            "sistema": None,
            "descripcion": description,
            "subtotal_pdf": line_subtotal,
        })
    return items


def _extract_conditions_from_pdf(text: str) -> str:
    match = re.search(r"CONDICIONES COMERCIALES\s*:(.*?)(?:Esperando contar con su preferencia|Atte\.|Ing\.)", text, re.IGNORECASE | re.DOTALL)
    if not match:
        return ""
    lines = []
    for raw in match.group(1).splitlines():
        clean = raw.strip().lstrip("-?* ").strip()
        if clean:
            lines.append(clean)
    return "\n".join(lines)


def build_import_payload_from_pdf(pdf_bytes: bytes, filename: str, responsable_hint: Optional[str] = None) -> dict:
    text, tables = _extract_pdf_text_and_tables(pdf_bytes)
    normalized_text = _normalize_text_for_match(text)

    # Regla principal: si pdfplumber detecta una tabla con encabezados reconocibles,
    # se respeta el mapeo directo de columnas y no se intenta adivinar.
    items = _extract_items_from_pdf_tables(tables)

    # Fallbacks solo cuando no hubo tabla reconocible.
    if not items:
        if "sistema descripcion unidad cantidad" in normalized_text and "p. unitario" in normalized_text:
            items = _extract_items_from_sistema_descripcion_pdf_text(text)
        elif "codigo" in normalized_text and "cantidad" in normalized_text and "importe" in normalized_text:
            items = _extract_items_from_pdf_text(text)
            if not items:
                items = _extract_items_from_pdf_block_regex(text)
        else:
            items = _extract_items_from_pdf_text(text)
            if not items:
                items = _extract_items_from_pdf_block_regex(text)

    if _looks_like_partida_numbers_as_quantity(items):
        text_items = []
        if "sistema descripcion unidad cantidad" in normalized_text and "p. unitario" in normalized_text:
            text_items = _extract_items_from_sistema_descripcion_pdf_text(text)
        elif "codigo" in normalized_text and "cantidad" in normalized_text and "importe" in normalized_text:
            text_items = _extract_items_from_pdf_text(text)
            if not text_items:
                text_items = _extract_items_from_pdf_block_regex(text)
        if text_items and not _looks_like_partida_numbers_as_quantity(text_items):
            items = text_items

    if not items:
        raise ValueError("No pude identificar conceptos importables dentro del PDF.")

    folio_match = re.search(r"Folio\s*:\s*([A-Z0-9\-]+)", text, re.IGNORECASE)
    folio = folio_match.group(1).strip() if folio_match else None

    fecha = _parse_spanish_date_from_pdf(text) or now_cdmx_naive()
    contacto = _extract_prefixed_line(text, "Con atencion a")
    empresa = _extract_prefixed_line(text, "Empresa")

    ubicacion = ""
    location_match = re.search(r"se realizaran\s+en\s+(.+?)(?:\.|\n)", _normalize_text_for_match(text), re.IGNORECASE)
    if location_match:
        ubicacion = _clean_pdf_text(location_match.group(1))

    iva_porc = 16.0
    iva_pct_match = re.search(r"IVA\s*(\d+(?:\.\d+)?)\s*%", text, re.IGNORECASE)
    if iva_pct_match:
        iva_porc = parse_float(iva_pct_match.group(1), 16.0)

    notas = "Importada desde PDF externo."
    conditions = _extract_conditions_from_pdf(text)
    if conditions:
        notas = _append_note(notas, conditions)

    total_match = re.search(r"Total\s*\$?\s*([\d,]+\.\d{2})", text, re.IGNORECASE)
    if total_match:
        notas = _append_note(notas, f"Total detectado en PDF: ${parse_float(total_match.group(1), 0.0):,.2f}")

    cliente_nombre = contacto or empresa or Path(filename).stem[:120]
    return {
        "folio": folio,
        "fecha": fecha.isoformat(sep=" "),
        "estatus": "PENDIENTE",
        "responsable": responsable_hint or "",
        "cliente": {
            "nombre_cliente": cliente_nombre,
            "empresa": empresa or None,
            "correo": None,
            "telefono": None,
            "direccion": ubicacion or None,
            "rfc": None,
        },
        "zona": "",
        "iva_porc": iva_porc,
        "notas": notas,
        "items": items,
    }


def _normalize_import_payload(payload: dict) -> dict:
    if not isinstance(payload, dict):
        raise ValueError("El JSON debe ser un objeto.")

    cliente_in = payload.get("cliente") or {}
    if not isinstance(cliente_in, dict):
        raise ValueError("'cliente' debe ser un objeto.")

    items_in = payload.get("items") or payload.get("conceptos") or payload.get("detalles") or []
    if not isinstance(items_in, list) or not items_in:
        raise ValueError("Debes enviar al menos un concepto en 'items'.")

    cliente = {
        "nombre_cliente": (cliente_in.get("nombre_cliente") or cliente_in.get("cliente") or payload.get("cliente_nombre") or payload.get("cliente") or "").strip(),
        "empresa": (cliente_in.get("empresa") or payload.get("empresa") or "").strip() or None,
        "correo": (cliente_in.get("correo") or payload.get("correo") or "").strip() or None,
        "telefono": (cliente_in.get("telefono") or payload.get("telefono") or "").strip() or None,
        "direccion": (cliente_in.get("direccion") or payload.get("direccion") or "").strip() or None,
        "rfc": (cliente_in.get("rfc") or payload.get("rfc") or "").strip() or None,
    }
    if not cliente["nombre_cliente"]:
        raise ValueError("Falta 'cliente.nombre_cliente'.")

    normalized_items = []
    for idx, item in enumerate(items_in, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"El concepto #{idx} debe ser un objeto.")
        nombre = (item.get("nombre_concepto") or item.get("concepto") or item.get("nombre") or "").strip()
        if not nombre:
            raise ValueError(f"El concepto #{idx} no tiene nombre.")
        normalized_items.append({
            "nombre_concepto": nombre,
            "unidad": (item.get("unidad") or "").strip(),
            "cantidad": parse_float(item.get("cantidad"), 1.0),
            "precio_unitario": parse_float(item.get("precio_unitario", item.get("precio")), 0.0),
            "sistema": (item.get("sistema") or "").strip() or None,
            "descripcion": (item.get("descripcion") or "").strip(),
            "subtotal_pdf": parse_float(item.get("subtotal_pdf", item.get("importe")), 0.0),
        })

    raw_estatus = (payload.get("estatus") or "").strip().upper()
    raw_aprobacion = (payload.get("estatus_aprobacion") or "").strip().upper()
    if raw_estatus in {"APROBADO", "APROBADA", "AUTORIZADO"}:
        raw_aprobacion = "APROBADA"
        raw_estatus = "PENDIENTE"
    elif raw_estatus in {"RECHAZADO", "RECHAZADA"}:
        raw_aprobacion = "RECHAZADA"
        raw_estatus = "PENDIENTE"
    elif raw_estatus in {"EN REVISION", "EN REVISIÓN"}:
        raw_aprobacion = "EN REVISIÓN"
        raw_estatus = "PENDIENTE"
    if raw_aprobacion == "APROBADO" or raw_aprobacion == "AUTORIZADO":
        raw_aprobacion = "APROBADA"
    elif raw_aprobacion == "RECHAZADO":
        raw_aprobacion = "RECHAZADA"
    elif raw_aprobacion == "EN REVISION":
        raw_aprobacion = "EN REVISIÓN"

    return {
        "folio": (payload.get("folio") or payload.get("folio_externo") or "").strip() or None,
        "fecha": parse_datetime_flexible(payload.get("fecha")) or now_cdmx_naive(),
        "estatus": raw_estatus if raw_estatus in VALID_ESTATUS_SEGUIMIENTO else "PENDIENTE",
        "estatus_aprobacion": raw_aprobacion if raw_aprobacion in VALID_ESTATUS_APROBACION else "EN REVISIÓN",
        "responsable": (payload.get("responsable") or "").strip() or None,
        "proyecto": (payload.get("proyecto") or payload.get("obra") or "").strip() or None,
        "especialidad": (payload.get("especialidad") or "").strip() or None,
        "cliente": cliente,
        "zona": (payload.get("zona") or "").strip(),
        "iva_porc": parse_float(payload.get("iva_porc"), 16.0),
        "notas": (payload.get("notas") or "").strip() or None,
        "items": normalized_items,
    }


def _find_or_create_cliente_import(cliente_data: dict, responsable_final: Optional[str]) -> Cliente:
    nombre_cliente = (cliente_data.get("nombre_cliente") or "").strip()
    empresa = (cliente_data.get("empresa") or "").strip()

    q = Cliente.query.filter(db.func.lower(Cliente.nombre_cliente) == nombre_cliente.lower())
    if empresa:
        q = q.filter(db.func.lower(Cliente.empresa) == empresa.lower())
    cliente = q.first()
    if cliente:
        return cliente

    cliente = Cliente(
        nombre_cliente=nombre_cliente,
        empresa=empresa or None,
        responsable=responsable_final,
        correo=cliente_data.get("correo"),
        telefono=cliente_data.get("telefono"),
        direccion=cliente_data.get("direccion"),
        rfc=cliente_data.get("rfc"),
    )
    db.session.add(cliente)
    db.session.flush()
    return cliente


def _pick_import_folio(preferred_folio: Optional[str]) -> str:
    preferred = (preferred_folio or "").strip()
    if preferred:
        exists = db.session.execute(text("SELECT 1 FROM cotizacion WHERE folio=:f LIMIT 1"), {"f": preferred}).fetchone()
        if not exists:
            return preferred
    return generar_folio()


def import_external_quote_payload(payload: dict, source_label: Optional[str] = None) -> Cotizacion:
    normalized = _normalize_import_payload(payload)
    responsable_final = normalized["responsable"] or None
    cliente = _find_or_create_cliente_import(normalized["cliente"], responsable_final)

    subtotal = 0.0
    detail_rows = []
    for item in normalized["items"]:
        line_subtotal = fmt(item.get("subtotal_pdf") or (item["cantidad"] * item["precio_unitario"]))
        subtotal += line_subtotal
        detail_rows.append((item, line_subtotal))

    zona = normalized["zona"]
    desc_porc = float({
        "Zona Norte": 10.0,
        "Zona Centro": 5.0,
        "Bajio": 10.0,
        "Zona Sur": 15.0,
        "Frontera": 8.0,
    }.get(zona, 0.0))
    descuento_total = subtotal * (desc_porc / 100.0)
    subtotal_desc = subtotal - descuento_total
    iva_monto = subtotal_desc * (normalized["iva_porc"] / 100.0)
    total = subtotal_desc + iva_monto

    notas = normalized["notas"]
    if source_label:
        notas = _append_note(notas, f"Importada desde: {source_label}")
    if normalized["folio"]:
        notas = _append_note(notas, f"Folio externo original: {normalized['folio']}")
    if zona and desc_porc > 0:
        notas = _append_note(notas, f"Zona: {zona} ({int(desc_porc)}% descuento)")

    cot = Cotizacion(
        folio=_pick_import_folio(normalized["folio"]),
        fecha=normalized["fecha"],
        cliente_id=cliente.id,
        estatus=normalized["estatus"],
        estatus_aprobacion=normalized.get("estatus_aprobacion") or "EN REVISIÓN",
        subtotal=fmt(subtotal),
        descuento_total=fmt(descuento_total),
        iva_porc=fmt(normalized["iva_porc"]),
        iva_monto=fmt(iva_monto),
        total=fmt(total),
        notas=notas,
        last_whatsapp_at=None,
        responsable=responsable_final,
        proyecto=normalized["proyecto"],
        especialidad=normalized.get("especialidad"),
    )
    db.session.add(cot)
    db.session.flush()

    for item, line_subtotal in detail_rows:
        concepto = Concepto.query.filter_by(nombre_concepto=item["nombre_concepto"]).first()
        if not concepto:
            concepto = Concepto(
                nombre_concepto=item["nombre_concepto"],
                unidad=item["unidad"] or None,
                precio_unitario=item["precio_unitario"],
                descripcion=item["descripcion"] or None,
            )
            db.session.add(concepto)
            db.session.flush()

        det = CotizacionDetalle(
            cotizacion_id=cot.id,
            concepto_id=concepto.id if concepto else None,
            nombre_concepto=item["nombre_concepto"],
            unidad=item["unidad"],
            cantidad=item["cantidad"],
            precio_unitario=item["precio_unitario"],
            sistema=item["sistema"],
            descripcion=item["descripcion"],
            subtotal=line_subtotal,
        )
        db.session.add(det)

    db.session.commit()
    _send_quote_created_notification(cot)
    _send_quote_review_email_safely(cot)
    return cot

def money(n: float) -> str:
    try:
        return "${:,.2f}".format(float(n or 0))
    except Exception:
        return "${:,.2f}".format(0)

def normalize_moneda(value: str | None) -> str:
    raw = (value or "").strip().upper()
    if raw in {"USD", "DOLAR", "DOLARES", "DÓLAR", "DÓLARES"}:
        return "USD"
    return "MXN"

def moneda_label(moneda: str | None) -> str:
    return "Dólares (USD)" if normalize_moneda(moneda) == "USD" else "Pesos (MXN)"

def money_currency(n: float, moneda: str | None = None) -> str:
    return f"{money(n)} {normalize_moneda(moneda)}"

def cantidad_en_letra(total: float, moneda: str | None = None) -> str:
    moneda_norm = normalize_moneda(moneda)
    try:
        from num2words import num2words
    except Exception:
        entero = int(total)
        cents = int(round((total - entero) * 100)) % 100
        unidad = "dólares" if moneda_norm == "USD" else "pesos"
        sufijo = "USD" if moneda_norm == "USD" else "M.N."
        return f"Cantidad en letra: {entero} {unidad} {cents:02d}/100 {sufijo}"
    entero = int(total)
    cents = int(round((total - entero) * 100)) % 100
    palabras = num2words(entero, lang="es").strip()
    if palabras.endswith(" uno"):
        palabras = palabras[:-4] + " un"
    if palabras:
        palabras = palabras[0].upper() + palabras[1:]
    unidad = "dólar" if moneda_norm == "USD" and entero == 1 else "dólares" if moneda_norm == "USD" else "peso" if entero == 1 else "pesos"
    sufijo = "USD" if moneda_norm == "USD" else "M.N."
    return f"Cantidad en letra: {palabras} {unidad} {cents:02d}/100 {sufijo}"

def cantidad_en_letra_mn(total: float) -> str:
    return cantidad_en_letra(total, "MXN")

def normalize_whatsapp(number: str) -> str:
    if not number:
        return ""
    n = number.strip()
    if n.startswith("whatsapp:"):
        return n
    if n.startswith("+"):
        return f"whatsapp:{n}"
    digits = "".join(ch for ch in n if ch.isdigit())
    if not digits:
        return ""
    # Si ya viene con 52, lo dejamos; si no, lo anteponemos
    if digits.startswith("52"):
        return f"whatsapp:+{digits}"
    return f"whatsapp:+52{digits}"

def can_send_whatsapp() -> bool:
    return bool(twilio_client and TWILIO_WHATSAPP and ADMIN_LIST)

def send_whatsapp_multi(to_list: Iterable[str], body: str) -> None:
    if not to_list:
        return
    if not can_send_whatsapp():
        print("[Twilio] Config incompleta; omito envío.")
        return
    for to in to_list:
        to_norm = normalize_whatsapp(to)
        if not to_norm:
            continue
        try:
            twilio_client.messages.create(from_=TWILIO_WHATSAPP, to=to_norm, body=body)
        except Exception as e:
            print(f"[Twilio] ERROR enviando a {to_norm}: {e}", file=sys.stderr)
            traceback.print_exc()

# ---------------------------------------------------------
# 🔐 Login / Logout
# ---------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        # Acepta varios names posibles del form
        nombre = (
            request.form.get("nombre")
            or request.form.get("username")
            or request.form.get("usuario")
            or request.form.get("user")
            or ""
        ).strip()

        password = (
            request.form.get("password")
            or request.form.get("clave")
            or request.form.get("pass")
            or ""
        ).strip()

        # DEBUG mínimo (se ve en logs de Render)
        print("[LOGIN] form keys:", list(request.form.keys()))
        print("[LOGIN] nombre recibido:", repr(nombre))

        # Match case-insensitive (Admin vs admin, Rafa vs rafa)
        u = Usuario.query.filter(db.func.lower(Usuario.nombre) == nombre.lower()).first()
        print("[LOGIN] usuario encontrado:", u)

        if not u:
            flash("Credenciales inválidas.", "danger")
            return redirect(url_for("login"))

        ok = u.check_password(password)
        print("[LOGIN] password ok:", ok)

        if not ok:
            flash("Credenciales inválidas.", "danger")
            return redirect(url_for("login"))

        remember = (request.form.get("remember") or "").lower() in {"1", "true", "on", "yes"}
        login_user(u, remember=remember)
        # Redirige a la página solicitada originalmente (si viene)
        nxt = request.args.get("next")
        if nxt:
            try:
                # Evita open-redirect (solo paths internos)
                p = urlparse(nxt)
                if p.netloc == "" and (nxt.startswith("/") or nxt.startswith("?")):
                    return redirect(nxt)
            except Exception:
                pass
        return redirect(url_for("index"))

    return render_template("login.html", title="Iniciar sesión")


PASSWORD_RESET_MAX_AGE = 30 * 60


def _password_reset_serializer() -> URLSafeTimedSerializer:
    return URLSafeTimedSerializer(app.secret_key, salt="mar-password-reset-v1")


def _password_reset_token(usuario: Usuario) -> str:
    # La huella del hash invalida automáticamente el enlace al cambiar la contraseña.
    payload = {"uid": usuario.id, "pwd": usuario.password_hash[-20:]}
    return _password_reset_serializer().dumps(payload)


def _password_reset_user(token: str) -> Usuario | None:
    try:
        payload = _password_reset_serializer().loads(token, max_age=PASSWORD_RESET_MAX_AGE)
        usuario = db.session.get(Usuario, int(payload.get("uid", 0)))
        if not usuario or payload.get("pwd") != usuario.password_hash[-20:]:
            return None
        return usuario
    except (BadSignature, SignatureExpired, TypeError, ValueError):
        return None


def _send_password_reset_email(usuario: Usuario, reset_url: str) -> None:
    msg = EmailMessage()
    msg["Subject"] = "Restablece tu acceso a Sistema MAR"
    msg["From"] = f"SISTEMA MAR <{SMTP_FROM or SMTP_USERNAME}>"
    msg["To"] = usuario.correo
    msg.set_content(
        "Recibimos una solicitud para restablecer tu contraseña de Sistema MAR.\n\n"
        f"Abre este enlace (válido durante 30 minutos):\n{reset_url}\n\n"
        "Si no hiciste esta solicitud, ignora este mensaje."
    )
    msg.add_alternative(
        f"""
        <div style="font-family:Arial,sans-serif;max-width:560px;margin:auto;color:#172033">
          <div style="background:#0c3c78;color:white;padding:22px 26px;border-radius:12px 12px 0 0">
            <strong style="font-size:20px">Sistema MAR · Poliutech</strong>
          </div>
          <div style="padding:28px 26px;border:1px solid #d9e2ec;border-top:0;border-radius:0 0 12px 12px">
            <h2 style="margin-top:0">Restablece tu contraseña</h2>
            <p>Hola {escape(usuario.nombre_representante)}, recibimos una solicitud para recuperar tu acceso.</p>
            <p style="margin:28px 0"><a href="{escape(reset_url)}" style="background:#0c3c78;color:white;text-decoration:none;padding:13px 22px;border-radius:8px;font-weight:bold">Crear nueva contraseña</a></p>
            <p style="color:#667085;font-size:14px">El enlace vence en 30 minutos y solo puede utilizarse una vez. Si no solicitaste el cambio, ignora este correo.</p>
          </div>
        </div>
        """,
        subtype="html",
    )
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.starttls()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg)


@app.route("/olvide-contrasena", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        identity = (request.form.get("identity") or "").strip().lower()
        usuario = Usuario.query.filter(
            db.or_(db.func.lower(Usuario.correo) == identity, db.func.lower(Usuario.nombre) == identity)
        ).first()
        if usuario and (usuario.correo or "").strip():
            try:
                reset_url = url_for("reset_password", token=_password_reset_token(usuario), _external=True)
                _send_password_reset_email(usuario, reset_url)
            except Exception:
                logger.exception("No se pudo enviar el correo de recuperación para usuario id=%s", usuario.id)
        flash("Si los datos coinciden con una cuenta, recibirás un enlace de recuperación en unos minutos.", "success")
        return redirect(url_for("forgot_password"))
    return render_template("forgot_password.html", title="Recuperar acceso")


@app.route("/restablecer-contrasena/<token>", methods=["GET", "POST"])
def reset_password(token: str):
    usuario = _password_reset_user(token)
    if not usuario:
        return render_template("reset_password.html", title="Enlace vencido", invalid_token=True), 400
    if request.method == "POST":
        password = request.form.get("password") or ""
        confirmation = request.form.get("password_confirmation") or ""
        if len(password) < 8:
            flash("La contraseña debe tener al menos 8 caracteres.", "danger")
        elif password != confirmation:
            flash("Las contraseñas no coinciden.", "danger")
        else:
            usuario.set_password(password)
            db.session.commit()
            flash("Tu contraseña se actualizó. Ya puedes iniciar sesión.", "success")
            return redirect(url_for("login"))
    return render_template("reset_password.html", title="Crear nueva contraseña", invalid_token=False, token=token)

@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))

# ---------------------------------------------------------
# Dashboard / Catálogos / Cotizador
# ---------------------------------------------------------
@app.route("/")
@login_required
def index():
    page = request.args.get("page", 1, type=int)
    per_page = 20
    desde = (request.args.get("desde") or "").strip()
    hasta = (request.args.get("hasta") or "").strip()
    estatus = (request.args.get("estatus") or "").strip()
    cliente = (request.args.get("cliente") or "").strip()
    especialidad = (request.args.get("especialidad") or "").strip()
    dashboard_filters = {
        "desde": desde,
        "hasta": hasta,
        "estatus": estatus,
        "cliente": cliente,
        "especialidad": especialidad,
    }

    try:
        base_query = _build_dashboard_cotizaciones_query(
            desde=desde,
            hasta=hasta,
            estatus=estatus,
            cliente=cliente,
            especialidad=especialidad,
        )
    except ValueError:
        base_query = _build_dashboard_cotizaciones_query()
        dashboard_filters = {"desde": "", "hasta": "", "estatus": "", "cliente": "", "especialidad": ""}

    total_cotizaciones = base_query.count()
    total_importe = (
        base_query.with_entities(db.func.coalesce(db.func.sum(Cotizacion.total), 0)).scalar()
        or 0
    )
    quotes_query = base_query.order_by(Cotizacion.fecha.desc())

    pagination = quotes_query.paginate(page=page, per_page=per_page, error_out=False)
    cotizaciones = pagination.items

    total_catalogo = Concepto.query.count()
    responsables = [
        row[0]
        for row in (
            _cotizaciones_base_query()
            .with_entities(Cotizacion.responsable)
            .filter(Cotizacion.responsable.isnot(None))
            .filter(db.func.trim(Cotizacion.responsable) != "")
            .distinct()
            .order_by(Cotizacion.responsable)
            .all()
        )
    ]

    return render_template(
        "dashboard.html",
        title="Sistema MAR",
        total_cotizaciones=total_cotizaciones,
        total_importe=float(total_importe),
        total_catalogo=total_catalogo,
        cotizaciones=cotizaciones,
        pagination=pagination,
        dashboard_filters=dashboard_filters,
        responsables=responsables,
        valid_estatus=VALID_ESTATUS_SEGUIMIENTO,
        valid_estatus_aprobacion=VALID_ESTATUS_APROBACION,
        show_splash=True
    )

@app.route("/cotizador")
@login_required
def cotizador():
    return render_template("cotizador.html", title="Nuevo - Sistema MAR", proyectos=_known_project_names())


@app.route("/proyectos")
@login_required
def proyectos():
    key_expr = _project_key_expr()
    name_expr = _project_display_expr()
    rows = (
        _cotizaciones_base_query()
        .with_entities(
            key_expr.label("key"),
            name_expr.label("nombre"),
            db.func.count(Cotizacion.id).label("cotizaciones"),
            db.func.coalesce(db.func.sum(Cotizacion.total), 0).label("total"),
            db.func.max(Cotizacion.fecha).label("ultima_fecha"),
        )
        .group_by(key_expr)
        .order_by(db.func.max(Cotizacion.fecha).desc())
        .all()
    )
    total_proyectos = len(rows)
    total_cotizaciones = sum(int(r.cotizaciones or 0) for r in rows)
    total_importe = sum(float(r.total or 0) for r in rows)
    return render_template(
        "proyectos.html",
        proyectos=rows,
        total_proyectos=total_proyectos,
        total_cotizaciones=total_cotizaciones,
        total_importe=total_importe,
        title="Proyectos - Sistema MAR",
    )


@app.route("/proyectos/detalle")
@login_required
def proyecto_detalle():
    nombre = (request.args.get("proyecto") or "Sin proyecto").strip() or "Sin proyecto"
    q = _cotizaciones_base_query()
    if nombre == "Sin proyecto":
        q = q.filter(or_(Cotizacion.proyecto.is_(None), db.func.trim(Cotizacion.proyecto) == ""))
    else:
        q = q.filter(db.func.lower(db.func.trim(Cotizacion.proyecto)) == nombre.lower())

    cotizaciones = q.order_by(Cotizacion.fecha.desc()).all()
    total_importe = sum(float(c.total or 0) for c in cotizaciones)
    promedio_importe = total_importe / len(cotizaciones) if cotizaciones else 0.0

    monthly_map = {}
    status_map = {estado: 0 for estado in VALID_ESTATUS}
    for cot in cotizaciones:
        if cot.fecha:
            key = cot.fecha.strftime("%Y-%m")
            label = cot.fecha.strftime("%b %Y")
        else:
            key = "0000-00"
            label = "Sin fecha"

        item = monthly_map.setdefault(key, {"label": label, "total": 0.0, "cotizaciones": 0})
        item["total"] += float(cot.total or 0)
        item["cotizaciones"] += 1

        estatus = (cot.estatus or "").strip().upper()
        if estatus in status_map:
            status_map[estatus] += 1

    monthly_series = [monthly_map[key] for key in sorted(monthly_map.keys())]
    status_series = {
        "labels": VALID_ESTATUS,
        "counts": [status_map.get(estado, 0) for estado in VALID_ESTATUS],
    }

    return render_template(
        "proyecto_detalle.html",
        proyecto=nombre,
        cotizaciones=cotizaciones,
        total_importe=total_importe,
        promedio_importe=promedio_importe,
        monthly_series=monthly_series,
        status_series=status_series,
        valid_estatus=VALID_ESTATUS,
        title=f"Proyecto {nombre} - Sistema MAR",
    )


@app.route("/cotizador/voz/transcribir", methods=["POST"])
@login_required
def cotizador_voice_transcribe():
    uploaded = request.files.get("audio")
    target = (request.form.get("target") or "comando").strip().lower()
    if not uploaded or not (uploaded.filename or "").strip():
        return jsonify({"ok": False, "error": "Adjunta un archivo de audio antes de transcribir."}), 400

    try:
        audio_bytes = uploaded.read()
        transcript = _voice_transcribe_audio_bytes(
            audio_bytes=audio_bytes,
            filename=uploaded.filename or "voz.webm",
            mime_type=uploaded.mimetype or mimetypes.guess_type(uploaded.filename or "")[0] or "application/octet-stream",
        )
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 503
    except Exception as exc:
        try:
            logger.exception("Error transcribiendo audio del cotizador web")
        except Exception:
            pass
        return jsonify({"ok": False, "error": f"No se pudo transcribir el audio: {exc}"}), 500

    return jsonify({"ok": True, "target": target, "transcript": transcript})


@app.route("/cotizador/voz/preview", methods=["POST"])
@login_required
def cotizador_voice_preview():
    payload = request.get_json(silent=True) or {}
    command_raw = str(payload.get("comando") or payload.get("transcript") or "").strip()
    client_override = str(payload.get("cliente") or "").strip()
    notes = str(payload.get("notas") or "").strip()
    conditions_raw = str(payload.get("condiciones") or payload.get("condiciones_raw") or "").strip()
    if not command_raw:
        return jsonify({"ok": False, "error": "Dicta o escribe un comando antes de continuar."}), 400

    try:
        user_obj = current_user._get_current_object() if hasattr(current_user, "_get_current_object") else current_user
        preview = _voice_preview_payload_for_mobile(
            command_raw=command_raw,
            user=user_obj,
            client_override=client_override,
            notes=notes,
            conditions_raw=conditions_raw,
        )
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    except Exception as exc:
        try:
            logger.exception("Error interpretando voz en cotizador web")
        except Exception:
            pass
        return jsonify({"ok": False, "error": f"No se pudo interpretar el comando: {exc}"}), 500

    return jsonify({"ok": True, "preview": preview})


@app.route("/altas", methods=["GET", "POST"])
@login_required
def altas_proveedores():
    if not is_admin():
        abort(403)

    rows = _load_provider_numbers()

    if request.method == "POST":
        numeros = request.form.getlist("numero[]")
        empresas = request.form.getlist("empresa[]")
        razones = request.form.getlist("razon_social_poliutech[]")
        relaciones = request.form.getlist("relacion[]")
        contactos = request.form.getlist("contacto[]")
        telefonos = request.form.getlist("telefono[]")
        correos = request.form.getlist("correo[]")

        total_rows = max(len(numeros), len(empresas), len(razones), len(relaciones), len(contactos), len(telefonos), len(correos), 0)
        rows: list[dict] = []
        for idx in range(total_rows):
            numero = (numeros[idx] if idx < len(numeros) else "").strip()
            empresa = (empresas[idx] if idx < len(empresas) else "").strip()
            razon_social = (razones[idx] if idx < len(razones) else "").strip()
            relacion = (relaciones[idx] if idx < len(relaciones) else "PROVEEDOR").strip().upper()
            if relacion not in {"CLIENTE", "PROVEEDOR"}:
                relacion = "PROVEEDOR"
            contacto = (contactos[idx] if idx < len(contactos) else "").strip()
            telefono = (telefonos[idx] if idx < len(telefonos) else "").strip()
            correo = (correos[idx] if idx < len(correos) else "").strip()

            if not any([numero, empresa, razon_social, contacto, telefono, correo]):
                continue

            if correo and not re.fullmatch(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", correo):
                flash(f"El correo '{correo}' no es valido.", "danger")
                return render_template(
                    "altas.html",
                    title="Altas de proveedores",
                    rows=[
                        _normalize_provider_row({
                            "numero": (numeros[pos] if pos < len(numeros) else "").strip(),
                            "empresa": (empresas[pos] if pos < len(empresas) else "").strip(),
                            "razon_social_poliutech": (razones[pos] if pos < len(razones) else "").strip(),
                            "relacion": (relaciones[pos] if pos < len(relaciones) else "PROVEEDOR").strip(),
                            "contacto": (contactos[pos] if pos < len(contactos) else "").strip(),
                            "telefono": (telefonos[pos] if pos < len(telefonos) else "").strip(),
                            "correo": (correos[pos] if pos < len(correos) else "").strip(),
                        }, pos + 1)
                        for pos in range(total_rows)
                    ],
                    filtered_rows=_filter_provider_rows(
                        [
                            _normalize_provider_row({
                                "numero": (numeros[pos] if pos < len(numeros) else "").strip(),
                                "empresa": (empresas[pos] if pos < len(empresas) else "").strip(),
                                "razon_social_poliutech": (razones[pos] if pos < len(razones) else "").strip(),
                                "relacion": (relaciones[pos] if pos < len(relaciones) else "PROVEEDOR").strip(),
                                "contacto": (contactos[pos] if pos < len(contactos) else "").strip(),
                                "telefono": (telefonos[pos] if pos < len(telefonos) else "").strip(),
                                "correo": (correos[pos] if pos < len(correos) else "").strip(),
                            }, pos + 1)
                            for pos in range(total_rows)
                        ],
                        _provider_filters_from_request(),
                    ),
                    filters=_provider_filters_from_request(),
                )

            rows.append({
                "id": len(rows) + 1,
                "numero": numero,
                "empresa": empresa,
                "razon_social_poliutech": razon_social,
                "relacion": relacion,
                "contacto": contacto,
                "telefono": telefono,
                "correo": correo,
            })

        _save_provider_numbers(rows)
        flash("Altas actualizadas correctamente.", "success")
        return redirect(url_for("altas_proveedores"))

    filters = _provider_filters_from_request()
    return render_template(
        "altas.html",
        title="Altas de proveedores",
        rows=rows,
        filtered_rows=_filter_provider_rows(rows, filters),
        filters=filters,
    )


@app.route("/altas/export.xlsx")
@login_required
def export_altas_proveedores_xlsx():
    if not is_admin():
        abort(403)

    filters = _provider_filters_from_request()
    rows = _filter_provider_rows(_load_provider_numbers(), filters)

    headers = [
        "NUMERO",
        "EMPRESA",
        "RAZON SOCIAL POLIUTECH",
        "RELACION",
        "CONTACTO",
        "TELEFONO",
        "CORREO",
    ]
    body_rows = []
    for row in rows:
        body_rows.append([
            row.get("numero", ""),
            row.get("empresa", ""),
            row.get("razon_social_poliutech", ""),
            row.get("relacion", ""),
            row.get("contacto", ""),
            row.get("telefono", ""),
            row.get("correo", ""),
        ])

    output_bytes = _build_simple_xlsx(
        "Altas",
        headers,
        body_rows,
        column_widths=[18, 28, 28, 18, 24, 18, 32],
    )

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Response(
        output_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="altas_proveedores_{stamp}.xlsx"'},
    )


@app.route("/altas/export.pdf")
@login_required
def export_altas_proveedores_pdf():
    if not is_admin():
        abort(403)

    filters = _provider_filters_from_request()
    rows = _filter_provider_rows(_load_provider_numbers(), filters)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=24 * mm,
        bottomMargin=38 * mm,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="EncabezadoAltas", fontName="Helvetica", fontSize=9, leading=12, spaceAfter=4, splitLongWords=False))
    styles.add(ParagraphStyle(name="AltasCell", fontName="Helvetica", fontSize=7.5, leading=9, splitLongWords=False))
    styles.add(ParagraphStyle(name="AltasCenter", fontName="Helvetica", fontSize=7.5, leading=9, alignment=1, splitLongWords=False))

    elems = []

    def encabezado(canv, doc_):
        canv.saveState()
        canv.setFillColor(colors.HexColor(MAR_BLUE))
        canv.rect(0, A4[1] - 40, A4[0], 40, stroke=0, fill=1)

        logo_path = os.path.join(app.static_folder or "static", "logo.png")
        if os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                max_w = 22.5 * mm
                scale = max_w / iw
                w = max_w
                h = ih * scale
                x_logo = 12
                y_logo = A4[1] - h - 8
                canv.drawImage(img, x_logo, y_logo, width=w, height=h, mask="auto")
            except Exception:
                pass

        canv.setFont("Helvetica-Bold", 14)
        canv.setFillColor(colors.white)
        canv.drawRightString(A4[0] - 12, A4[1] - 18, "ALTAS DE PROVEEDORES")
        canv.setFont("Helvetica", 10)
        canv.drawRightString(A4[0] - 12, A4[1] - 31, "Recubrimientos Especializados")
        canv.restoreState()

    def footer(canv, doc_):
        canv.saveState()
        division_path = os.path.join(app.static_folder or "static", "division.png")
        if os.path.exists(division_path):
            try:
                canv.drawImage(division_path, (A4[0] - 155 * mm) / 2, 45, width=155 * mm, height=3 * mm, mask="auto")
            except Exception:
                pass

        canv.setFont("Helvetica-Bold", 9)
        canv.setFillColor(colors.HexColor(MAR_BLUE))
        canv.drawCentredString(A4[0] / 2, 35, "POLIUTECH - Recubrimientos Especializados")

        canv.setFont("Helvetica", 8)
        canv.setFillColor(colors.HexColor("#333333"))
        canv.drawCentredString(A4[0] / 2, 25, "Campos Eliseos 223 Oficina 602 - Col. Polanco V Seccion - Miguel Hidalgo, CDMX 11560")
        canv.drawCentredString(A4[0] / 2, 15, "Tel: 55 5938 6530 / 55 5938 0536 - info@poliutech.com - www.poliutech.com")

        try:
            canv.setTitle("Altas de proveedores")
        except Exception:
            pass

        canv.restoreState()

    filtro_razon = (filters.get("razon_social_poliutech") or "").strip()
    filtro_relacion = (filters.get("relacion") or "").strip()
    generated_at = now_cdmx_naive().strftime("%d/%m/%Y %H:%M")
    meta_data = [
        [
            Paragraph(f"<b>Fecha de exportación:</b> {generated_at}", styles["EncabezadoAltas"]),
            Paragraph(f"<b>Total de registros:</b> {len(rows)}", styles["EncabezadoAltas"]),
        ],
        [
            Paragraph("<b>Filtro aplicado:</b> Razón social Poliutech", styles["EncabezadoAltas"]),
            Paragraph(filtro_razon or "Todos", styles["EncabezadoAltas"]),
        ],
        [
            Paragraph("<b>Relación:</b>", styles["EncabezadoAltas"]),
            Paragraph(filtro_relacion or "Todas", styles["EncabezadoAltas"]),
        ],
    ]
    meta_tbl = Table(meta_data, colWidths=[95 * mm, 95 * mm], hAlign="LEFT")
    meta_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))
    elems.append(meta_tbl)
    elems.append(Spacer(1, 6))

    data = [[
        "NUMERO",
        "EMPRESA",
        "RAZON SOCIAL POLIUTECH",
        "RELACION",
        "CONTACTO",
        "TELEFONO",
        "CORREO",
    ]]
    for row in rows:
        data.append([
            Paragraph(_truncate_pdf_text(row.get("numero", ""), 24), styles["AltasCenter"]),
            Paragraph(_truncate_pdf_text(row.get("empresa", ""), 48), styles["AltasCell"]),
            Paragraph(_truncate_pdf_text(row.get("razon_social_poliutech", ""), 52), styles["AltasCell"]),
            Paragraph(_truncate_pdf_text(row.get("relacion", ""), 12), styles["AltasCenter"]),
            Paragraph(_truncate_pdf_text(row.get("contacto", ""), 38), styles["AltasCell"]),
            Paragraph(_truncate_pdf_text(row.get("telefono", ""), 24), styles["AltasCenter"]),
            Paragraph(_truncate_pdf_text(row.get("correo", ""), 42), styles["AltasCell"]),
        ])

    if len(data) == 1:
        data.append([
            Paragraph("-", styles["AltasCenter"]),
            Paragraph("No hay registros para exportar con el filtro actual.", styles["AltasCell"]),
            Paragraph("-", styles["AltasCenter"]),
            Paragraph("-", styles["AltasCenter"]),
            Paragraph("-", styles["AltasCenter"]),
            Paragraph("-", styles["AltasCenter"]),
            Paragraph("-", styles["AltasCenter"]),
        ])

    tbl = Table(
        data,
        colWidths=[15 * mm, 31 * mm, 40 * mm, 23 * mm, 25 * mm, 21 * mm, 35 * mm],
        repeatRows=1,
        hAlign="CENTER",
    )
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(MAR_BLUE)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (3, 0), (3, -1), "CENTER"),
        ("ALIGN", (5, 0), (5, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("WORDWRAP", (0, 0), (-1, -1), True),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elems.append(tbl)

    doc.build(
        elems,
        onFirstPage=lambda canv, d: (draw_watermark(canv, app), encabezado(canv, d), footer(canv, d)),
        onLaterPages=lambda canv, d: (draw_watermark(canv, app), encabezado(canv, d), footer(canv, d)),
    )

    buf.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response = Response(
        buf.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f'inline; filename="altas_proveedores_{stamp}.pdf"'},
    )
    response.direct_passthrough = False
    return response


@app.route("/prospectos", methods=["GET", "POST"])
@login_required
def prospectos():
    if request.method == "POST":
        action = (request.form.get("action") or "").strip().lower()

        if action == "add":
            titulo = (request.form.get("titulo") or "").strip()
            descripcion = (request.form.get("descripcion") or "").strip()
            contacto = (request.form.get("contacto") or "").strip()
            telefono = (request.form.get("telefono") or "").strip()
            correo = (request.form.get("correo") or "").strip()
            status = _normalize_prospecto_status(request.form.get("status"))
            responsable = (request.form.get("responsable") or "").strip() if is_admin() else (responsable_actual() or "").strip()

            if not titulo:
                flash("Captura el título del prospecto.", "warning")
                return redirect(url_for("prospectos"))
            if correo and not re.fullmatch(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", correo):
                flash(f"El correo '{correo}' no es valido.", "danger")
                return redirect(url_for("prospectos"))

            prospecto = Prospecto(
                titulo=titulo,
                descripcion=descripcion or None,
                contacto=contacto or None,
                telefono=telefono or None,
                correo=correo or None,
                status=status,
                responsable=responsable or None,
            )
            db.session.add(prospecto)
            db.session.commit()
            flash("Prospecto agregado correctamente.", "success")
            return redirect(url_for("prospectos"))

        if action == "import":
            uploaded = request.files.get("import_file")
            if not uploaded or not (uploaded.filename or "").strip():
                flash("Selecciona un archivo Excel antes de importar.", "warning")
                return redirect(url_for("prospectos"))

            filename = (uploaded.filename or "").strip().lower()
            if not filename.endswith(".xlsx"):
                flash("Solo se permite importar archivos .xlsx.", "danger")
                return redirect(url_for("prospectos"))

            try:
                file_bytes = uploaded.read()
                if not file_bytes:
                    raise ValueError("El archivo Excel llegó vacío.")
                imported_rows = _load_prospectos_from_xlsx(file_bytes)
            except Exception as exc:
                flash(f"No pude leer el Excel: {exc}", "danger")
                return redirect(url_for("prospectos"))

            if not imported_rows:
                flash("No se encontraron prospectos válidos en el Excel.", "warning")
                return redirect(url_for("prospectos"))

            inserted = 0
            updated = 0
            for row in imported_rows:
                titulo = (row.get("titulo") or "").strip()
                descripcion = (row.get("descripcion") or "").strip()
                contacto = (row.get("contacto") or "").strip()
                telefono = (row.get("telefono") or "").strip()
                correo = (row.get("correo") or "").strip()

                if correo and not re.fullmatch(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", correo):
                    flash(f"El correo '{correo}' no es valido.", "danger")
                    return redirect(url_for("prospectos"))

                existing = Prospecto.query.filter(
                    db.func.lower(Prospecto.titulo) == titulo.lower(),
                    db.func.lower(db.func.coalesce(Prospecto.correo, "")) == correo.lower(),
                ).first()

                if existing:
                    existing.descripcion = descripcion or existing.descripcion
                    existing.contacto = contacto or existing.contacto
                    existing.telefono = telefono or existing.telefono
                    existing.correo = correo or existing.correo
                    if not (existing.status or "").strip():
                        existing.status = "PENDIENTE"
                    updated += 1
                    continue

                db.session.add(Prospecto(
                    titulo=titulo,
                    descripcion=descripcion or None,
                    contacto=contacto or None,
                    telefono=telefono or None,
                    correo=correo or None,
                    status="PENDIENTE",
                ))
                inserted += 1

            db.session.commit()
            flash(f"Importación completada. Nuevos: {inserted}. Actualizados: {updated}.", "success")
            return redirect(url_for("prospectos"))

        if action == "update":
            row_ids = request.form.getlist("row_id[]")
            titulos = request.form.getlist("titulo[]")
            descripciones = request.form.getlist("descripcion[]")
            contactos = request.form.getlist("contacto[]")
            telefonos = request.form.getlist("telefono[]")
            correos = request.form.getlist("correo[]")
            statuses = request.form.getlist("status[]")

            updated = 0
            for idx, row_id in enumerate(row_ids):
                if not str(row_id).strip().isdigit():
                    continue
                prospecto = db.session.get(Prospecto, int(row_id))
                if not prospecto:
                    continue
                require_prospecto_owner_or_admin(prospecto)

                titulo = (titulos[idx] if idx < len(titulos) else "").strip()
                descripcion = (descripciones[idx] if idx < len(descripciones) else "").strip()
                contacto = (contactos[idx] if idx < len(contactos) else "").strip()
                telefono = (telefonos[idx] if idx < len(telefonos) else "").strip()
                correo = (correos[idx] if idx < len(correos) else "").strip()
                status = _normalize_prospecto_status(statuses[idx] if idx < len(statuses) else "")

                if not titulo:
                    flash("Todos los prospectos deben tener título.", "warning")
                    return redirect(url_for("prospectos"))
                if correo and not re.fullmatch(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", correo):
                    flash(f"El correo '{correo}' no es valido.", "danger")
                    return redirect(url_for("prospectos"))

                prospecto.titulo = titulo
                prospecto.descripcion = descripcion or None
                prospecto.contacto = contacto or None
                prospecto.telefono = telefono or None
                prospecto.correo = correo or None
                prospecto.status = status
                if not prospecto.responsable:
                    prospecto.responsable = responsable_actual() or prospecto.responsable
                updated += 1

            db.session.commit()
            flash(f"Se actualizaron {updated} prospecto(s).", "success")
            return redirect(url_for("prospectos"))

        if action == "delete":
            selected_ids = [int(value) for value in request.form.getlist("selected_ids[]") if str(value).strip().isdigit()]
            if not selected_ids:
                flash("Selecciona al menos un prospecto para eliminar.", "warning")
                return redirect(url_for("prospectos"))

            items = Prospecto.query.filter(Prospecto.id.in_(selected_ids)).all()
            deleted = 0
            for prospecto in items:
                require_prospecto_owner_or_admin(prospecto)
                db.session.delete(prospecto)
                deleted += 1
            db.session.commit()
            flash(f"Se eliminaron {deleted} prospecto(s).", "success")
            return redirect(url_for("prospectos"))

        flash("Acción no válida para prospectos.", "danger")
        return redirect(url_for("prospectos"))

    rows = _load_prospectos()
    filters = _prospectos_filters_from_request()
    filtered_rows = _filter_prospectos(rows, filters)
    return render_template(
        "prospectos.html",
        title="Prospectos",
        rows=rows,
        filtered_rows=filtered_rows,
        filters=filters,
        status_options=PROSPECT_STATUS_OPTIONS,
        default_responsable=responsable_actual() or "",
    )


@app.route("/prospectos/export.xls")
@login_required
def export_prospectos_xls():
    rows = _filter_prospectos(_load_prospectos(), _prospectos_filters_from_request())
    headers = ["TITULO", "DESCRIPCION", "CONTACTO", "TELEFONO", "CORREO", "STATUS"]
    body_rows = [
        [
            row.get("titulo", ""),
            row.get("descripcion", ""),
            row.get("contacto", ""),
            row.get("telefono", ""),
            row.get("correo", ""),
            row.get("status", ""),
        ]
        for row in rows
    ]
    output_bytes = _build_simple_xls("Prospectos", headers, body_rows)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Response(
        output_bytes,
        mimetype="application/vnd.ms-excel",
        headers={"Content-Disposition": f'attachment; filename="prospectos_{stamp}.xls"'},
    )


@app.route("/prospectos/<int:prospecto_id>/seguimiento")
@login_required
def prospecto_seguimiento(prospecto_id: int):
    prospecto = Prospecto.query.get_or_404(prospecto_id)
    return render_template(
        "prospecto_seguimiento.html",
        prospecto=prospecto,
        cliente_info=_cliente_seguimiento_payload(
            nombre=prospecto.contacto or prospecto.titulo,
            correo=prospecto.correo,
            telefono=prospecto.telefono,
            responsable=prospecto.responsable,
            titulo="Datos del cliente / prospecto",
            extras=[{"label": "Prospecto", "value": prospecto.titulo}],
        ),
        seguimientos=prospecto.seguimientos,
        mention_users=_usuarios_menciones_payload(),
        title=f"Seguimiento prospecto {prospecto.titulo}",
    )


@app.route("/prospectos/<int:prospecto_id>/seguimiento", methods=["POST"])
@login_required
def crear_prospecto_seguimiento(prospecto_id: int):
    prospecto = Prospecto.query.get_or_404(prospecto_id)
    comentario = (request.form.get("comentario") or "").strip()
    nuevo_status = _normalize_prospecto_status(request.form.get("status"))
    tagged_users = _usuarios_mencionados_en_comentario(comentario)

    if not comentario:
        flash("Escribe un comentario de seguimiento.", "warning")
        return redirect(url_for("prospecto_seguimiento", prospecto_id=prospecto.id))

    prospecto.status = nuevo_status
    autor = (responsable_actual() or "Sistema").strip()
    seg = ProspectoSeguimiento(
        prospecto_id=prospecto.id,
        usuario_id=getattr(current_user, "id", None),
        autor=autor,
        comentario=comentario,
        fecha_seguimiento=now_cdmx_naive(),
    )
    db.session.add(seg)
    db.session.commit()
    try:
        _notify_tagged_followup(
            tagged_users=tagged_users,
            module_label="Prospectos",
            item_label=prospecto.titulo or f"Prospecto #{prospecto.id}",
            autor=autor,
            comentario=comentario,
            view_endpoint="prospecto_seguimiento",
            view_params={"prospecto_id": prospecto.id, "_anchor": f"seguimiento-{seg.id}"},
        )
    except Exception as exc:
        logger.exception("No se pudo notificar etiquetas del prospecto %s", prospecto.id)
        flash(f"Seguimiento guardado, pero no se pudo enviar correo a etiquetados: {exc}", "warning")
        return redirect(url_for("prospecto_seguimiento", prospecto_id=prospecto.id, _anchor=f"seguimiento-{seg.id}"))
    flash("Seguimiento guardado correctamente.", "success")
    return redirect(url_for("prospecto_seguimiento", prospecto_id=prospecto.id, _anchor=f"seguimiento-{seg.id}"))


@app.route("/prospectos/<int:prospecto_id>/seguimiento/<int:seg_id>/editar", methods=["POST"])
@login_required
def editar_prospecto_seguimiento(prospecto_id: int, seg_id: int):
    prospecto = Prospecto.query.get_or_404(prospecto_id)
    seg = ProspectoSeguimiento.query.filter_by(id=seg_id, prospecto_id=prospecto.id).first_or_404()
    require_prospecto_followup_author_or_admin(seg)

    comentario = (request.form.get("comentario") or "").strip()
    if not comentario:
        flash("El comentario no puede quedar vacío.", "warning")
        return redirect(url_for("prospecto_seguimiento", prospecto_id=prospecto.id))

    seg.comentario = comentario
    seg.actualizado_en = now_cdmx_naive()
    db.session.commit()
    flash("Seguimiento actualizado.", "success")
    return redirect(url_for("prospecto_seguimiento", prospecto_id=prospecto.id, _anchor=f"seguimiento-{seg.id}"))


@app.route("/prospectos/<int:prospecto_id>/seguimiento/<int:seg_id>/eliminar", methods=["POST"])
@login_required
def eliminar_prospecto_seguimiento(prospecto_id: int, seg_id: int):
    prospecto = Prospecto.query.get_or_404(prospecto_id)
    seg = ProspectoSeguimiento.query.filter_by(id=seg_id, prospecto_id=prospecto.id).first_or_404()
    require_prospecto_followup_author_or_admin(seg)
    db.session.delete(seg)
    db.session.commit()
    flash("Seguimiento eliminado.", "success")
    return redirect(url_for("prospecto_seguimiento", prospecto_id=prospecto.id))


@app.route("/soporte")
@login_required
def soporte_tickets():
    filters = _ticket_filters_from_request()
    rows = _load_ticket_rows(filters)
    total_abiertos = sum(1 for row in rows if row["estado"] not in {"RESUELTO", "CERRADO"})
    total_urgentes = sum(1 for row in rows if row["prioridad"] == "URGENTE")
    return render_template(
        "soporte_tickets.html",
        title="Soporte - Tickets",
        rows=rows,
        filters=filters,
        status_options=TICKET_STATUS_OPTIONS,
        priority_options=TICKET_PRIORITY_OPTIONS,
        category_options=TICKET_CATEGORY_OPTIONS,
        total_abiertos=total_abiertos,
        total_urgentes=total_urgentes,
        default_responsable=responsable_actual() or "",
    )


@app.route("/soporte/nuevo", methods=["GET", "POST"])
@login_required
def soporte_ticket_nuevo():
    if request.method == "POST":
        asunto = (request.form.get("asunto") or "").strip()
        descripcion = (request.form.get("descripcion") or "").strip()
        solicitante = (request.form.get("solicitante") or "").strip()
        correo = (request.form.get("correo") or "").strip()
        categoria = _normalize_ticket_category(request.form.get("categoria"))
        prioridad = _normalize_ticket_priority(request.form.get("prioridad"))
        responsable = (request.form.get("responsable") or "").strip() if is_admin() else (responsable_actual() or "").strip()

        if not asunto or not descripcion or not solicitante:
            flash("Captura asunto, descripción y solicitante.", "warning")
            return redirect(url_for("soporte_ticket_nuevo"))
        if correo and not re.fullmatch(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", correo):
            flash(f"El correo '{correo}' no es valido.", "danger")
            return redirect(url_for("soporte_ticket_nuevo"))

        ticket = TicketSoporte(
            asunto=asunto,
            descripcion=descripcion,
            solicitante=solicitante,
            correo=correo or None,
            telefono=None,
            empresa=None,
            categoria=categoria,
            prioridad=prioridad,
            estado="NUEVO",
            responsable=responsable or None,
            creado_por_id=getattr(current_user, "id", None),
            creado_en=now_cdmx_naive(),
            actualizado_en=now_cdmx_naive(),
        )
        db.session.add(ticket)
        db.session.flush()
        ticket.folio = f"TCK-{ticket.id:06d}"

        try:
            saved = _save_ticket_attachments(ticket, request.files.getlist("adjuntos"))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
            return redirect(url_for("soporte_ticket_nuevo"))

        db.session.commit()
        try:
            _send_support_ticket_email(ticket)
            flash(f"Ticket {ticket.folio} creado correctamente y notificado a sistemas. Adjuntos: {saved}.", "success")
        except Exception as exc:
            try:
                logger.exception("No se pudo enviar correo de soporte %s", ticket.folio or ticket.id)
            except Exception:
                pass
            flash(f"Ticket {ticket.folio} creado, pero no se pudo enviar el correo: {exc}", "warning")
        return redirect(url_for("soporte_ticket_detalle", ticket_id=ticket.id))

    return render_template(
        "soporte_ticket_form.html",
        title="Nuevo ticket de soporte",
        status_options=TICKET_STATUS_OPTIONS,
        priority_options=TICKET_PRIORITY_OPTIONS,
        category_options=TICKET_CATEGORY_OPTIONS,
        default_responsable=responsable_actual() or "",
    )


@app.route("/soporte/<int:ticket_id>", methods=["GET", "POST"])
@login_required
def soporte_ticket_detalle(ticket_id: int):
    ticket = TicketSoporte.query.get_or_404(ticket_id)
    require_ticket_owner_or_admin(ticket)

    if request.method == "POST":
        action = (request.form.get("action") or "").strip().lower()

        if action == "update":
            ticket.estado = _normalize_ticket_status(request.form.get("estado"))
            ticket.prioridad = _normalize_ticket_priority(request.form.get("prioridad"))
            ticket.categoria = _normalize_ticket_category(request.form.get("categoria"))
            if is_admin():
                ticket.responsable = (request.form.get("responsable") or "").strip() or None
            ticket.actualizado_en = now_cdmx_naive()
            ticket.cerrado_en = now_cdmx_naive() if _ticket_is_closed(ticket.estado) else None
            db.session.commit()
            flash("Ticket actualizado.", "success")
            return redirect(url_for("soporte_ticket_detalle", ticket_id=ticket.id))

        if action == "comment":
            comentario = (request.form.get("comentario") or "").strip()
            tagged_users = _usuarios_mencionados_en_comentario(comentario)
            if not comentario and not any((f.filename or "").strip() for f in request.files.getlist("adjuntos")):
                flash("Escribe un comentario o adjunta una captura.", "warning")
                return redirect(url_for("soporte_ticket_detalle", ticket_id=ticket.id))

            autor = (responsable_actual() or "Soporte").strip()
            comentario_final = comentario or "Adjuntos agregados."
            seg = TicketComentario(
                ticket_id=ticket.id,
                usuario_id=getattr(current_user, "id", None),
                autor=autor,
                comentario=comentario_final,
                es_interno=bool(request.form.get("es_interno")) and is_admin(),
                creado_en=now_cdmx_naive(),
            )
            db.session.add(seg)
            db.session.flush()
            try:
                _save_ticket_attachments(ticket, request.files.getlist("adjuntos"), comentario=seg)
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
                return redirect(url_for("soporte_ticket_detalle", ticket_id=ticket.id))

            if ticket.estado == "NUEVO":
                ticket.estado = "EN REVISION"
            ticket.actualizado_en = now_cdmx_naive()
            db.session.commit()
            try:
                _notify_tagged_followup(
                    tagged_users=tagged_users,
                    module_label="Soporte",
                    item_label=f"{ticket.folio or ('TCK-%06d' % ticket.id)} - {ticket.asunto}",
                    autor=autor,
                    comentario=comentario_final,
                    view_endpoint="soporte_ticket_detalle",
                    view_params={"ticket_id": ticket.id, "_anchor": f"comentario-{seg.id}"},
                )
            except Exception as exc:
                logger.exception("No se pudo notificar etiquetas del ticket %s", ticket.id)
                flash(f"Comentario guardado, pero no se pudo enviar correo a etiquetados: {exc}", "warning")
                return redirect(url_for("soporte_ticket_detalle", ticket_id=ticket.id, _anchor=f"comentario-{seg.id}"))
            flash("Comentario guardado.", "success")
            return redirect(url_for("soporte_ticket_detalle", ticket_id=ticket.id, _anchor=f"comentario-{seg.id}"))

        flash("Acción no válida para el ticket.", "danger")
        return redirect(url_for("soporte_ticket_detalle", ticket_id=ticket.id))

    return render_template(
        "soporte_ticket_detalle.html",
        title=f"Ticket {ticket.folio or ticket.id}",
        ticket=ticket,
        cliente_info=_cliente_seguimiento_payload(
            nombre=ticket.solicitante,
            empresa=ticket.empresa,
            correo=ticket.correo,
            telefono=ticket.telefono,
            responsable=ticket.responsable,
            titulo="Datos del cliente / solicitante",
            extras=[{"label": "Asunto", "value": ticket.asunto}],
        ),
        comentarios=ticket.comentarios,
        adjuntos=ticket.adjuntos,
        status_options=TICKET_STATUS_OPTIONS,
        priority_options=TICKET_PRIORITY_OPTIONS,
        category_options=TICKET_CATEGORY_OPTIONS,
        mention_users=_usuarios_menciones_payload(),
        ticket_public_url=_ticket_public_url,
    )


@app.route("/prospectos/export.pdf")
@login_required
def export_prospectos_pdf():
    filters = _prospectos_filters_from_request()
    rows = _filter_prospectos(_load_prospectos(), filters)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=24 * mm,
        bottomMargin=38 * mm,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="EncabezadoProspectos", fontName="Helvetica", fontSize=9, leading=12, spaceAfter=4, splitLongWords=False))
    styles.add(ParagraphStyle(name="ProspectosCell", fontName="Helvetica", fontSize=7.2, leading=8.5, splitLongWords=False))
    styles.add(ParagraphStyle(name="ProspectosCenter", fontName="Helvetica", fontSize=7.2, leading=8.5, alignment=1, splitLongWords=False))

    elems = []

    def encabezado(canv, doc_):
        canv.saveState()
        canv.setFillColor(colors.HexColor(MAR_BLUE))
        canv.rect(0, A4[1] - 40, A4[0], 40, stroke=0, fill=1)

        logo_path = os.path.join(app.static_folder or "static", "logo.png")
        if os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                max_w = 22.5 * mm
                scale = max_w / iw
                canv.drawImage(img, 12, A4[1] - (ih * scale) - 8, width=max_w, height=ih * scale, mask="auto")
            except Exception:
                pass

        canv.setFont("Helvetica-Bold", 14)
        canv.setFillColor(colors.white)
        canv.drawRightString(A4[0] - 12, A4[1] - 18, "PROSPECTOS")
        canv.setFont("Helvetica", 10)
        canv.drawRightString(A4[0] - 12, A4[1] - 31, "Recubrimientos Especializados")
        canv.restoreState()

    def footer(canv, doc_):
        canv.saveState()
        division_path = os.path.join(app.static_folder or "static", "division.png")
        if os.path.exists(division_path):
            try:
                canv.drawImage(division_path, (A4[0] - 155 * mm) / 2, 45, width=155 * mm, height=3 * mm, mask="auto")
            except Exception:
                pass

        canv.setFont("Helvetica-Bold", 9)
        canv.setFillColor(colors.HexColor(MAR_BLUE))
        canv.drawCentredString(A4[0] / 2, 35, "POLIUTECH - Recubrimientos Especializados")
        canv.setFont("Helvetica", 8)
        canv.setFillColor(colors.HexColor("#333333"))
        canv.drawCentredString(A4[0] / 2, 25, "Campos Eliseos 223 Oficina 602 - Col. Polanco V Seccion - Miguel Hidalgo, CDMX 11560")
        canv.drawCentredString(A4[0] / 2, 15, "Tel: 55 5938 6530 / 55 5938 0536 - info@poliutech.com - www.poliutech.com")
        canv.restoreState()

    generated_at = now_cdmx_naive().strftime("%d/%m/%Y %H:%M")
    meta_data = [
        [
            Paragraph(f"<b>Fecha de exportación:</b> {generated_at}", styles["EncabezadoProspectos"]),
            Paragraph(f"<b>Total de registros:</b> {len(rows)}", styles["EncabezadoProspectos"]),
        ],
        [
            Paragraph("<b>Filtro título/contacto:</b>", styles["EncabezadoProspectos"]),
            Paragraph((filters.get("titulo") or filters.get("contacto") or "Todos").upper() if (filters.get("titulo") or filters.get("contacto")) else "Todos", styles["EncabezadoProspectos"]),
        ],
        [
            Paragraph("<b>Filtro status:</b>", styles["EncabezadoProspectos"]),
            Paragraph(filters.get("status") or "Todos", styles["EncabezadoProspectos"]),
        ],
    ]
    meta_tbl = Table(meta_data, colWidths=[95 * mm, 95 * mm], hAlign="LEFT")
    meta_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))
    elems.append(meta_tbl)
    elems.append(Spacer(1, 6))

    data = [[
        "TITULO",
        "DESCRIPCION",
        "CONTACTO",
        "TELEFONO",
        "CORREO",
        "STATUS",
    ]]
    for row in rows:
        data.append([
            Paragraph(_truncate_pdf_text(row.get("titulo", ""), 38), styles["ProspectosCell"]),
            Paragraph(_truncate_pdf_text(row.get("descripcion", ""), 78), styles["ProspectosCell"]),
            Paragraph(_truncate_pdf_text(row.get("contacto", ""), 30), styles["ProspectosCell"]),
            Paragraph(_truncate_pdf_text(row.get("telefono", ""), 22), styles["ProspectosCenter"]),
            Paragraph(_truncate_pdf_text(row.get("correo", ""), 34), styles["ProspectosCell"]),
            Paragraph(_truncate_pdf_text(row.get("status", ""), 18), styles["ProspectosCenter"]),
        ])

    if len(data) == 1:
        data.append([
            Paragraph("-", styles["ProspectosCenter"]),
            Paragraph("No hay registros para exportar con el filtro actual.", styles["ProspectosCell"]),
            Paragraph("-", styles["ProspectosCenter"]),
            Paragraph("-", styles["ProspectosCenter"]),
            Paragraph("-", styles["ProspectosCenter"]),
            Paragraph("-", styles["ProspectosCenter"]),
        ])

    tbl = Table(
        data,
        colWidths=[28 * mm, 64 * mm, 26 * mm, 22 * mm, 34 * mm, 20 * mm],
        repeatRows=1,
        hAlign="CENTER",
    )
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(MAR_BLUE)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (3, 0), (3, -1), "CENTER"),
        ("ALIGN", (5, 0), (5, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 7.2),
        ("WORDWRAP", (0, 0), (-1, -1), True),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elems.append(tbl)

    doc.build(
        elems,
        onFirstPage=lambda canv, d: (draw_watermark(canv, app), encabezado(canv, d), footer(canv, d)),
        onLaterPages=lambda canv, d: (draw_watermark(canv, app), encabezado(canv, d), footer(canv, d)),
    )

    buf.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response = Response(
        buf.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f'inline; filename="prospectos_{stamp}.pdf"'},
    )
    response.direct_passthrough = False
    return response


@app.route("/registro-obras", methods=["GET", "POST"])
@login_required
def registro_obras():
    rows = _load_registro_obras()

    if request.method == "POST":
        action = (request.form.get("action") or "").strip().lower()
        pending_email_rows: list[dict] = []
        if action == "add":
            send_email = (request.form.get("send_email") or "").strip().lower() in {"1", "true", "on", "yes"}
            row = _normalize_registro_obra_row({
                "numero": "",
                "obra": request.form.get("obra"),
                "ubicacion": request.form.get("ubicacion"),
                "encargado": request.form.get("encargado"),
                "puesto": request.form.get("puesto"),
                "telefono": request.form.get("telefono"),
                "correo": request.form.get("correo"),
                "responsable": request.form.get("responsable"),
            }, len(rows) + 1)
            if not any([row["obra"], row["ubicacion"], row["encargado"], row["puesto"], row["telefono"], row["correo"], row["responsable"]]):
                flash("Captura al menos un dato antes de agregar el registro.", "warning")
                return redirect(url_for("registro_obras"))
            if send_email and not row["correo"]:
                flash("Debes capturar un correo si activas ENVIAR CORREO.", "danger")
                return redirect(url_for("registro_obras"))
            try:
                _parse_email_list(row["correo"])
            except ValueError as e:
                flash(str(e), "danger")
                return redirect(url_for("registro_obras"))
            if not is_admin():
                row["responsable"] = responsable_actual() or row["responsable"]
            row["numero"] = str(len(rows) + 1)
            rows.append(row)
            if send_email:
                pending_email_rows.append(row)
            if send_email:
                flash("Registro agregado.", "success")
            else:
                flash("Registro agregado.", "success")
        elif action == "update":
            row_ids = request.form.getlist("row_id[]")
            obras = request.form.getlist("obra[]")
            ubicaciones = request.form.getlist("ubicacion[]")
            encargados = request.form.getlist("encargado[]")
            puestos = request.form.getlist("puesto[]")
            telefonos = request.form.getlist("telefono[]")
            correos = request.form.getlist("correo[]")
            responsables = request.form.getlist("responsable[]")
            updated_rows = []
            for idx, row_id in enumerate(row_ids):
                numeric_row_id = int(row_id or idx + 1)
                row = _normalize_registro_obra_row({
                    "id": numeric_row_id,
                    "numero": "",
                    "obra": obras[idx] if idx < len(obras) else "",
                    "ubicacion": ubicaciones[idx] if idx < len(ubicaciones) else "",
                    "encargado": encargados[idx] if idx < len(encargados) else "",
                    "puesto": puestos[idx] if idx < len(puestos) else "",
                    "telefono": telefonos[idx] if idx < len(telefonos) else "",
                    "correo": correos[idx] if idx < len(correos) else "",
                    "responsable": responsables[idx] if idx < len(responsables) else "",
                }, numeric_row_id)
                if not any([row["obra"], row["ubicacion"], row["encargado"], row["puesto"], row["telefono"], row["correo"], row["responsable"]]):
                    continue
                try:
                    _parse_email_list(row["correo"])
                except ValueError as e:
                    flash(str(e), "danger")
                    return redirect(url_for("registro_obras"))
                if not is_admin():
                    row["responsable"] = responsable_actual() or row["responsable"]
                updated_rows.append(row)
            rows = updated_rows
            flash("Registros actualizados.", "success")
        elif action == "import":
            uploaded = request.files.get("import_file")
            if not uploaded or not (uploaded.filename or "").strip():
                flash("Selecciona un archivo Excel para importar.", "warning")
                return redirect(url_for("registro_obras"))

            import_responsable = (request.form.get("import_responsable") or "").strip()
            if not is_admin():
                import_responsable = responsable_actual() or import_responsable

            try:
                imported_rows = _load_registro_obras_from_xlsx(uploaded.read(), default_responsable=import_responsable)
            except Exception:
                imported_rows = []

            if not imported_rows:
                flash("No se encontraron registros válidos en el Excel.", "warning")
                return redirect(url_for("registro_obras"))

            existing_keys = {_registro_obra_duplicate_key(row) for row in rows}
            accepted_rows = []
            skipped_duplicates = 0
            for imported in imported_rows:
                try:
                    _parse_email_list(imported["correo"])
                except ValueError as e:
                    flash(f"{e} en el archivo importado.", "danger")
                    return redirect(url_for("registro_obras"))
                if not is_admin():
                    imported["responsable"] = responsable_actual() or imported["responsable"]
                duplicate_key = _registro_obra_duplicate_key(imported)
                if duplicate_key in existing_keys:
                    skipped_duplicates += 1
                    continue
                existing_keys.add(duplicate_key)
                accepted_rows.append(imported)
                rows.append(imported)

            if not accepted_rows:
                flash("No se importaron registros nuevos; todos ya existían.", "warning")
                return redirect(url_for("registro_obras"))

            message = f"Se importaron {len(accepted_rows)} registros desde Excel."
            if skipped_duplicates:
                message += f" Se omitieron {skipped_duplicates} duplicados."
            flash(message, "success")
        elif action == "delete":
            selected_ids = {int(value) for value in request.form.getlist("selected_ids[]") if str(value).strip().isdigit()}
            rows = [row for row in rows if int(row.get("id", 0) or 0) not in selected_ids]
            flash("Registros eliminados.", "success")

        for idx, row in enumerate(rows, start=1):
            row["numero"] = str(idx)
        _save_registro_obras(rows)
        for row in rows:
            _sync_cliente_from_registro_obra(row)
        db.session.commit()
        if pending_email_rows:
            email_errors = []
            email_sent = 0
            for row in pending_email_rows:
                try:
                    _send_registro_obra_email(row)
                    email_sent += 1
                except Exception as e:
                    email_errors.append(f"{row.get('obra') or row.get('encargado') or row.get('correo')}: {e}")
            if email_sent:
                flash("Envío de correo exitoso.", "success")
            if email_errors:
                flash("No se pudo enviar: " + " | ".join(email_errors), "warning")
        return redirect(url_for("registro_obras"))

    filters = _registro_obras_filters_from_request()
    filtered_rows = _filter_registro_obras(rows, filters)
    return render_template(
        "registro_obras.html",
        title="Registro de obras",
        rows=rows,
        filtered_rows=filtered_rows,
        filters=filters,
        default_responsable=responsable_actual() or "",
        default_numero=str(len(rows) + 1),
    )


@app.route("/registro-obras/<int:registro_id>/seguimiento")
@login_required
def registro_obra_seguimiento(registro_id: int):
    registro = RegistroObra.query.get_or_404(registro_id)
    require_registro_obra_owner_or_admin(registro)
    return render_template(
        "registro_obra_seguimiento.html",
        registro=registro,
        cliente_info=_cliente_seguimiento_payload(
            nombre=registro.encargado or registro.obra,
            correo=registro.correo,
            telefono=registro.telefono,
            direccion=registro.ubicacion,
            responsable=registro.responsable,
            titulo="Datos del cliente / obra",
            extras=[
                {"label": "Obra", "value": registro.obra},
                {"label": "Puesto", "value": registro.puesto},
            ],
        ),
        seguimientos=registro.seguimientos,
        mention_users=_usuarios_menciones_payload(),
        title=f"Seguimiento obra {registro.obra}",
    )


@app.route("/registro-obras/<int:registro_id>/seguimiento", methods=["POST"])
@login_required
def crear_registro_obra_seguimiento(registro_id: int):
    registro = RegistroObra.query.get_or_404(registro_id)
    require_registro_obra_owner_or_admin(registro)
    comentario = (request.form.get("comentario") or "").strip()
    tagged_users = _usuarios_mencionados_en_comentario(comentario)

    if not comentario:
        flash("Escribe un comentario de seguimiento.", "warning")
        return redirect(url_for("registro_obra_seguimiento", registro_id=registro.id))

    autor = (responsable_actual() or "Sistema").strip()
    seg = RegistroObraSeguimiento(
        registro_obra_id=registro.id,
        usuario_id=getattr(current_user, "id", None),
        autor=autor,
        comentario=comentario,
        fecha_seguimiento=now_cdmx_naive(),
        actualizado_en=now_cdmx_naive(),
    )
    db.session.add(seg)
    db.session.commit()
    try:
        _notify_tagged_followup(
            tagged_users=tagged_users,
            module_label="Registro de obras",
            item_label=registro.obra or f"Obra #{registro.id}",
            autor=autor,
            comentario=comentario,
            view_endpoint="registro_obra_seguimiento",
            view_params={"registro_id": registro.id, "_anchor": f"seguimiento-{seg.id}"},
        )
    except Exception as exc:
        logger.exception("No se pudo notificar etiquetas de obra %s", registro.id)
        flash(f"Seguimiento guardado, pero no se pudo enviar correo a etiquetados: {exc}", "warning")
        return redirect(url_for("registro_obra_seguimiento", registro_id=registro.id, _anchor=f"seguimiento-{seg.id}"))
    flash("Seguimiento de obra guardado correctamente.", "success")
    return redirect(url_for("registro_obra_seguimiento", registro_id=registro.id, _anchor=f"seguimiento-{seg.id}"))


@app.route("/registro-obras/<int:registro_id>/seguimiento/<int:seg_id>/editar", methods=["POST"])
@login_required
def editar_registro_obra_seguimiento(registro_id: int, seg_id: int):
    registro = RegistroObra.query.get_or_404(registro_id)
    require_registro_obra_owner_or_admin(registro)
    seg = RegistroObraSeguimiento.query.filter_by(id=seg_id, registro_obra_id=registro.id).first_or_404()
    require_registro_obra_followup_author_or_admin(seg)

    comentario = (request.form.get("comentario") or "").strip()
    if not comentario:
        flash("El comentario no puede quedar vacío.", "warning")
        return redirect(url_for("registro_obra_seguimiento", registro_id=registro.id))

    seg.comentario = comentario
    seg.actualizado_en = now_cdmx_naive()
    db.session.commit()
    flash("Seguimiento actualizado.", "success")
    return redirect(url_for("registro_obra_seguimiento", registro_id=registro.id, _anchor=f"seguimiento-{seg.id}"))


@app.route("/registro-obras/<int:registro_id>/seguimiento/<int:seg_id>/eliminar", methods=["POST"])
@login_required
def eliminar_registro_obra_seguimiento(registro_id: int, seg_id: int):
    registro = RegistroObra.query.get_or_404(registro_id)
    require_registro_obra_owner_or_admin(registro)
    seg = RegistroObraSeguimiento.query.filter_by(id=seg_id, registro_obra_id=registro.id).first_or_404()
    require_registro_obra_followup_author_or_admin(seg)
    db.session.delete(seg)
    db.session.commit()
    flash("Seguimiento eliminado.", "success")
    return redirect(url_for("registro_obra_seguimiento", registro_id=registro.id))


@app.route("/registro-obras/export.xlsx")
@login_required
def export_registro_obras_xlsx():
    filters = _registro_obras_filters_from_request()
    rows = _filter_registro_obras(_load_registro_obras(), filters)
    body_rows = [
        ["", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", ""],
        ["N°", "OBRA", "UBICACIÓN", "ENCARGADO", "PUESTO", "TELEFONO", "CORREO", "RESPONSABLE"],
    ]
    for row in rows:
        body_rows.append([
            row.get("numero", ""),
            row.get("obra", ""),
            row.get("ubicacion", ""),
            row.get("encargado", ""),
            row.get("puesto", ""),
            row.get("telefono", ""),
            row.get("correo", ""),
            row.get("responsable", ""),
        ])

    output_bytes = _build_matrix_xlsx("Registro Obras", body_rows, column_widths=[10, 34, 28, 24, 20, 18, 32, 18])
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Response(
        output_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="registro_obras_{stamp}.xlsx"'},
    )


@app.route("/api/mobile/login", methods=["POST"])
def api_mobile_login():
    payload = request.get_json(silent=True) or {}
    nombre = (payload.get("nombre") or payload.get("usuario") or "").strip()
    password = str(payload.get("password") or "").strip()
    if not nombre or not password:
        return _mobile_json_error("Faltan credenciales.", 400)

    user = Usuario.query.filter(db.func.lower(Usuario.nombre) == nombre.lower()).first()
    if not user or not user.check_password(password):
        return _mobile_json_error("Credenciales incorrectas.", 401)

    return jsonify({
        "ok": True,
        "token": _issue_mobile_token(user),
        "user": {
            "id": user.id,
            "nombre": user.nombre,
            "rol": user.rol,
            "email": user.correo or "",
            "correo": user.correo or "",
            "responsable": _mobile_user_responsable(user),
        },
    })


@app.route("/api/mobile/push-token", methods=["POST"])
@require_mobile_auth
def api_mobile_push_token_register():
    payload = request.get_json(silent=True) or {}
    token = (payload.get("token") or "").strip()
    if not token:
        return _mobile_json_error("Falta el token push.", 400)

    user = g.mobile_user
    device = _upsert_mobile_device(
        user,
        token=token,
        plataforma=(payload.get("platform") or "android"),
        device_name=(payload.get("device_name") or ""),
        app_version=(payload.get("app_version") or ""),
    )
    return jsonify({
        "ok": True,
        "device": {
            "id": device.id,
            "platform": device.plataforma,
            "active": bool(device.is_active),
        },
        "firebase_configured": bool(_firebase_is_configured()),
    })


@app.route("/api/mobile/session-token", methods=["GET"])
@login_required
def api_mobile_session_token():
    user = current_user
    return jsonify({
        "ok": True,
        "token": _issue_mobile_token(user),
        "user": {
            "id": user.id,
            "nombre": user.nombre,
            "rol": user.rol,
            "email": user.correo or "",
            "correo": user.correo or "",
            "responsable": _mobile_user_responsable(user),
        },
    })


@app.route("/api/mobile/push-token", methods=["DELETE"])
@require_mobile_auth
def api_mobile_push_token_unregister():
    payload = request.get_json(silent=True) or {}
    token = (payload.get("token") or "").strip()
    if not token:
        return _mobile_json_error("Falta el token push.", 400)
    _deactivate_mobile_device(token)
    return jsonify({"ok": True})


@app.route("/api/mobile/registro-obras", methods=["GET"])
@require_mobile_auth
def api_mobile_registro_obras_list():
    user = g.mobile_user
    rows = _load_registro_obras()
    items = _filter_registro_obras_for_mobile(
        rows,
        user,
        obra=request.args.get("obra", ""),
        responsable=request.args.get("responsable", ""),
    )
    return jsonify({"ok": True, "items": items})


@app.route("/api/mobile/cotizaciones/pendientes", methods=["GET"])
@require_mobile_auth
def api_mobile_pending_quotes():
    query = Cotizacion.query.outerjoin(Cliente, Cotizacion.cliente_id == Cliente.id)
    query = query.filter(Cotizacion.eliminada_en.is_(None))
    query = query.filter(db.func.upper(Cotizacion.estatus_aprobacion).in_(["EN REVISIÓN", "EN REVISION"]))

    items = []
    for cot in query.order_by(Cotizacion.fecha.desc()).limit(100).all():
        items.append({
            "id": cot.id,
            "folio": cot.folio or "",
            "fecha": cot.fecha.isoformat() if cot.fecha else "",
            "estatus": cot.estatus or "",
            "estatus_aprobacion": cot.estatus_aprobacion or "EN REVISIÓN",
            "especialidad": cot.especialidad or "",
            "total": cot.total or 0,
            "responsable": cot.responsable or "",
            "proyecto": cot.proyecto or "",
            "cliente": cot.cliente.nombre_cliente if cot.cliente else "",
            "pdf_url": _mobile_quote_pdf_url(cot.id),
        })
    return jsonify({"ok": True, "items": items})


@app.route("/api/mobile/dashboard/summary", methods=["GET"])
@require_mobile_auth
def api_mobile_dashboard_summary():
    query = _cotizaciones_activas_query()

    total_cotizaciones = query.count()
    total_importe = float(
        query.with_entities(db.func.coalesce(db.func.sum(Cotizacion.total), 0)).scalar() or 0
    )
    rows = (
        query.with_entities(Cotizacion.estatus, db.func.count(Cotizacion.id))
        .group_by(Cotizacion.estatus)
        .all()
    )
    by_status = {status: 0 for status in VALID_ESTATUS}
    for status, count in rows:
        by_status[(status or "").strip().upper()] = int(count or 0)

    return jsonify({
        "ok": True,
        "kpis": {
            "total_cotizaciones": int(total_cotizaciones),
            "total_importe": total_importe,
        },
        "status_breakdown": by_status,
        "valid_estatus": VALID_ESTATUS,
    })


@app.route("/api/mobile/cotizaciones", methods=["GET"])
@require_mobile_auth
def api_mobile_quotes():
    estatus = (request.args.get("estatus") or "").strip().upper()
    query = Cotizacion.query.outerjoin(Cliente, Cotizacion.cliente_id == Cliente.id)
    query = query.filter(Cotizacion.eliminada_en.is_(None))
    if estatus:
        query = query.filter(Cotizacion.estatus == estatus)

    items = []
    for cot in query.order_by(Cotizacion.fecha.desc()).all():
        items.append({
            "id": cot.id,
            "folio": cot.folio or "",
            "fecha": cot.fecha.isoformat() if cot.fecha else "",
            "estatus": cot.estatus or "",
            "estatus_aprobacion": cot.estatus_aprobacion or "EN REVISIÓN",
            "especialidad": cot.especialidad or "",
            "total": cot.total or 0,
            "responsable": cot.responsable or "",
            "proyecto": cot.proyecto or "",
            "cliente": cot.cliente.nombre_cliente if cot.cliente else "",
            "pdf_url": _mobile_quote_pdf_url(cot.id),
        })
    return jsonify({"ok": True, "items": items, "valid_estatus": VALID_ESTATUS})


@app.route("/api/mobile/cotizaciones/voz", methods=["POST"])
@require_mobile_auth
def api_mobile_voice_quote():
    payload = request.get_json(silent=True) or {}
    command_raw = str(payload.get("comando") or payload.get("transcript") or "").strip()
    client_override = str(payload.get("cliente") or "").strip()
    notes = str(payload.get("notas") or "").strip()
    conditions_raw = str(payload.get("condiciones") or payload.get("condiciones_raw") or "").strip()
    confirmar = bool(payload.get("confirmar"))
    if not command_raw:
        return _mobile_json_error("Dicta o escribe un comando antes de continuar.", 400)

    try:
        preview = _voice_preview_payload_for_mobile(
            command_raw=command_raw,
            user=g.mobile_user,
            client_override=client_override,
            notes=notes,
            conditions_raw=conditions_raw,
        )
    except ValueError as exc:
        return _mobile_json_error(str(exc), 400)
    except Exception as exc:
        try:
            logger.exception("Error interpretando cotización por voz")
        except Exception:
            pass
        return _mobile_json_error(f"No se pudo interpretar el comando: {exc}", 500)

    if not confirmar:
        return jsonify({
            "ok": True,
            "modo": "preview",
            "preview": preview,
        })

    cot = _create_mobile_voice_quote(preview, g.mobile_user)
    return jsonify({
        "ok": True,
        "modo": "created",
        "preview": preview,
        "cotizacion": {
            "id": cot.id,
            "folio": cot.folio or "",
            "estatus": cot.estatus or "",
            "estatus_aprobacion": cot.estatus_aprobacion or "EN REVISIÓN",
            "especialidad": cot.especialidad or "",
            "total": float(cot.total or 0),
            "cliente": cot.cliente.nombre_cliente if cot.cliente else "",
            "pdf_url": _mobile_quote_pdf_url(cot.id),
        },
        "mensaje": f"Cotización {cot.folio} creada desde comando de voz.",
    }), 201


@app.route("/api/mobile/cotizaciones/voz/transcribir", methods=["POST"])
@require_mobile_auth
def api_mobile_voice_transcribe():
    uploaded = request.files.get("audio")
    target = (request.form.get("target") or request.args.get("target") or "comando").strip().lower()
    if not uploaded or not (uploaded.filename or "").strip():
        return _mobile_json_error("Adjunta un archivo de audio antes de transcribir.", 400)

    try:
        audio_bytes = uploaded.read()
        transcript = _voice_transcribe_audio_bytes(
            audio_bytes=audio_bytes,
            filename=uploaded.filename or "voz.m4a",
            mime_type=uploaded.mimetype or mimetypes.guess_type(uploaded.filename or "")[0] or "application/octet-stream",
        )
    except ValueError as exc:
        return _mobile_json_error(str(exc), 400)
    except RuntimeError as exc:
        return _mobile_json_error(str(exc), 503)
    except Exception as exc:
        try:
            logger.exception("Error transcribiendo audio de voz")
        except Exception:
            pass
        return _mobile_json_error(f"No se pudo transcribir el audio: {exc}", 500)

    return jsonify({
        "ok": True,
        "target": target,
        "transcript": transcript,
    })


@app.route("/api/mobile/cotizaciones/<int:cot_id>/estatus", methods=["POST"])
@require_mobile_auth
def api_mobile_update_quote_status(cot_id: int):
    user = g.mobile_user
    cot = _cotizacion_activa_or_404(cot_id)
    if not _mobile_user_is_admin(user):
        if (cot.responsable or "").strip().lower() != _mobile_user_responsable(user).lower():
            return _mobile_json_error("No autorizado para esta cotización.", 403)

    payload = request.get_json(silent=True) or {}
    nuevo = (payload.get("estatus") or "").strip().upper()
    if nuevo not in VALID_ESTATUS:
        return _mobile_json_error("Estatus inválido.", 400)

    anterior = (cot.estatus or "").strip().upper()
    if nuevo == anterior:
        return jsonify({"ok": True, "folio": cot.folio or "", "estatus": nuevo, "mensaje": "Sin cambios."})

    cot.estatus = nuevo
    db.session.commit()

    try:
        _send_quote_status_push(cot, anterior, nuevo)
    except Exception as exc:
        logger.warning("Push de estatus móvil fallida: %s", exc)

    try:
        body = (
            f"🔄 *Actualización de estatus*\\n"
            f"Folio: *{cot.folio}*\\n"
            f"Anterior: {anterior}\\n"
            f"Nuevo: *{nuevo}*\\n"
            f"Total: {money(cot.total)}"
        )
        send_whatsapp_multi(ADMIN_LIST, body)
    except Exception as exc:
        logger.warning("WhatsApp de estatus móvil falló: %s", exc)

    return jsonify({
        "ok": True,
        "folio": cot.folio or "",
        "estatus": nuevo,
        "mensaje": f"Estatus de la cotización {cot.folio} actualizado a {nuevo}.",
    })


@app.route("/api/mobile/cotizaciones/<int:cot_id>/revision", methods=["POST"])
@require_mobile_auth
def api_mobile_quote_review_decision(cot_id: int):
    user = g.mobile_user
    user_name = (_mobile_user_responsable(user) or getattr(user, "nombre", "") or "").strip().lower()
    user_email = (getattr(user, "correo", "") or "").strip().lower()
    if user.id != 18 and not user_name.startswith("hansel") and user_email != "hjaramillo@poliutech.com":
        return _mobile_json_error("Solo Hansel puede aprobar o rechazar cotizaciones.", 403)

    cot = _cotizacion_activa_or_404(cot_id)
    payload = request.get_json(silent=True) or {}
    action = (payload.get("action") or "").strip().lower()
    reason = (payload.get("reason") or payload.get("motivo") or "").strip()
    status_by_action = {
        "approve": "APROBADA",
        "approved": "APROBADA",
        "aprobar": "APROBADA",
        "reject": "RECHAZADA",
        "rejected": "RECHAZADA",
        "rechazar": "RECHAZADA",
    }
    selected_status = status_by_action.get(action)
    if not selected_status:
        return _mobile_json_error("Acción inválida.", 400)
    if selected_status == "RECHAZADA" and not reason:
        return _mobile_json_error("Captura el motivo del rechazo.", 400)

    author = _mobile_user_responsable(user) or getattr(user, "nombre", "") or "Hansel"
    seg = _apply_quote_review_decision(
        cot,
        selected_status,
        reason,
        actor=user,
        author_label=f"Revision movil: {author}",
    )
    return jsonify({
        "ok": True,
        "cotizacion": _mobile_quote_to_json(cot),
        "seguimiento_id": seg.id,
        "estatus": selected_status,
        "reason": reason,
    })


@app.route("/api/mobile/cotizaciones/<int:cot_id>/seguimiento/<int:seg_id>", methods=["GET"])
@require_mobile_auth
def api_mobile_quote_followup_detail(cot_id: int, seg_id: int):
    user = g.mobile_user
    cot = _cotizacion_activa_or_404(cot_id)
    if not _mobile_user_can_access_quote(user, cot):
        return _mobile_json_error("No autorizado para esta cotización.", 403)

    seg = CotizacionSeguimiento.query.filter_by(id=seg_id, cotizacion_id=cot.id).first()
    if not seg:
        return _mobile_json_error("Seguimiento no encontrado.", 404)

    return jsonify({
        "ok": True,
        "cotizacion": {
            "id": cot.id,
            "folio": cot.folio or "",
            "estatus": cot.estatus or "",
            "estatus_aprobacion": cot.estatus_aprobacion or "EN REVISIÓN",
            "especialidad": cot.especialidad or "",
            "responsable": cot.responsable or "",
            "cliente": cot.cliente.nombre_cliente if cot.cliente else "",
        },
        "seguimiento": {
            "id": seg.id,
            "autor": seg.autor or "",
            "comentario": seg.comentario or "",
            "fecha": seg.fecha_seguimiento.isoformat() if seg.fecha_seguimiento else "",
            "actualizado_en": seg.actualizado_en.isoformat() if seg.actualizado_en else "",
        },
    })


@app.route("/api/mobile/cotizaciones/<int:cot_id>/pdf", methods=["GET"])
@require_mobile_auth
def api_mobile_quote_pdf(cot_id: int):
    cot = _cotizacion_activa_or_404(cot_id)
    if not _mobile_user_can_access_quote(g.mobile_user, cot):
        return _mobile_json_error("No autorizado para esta cotización.", 403)
    return _build_cotizacion_pdf_response(cot)


@app.route("/api/mobile/registro-obras", methods=["POST"])
@require_mobile_auth
def api_mobile_registro_obras_create():
    user = g.mobile_user
    rows = _load_registro_obras()
    payload = request.get_json(silent=True) or {}
    send_email = bool(payload.get("send_email"))
    row = _normalize_registro_obra_row({
        "numero": "",
        "obra": payload.get("obra"),
        "ubicacion": payload.get("ubicacion"),
        "encargado": payload.get("encargado"),
        "puesto": payload.get("puesto"),
        "telefono": payload.get("telefono"),
        "correo": payload.get("correo"),
        "responsable": payload.get("responsable"),
    }, len(rows) + 1)

    if not row["obra"]:
        return _mobile_json_error("El campo 'obra' es obligatorio.", 400)
    if send_email and not row["correo"]:
        return _mobile_json_error("Debes capturar un correo si activas ENVIAR CORREO.", 400)
    try:
        _parse_email_list(row["correo"])
    except ValueError as e:
        return _mobile_json_error(str(e), 400)
    if not _mobile_user_is_admin(user):
        row["responsable"] = _mobile_user_responsable(user)

    row["numero"] = str(len(rows) + 1)
    rows.append(row)
    _save_registro_obras(rows)
    _sync_cliente_from_registro_obra(row)
    db.session.commit()
    if send_email:
        try:
            _send_registro_obra_email(row)
        except Exception as e:
            return jsonify({"ok": True, "item": row, "email_warning": str(e)}), 201
    return jsonify({"ok": True, "item": row}), 201


@app.route("/api/mobile/registro-obras/<int:item_id>", methods=["PUT"])
@require_mobile_auth
def api_mobile_registro_obras_update(item_id: int):
    user = g.mobile_user
    rows = _load_registro_obras()
    target = next((row for row in rows if int(row.get("id", 0) or 0) == item_id), None)
    if not target:
        return _mobile_json_error("Registro no encontrado.", 404)

    owner = (target.get("responsable") or "").strip().lower()
    if not _mobile_user_is_admin(user) and owner != _mobile_user_responsable(user).lower():
        return _mobile_json_error("No autorizado.", 403)

    payload = request.get_json(silent=True) or {}
    send_email = bool(payload.get("send_email"))
    updated = _normalize_registro_obra_row({
        "id": target.get("id"),
        "numero": target.get("numero"),
        "obra": payload.get("obra"),
        "ubicacion": payload.get("ubicacion"),
        "encargado": payload.get("encargado"),
        "puesto": payload.get("puesto"),
        "telefono": payload.get("telefono"),
        "correo": payload.get("correo"),
        "responsable": payload.get("responsable"),
    }, item_id)
    if not updated["obra"]:
        return _mobile_json_error("El campo 'obra' es obligatorio.", 400)
    if send_email and not updated["correo"]:
        return _mobile_json_error("Debes capturar un correo si activas ENVIAR CORREO.", 400)
    try:
        _parse_email_list(updated["correo"])
    except ValueError as e:
        return _mobile_json_error(str(e), 400)
    if not _mobile_user_is_admin(user):
        updated["responsable"] = _mobile_user_responsable(user)

    target.update(updated)
    for idx, row in enumerate(rows, start=1):
        row["numero"] = str(idx)
    _save_registro_obras(rows)
    _sync_cliente_from_registro_obra(target)
    db.session.commit()
    if send_email:
        try:
            _send_registro_obra_email(target)
        except Exception as e:
            return jsonify({"ok": True, "item": target, "email_warning": str(e)})
    return jsonify({"ok": True, "item": target})


@app.route("/api/mobile/registro-obras/bulk-delete", methods=["POST"])
@require_mobile_auth
def api_mobile_registro_obras_bulk_delete():
    user = g.mobile_user
    rows = _load_registro_obras()
    payload = request.get_json(silent=True) or {}
    raw_ids = payload.get("ids") or []
    if not isinstance(raw_ids, list):
        return _mobile_json_error("El campo 'ids' debe ser una lista.", 400)
    selected_ids = {int(value) for value in raw_ids if str(value).strip().isdigit()}
    if not selected_ids:
        return _mobile_json_error("No se enviaron ids válidos.", 400)

    kept = []
    deleted = 0
    for row in rows:
        row_id = int(row.get("id", 0) or 0)
        owner = (row.get("responsable") or "").strip().lower()
        can_delete = _mobile_user_is_admin(user) or owner == _mobile_user_responsable(user).lower()
        if row_id in selected_ids and can_delete:
            deleted += 1
            continue
        kept.append(row)

    for idx, row in enumerate(kept, start=1):
        row["numero"] = str(idx)
    _save_registro_obras(kept)
    db.session.commit()
    return jsonify({"ok": True, "deleted": deleted})


@app.route("/admin/cotizaciones/importar", methods=["GET", "POST"])
@login_required
def importar_cotizacion_externa():
    if not is_admin():
        abort(403)

    detected = None

    if request.method == "POST":
        uploaded = request.files.get("cotizacion_pdf")
        responsable_destino = (request.form.get("responsable_destino") or "").strip() or responsable_actual()

        if not uploaded or not (uploaded.filename or "").strip():
            flash("Selecciona un PDF antes de importar.", "danger")
        else:
            try:
                pdf_bytes = uploaded.read()
                if not pdf_bytes:
                    raise ValueError("El archivo PDF llego vacio.")

                payload = build_import_payload_from_pdf(
                    pdf_bytes,
                    uploaded.filename or "cotizacion.pdf",
                    responsable_hint=responsable_destino,
                )
                detected = _normalize_import_payload(payload)
                subtotal_detectado = sum((it.get("cantidad") or 0) * (it.get("precio_unitario") or 0) for it in detected["items"])
                total_detectado = subtotal_detectado * (1 + ((detected.get("iva_porc") or 0) / 100.0))
                detected["total_calculado"] = fmt(total_detectado)
                cot = import_external_quote_payload(payload, source_label=uploaded.filename or "cotizacion.pdf")
                flash(f"Cotizacion importada correctamente: {cot.folio}", "success")
                return redirect(url_for("view_cotizacion", cot_id=cot.id))
            except Exception as e:
                try:
                    print(f"[IMPORTADOR PDF] ERROR: {e}", file=sys.stderr)
                    traceback.print_exc()
                except Exception:
                    pass
                flash(f"No se pudo importar la cotizacion: {e}", "danger")

    return render_template(
        "cotizacion_import.html",
        title="Importar cotizacion - Sistema MAR",
        detected=detected,
    )
@app.route("/admin/catalogos")
@login_required
def admin_catalogos():
    page_clientes = request.args.get("page_clientes", 1, type=int)
    page_conceptos = request.args.get("page_conceptos", 1, type=int)
    q_clientes = (request.args.get("q_clientes") or "").strip()
    q_conceptos = (request.args.get("q_conceptos") or "").strip()

    qc = Cliente.query
    if not is_admin():
        qc = qc.filter(Cliente.responsable == responsable_actual())
    if q_clientes:
        like_clientes = f"%{q_clientes}%"
        qc = qc.filter(or_(
            Cliente.nombre_cliente.ilike(like_clientes),
            Cliente.empresa.ilike(like_clientes),
            Cliente.responsable.ilike(like_clientes),
            Cliente.correo.ilike(like_clientes),
            Cliente.telefono.ilike(like_clientes),
            Cliente.rfc.ilike(like_clientes),
        ))

    conceptos_q = Concepto.query
    if q_conceptos:
        like_conceptos = f"%{q_conceptos}%"
        conceptos_q = conceptos_q.filter(or_(
            Concepto.nombre_concepto.ilike(like_conceptos),
            Concepto.unidad.ilike(like_conceptos),
            Concepto.sistema.ilike(like_conceptos),
            Concepto.descripcion.ilike(like_conceptos),
        ))

    clientes_pag = qc.order_by(Cliente.id.desc()).paginate(page=page_clientes, per_page=10, error_out=False)
    conceptos_pag = conceptos_q.order_by(Concepto.id.desc()).paginate(page=page_conceptos, per_page=10, error_out=False)

    return render_template(
        "admin_catalogos.html",
        title="Admin Catálogos",
        clientes=clientes_pag.items,
        clientes_pag=clientes_pag,
        conceptos=conceptos_pag.items,
        conceptos_pag=conceptos_pag,
        q_clientes=q_clientes,
        q_conceptos=q_conceptos,
    )

# ---------------------------------------------------------
# Autocompletar (con filtro por responsable en clientes)
# ---------------------------------------------------------
@app.route("/api/clientes/suggest")
@login_required
def api_clientes_suggest():
    q = (request.args.get("q", "")).strip()
    if len(q) < 1:
        return jsonify([])

    resq = (Cliente.query
            .filter(
                (Cliente.nombre_cliente.ilike(f"%{q}%")) |
                (Cliente.empresa.ilike(f"%{q}%"))
            ))

    if not is_admin():
        resq = resq.filter(Cliente.responsable == responsable_actual())

    res = (resq.order_by(Cliente.nombre_cliente).limit(10).all())

    return jsonify([{
        "label": f"{c.nombre_cliente} · {c.empresa}" if c.empresa else c.nombre_cliente,
        "nombre_cliente": c.nombre_cliente,
        "empresa": c.empresa,
        "responsable": c.responsable,
        "correo": c.correo,
        "telefono": c.telefono,
        "direccion": c.direccion,
        "rfc": c.rfc,
        "sistema": getattr(c, "sistema", "") or ""
    } for c in res])

@app.route("/api/conceptos/suggest")
@login_required
def api_conceptos_suggest():
    q = (request.args.get("q", "")).strip()
    if len(q) < 1:
        return jsonify([])
    res = (Concepto.query
           .filter(Concepto.nombre_concepto.ilike(f"%{q}%"))
           .order_by(Concepto.nombre_concepto).limit(10).all())
    return jsonify([{
        "label": c.nombre_concepto,
        "nombre_concepto": c.nombre_concepto,
        "unidad": c.unidad,
        "precio_unitario": c.precio_unitario,
        "descripcion": c.descripcion
    } for c in res])

# ---------------------------------------------------------
# Crear/Editar/Ver/Exportar Cotizaciones
# ---------------------------------------------------------
@app.route("/cotizaciones/crear", methods=["POST"])
@login_required
def crear_cotizacion():
    f = request.form

    nombre_cliente = (f.get("cliente") or f.get("cliente_nombre") or "").strip()
    empresa = (f.get("empresa") or "").strip()
    proyecto = (f.get("proyecto") or "").strip() or None
    ciudad_trabajo = (f.get("ciudad_trabajo") or "").strip().upper() or None
    moneda = normalize_moneda(f.get("moneda"))

    # === responsable_final ===
    # USER: siempre su nombre (primer nombre)
    # ADMIN: puede mandar responsable desde form; si no manda, queda vacío
    if is_admin():
        responsable_final = (f.get("responsable") or "").strip()
        # si admin dejó vacío, NO inventamos; queda None
    else:
        responsable_final = responsable_actual()

    responsable_final = responsable_final or None

    # --- CREAR O BUSCAR CLIENTE ---
    cliente = None
    if nombre_cliente:
        q = Cliente.query.filter(db.func.lower(Cliente.nombre_cliente) == nombre_cliente.lower())
        if empresa:
            q = q.filter(db.func.lower(Cliente.empresa) == empresa.lower())

        if not is_admin():
            q = q.filter(Cliente.responsable == (responsable_final or ""))

        cliente = q.first()

        if not cliente:
            cliente = Cliente(
                nombre_cliente=nombre_cliente.strip(),
                empresa=empresa.strip() or None,
                responsable=responsable_final,
                correo=(f.get("correo") or "").strip() or None,
                telefono=(f.get("telefono") or "").strip() or None,
                direccion=(f.get("direccion") or "").strip() or None,
                rfc=(f.get("rfc") or "").strip() or None,  # en BD, aunque en PDF ya no lo mostramos
            )
            db.session.add(cliente)
            db.session.flush()

    iva_porc = parse_float(f.get("iva_porc"), 16.0)

    # --- Zona (descuento) ---
    zona = (f.get("zona") or "").strip()
    ZONA_PORC = {
        "Zona Norte": 10.0,
        "Zona Centro": 5.0,
        "Bajío": 10.0,
        "Zona Sur": 15.0,
        "Frontera": 8.0,
    }
    desc_porc = float(ZONA_PORC.get(zona, 0.0))

    cot = Cotizacion(
        folio=generar_folio(),
        fecha=now_cdmx_naive(),
        cliente_id=cliente.id if cliente else None,
        estatus=(f.get("estatus") or "PENDIENTE").upper(),
        estatus_aprobacion=(f.get("estatus_aprobacion") or "EN REVISIÓN").upper(),
        especialidad=(f.get("especialidad") or "").strip() or None,
        notas=(f.get("notas") or "").strip() or None,
        last_whatsapp_at=None,
        responsable=responsable_final,
        proyecto=proyecto,
        ciudad_trabajo=ciudad_trabajo,
        moneda=moneda,
    )
    db.session.add(cot)
    db.session.flush()

    nombres = f.getlist("item_nombre_concepto[]")
    unidades = f.getlist("item_unidad[]")
    capitulos = f.getlist("item_capitulo[]")
    cantidades = f.getlist("item_cantidad[]")
    precios = f.getlist("item_precio[]")
    sistemas = f.getlist("item_sistema[]")
    descripciones = f.getlist("item_descripcion[]")

    subtotal = 0.0
    n = max(len(nombres), len(unidades), len(cantidades), len(precios))
    for i in range(n):
        nom = (nombres[i] if i < len(nombres) else "").strip()
        if not nom:
            continue
        uni = (unidades[i] if i < len(unidades) else "").strip()
        cap = (capitulos[i] if i < len(capitulos) else "").strip() or None
        cant = parse_float(cantidades[i] if i < len(cantidades) else 0, 0.0)
        pu   = parse_float(precios[i] if i < len(precios) else 0, 0.0)
        sis  = (sistemas[i] if i < len(sistemas) else "").strip()
        desc = (descripciones[i] if i < len(descripciones) else "") or ""

        line_subtotal = cant * pu
        subtotal += line_subtotal

        concepto = Concepto.query.filter_by(nombre_concepto=nom).first()
        if not concepto:
            concepto = Concepto(
                nombre_concepto=nom,
                unidad=uni or None,
                precio_unitario=pu,
                descripcion=desc or None
            )
            db.session.add(concepto)
            db.session.flush()

        det = CotizacionDetalle(**_safe_detalle_kwargs(
            cotizacion_id=cot.id,
            concepto_id=concepto.id if concepto else None,
            nombre_concepto=nom,
            unidad=uni,
            capitulo=cap,
            cantidad=cant,
            precio_unitario=pu,
            sistema=sis or None,
            descripcion=desc,
            subtotal=line_subtotal,
        ))
        db.session.add(det)

    # --- aplicar descuento porcentual antes de IVA ---
    descuento_porc_capturado = parse_float(f.get("descuento_total"), desc_porc)
    descuento_porc_aplicado = min(max(descuento_porc_capturado, 0.0), 100.0)
    descuento_total = subtotal * (descuento_porc_aplicado / 100.0)
    subtotal_desc = subtotal - descuento_total
    iva_monto = subtotal_desc * (iva_porc / 100.0)
    total = subtotal_desc + iva_monto

    # --- trazabilidad de zona en Condiciones Comerciales (notas) ---
    if zona:
        zona_line = f"Zona: {zona} ({descuento_porc_aplicado:g}% descuento)"
        notas = (cot.notas or "").strip()
        # elimina cualquier línea previa de Zona:
        notas_lines = [ln for ln in notas.splitlines() if ln.strip() and not ln.strip().lower().startswith("zona:")]
        notas_lines.append(zona_line)
        cot.notas = "\n".join(notas_lines).strip()

    cot.subtotal = fmt(subtotal)
    cot.descuento_total = fmt(descuento_total)
    cot.iva_porc = fmt(iva_porc)
    cot.iva_monto = fmt(iva_monto)
    cot.total = fmt(total)
    db.session.commit()

    _send_quote_created_notification(cot)
    _send_quote_review_email_safely(cot)

    # --- Apertura automática del PDF ---
    pdf_url = url_for("export_cotizacion_pdf", cot_id=cot.id)
    volver = url_for("cotizador")

    return f"""<!DOCTYPE html>
    <html>
    <head>
    <meta charset="utf-8">
    <title>{cot.folio}</title>
    <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
    </head>
    <body>
    <script>
      Swal.fire({{
        icon: 'success',
        title: 'Cotización creada con éxito',
        html: 'Folio: <b>{cot.folio}</b><br>Se abrirá el PDF automáticamente.',
        timer: 2500,
        timerProgressBar: true,
        showConfirmButton: false,
        didOpen: () => {{
          window.open("{pdf_url}", "_blank");
          setTimeout(() => {{
            window.location.href = "{volver}";
          }}, 2500);
        }}
      }});
    </script>
    </body>
    </html>"""

@app.route("/cotizaciones/<int:cot_id>/editar")
@login_required
def editar_cotizacion(cot_id: int):
    c = _cotizacion_activa_or_404(cot_id)
    require_owner_or_admin(c)
    # zona actual (si existe) viene persistida en notas como: "Zona: ... (X% descuento)"
    zona_actual = ""
    try:
        if c.notas:
            for ln in str(c.notas).splitlines():
                if ln.strip().lower().startswith("zona:"):
                    # Zona: <NOMBRE> (..)
                    tmp = ln.split(":", 1)[1].strip()
                    zona_actual = tmp.split("(", 1)[0].strip()
                    break
    except Exception:
        zona_actual = ""
    notas_adicionales, _ = _split_notas_y_zona(c.notas or "")
    descuento_porc_actual = 0.0
    if float(c.subtotal or 0) > 0:
        descuento_porc_actual = (float(c.descuento_total or 0) / float(c.subtotal or 0)) * 100.0
    return render_template("cotizacion_edit.html", c=c, zona_actual=zona_actual, notas_adicionales=notas_adicionales, descuento_porc_actual=descuento_porc_actual, proyectos=_known_project_names(), valid_estatus=VALID_ESTATUS_SEGUIMIENTO, valid_estatus_aprobacion=VALID_ESTATUS_APROBACION, title=f"Editar {c.folio}")

@app.route("/cotizaciones/<int:cot_id>/actualizar", methods=["POST"])
@login_required
def actualizar_cotizacion(cot_id: int):
    c = _cotizacion_activa_or_404(cot_id)
    require_owner_or_admin(c)

    f = request.form

    # === CLIENTE ===
    cliente_nombre = (f.get("cliente") or f.get("cliente_nombre") or "").strip()
    empresa = (f.get("empresa") or "").strip()

    # solo admin puede reasignar responsable
    if is_admin():
        responsable_form = (f.get("responsable") or "").strip()
        responsable_final = responsable_form or c.responsable
    else:
        responsable_final = responsable_actual() or c.responsable

    correo = (f.get("correo") or "").strip()
    telefono = (f.get("telefono") or "").strip()
    direccion = (f.get("direccion") or "").strip()
    rfc = (f.get("rfc") or "").strip()

    cliente = None
    if cliente_nombre:
        cliente = None
        if c.cliente and (c.cliente.nombre_cliente or "").strip().lower() == cliente_nombre.lower():
            cliente = c.cliente
        else:
            cliente = Cliente.query.filter_by(nombre_cliente=cliente_nombre).first()
        if cliente and not is_admin():
            require_cliente_owner_or_admin(cliente)

        if not cliente:
            cliente = Cliente(
                nombre_cliente=cliente_nombre,
                empresa=empresa or None,
                responsable=responsable_final or None,
                correo=correo or None,
                telefono=telefono or None,
                direccion=direccion or None,
                rfc=rfc or None,
            )
            db.session.add(cliente)
            db.session.flush()
            print(f"[INFO] Nuevo cliente agregado (en actualización): {cliente_nombre}")
        else:
            cliente.empresa = empresa or None
            cliente.responsable = responsable_final or None
            cliente.correo = correo or None
            cliente.telefono = telefono or None
            cliente.direccion = direccion or None
            cliente.rfc = rfc or None
        c.cliente_id = cliente.id

    # === ENCABEZADO ===
    estatus_form = (f.get("estatus") or c.estatus or "PENDIENTE").upper()
    if estatus_form not in VALID_ESTATUS_SEGUIMIENTO:
        flash("Selecciona un estatus de seguimiento válido.", "danger")
        return redirect(url_for("editar_cotizacion", cot_id=c.id))
    c.estatus = estatus_form
    c.estatus_aprobacion = "EN REVISIÓN"
    c.especialidad = (f.get("especialidad") or "").strip() or None
    c.notas = (f.get("notas") or "").strip()
    c.responsable = (responsable_final or c.responsable)
    c.proyecto = (f.get("proyecto") or "").strip() or None
    c.ciudad_trabajo = (f.get("ciudad_trabajo") or "").strip().upper() or None
    c.moneda = normalize_moneda(f.get("moneda") or getattr(c, "moneda", None))
    iva_porc = parse_float(f.get("iva_porc"), c.iva_porc or 16.0)

    # --- Zona (descuento) ---
    zona = (f.get("zona") or "").strip()
    ZONA_PORC = {
        "Zona Norte": 10.0,
        "Zona Centro": 5.0,
        "Bajío": 10.0,
        "Zona Sur": 15.0,
        "Frontera": 8.0,
    }
    desc_porc = float(ZONA_PORC.get(zona, 0.0))

    # === LIMPIAR DETALLES EXISTENTES ===
    for d in list(c.detalles):
        db.session.delete(d)

    # === DETALLES NUEVOS ===
    nombres = f.getlist("item_nombre_concepto[]")
    unidades = f.getlist("item_unidad[]")
    capitulos = f.getlist("item_capitulo[]")
    cantidades = f.getlist("item_cantidad[]")
    precios = f.getlist("item_precio[]")
    sistemas = f.getlist("item_sistema[]")
    descripciones = f.getlist("item_descripcion[]")

    subtotal = 0.0
    n = max(len(nombres), len(unidades), len(cantidades), len(precios))
    for i in range(n):
        nombre = (nombres[i] if i < len(nombres) else "").strip()
        if not nombre:
            continue
        unidad = (unidades[i] if i < len(unidades) else "").strip()
        capitulo = (capitulos[i] if i < len(capitulos) else "").strip() or None
        cantidad = parse_float(cantidades[i] if i < len(cantidades) else 0, 0.0)
        precio = parse_float(precios[i] if i < len(precios) else 0, 0.0)
        sistema = (sistemas[i] if i < len(sistemas) else "").strip()
        descripcion = (descripciones[i] if i < len(descripciones) else "").strip()

        linea_subtotal = cantidad * precio
        subtotal += linea_subtotal

        concepto = Concepto.query.filter_by(nombre_concepto=nombre).first()
        if not concepto:
            concepto = Concepto(
                nombre_concepto=nombre,
                unidad=unidad or None,
                precio_unitario=precio,
                descripcion=descripcion or None,
            )
            db.session.add(concepto)
            db.session.flush()
            print(f"[INFO] Nuevo concepto agregado (en actualización): {nombre}")

        det = CotizacionDetalle(**_safe_detalle_kwargs(
            cotizacion_id=c.id,
            concepto_id=concepto.id,
            nombre_concepto=nombre,
            unidad=unidad,
            capitulo=capitulo,
            cantidad=cantidad,
            precio_unitario=precio,
            sistema=sistema or None,
            descripcion=descripcion,
            subtotal=linea_subtotal,
        ))
        db.session.add(det)

    # === TOTALES ===
    descuento_porc_capturado = parse_float(f.get("descuento_total"), desc_porc)
    descuento_porc_aplicado = min(max(descuento_porc_capturado, 0.0), 100.0)
    descuento_total = subtotal * (descuento_porc_aplicado / 100.0)
    subtotal_desc = subtotal - descuento_total
    iva_monto = subtotal_desc * (iva_porc / 100.0)
    total = subtotal_desc + iva_monto

    if zona:
        zona_line = f"Zona: {zona} ({descuento_porc_aplicado:g}% descuento)"
        notas = (c.notas or "").strip()
        notas_lines = [ln for ln in notas.splitlines() if ln.strip() and not ln.strip().lower().startswith("zona:")]
        notas_lines.append(zona_line)
        c.notas = "\n".join(notas_lines).strip()

    c.subtotal = fmt(subtotal)
    c.descuento_total = fmt(descuento_total)
    c.iva_porc = fmt(iva_porc)
    c.iva_monto = fmt(iva_monto)
    c.total = fmt(total)

    db.session.commit()

    # --- WhatsApp en actualización ---
    try:
        body = (
            "🔄 *Actualización de Cotización*\\n"
            f"Folio: *{c.folio}*\\n"
            f"Estatus seguimiento: *{c.estatus}*\\n"
            f"Estatus aprobación: *{c.estatus_aprobacion}*\\n"
            f"Total: {money(c.total)}"
        )
        send_whatsapp_multi(ADMIN_LIST, body)
    except Exception as e:
        print(f"[Twilio] Error en actualización: {e}", file=sys.stderr)

    try:
        _send_quote_updated_email(c)
    except Exception as e:
        logger.warning("Correo de edición de cotizacion falló: %s", e)

    try:
        _send_quote_updated_push(c)
    except Exception as e:
        logger.warning("Push de edición de cotizacion falló: %s", e)

    pdf_url = url_for("export_cotizacion_pdf", cot_id=c.id)
    detalle = url_for("view_cotizacion", cot_id=c.id)
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Actualizada {c.folio}</title></head>
<body>
<script>
window.open("{pdf_url}", "_blank");
window.location.href = "{detalle}";
</script>
<p>Abrir PDF: <a href="{pdf_url}" target="_blank">aquí</a>. Ver detalle: <a href="{detalle}">cotización</a>.</p>
</body></html>"""

@app.route("/cotizaciones/<int:cot_id>/eliminar")
@login_required
def eliminar_cotizacion(cot_id):
    cot = _cotizacion_activa_or_404(cot_id)
    # ✅ Solo ADMIN puede eliminar
    if not is_admin():
        abort(403)

    try:
        _soft_delete_cotizacion(cot)
        db.session.commit()
        flash(f"Cotización {cot.folio} enviada a papelera por {COTIZACION_TRASH_RETENTION_DAYS} días.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error al eliminar la cotización: {str(e)}", "danger")
    return redirect(url_for("index"))


@app.route("/cotizaciones/bulk-eliminar", methods=["POST"])
@login_required
def bulk_eliminar_cotizaciones():
    """Elimina múltiples cotizaciones seleccionadas desde el dashboard.

    ✅ Solo ADMIN.
    """
    if not is_admin():
        return jsonify({"error": "Solo el administrador puede eliminar cotizaciones."}), 403
    payload = request.get_json(silent=True) or {}
    ids = payload.get("ids")
    if not isinstance(ids, list):
        # también soporta form-data: ids[]=1&ids[]=2
        ids = request.form.getlist("ids")

    # Normalizar
    norm_ids: List[int] = []
    for x in ids or []:
        try:
            norm_ids.append(int(x))
        except Exception:
            continue

    # limitar para evitar borrados accidentales enormes
    norm_ids = list(dict.fromkeys(norm_ids))[:500]
    if not norm_ids:
        return jsonify({"error": "No se recibieron IDs válidos"}), 400

    deleted_ids: List[int] = []
    skipped = 0

    try:
        for cot_id in norm_ids:
            cot = _cotizaciones_activas_query().filter(Cotizacion.id == cot_id).first()
            if not cot:
                skipped += 1
                continue

            # (Admin-only) — no validación de ownership

            _soft_delete_cotizacion(cot)
            deleted_ids.append(cot_id)

        db.session.commit()
        return jsonify({"deleted": len(deleted_ids), "skipped": skipped, "deleted_ids": deleted_ids})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@app.route("/cotizaciones/bulk-eliminar-filtradas", methods=["POST"])
@login_required
def bulk_eliminar_filtradas():
    """Elimina cotizaciones visibles por filtros del dashboard.

    ✅ Solo ADMIN.
    Recibe JSON: { filters: { desde:'YYYY-MM-DD', hasta:'YYYY-MM-DD', estatus:'', cliente:'', especialidad:'' } }
    """
    if not is_admin():
        return jsonify({"error": "Solo el administrador puede eliminar cotizaciones."}), 403

    payload = request.get_json(silent=True) or {}
    filters = payload.get("filters") or {}

    desde_s = (filters.get("desde") or "").strip()
    hasta_s = (filters.get("hasta") or "").strip()
    estatus_s = (filters.get("estatus") or "").strip()
    cliente_s = (filters.get("cliente") or "").strip().lower()
    especialidad_s = (filters.get("especialidad") or "").strip()

    try:
        q = _build_dashboard_cotizaciones_query(
            desde=desde_s,
            hasta=hasta_s,
            estatus=estatus_s,
            cliente=cliente_s,
            especialidad=especialidad_s,
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    q = q.order_by(Cotizacion.fecha.desc())

    MAX_DELETE = 2000
    items = q.limit(MAX_DELETE + 1).all()
    if len(items) > MAX_DELETE:
        return jsonify({
            "error": f"Demasiadas cotizaciones para eliminar ({MAX_DELETE}+). Ajusta filtros y vuelve a intentar."
        }), 400

    if not items:
        return jsonify({"deleted": 0, "deleted_ids": []})

    deleted_ids: List[int] = []
    try:
        for cot in items:
            cot_id = cot.id
            _soft_delete_cotizacion(cot)
            deleted_ids.append(cot_id)

        db.session.commit()
        return jsonify({"deleted": len(deleted_ids), "deleted_ids": deleted_ids})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route("/cotizaciones/papelera")
@login_required
def papelera_cotizaciones():
    if not is_admin():
        abort(403)
    try:
        _purge_expired_cotizacion_trash()
    except Exception as e:
        db.session.rollback()
        flash(f"No se pudo limpiar la papelera vencida: {e}", "warning")

    items = (
        Cotizacion.query
        .filter(Cotizacion.eliminada_en.isnot(None))
        .order_by(Cotizacion.eliminada_en.desc())
        .all()
    )
    return render_template(
        "cotizaciones_papelera.html",
        items=items,
        retention_days=COTIZACION_TRASH_RETENTION_DAYS,
        now=now_cdmx_naive(),
        title="Papelera de cotizaciones - Sistema MAR",
    )

@app.route("/cotizaciones/<int:cot_id>/restaurar", methods=["POST"])
@login_required
def restaurar_cotizacion(cot_id: int):
    if not is_admin():
        abort(403)
    cot = Cotizacion.query.filter(Cotizacion.id == cot_id, Cotizacion.eliminada_en.isnot(None)).first_or_404()
    try:
        _restore_cotizacion(cot)
        db.session.commit()
        flash(f"Cotización {cot.folio} restaurada correctamente.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"No se pudo restaurar la cotización: {e}", "danger")
    return redirect(url_for("papelera_cotizaciones"))

@app.route("/cotizaciones")
@login_required
def list_cotizaciones():
    page = int(request.args.get("p", 1) or 1)
    per_page = 25

    q = _cotizaciones_activas_query()
    if not is_admin():
        q = q.filter(Cotizacion.responsable == responsable_actual())

    q = q.order_by(Cotizacion.fecha.desc())

    total = q.count()
    pages = max(1, math.ceil(total / per_page))
    page = max(1, min(page, pages))
    items = q.offset((page-1)*per_page).limit(per_page).all()

    return render_template(
        "cotizaciones_list.html",
        items=items, page=page, pages=pages, total=total,
        title="Cotizaciones · Sistema MAR"
    )

@app.route("/cotizaciones/<int:cot_id>")
@login_required
def view_cotizacion(cot_id: int):
    c = _cotizacion_activa_or_404(cot_id)
    require_owner_or_admin(c)
    zona_actual = ""
    try:
        if c.notas:
            for ln in str(c.notas).splitlines():
                if ln.strip().lower().startswith("zona:"):
                    tmp = ln.split(":", 1)[1].strip()
                    zona_actual = tmp.split("(", 1)[0].strip()
                    break
    except Exception:
        zona_actual = ""
    condiciones_finales = _condiciones_comerciales_finales(c.notas or "")
    notas_adicionales, _ = _split_notas_y_zona(c.notas or "")
    return render_template("cotizacion_view.html", c=c, zona_actual=zona_actual, condiciones_finales=condiciones_finales, notas_adicionales=notas_adicionales, title=f"Ver {c.folio}")

@app.route("/cotizaciones/<int:cot_id>/seguimiento")
@login_required
def cotizacion_seguimiento(cot_id: int):
    c = _cotizacion_activa_or_404(cot_id)
    require_owner_or_admin(c)
    return render_template(
        "cotizacion_seguimiento.html",
        c=c,
        cliente_info=_cliente_seguimiento_payload(
            cliente=c.cliente,
            responsable=c.responsable,
            extras=[
                {"label": "Proyecto", "value": c.proyecto},
                {"label": "Ciudad de trabajo", "value": c.ciudad_trabajo},
            ],
        ),
        seguimientos=c.seguimientos,
        valid_estatus=VALID_ESTATUS,
        mention_users=_usuarios_menciones_payload(),
        title=f"Seguimiento {c.folio}",
    )

@app.route("/cotizaciones/<int:cot_id>/seguimiento", methods=["POST"])
@login_required
def crear_cotizacion_seguimiento(cot_id: int):
    c = _cotizacion_activa_or_404(cot_id)
    require_owner_or_admin(c)

    nuevo_estatus = (request.form.get("estatus") or "").strip().upper()
    nuevo_responsable = (request.form.get("responsable") or "").strip()
    comentario = (request.form.get("comentario") or "").strip()
    tagged_users = _usuarios_mencionados_en_comentario(comentario)
    hubo_cambio = False

    if nuevo_estatus:
        if nuevo_estatus not in VALID_ESTATUS:
            flash("Selecciona un estatus válido.", "danger")
            return redirect(url_for("cotizacion_seguimiento", cot_id=c.id))
        if (c.estatus or "").strip().upper() != nuevo_estatus:
            c.estatus = nuevo_estatus
            hubo_cambio = True

    if nuevo_responsable != (c.responsable or "").strip():
        c.responsable = nuevo_responsable
        hubo_cambio = True

    if not comentario and not hubo_cambio:
        flash("Haz un cambio de estatus/responsable o escribe un comentario para guardar.", "warning")
        return redirect(url_for("cotizacion_seguimiento", cot_id=c.id))

    seg = None
    autor = (responsable_actual() or "Sistema").strip()
    if comentario:
        seg = CotizacionSeguimiento(
            cotizacion_id=c.id,
            usuario_id=getattr(current_user, "id", None),
            autor=autor,
            comentario=comentario,
            fecha_seguimiento=now_cdmx_naive(),
            actualizado_en=now_cdmx_naive(),
        )
        db.session.add(seg)

    db.session.commit()

    if seg is not None:
        try:
            _send_quote_followup_push(c, seg)
        except Exception as exc:
            logger.warning("Push de seguimiento fallida: %s", exc)
        try:
            _notify_tagged_followup(
                tagged_users=tagged_users,
                module_label="Cotizaciones",
                item_label=c.folio or f"Cotización #{c.id}",
                autor=autor,
                comentario=comentario,
                view_endpoint="cotizacion_seguimiento",
                view_params={"cot_id": c.id, "_anchor": f"seguimiento-{seg.id}"},
            )
        except Exception as exc:
            logger.exception("No se pudo notificar etiquetas de cotizacion %s", c.id)
            flash(f"Seguimiento guardado, pero no se pudo enviar correo a etiquetados: {exc}", "warning")
            return redirect(url_for("cotizacion_seguimiento", cot_id=c.id, _anchor=f"seguimiento-{seg.id}"))

    if seg is not None and hubo_cambio:
        flash("Se guardó el seguimiento y también se actualizó la cotización.", "success")
    elif seg is not None:
        flash("Seguimiento guardado correctamente.", "success")
    else:
        flash("Cambios de la cotización guardados.", "success")

    if seg is not None:
        return redirect(url_for("cotizacion_seguimiento", cot_id=c.id, _anchor=f"seguimiento-{seg.id}"))
    return redirect(url_for("cotizacion_seguimiento", cot_id=c.id))

@app.route("/cotizaciones/<int:cot_id>/seguimiento/<int:seg_id>/editar", methods=["POST"])
@login_required
def editar_cotizacion_seguimiento(cot_id: int, seg_id: int):
    c = _cotizacion_activa_or_404(cot_id)
    require_owner_or_admin(c)
    seg = CotizacionSeguimiento.query.filter_by(id=seg_id, cotizacion_id=c.id).first_or_404()
    require_followup_author_or_admin(seg)

    comentario = (request.form.get("comentario") or "").strip()
    if not comentario:
        flash("El comentario no puede quedar vacío.", "warning")
        return redirect(url_for("cotizacion_seguimiento", cot_id=c.id))

    seg.comentario = comentario
    seg.actualizado_en = now_cdmx_naive()
    db.session.commit()
    flash("Seguimiento actualizado.", "success")
    return redirect(url_for("cotizacion_seguimiento", cot_id=c.id))

@app.route("/cotizaciones/<int:cot_id>/seguimiento/<int:seg_id>/eliminar", methods=["POST"])
@login_required
def eliminar_cotizacion_seguimiento(cot_id: int, seg_id: int):
    c = _cotizacion_activa_or_404(cot_id)
    require_owner_or_admin(c)
    seg = CotizacionSeguimiento.query.filter_by(id=seg_id, cotizacion_id=c.id).first_or_404()
    require_followup_author_or_admin(seg)

    db.session.delete(seg)
    db.session.commit()
    flash("Seguimiento eliminado.", "success")
    return redirect(url_for("cotizacion_seguimiento", cot_id=c.id))

@app.route("/cotizaciones/<int:cot_id>/ver")
@login_required
def ver_cotizacion(cot_id: int):
    cot = _cotizacion_activa_or_404(cot_id)
    require_owner_or_admin(cot)
    condiciones_finales = _condiciones_comerciales_finales(cot.notas or "")
    notas_adicionales, _ = _split_notas_y_zona(cot.notas or "")
    return render_template("cotizacion_view.html", c=cot, condiciones_finales=condiciones_finales, notas_adicionales=notas_adicionales, title=f"Vista de {cot.folio}")

# ---------------------------------------------------------
# API: actualizar estatus (inline) + WhatsApp
# ---------------------------------------------------------
@app.route("/api/cotizaciones/<int:cot_id>/estatus", methods=["POST"])
@login_required
def api_update_estatus(cot_id):
    c = _cotizacion_activa_or_404(cot_id)
    require_owner_or_admin(c)

    ct = request.headers.get("Content-Type", "")
    if "application/json" in ct:
        data = request.get_json(silent=True) or {}
        nuevo = (data.get("estatus") or "").upper().strip()
    else:
        nuevo = (request.form.get("estatus") or "").upper().strip()

    if nuevo not in VALID_ESTATUS:
        return jsonify({"ok": False, "error": "Estatus inválido"}), 400

    anterior = c.estatus
    if nuevo == anterior:
        return jsonify({"ok": True, "folio": c.folio, "estatus": nuevo, "mensaje": "Sin cambios."})

    c.estatus = nuevo
    db.session.commit()

    if nuevo in {"APROBADO", "AUTORIZADO", "RECHAZADO"}:
        try:
            _send_quote_review_response_email(c, nuevo)
        except Exception as e:
            logger.warning("Correo de respuesta de cotizacion fallido: %s", e)

    try:
        _send_quote_status_push(c, anterior, nuevo)
    except Exception as e:
        logger.warning("Push de estatus fallida: %s", e)

    try:
        body = (
            f"🔄 *Actualización de estatus*\\n"
            f"Folio: *{c.folio}*\\n"
            f"Anterior: {anterior}\\n"
            f"Nuevo: *{nuevo}*\\n"
            f"Total: {money(c.total)}"
        )
        send_whatsapp_multi(ADMIN_LIST, body)
    except Exception as e:
        print(f"[Twilio] Error al enviar notificación de estatus: {e}", file=sys.stderr)

    return jsonify({
        "ok": True,
        "folio": c.folio,
        "estatus": nuevo,
        "mensaje": f"Estatus de la cotización {c.folio} actualizado a {nuevo}."
    })

@app.route("/api/cotizaciones/<int:cot_id>/estatus-aprobacion", methods=["POST"])
@login_required
def api_update_estatus_aprobacion(cot_id):
    c = _cotizacion_activa_or_404(cot_id)
    require_owner_or_admin(c)

    ct = request.headers.get("Content-Type", "")
    if "application/json" in ct:
        data = request.get_json(silent=True) or {}
        nuevo = (data.get("estatus_aprobacion") or data.get("estatus") or "").upper().strip()
    else:
        nuevo = (request.form.get("estatus_aprobacion") or request.form.get("estatus") or "").upper().strip()
    if nuevo == "APROBADO":
        nuevo = "APROBADA"
    elif nuevo == "RECHAZADO":
        nuevo = "RECHAZADA"
    elif nuevo == "EN REVISION":
        nuevo = "EN REVISIÓN"

    if nuevo not in VALID_ESTATUS_APROBACION:
        return jsonify({"ok": False, "error": "Estatus de aprobación inválido"}), 400

    anterior = (c.estatus_aprobacion or "EN REVISIÓN").strip().upper()
    if anterior == "EN REVISION":
        anterior = "EN REVISIÓN"
    if nuevo == anterior:
        return jsonify({"ok": True, "folio": c.folio, "estatus_aprobacion": nuevo, "mensaje": "Sin cambios."})

    c.estatus_aprobacion = nuevo
    db.session.commit()

    try:
        _send_quote_review_response_email(c, nuevo)
    except Exception as e:
        logger.warning("Correo de respuesta de aprobación de cotizacion fallido: %s", e)

    try:
        _send_quote_review_result_push(c, nuevo)
    except Exception as e:
        logger.warning("Push de aprobación fallida: %s", e)

    return jsonify({
        "ok": True,
        "folio": c.folio,
        "estatus_aprobacion": nuevo,
        "mensaje": f"Estatus de aprobación de {c.folio} actualizado a {nuevo}."
    })

# ---------------------------------------------------------
# Exportaciones CSV / Excel
# ---------------------------------------------------------
@app.route("/cotizaciones/<int:cot_id>/export.csv")
@login_required
def export_cotizacion_csv(cot_id: int):
    c = _cotizacion_activa_or_404(cot_id)
    require_owner_or_admin(c)

    output = io.StringIO()
    w = csv.writer(output)

    w.writerow(["Folio","Fecha","Estatus","Representante","Cliente","Empresa","Subtotal","IVA %","IVA $","Total","Notas"])
    w.writerow([
        c.folio, c.fecha.strftime("%Y-%m-%d %H:%M"), c.estatus, (c.responsable or ""),
        c.cliente.nombre_cliente if c.cliente else "",
        c.cliente.empresa if c.cliente else "",
        f"{c.subtotal:.2f}",
        f"{c.iva_porc:.2f}", f"{c.iva_monto:.2f}",
        f"{c.total:.2f}", (c.notas or "")
    ])
    w.writerow([])
    w.writerow(["Capitulo","Cant","Unidad","Concepto","Sistema","PU","Subtotal","Descripción"])
    for d in c.detalles:
        w.writerow([
            getattr(d, "capitulo", "") or "", d.cantidad, d.unidad or "", d.nombre_concepto, d.sistema or "",
            f"{d.precio_unitario:.2f}", f"{d.subtotal:.2f}", (d.descripcion or "")
        ])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={'Content-Disposition': f'attachment; filename="{c.folio or "cotizacion"}.csv"'}
    )

@app.route("/cotizaciones/<int:cot_id>/export.xlsx")
@login_required
def export_cotizacion_xlsx(cot_id: int):
    if Workbook is None:
        abort(501, description="openpyxl no instalado en el servidor.")
    c = _cotizacion_activa_or_404(cot_id)
    require_owner_or_admin(c)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    header_fill = PatternFill("solid", fgColor=MAR_BLUE_XLSX)
    white = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:G1"); ws["A1"] = f"COTIZACIÓN {c.folio}"
    ws["A1"].font = Font(bold=True, size=14); ws["A1"].alignment = center

    ws.append(["Folio", c.folio, "", "Fecha", c.fecha.strftime("%d/%m/%Y %H:%M"), ""])
    ws.append(["Cliente", (c.cliente.nombre_cliente if c.cliente else ""), "", "Empresa", (c.cliente.empresa if c.cliente else ""), ""])
    ws.append(["Representante", c.responsable or "", "", "Estatus", c.estatus, ""])
    ws.append([])

    headers = ["Capitulo", "Cant", "Unidad", "Concepto", "Sistema", "Precio Unit.", "Subtotal"]
    ws.append(headers)
    for col in range(1, 8):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = header_fill; cell.font = white; cell.alignment = center; cell.border = border

    for d in c.detalles:
        ws.append([getattr(d, "capitulo", "") or "", d.cantidad, d.unidad or "", d.nombre_concepto, d.sistema or "",
                   float(d.precio_unitario or 0), float(d.subtotal or 0)])
        r = ws.max_row
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = border
        ws.cell(row=r, column=2).number_format = '0.00'
        ws.cell(row=r, column=6).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=7).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=4).alignment = left

    ws.append([])
    ws.append(["", "", cantidad_en_letra_mn(c.total)])
    ws.append(["", "Subtotal:", float(c.subtotal or 0)])
    ws.append(["", f"IVA ({c.iva_porc:.2f}%):", float(c.iva_monto or 0)])
    ws.append(["", "Total:", float(c.total or 0)])
    for r in range(ws.max_row-2, ws.max_row+1):
        ws.cell(row=r, column=2).font = bold
        ws.cell(row=r, column=3).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=3).alignment = right

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    bio = io.BytesIO()
    wb.save(bio); bio.seek(0)
    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{c.folio}.xlsx"'}
    )


def _email_body_cotizacion(c: Cotizacion) -> str:
    cli = c.cliente
    atencion = ""
    if cli:
        atencion = (cli.nombre_cliente or cli.empresa or "").strip()

    return (
        f"Con atención a: {atencion}\n\n"
        "Buenas tardes, por medio de la presente hacemos llegar la cotización requerida.\n\n"
        "Cualquier duda, estamos a sus órdenes.\n"
        "Saludos cordiales.\n"
    )


def _email_signature_text() -> str:
    return (
        "\n"
        "POLIUTECH RECUBRIMIENTOS ESPECIALIZADOS\n"
        "oficinas: 5559380536, 5559386530\n"
        "Número celular: 5534662836\n"
        "Correo electrónico: cotizaciones@poliutech.com\n"
        "www.poliutech.com\n"
    )


def _email_body_cotizacion_html(c: Cotizacion) -> str:
    cli = c.cliente
    atencion = ""
    if cli:
        atencion = escape((cli.nombre_cliente or cli.empresa or "").strip())

    return f"""
    <html>
      <body style="font-family: Arial, sans-serif; color: #222; line-height: 1.45;">
        <p style="margin: 0 0 16px 0;">Con atención a: {atencion}</p>
        <p style="margin: 0 0 16px 0;">Buenas tardes, por medio de la presente hacemos llegar la cotización requerida.</p>
        <p style="margin: 0 0 22px 0;">Cualquier duda, estamos a sus órdenes.<br>Saludos cordiales.</p>

        <div style="padding-top: 14px; border-top: 1px solid #cfcfcf; max-width: 620px;">
          <div style="font-size: 14px; margin-bottom: 14px;">
            <div style="font-weight: 700;">POLIUTECH RECUBRIMIENTOS ESPECIALIZADOS</div>
            <div>oficinas:. 5559380536, 5559386530</div>
            <div>Número celular. 5534662836</div>
            <div>Correo electrónico : <a href="mailto:cotizaciones@poliutech.com">cotizaciones@poliutech.com</a></div>
            <div><a href="https://www.poliutech.com" target="_blank">www.poliutech.com</a></div>
          </div>
          <div>
            <img src="cid:poliutech-logo" alt="Poliutech" style="display:block; width:280px; height:auto; border:0;">
          </div>
        </div>
      </body>
    </html>
    """.strip()


def _parse_email_list(raw: str | list[str] | tuple[str, ...] | None) -> list[str]:
    if isinstance(raw, (list, tuple)):
        parts = [str(item or "").strip() for item in raw]
        candidate = ",".join([part for part in parts if part])
    else:
        candidate = str(raw or "").strip()

    if not candidate:
        return []

    emails: list[str] = []
    for _, addr in getaddresses([candidate]):
        addr = (addr or "").strip()
        if not addr:
            continue
        if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", addr):
            raise ValueError(f"Correo inválido: {addr}")
        emails.append(addr)
    return emails


def _unique_emails(*groups: list[str]) -> list[str]:
    unique: list[str] = []
    seen: set[str] = set()
    for group in groups:
        for email in group:
            key = email.lower()
            if key not in seen:
                seen.add(key)
                unique.append(email)
    return unique


def _finanzas_auth_notify_recipients() -> list[str]:
    return _parse_email_list(FINANZAS_AUTH_NOTIFY_EMAILS)


def _mobile_push_user_ids_for_finanzas_auth_notify() -> list[int]:
    target_emails = {email.lower() for email in _finanzas_auth_notify_recipients()}
    aliases = {"mescalera", "mesacalera", "miguel", "miguele"}
    user_ids: set[int] = set()
    for user in Usuario.query.all():
        user_name = (getattr(user, "nombre", "") or "").strip().lower()
        visible_name = (_mobile_user_responsable(user) or "").strip().lower()
        raw_visible_name = (getattr(user, "nombre_visible", "") or "").strip().lower()
        user_email = (getattr(user, "correo", "") or "").strip().lower()
        identity_parts = {user_name, visible_name, raw_visible_name}
        if (
            user_email in target_emails
            or any(part in aliases or part.startswith("mescalera ") or part.startswith("mesacalera ") or part.startswith("miguel ") for part in identity_parts if part)
        ):
            if user.id:
                user_ids.add(user.id)
    if not user_ids:
        logger.warning("Push finanzas autorizaciones: no se encontraron usuarios para %s.", sorted(target_emails))
    return list(user_ids)


def _quote_responsible_email(c: Cotizacion) -> list[str]:
    responsable = (c.responsable or "").strip()
    if not responsable:
        return []

    user = Usuario.query.filter(
        or_(
            db.func.lower(Usuario.nombre) == responsable.lower(),
            db.func.lower(db.func.coalesce(Usuario.nombre_visible, "")) == responsable.lower(),
        )
    ).first()
    if not user or not (user.correo or "").strip():
        return []
    return _parse_email_list(user.correo)


def _usuarios_etiquetables() -> list[Usuario]:
    return Usuario.query.order_by(Usuario.nombre.asc()).all()


def _usuarios_menciones_payload() -> list[dict]:
    return [
        {
            "id": usuario.id,
            "nombre": usuario.nombre or "",
            "correo": usuario.correo or "",
        }
        for usuario in _usuarios_etiquetables()
        if usuario.nombre and usuario.correo
    ]


def _normalize_mention_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip().lower()


def _usuarios_mencionados_en_comentario(comentario: str) -> list[Usuario]:
    normalized_comment = _normalize_mention_text(comentario)
    if "@" not in normalized_comment:
        return []

    mentioned: list[Usuario] = []
    seen_ids: set[int] = set()
    users = sorted(_usuarios_etiquetables(), key=lambda user: len(user.nombre or ""), reverse=True)
    for usuario in users:
        nombre = _normalize_mention_text(usuario.nombre or "")
        if not nombre:
            continue
        pattern = rf"@\s*{re.escape(nombre)}(?=$|[\s.,;:!?\)\]\}}])"
        if re.search(pattern, normalized_comment) and usuario.id not in seen_ids:
            mentioned.append(usuario)
            seen_ids.add(usuario.id)
    return mentioned


def _followup_tag_email_html(
    *,
    usuario: Usuario,
    module_label: str,
    item_label: str,
    autor: str,
    comentario: str,
    view_url: str,
) -> str:
    usuario_nombre = escape(usuario.nombre or "Usuario")
    module_label = escape(module_label)
    item_label = escape(item_label)
    autor = escape(autor or "Sistema")
    comentario_html = escape(comentario or "").replace("\n", "<br>")
    view_url = escape(view_url)
    return f"""
    <div style="font-family:Arial,sans-serif;background:#f5f7fb;padding:24px;color:#172033;">
      <div style="max-width:680px;margin:0 auto;background:#ffffff;border:1px solid #dde3ea;border-radius:10px;overflow:hidden;">
        <div style="background:#0C3C78;color:#ffffff;padding:18px 22px;">
          <h2 style="margin:0;font-size:20px;">Te etiquetaron en un seguimiento</h2>
          <div style="font-size:13px;opacity:.92;margin-top:4px;">{module_label}</div>
        </div>
        <div style="padding:22px;">
          <p style="margin:0 0 12px 0;">Hola <b>{usuario_nombre}</b>,</p>
          <p style="margin:0 0 18px 0;"><b>{autor}</b> te etiquetó en <b>{item_label}</b>.</p>
          <div style="border:1px solid #dde3ea;background:#f8fafc;border-radius:8px;padding:14px 16px;line-height:1.5;">
            {comentario_html or "Sin comentario."}
          </div>
          <table role="presentation" cellpadding="0" cellspacing="0" border="0" align="center" style="margin:28px auto 8px auto;">
            <tr>
              <td align="center" bgcolor="#0C3C78" style="border-radius:10px;box-shadow:0 8px 18px rgba(12,60,120,0.28);">
                <a href="{view_url}" target="_blank" style="display:inline-block;min-width:190px;padding:15px 28px;border-radius:10px;background:#0C3C78;color:#ffffff !important;text-decoration:none;font-size:17px;font-weight:800;letter-spacing:.4px;text-align:center;">
                  VER SEGUIMIENTO
                </a>
              </td>
            </tr>
          </table>
          <p style="margin:20px 0 0 0;color:#64748b;font-size:12px;">Si el botón no abre, copia este enlace en tu navegador:<br>{view_url}</p>
        </div>
      </div>
    </div>
    """


def _send_followup_tag_emails(
    *,
    tagged_users: list[Usuario],
    module_label: str,
    item_label: str,
    autor: str,
    comentario: str,
    view_url: str,
) -> int:
    sent = 0
    for usuario in tagged_users:
        recipients = _parse_email_list(getattr(usuario, "correo", None))
        if not recipients:
            continue
        msg = EmailMessage()
        msg["Subject"] = f"Te etiquetaron en {module_label}: {item_label}"
        msg["From"] = f"SISTEMA MAR <{SMTP_FROM or SMTP_USERNAME}>"
        msg["To"] = ", ".join(recipients)
        msg.set_content(
            f"Hola {usuario.nombre or 'Usuario'},\n\n"
            f"{autor or 'Sistema'} te etiquetó en {module_label}: {item_label}.\n\n"
            f"Seguimiento:\n{comentario or 'Sin comentario.'}\n\n"
            f"VER: {view_url}\n"
        )
        msg.add_alternative(
            _followup_tag_email_html(
                usuario=usuario,
                module_label=module_label,
                item_label=item_label,
                autor=autor,
                comentario=comentario,
                view_url=view_url,
            ),
            subtype="html",
        )
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
            smtp.ehlo()
            smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
            smtp.send_message(msg, to_addrs=recipients)
        sent += 1
    return sent


def _notify_tagged_followup(
    *,
    tagged_users: list[Usuario],
    module_label: str,
    item_label: str,
    autor: str,
    comentario: str,
    view_endpoint: str,
    view_params: dict,
) -> int:
    if not tagged_users:
        return 0
    view_url = url_for(view_endpoint, _external=True, **view_params)
    return _send_followup_tag_emails(
        tagged_users=tagged_users,
        module_label=module_label,
        item_label=item_label,
        autor=autor,
        comentario=comentario,
        view_url=view_url,
    )


def _send_cotizacion_email(c: Cotizacion, recipients: list[str], cc: list[str] | None = None, bcc: list[str] | None = None) -> None:
    cc = cc or []
    bcc = bcc or []
    pdf_response = export_cotizacion_pdf(c.id)
    pdf_response.direct_passthrough = False
    pdf_bytes = pdf_response.get_data()

    msg = EmailMessage()
    msg["Subject"] = f"Cotización {c.folio}"
    msg["From"] = SMTP_FROM
    msg["To"] = ", ".join(recipients)
    if cc:
        msg["Cc"] = ", ".join(cc)
    msg.set_content(_email_body_cotizacion(c) + _email_signature_text())
    msg.add_alternative(_email_body_cotizacion_html(c), subtype="html")

    logo_path = Path(app.static_folder or "static") / "logo.png"
    if logo_path.exists():
        logo_bytes = logo_path.read_bytes()
        mime_type, _ = mimetypes.guess_type(str(logo_path))
        maintype, subtype = ("image", "jpeg")
        if mime_type and "/" in mime_type:
            maintype, subtype = mime_type.split("/", 1)
        html_part = msg.get_body(preferencelist=("html",))
        if html_part is not None:
            html_part.add_related(
                logo_bytes,
                maintype=maintype,
                subtype=subtype,
                cid="<poliutech-logo>",
            )

    msg.add_attachment(
        pdf_bytes,
        maintype="application",
        subtype="pdf",
        filename=f"{c.folio}.pdf",
    )

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg, to_addrs=[*recipients, *cc, *bcc])


def _quote_updated_mail_html(c: Cotizacion, view_url: str, approve_url: str, reject_url: str) -> str:
    cli = c.cliente
    folio = escape(c.folio or f"#{c.id}")
    cliente = escape(cli.nombre_cliente if cli else "Sin cliente")
    empresa = escape(cli.empresa if cli and cli.empresa else "Sin empresa")
    proyecto = escape(c.proyecto or "Sin proyecto")
    responsable = escape(c.responsable or "Sin responsable")
    estatus = escape(c.estatus or "Sin estatus")
    aprobacion = escape(c.estatus_aprobacion or "Sin estatus")
    total = f"{money(c.total)} {escape(c.moneda or 'MXN')}"
    button_base = (
        "display:inline-block;min-width:142px;text-align:center;padding:13px 18px;"
        "border-radius:8px;text-decoration:none;font-weight:800;font-size:15px;"
        "margin:0 8px 10px 0;"
    )
    return f"""
    <html>
      <body style="margin:0;padding:0;background:#eef2f7;font-family:Arial,Helvetica,sans-serif;color:#1f2937;">
        <div style="max-width:760px;margin:0 auto;padding:30px 16px;">
          <div style="background:#ffffff;border:1px solid #d9e2ec;border-radius:10px;overflow:hidden;">
            <div style="background:#0C3C78;color:#ffffff;padding:24px 28px;">
              <div style="font-size:12px;font-weight:700;letter-spacing:.9px;text-transform:uppercase;">MAR · Poliutech</div>
              <div style="font-size:24px;font-weight:900;margin-top:6px;">Cotizacion editada</div>
              <div style="font-size:14px;opacity:.95;margin-top:7px;">{folio}</div>
            </div>
            <div style="padding:28px;">
              <p style="margin:0 0 22px 0;font-size:15px;color:#475569;">Se guardaron cambios en esta cotizacion y vuelve a quedar pendiente de aprobacion. Se adjunta el PDF actualizado.</p>
              <table style="border-collapse:collapse;width:100%;background:#ffffff;border:1px solid #dbe4ef;border-radius:10px;overflow:hidden;">
                <tr><td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:800;">Cliente</td><td style="padding:13px 16px;border-bottom:1px solid #edf2f7;">{cliente}</td></tr>
                <tr><td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:800;">Empresa</td><td style="padding:13px 16px;border-bottom:1px solid #edf2f7;">{empresa}</td></tr>
                <tr><td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:800;">Proyecto</td><td style="padding:13px 16px;border-bottom:1px solid #edf2f7;">{proyecto}</td></tr>
                <tr><td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:800;">Responsable</td><td style="padding:13px 16px;border-bottom:1px solid #edf2f7;">{responsable}</td></tr>
                <tr><td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:800;">Seguimiento</td><td style="padding:13px 16px;border-bottom:1px solid #edf2f7;">{estatus}</td></tr>
                <tr><td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:800;">Aprobacion</td><td style="padding:13px 16px;border-bottom:1px solid #edf2f7;">{aprobacion}</td></tr>
                <tr><td style="padding:13px 16px;color:#64748b;font-weight:800;">Total</td><td style="padding:13px 16px;color:#0C3C78;font-size:20px;font-weight:900;">{total}</td></tr>
              </table>
              <div style="margin-top:24px;">
                <a href="{approve_url}" style="{button_base}background:#16854f;color:#ffffff;border:1px solid #16854f;">APROBAR</a>
                <a href="{reject_url}" style="{button_base}background:#c62828;color:#ffffff;border:1px solid #c62828;">RECHAZAR</a>
                <a href="{view_url}" style="{button_base}background:#0C3C78;color:#ffffff;border:1px solid #0C3C78;">VER COTIZACION</a>
              </div>
            </div>
          </div>
        </div>
      </body>
    </html>
    """.strip()


def _send_quote_updated_email(c: Cotizacion) -> None:
    recipients = _unique_emails(
        _quote_responsible_email(c),
        _parse_email_list(COTIZACION_REVIEW_EMAIL),
        _parse_email_list(COTIZACION_RESPONSE_EMAIL),
        _parse_email_list(COTIZACION_APPROVALS_EMAIL),
        _parse_email_list(COTIZACION_REVIEW_RESULT_AAZCONA_EMAIL),
    )
    bcc = _parse_email_list(COTIZACION_REVIEW_BCC_EMAIL)
    if not recipients:
        raise ValueError("No hay correo configurado para edición de cotizaciones.")

    pdf_response = export_cotizacion_pdf(c.id)
    pdf_response.direct_passthrough = False
    pdf_bytes = pdf_response.get_data()
    view_url = url_for("view_cotizacion", cot_id=c.id, _external=True)
    approve_url = url_for("cotizacion_revision_decidir", cot_id=c.id, action="approve", token=_quote_review_token(c, "approve"), _external=True)
    reject_url = url_for("cotizacion_revision_decidir", cot_id=c.id, action="reject", token=_quote_review_token(c, "reject"), _external=True)

    msg = EmailMessage()
    msg["Subject"] = f"Cotizacion editada {c.folio or c.id}"
    msg["From"] = f"COTIZACIONES POLIUTECH <{SMTP_FROM or SMTP_USERNAME}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content(
        f"Se edito la cotizacion {c.folio or c.id}.\n"
        f"Cliente: {c.cliente.nombre_cliente if c.cliente else 'Sin cliente'}\n"
        f"Proyecto: {c.proyecto or 'Sin proyecto'}\n"
        f"Estatus seguimiento: {c.estatus or 'Sin estatus'}\n"
        f"Estatus aprobacion: {c.estatus_aprobacion or 'Sin estatus'}\n"
        f"Total: {money(c.total)} {c.moneda or 'MXN'}\n\n"
        f"Aprobar: {approve_url}\n"
        f"Rechazar: {reject_url}\n"
        f"Ver cotizacion: {view_url}\n"
    )
    msg.add_alternative(_quote_updated_mail_html(c, view_url, approve_url, reject_url), subtype="html")
    msg.add_attachment(
        pdf_bytes,
        maintype="application",
        subtype="pdf",
        filename=f"{c.folio or c.id}.pdf",
    )

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg, to_addrs=[*recipients, *bcc])


def _quote_review_serializer() -> URLSafeTimedSerializer:
    return URLSafeTimedSerializer(app.secret_key, salt="cotizacion-review")


def _quote_review_token(cot: Cotizacion, action: str) -> str:
    return _quote_review_serializer().dumps({"cotizacion_id": cot.id, "action": action})


def _quote_review_load_from_token(cot_id: int, token: str, action: str) -> Cotizacion:
    try:
        payload = _quote_review_serializer().loads(token or "", max_age=60 * 60 * 24 * 45)
    except (BadSignature, SignatureExpired):
        abort(403)
    if int(payload.get("cotizacion_id") or 0) != int(cot_id):
        abort(403)
    token_action = (payload.get("action") or "").strip()
    if token_action != action:
        abort(403)
    return _cotizacion_activa_or_404(cot_id)


def _quote_status_flag_class(status: str) -> str:
    normalized = (status or "").strip().upper()
    if normalized in {"APROBADA"}:
        return "quote-flag-green"
    if normalized in {"RECHAZADA"}:
        return "quote-flag-red"
    if normalized in {"EN REVISIÓN", "EN REVISION"}:
        return "quote-flag-yellow"
    return "quote-flag-muted"


def _quote_review_mail_html(c: Cotizacion, approve_url: str, reject_url: str, review_url: str) -> str:
    cli = c.cliente
    cliente = escape(cli.nombre_cliente if cli else "Sin cliente")
    empresa = escape(cli.empresa if cli and cli.empresa else "")
    proyecto = escape(c.proyecto or "Sin proyecto")
    folio = escape(c.folio or f"#{c.id}")
    total = f"${float(c.total or 0):,.2f} {escape(c.moneda or 'MXN')}"
    responsable = escape(c.responsable or "Sin responsable")
    button_base = (
        "display:inline-block;min-width:142px;text-align:center;padding:14px 20px;"
        "border-radius:8px;text-decoration:none;font-weight:800;font-size:15px;"
        "margin:0 8px 10px 0;color:#ffffff;"
    )
    return f"""
    <html>
      <body style="margin:0;padding:0;background:#f3f6fb;font-family:Arial,Helvetica,sans-serif;color:#1f2937;">
        <div style="max-width:760px;margin:0 auto;padding:28px 16px;">
          <div style="background:#ffffff;border:1px solid #dbe4ef;border-radius:10px;overflow:hidden;">
            <div style="background:#0C3C78;color:#ffffff;padding:22px 26px;">
              <div style="font-size:12px;font-weight:700;letter-spacing:.9px;text-transform:uppercase;">MAR · Poliutech</div>
              <div style="font-size:23px;font-weight:800;margin-top:5px;">Cotizacion pendiente de revision</div>
              <div style="font-size:14px;opacity:.92;margin-top:6px;">{folio}</div>
            </div>
            <div style="padding:26px;">
              <p style="margin:0 0 18px 0;">Se genero una nueva cotizacion con estatus <b>EN REVISIÓN</b>.</p>
              <table style="border-collapse:collapse;width:100%;margin-bottom:22px;">
                <tr><td style="padding:10px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:700;">Cliente</td><td style="padding:10px;border-bottom:1px solid #edf2f7;">{cliente}</td></tr>
                <tr><td style="padding:10px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:700;">Empresa</td><td style="padding:10px;border-bottom:1px solid #edf2f7;">{empresa}</td></tr>
                <tr><td style="padding:10px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:700;">Proyecto</td><td style="padding:10px;border-bottom:1px solid #edf2f7;">{proyecto}</td></tr>
                <tr><td style="padding:10px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:700;">Responsable</td><td style="padding:10px;border-bottom:1px solid #edf2f7;">{responsable}</td></tr>
                <tr><td style="padding:10px;color:#64748b;font-weight:700;">Total</td><td style="padding:10px;font-size:20px;font-weight:900;color:#0C3C78;">{total}</td></tr>
              </table>
              <div>
                <a href="{reject_url}" style="{button_base}background:#c62828;border:1px solid #c62828;">RECHAZADA</a>
                <a href="{approve_url}" style="{button_base}background:#16854f;border:1px solid #16854f;">APROBADA</a>
                <a href="{review_url}" style="{button_base}background:#f0ad00;border:1px solid #f0ad00;color:#1f2937;">EN REVISIÓN</a>
              </div>
              <p style="margin:16px 0 0 0;color:#64748b;font-size:12px;">Si un boton no abre, copia el enlace desde el correo en tu navegador.</p>
            </div>
          </div>
        </div>
      </body>
    </html>
    """.strip()


def _send_quote_review_email(c: Cotizacion) -> None:
    recipients = _parse_email_list(COTIZACION_REVIEW_EMAIL)
    bcc = _parse_email_list(COTIZACION_REVIEW_BCC_EMAIL)
    if not recipients:
        raise ValueError("No hay correo configurado para revision de cotizaciones.")
    pdf_response = export_cotizacion_pdf(c.id)
    pdf_response.direct_passthrough = False
    pdf_bytes = pdf_response.get_data()
    approve_url = url_for("cotizacion_revision_decidir", cot_id=c.id, action="approve", token=_quote_review_token(c, "approve"), _external=True)
    reject_url = url_for("cotizacion_revision_decidir", cot_id=c.id, action="reject", token=_quote_review_token(c, "reject"), _external=True)
    review_url = url_for("cotizacion_revision_decidir", cot_id=c.id, action="review", token=_quote_review_token(c, "review"), _external=True)

    msg = EmailMessage()
    msg["Subject"] = f"Revision de cotizacion {c.folio or c.id}"
    msg["From"] = f"COTIZACIONES POLIUTECH <{SMTP_FROM or SMTP_USERNAME}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content(
        f"Nueva cotizacion {c.folio or c.id}\n"
        f"Estatus: EN REVISIÓN\n"
        f"Total: {money(c.total)} {c.moneda or 'MXN'}\n\n"
        "Abre este correo en vista HTML para usar los botones.\n"
        f"Rechazado: {reject_url}\n"
        f"Aprobar: {approve_url}\n"
        f"En revision: {review_url}\n"
    )
    msg.add_alternative(_quote_review_mail_html(c, approve_url, reject_url, review_url), subtype="html")
    msg.add_attachment(
        pdf_bytes,
        maintype="application",
        subtype="pdf",
        filename=f"{c.folio or c.id}.pdf",
    )

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg, to_addrs=[*recipients, *bcc])


def _quote_review_response_mail_html(c: Cotizacion, selected_status: str, reason: str = "") -> str:
    normalized = (selected_status or "").strip().upper()
    if normalized in {"APROBADO", "APROBADA", "AUTORIZADO"}:
        accent = "#16854f"
        bg_soft = "#eaf7f0"
        title = "Cotizacion autorizada"
    elif normalized in {"RECHAZADO", "RECHAZADA"}:
        accent = "#c62828"
        bg_soft = "#fdecee"
        title = "Cotizacion rechazada"
    else:
        accent = "#f0ad00"
        bg_soft = "#fff7df"
        title = "Cotizacion en revision"

    cli = c.cliente
    folio = escape(c.folio or f"#{c.id}")
    cliente = escape(cli.nombre_cliente if cli else "Sin cliente")
    empresa = escape(cli.empresa if cli and cli.empresa else "Sin empresa")
    proyecto = escape(c.proyecto or "Sin proyecto")
    responsable = escape(c.responsable or "Sin responsable")
    total = f"{money(c.total)} {escape(c.moneda or 'MXN')}"
    motivo_html = ""
    if reason.strip():
        motivo_html = f"""
          <div style="margin-top:22px;padding:18px 20px;background:#fff;border:1px solid #f1c4c9;border-left:6px solid #c62828;border-radius:8px;">
            <div style="font-size:12px;text-transform:uppercase;letter-spacing:.7px;color:#9f1d1d;font-weight:800;margin-bottom:8px;">Motivo de rechazo</div>
            <div style="font-size:15px;line-height:1.55;color:#1f2937;white-space:pre-wrap;">{escape(reason.strip())}</div>
          </div>
        """

    return f"""
    <html>
      <body style="margin:0;padding:0;background:#eef2f7;font-family:Arial,Helvetica,sans-serif;color:#1f2937;">
        <div style="max-width:760px;margin:0 auto;padding:30px 16px;">
          <div style="background:#ffffff;border:1px solid #d9e2ec;border-radius:10px;overflow:hidden;box-shadow:0 8px 24px rgba(15,45,80,.08);">
            <div style="background:{accent};color:#ffffff;padding:24px 28px;">
              <div style="font-size:12px;font-weight:700;letter-spacing:.9px;text-transform:uppercase;opacity:.92;">MAR · Poliutech</div>
              <div style="font-size:24px;font-weight:900;margin-top:6px;">{title}</div>
              <div style="font-size:14px;opacity:.95;margin-top:7px;">Respuesta registrada para {folio}</div>
            </div>
            <div style="padding:28px;">
              <div style="display:inline-block;background:{bg_soft};border:1px solid {accent};border-radius:999px;padding:9px 16px;color:{accent};font-weight:900;font-size:13px;letter-spacing:.5px;text-transform:uppercase;">
                {escape(normalized)}
              </div>
              <p style="margin:20px 0 22px 0;font-size:15px;color:#475569;">Se registro una respuesta de revision para esta cotizacion.</p>
              <div style="border:1px solid #dbe4ef;border-radius:10px;overflow:hidden;">
                <div style="background:#f8fafc;padding:15px 18px;border-bottom:1px solid #dbe4ef;">
                  <div style="font-size:12px;text-transform:uppercase;letter-spacing:.7px;color:#64748b;font-weight:800;">Folio</div>
                  <div style="font-size:21px;font-weight:900;color:{accent};margin-top:2px;">{folio}</div>
                </div>
                <table style="border-collapse:collapse;width:100%;background:#ffffff;">
                  <tr>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;width:34%;color:#64748b;font-weight:800;">Cliente</td>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#111827;font-weight:700;">{cliente}</td>
                  </tr>
                  <tr>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:800;">Empresa</td>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#111827;">{empresa}</td>
                  </tr>
                  <tr>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:800;">Proyecto</td>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#111827;">{proyecto}</td>
                  </tr>
                  <tr>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:800;">Responsable</td>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#111827;">{responsable}</td>
                  </tr>
                  <tr>
                    <td style="padding:13px 16px;color:#64748b;font-weight:800;">Total</td>
                    <td style="padding:13px 16px;color:{accent};font-size:20px;font-weight:900;">{total}</td>
                  </tr>
                </table>
              </div>
              {motivo_html}
              <div style="margin-top:22px;padding-top:16px;border-top:1px solid #e5e7eb;color:#64748b;font-size:12px;">
                Este mensaje fue generado automaticamente por MAR. El seguimiento de la cotizacion ya fue actualizado.
              </div>
            </div>
          </div>
        </div>
      </body>
    </html>
    """.strip()


def _send_quote_review_response_email(c: Cotizacion, selected_status: str, reason: str = "") -> None:
    recipients = _unique_emails(
        _quote_responsible_email(c),
        _parse_email_list(COTIZACION_RESPONSE_EMAIL),
        _parse_email_list(COTIZACION_APPROVALS_EMAIL),
        _parse_email_list(COTIZACION_REVIEW_RESULT_AAZCONA_EMAIL),
    )
    if not recipients:
        raise ValueError("No hay correo configurado para respuesta de cotizaciones.")
    motivo_line = f"\nMotivo de rechazo: {reason.strip()}" if reason.strip() else ""
    msg = EmailMessage()
    msg["Subject"] = f"Respuesta cotizacion {c.folio or c.id}: {selected_status}"
    msg["From"] = f"COTIZACIONES POLIUTECH <{SMTP_FROM or SMTP_USERNAME}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content(
        f"Se registro una respuesta para la cotizacion {c.folio or c.id}.\n"
        f"Estatus seleccionado: {selected_status}\n"
        f"Cliente: {c.cliente.nombre_cliente if c.cliente else 'Sin cliente'}\n"
        f"Proyecto: {c.proyecto or 'Sin proyecto'}\n"
        f"Total: {money(c.total)} {c.moneda or 'MXN'}"
        f"{motivo_line}\n"
    )
    msg.add_alternative(_quote_review_response_mail_html(c, selected_status, reason), subtype="html")
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg, to_addrs=recipients)


def _send_quote_review_email_safely(c: Cotizacion) -> None:
    try:
        _send_quote_review_email(c)
    except Exception as exc:
        try:
            logger.exception("No se pudo enviar correo de revision de cotizacion %s", c.folio or c.id)
        except Exception:
            pass
        print(f"[Cotizaciones] Error enviando revision {c.folio or c.id}: {exc}", file=sys.stderr)


def _apply_quote_review_decision(
    c: Cotizacion,
    selected_status: str,
    reason: str = "",
    actor: Optional[Usuario] = None,
    author_label: str = "Revision por correo",
) -> CotizacionSeguimiento:
    selected_status = (selected_status or "").strip().upper()
    if selected_status == "APROBADO":
        selected_status = "APROBADA"
    elif selected_status == "RECHAZADO":
        selected_status = "RECHAZADA"
    elif selected_status == "EN REVISION":
        selected_status = "EN REVISIÓN"
    if selected_status not in set(VALID_ESTATUS_APROBACION):
        abort(400)
    if selected_status == "RECHAZADA" and not reason.strip():
        abort(400)

    previous_status = c.estatus_aprobacion
    c.estatus_aprobacion = selected_status
    comentario = f"Revision de cotizacion: {selected_status}."
    if reason.strip():
        comentario += f"\nMotivo de rechazo: {reason.strip()}"
    seg = CotizacionSeguimiento(
        cotizacion_id=c.id,
        usuario_id=getattr(actor, "id", None),
        autor=author_label or "Revision por correo",
        comentario=comentario,
        fecha_seguimiento=now_cdmx_naive(),
        actualizado_en=now_cdmx_naive(),
    )
    db.session.add(seg)
    db.session.commit()
    _send_quote_review_response_email(c, selected_status, reason)
    try:
        _send_quote_review_result_push(c, selected_status, reason)
    except Exception as exc:
        logger.warning("Push de respuesta de revisión falló: %s", exc)
    return seg


@app.route("/cotizaciones/revision/<int:cot_id>/<action>", methods=["GET", "POST"])
def cotizacion_revision_decidir(cot_id: int, action: str):
    action = (action or "").strip().lower()
    if action not in {"approve", "reject", "review"}:
        abort(404)
    c = _quote_review_load_from_token(cot_id, request.args.get("token"), action)
    token = request.args.get("token") or ""

    status_by_action = {
        "approve": "APROBADA",
        "review": "EN REVISIÓN",
        "reject": "RECHAZADA",
    }
    selected_status = status_by_action[action]

    if action == "reject":
        if request.method == "POST":
            reason = (request.form.get("motivo") or "").strip()
            if not reason:
                return render_template_string(
                    _quote_reject_form_html(c, token, "Captura el motivo del rechazo."),
                    title=f"Rechazar {c.folio}",
                ), 400
            _apply_quote_review_decision(c, selected_status, reason)
            return redirect(url_for("view_cotizacion", cot_id=c.id, quote_review_done="1"))

        return render_template_string(_quote_reject_form_html(c, token), title=f"Rechazar {c.folio}")

    _apply_quote_review_decision(c, selected_status)
    return redirect(url_for("view_cotizacion", cot_id=c.id, quote_review_done="1"))


def _quote_reject_form_html(c: Cotizacion, token: str, error: str = "") -> str:
    error_html = f'<div class="alert alert-danger">{escape(error)}</div>' if error else ""
    action_url = url_for("cotizacion_revision_decidir", cot_id=c.id, action="reject", token=token)
    return f"""
    <!doctype html>
    <html lang="es">
      <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>Rechazar {escape(c.folio or str(c.id))}</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
        <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
      </head>
      <body style="background:#f1f3f5;">
        <main class="container py-5" style="max-width:760px;">
          <div class="card shadow-sm" style="border:2px solid #d9dee5;border-radius:8px;overflow:hidden;">
            <div class="card-header bg-danger text-white">
              <h1 class="h5 mb-0">Rechazar cotización {escape(c.folio or str(c.id))}</h1>
            </div>
            <div class="card-body">
              {error_html}
              <p class="text-muted">Escribe el motivo del rechazo. Se guardará en el seguimiento de la cotización y se notificará por correo.</p>
              <form id="quote-reject-form" method="post" action="{action_url}">
                <label for="motivo" class="form-label fw-bold">Motivo de rechazo</label>
                <textarea id="motivo" name="motivo" class="form-control" rows="5" style="border:2px solid #b8c0cc;box-shadow:inset 0 1px 2px rgba(0,0,0,.04);" required></textarea>
                <div class="d-flex justify-content-end mt-3">
                  <button type="submit" class="btn btn-danger">Guardar rechazo</button>
                </div>
              </form>
            </div>
          </div>
        </main>
        <script>
          const rejectForm = document.getElementById("quote-reject-form");
          if (rejectForm) {{
            rejectForm.addEventListener("submit", () => {{
              const btn = rejectForm.querySelector('button[type="submit"]');
              if (btn) btn.disabled = true;
              Swal.fire({{
                title: "Enviando motivo...",
                text: "Se está guardando el rechazo y notificando por correo.",
                allowOutsideClick: false,
                allowEscapeKey: false,
                showConfirmButton: false,
                didOpen: () => Swal.showLoading(),
              }});
            }});
          }}
        </script>
      </body>
    </html>
    """.strip()


def _quote_decision_result_html(c: Cotizacion, selected_status: str, reason: str = "") -> str:
    reason_html = f"<p><b>Motivo:</b> {escape(reason)}</p>" if reason else ""
    return f"""
    <!doctype html>
    <html lang="es">
      <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>{escape(c.folio or str(c.id))}</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
      </head>
      <body class="bg-light">
        <main class="container py-5" style="max-width:760px;">
          <div class="card shadow-sm border-0">
            <div class="card-header bg-primary text-white">
              <h1 class="h5 mb-0">Respuesta registrada</h1>
            </div>
            <div class="card-body">
              <p class="mb-2">La cotización <b>{escape(c.folio or str(c.id))}</b> quedó con estatus <b>{escape(selected_status)}</b>.</p>
              {reason_html}
              <p class="text-muted mb-0">El seguimiento fue actualizado y se envió la respuesta por correo.</p>
            </div>
          </div>
        </main>
      </body>
    </html>
    """.strip()


@app.route("/api/cotizaciones/<int:cot_id>/send-email", methods=["POST"])
@login_required
def api_send_cotizacion_email(cot_id: int):
    c = _cotizacion_activa_or_404(cot_id)
    require_owner_or_admin(c)

    data = request.get_json(silent=True) or {}
    recipient = (data.get("to") or "").strip()
    if not recipient and c.cliente:
        recipient = (c.cliente.correo or "").strip()
    cc_raw = data.get("cc")
    bcc_raw = data.get("bcc")

    if not recipient:
        return jsonify({"ok": False, "error": "La cotización no tiene un correo destino."}), 400

    try:
        recipients = _parse_email_list(recipient)
        if not recipients:
            return jsonify({"ok": False, "error": "La cotización no tiene un correo destino."}), 400
        cc = _parse_email_list(cc_raw)
        bcc = _parse_email_list(bcc_raw)
        _send_cotizacion_email(c, recipients, cc=cc, bcc=bcc)
        to_display = ", ".join(recipients)
        return jsonify({
            "ok": True,
            "folio": c.folio,
            "to": to_display,
            "cc": cc,
            "bcc_count": len(bcc),
            "message": f"Cotización {c.folio} enviada a {to_display}."
        })
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        print(f"[MAIL] Error enviando cotización {c.folio} a {recipient}: {e}", file=sys.stderr)
        return jsonify({"ok": False, "error": f"No se pudo enviar el correo: {e}"}), 500

# ---------------------------------------------------------

@app.route("/cotizaciones/export/dashboard.xlsx")
@login_required
def export_dashboard_cotizaciones_xlsx():
    if Workbook is None:
        abort(501, description="openpyxl no instalado en el servidor.")

    desde = (request.args.get("desde") or "").strip()
    hasta = (request.args.get("hasta") or "").strip()
    estatus = (request.args.get("estatus") or "").strip()
    cliente = (request.args.get("cliente") or "").strip()
    especialidad = (request.args.get("especialidad") or "").strip()

    especialidad = (request.args.get("especialidad") or "").strip()

    try:
        cotizaciones = (_build_dashboard_cotizaciones_query(
            desde=desde,
            hasta=hasta,
            estatus=estatus,
            cliente=cliente,
            especialidad=especialidad,
        ).order_by(Cotizacion.fecha.desc()).all())
    except ValueError as exc:
        abort(400, description=str(exc))

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotizaciones"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    header_fill = PatternFill("solid", fgColor=MAR_BLUE_XLSX)
    white = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:K1")
    ws["A1"] = "REPORTE DE COTIZACIONES"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    filtros_texto = []
    if desde:
        filtros_texto.append(f"Desde: {desde}")
    if hasta:
        filtros_texto.append(f"Hasta: {hasta}")
    if estatus:
        filtros_texto.append(f"Estatus: {estatus}")
    if cliente:
        filtros_texto.append(f"Cliente/Empresa: {cliente}")
    if especialidad:
        filtros_texto.append(f"Especialidad: {especialidad}")
    if not filtros_texto:
        filtros_texto.append("Sin filtros")

    ws.merge_cells("A2:M2")
    ws["A2"] = " | ".join(filtros_texto)
    ws["A2"].alignment = left

    headers = ["Folio", "Fecha", "Cliente", "Empresa", "Telefono", "Responsable", "Especialidad", "Aprobacion", "Seguimiento", "Subtotal", "IVA %", "IVA $", "Total"]
    ws.append([])
    ws.append(headers)

    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = white
        cell.alignment = center
        cell.border = border

    for c in cotizaciones:
        ws.append([
            c.folio or "",
            c.fecha.strftime("%Y-%m-%d %H:%M") if c.fecha else "",
            c.cliente.nombre_cliente if c.cliente else "",
            c.cliente.empresa if c.cliente else "",
            c.cliente.telefono if c.cliente and c.cliente.telefono else "",
            c.responsable or "",
            c.especialidad or "",
            c.estatus_aprobacion or "EN REVISIÓN",
            c.estatus or "",
            float(c.subtotal or 0),
            float(c.iva_porc or 0),
            float(c.iva_monto or 0),
            float(c.total or 0),
        ])
        row = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = border
        for col in (10, 12, 13):
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=11).number_format = '0.00'
        ws.cell(row=row, column=1).alignment = left
        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=3).alignment = left
        ws.cell(row=row, column=4).alignment = left
        ws.cell(row=row, column=5).alignment = left

    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=12, value="Total exportado:").font = bold
    ws.cell(row=total_row, column=13, value=f"=SUM(M{header_row + 1}:M{ws.max_row})")
    ws.cell(row=total_row, column=13).font = bold
    ws.cell(row=total_row, column=13).number_format = '"$"#,##0.00'

    ws.auto_filter.ref = f"A{header_row}:M{max(header_row, ws.max_row)}"
    ws.freeze_panes = f"A{header_row + 1}"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 14

    # Las graficas se construyen con la misma lista ya filtrada que alimenta
    # la tabla. De esta forma el archivo siempre representa exactamente lo que
    # el usuario estaba viendo al momento de exportar.
    monthly = {}
    status_counts = {status: 0 for status in VALID_ESTATUS}
    for cot in cotizaciones:
        if cot.fecha:
            month_key = cot.fecha.strftime("%Y-%m")
            month_values = monthly.setdefault(month_key, {"count": 0, "total": 0.0})
            month_values["count"] += 1
            month_values["total"] += float(cot.total or 0)

        status_key = (cot.estatus or "").strip().upper()
        if status_key in status_counts:
            status_counts[status_key] += 1

    charts_ws = wb.create_sheet("Graficas")
    charts_ws.sheet_view.showGridLines = False
    charts_ws.merge_cells("A1:N1")
    charts_ws["A1"] = "GRAFICAS DE COTIZACIONES"
    charts_ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    charts_ws["A1"].fill = header_fill
    charts_ws["A1"].alignment = center
    charts_ws.merge_cells("A2:N2")
    charts_ws["A2"] = " | ".join(filtros_texto)
    charts_ws["A2"].alignment = left

    charts_ws["A4"] = "Indicador"
    charts_ws["B4"] = "Valor"
    charts_ws["A5"] = "Cotizaciones"
    charts_ws["B5"] = len(cotizaciones)
    charts_ws["A6"] = "Importe total"
    charts_ws["B6"] = sum(float(cot.total or 0) for cot in cotizaciones)
    charts_ws["B6"].number_format = '"$"#,##0.00'
    for row in charts_ws.iter_rows(min_row=4, max_row=6, min_col=1, max_col=2):
        for cell in row:
            cell.border = border
    for cell in charts_ws[4]:
        if cell.column <= 2:
            cell.fill = header_fill
            cell.font = white
            cell.alignment = center

    monthly_header_row = 9
    charts_ws.cell(monthly_header_row, 1, "Mes")
    charts_ws.cell(monthly_header_row, 2, "Cotizaciones")
    charts_ws.cell(monthly_header_row, 3, "Importe")
    for col in range(1, 4):
        cell = charts_ws.cell(monthly_header_row, col)
        cell.fill = header_fill
        cell.font = white
        cell.alignment = center
        cell.border = border

    for month_key in sorted(monthly):
        values = monthly[month_key]
        charts_ws.append([month_key, values["count"], values["total"]])
        charts_ws.cell(charts_ws.max_row, 3).number_format = '"$"#,##0.00'

    monthly_last_row = charts_ws.max_row
    if monthly_last_row > monthly_header_row:
        amount_chart = BarChart()
        amount_chart.type = "col"
        amount_chart.style = 10
        amount_chart.title = "Importe y cotizaciones por mes"
        amount_chart.y_axis.title = "Importe"
        amount_chart.x_axis.title = "Mes"
        amount_chart.height = 9
        amount_chart.width = 18
        amount_chart.add_data(
            Reference(charts_ws, min_col=3, min_row=monthly_header_row, max_row=monthly_last_row),
            titles_from_data=True,
        )
        amount_chart.set_categories(
            Reference(charts_ws, min_col=1, min_row=monthly_header_row + 1, max_row=monthly_last_row)
        )

        count_chart = LineChart()
        count_chart.add_data(
            Reference(charts_ws, min_col=2, min_row=monthly_header_row, max_row=monthly_last_row),
            titles_from_data=True,
        )
        count_chart.y_axis.title = "Cotizaciones"
        count_chart.y_axis.axId = 200
        count_chart.y_axis.crosses = "max"
        count_chart.graphicalProperties = None
        amount_chart += count_chart
        charts_ws.add_chart(amount_chart, "E4")

    status_header_row = max(monthly_last_row + 3, 18)
    charts_ws.cell(status_header_row, 1, "Estatus")
    charts_ws.cell(status_header_row, 2, "Cotizaciones")
    for col in range(1, 3):
        cell = charts_ws.cell(status_header_row, col)
        cell.fill = header_fill
        cell.font = white
        cell.alignment = center
        cell.border = border
    for status, count in status_counts.items():
        charts_ws.append([status, count])

    status_last_row = charts_ws.max_row
    if sum(status_counts.values()) > 0:
        status_chart = PieChart()
        status_chart.title = "Distribucion por estatus"
        status_chart.height = 9
        status_chart.width = 15
        status_chart.add_data(
            Reference(charts_ws, min_col=2, min_row=status_header_row, max_row=status_last_row),
            titles_from_data=True,
        )
        status_chart.set_categories(
            Reference(charts_ws, min_col=1, min_row=status_header_row + 1, max_row=status_last_row)
        )
        status_chart.dataLabels = DataLabelList()
        status_chart.dataLabels.showPercent = True
        status_chart.dataLabels.showLeaderLines = True
        charts_ws.add_chart(status_chart, f"E{status_header_row}")

    charts_ws.column_dimensions["A"].width = 24
    charts_ws.column_dimensions["B"].width = 16
    charts_ws.column_dimensions["C"].width = 18
    charts_ws.freeze_panes = "A4"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    stamp = now_cdmx_naive().strftime("%Y%m%d_%H%M%S")
    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="cotizaciones_dashboard_{stamp}.xlsx"'}
    )


@app.route("/cotizaciones/export/seguimientos.pdf")
@login_required
def export_dashboard_followups_pdf():
    desde = (request.args.get("desde") or "").strip()
    hasta = (request.args.get("hasta") or "").strip()
    estatus = (request.args.get("estatus") or "").strip()
    cliente = (request.args.get("cliente") or "").strip()
    especialidad = (request.args.get("especialidad") or "").strip()

    try:
        cotizaciones = (
            _build_dashboard_cotizaciones_query(
                desde=desde,
                hasta=hasta,
                estatus=estatus,
                cliente=cliente,
                especialidad=especialidad,
            )
            .order_by(Cotizacion.fecha.desc())
            .all()
        )
    except ValueError as exc:
        abort(400, description=str(exc))

    total_con_seguimiento = sum(1 for cot in cotizaciones if cot.seguimientos)
    total_sin_seguimiento = max(0, len(cotizaciones) - total_con_seguimiento)
    total_importe = sum(float(c.total or 0) for c in cotizaciones)
    estatus_counts = {status: 0 for status in VALID_ESTATUS}
    for cot in cotizaciones:
        status = (cot.estatus or "").strip().upper()
        if status in estatus_counts:
            estatus_counts[status] += 1

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=22 * mm,
        bottomMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="FollowupHeading", fontName="Helvetica-Bold", fontSize=11, leading=14, textColor=colors.HexColor(MAR_BLUE), spaceAfter=4))
    styles.add(ParagraphStyle(name="FollowupBody", fontName="Helvetica", fontSize=9, leading=12, textColor=colors.HexColor("#222222"), spaceAfter=2))
    styles.add(ParagraphStyle(name="FollowupMeta", fontName="Helvetica", fontSize=8.3, leading=10.5, textColor=colors.HexColor("#5f6b7a"), spaceAfter=2))
    styles.add(ParagraphStyle(name="FollowupComment", fontName="Helvetica", fontSize=9, leading=12, textColor=colors.HexColor("#222222"), spaceAfter=4))

    elems = []

    def _header_footer(canv, doc_):
        canv.saveState()
        canv.setFillColor(colors.HexColor(MAR_BLUE))
        canv.rect(0, A4[1] - 34, A4[0], 34, stroke=0, fill=1)

        logo_path = os.path.join(app.static_folder or "static", "logo.png")
        if os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                max_w = 18 * mm
                scale = max_w / iw
                canv.drawImage(img, 12, A4[1] - (ih * scale) - 8, width=max_w, height=ih * scale, mask="auto")
            except Exception:
                pass

        canv.setFont("Helvetica-Bold", 13)
        canv.setFillColor(colors.white)
        canv.drawRightString(A4[0] - 12, A4[1] - 14, "BITACORA DE SEGUIMIENTO")
        canv.setFont("Helvetica", 8.5)
        canv.drawRightString(A4[0] - 12, A4[1] - 25, "Comentarios de cotizaciones")

        canv.setFont("Helvetica", 8)
        canv.setFillColor(colors.HexColor("#555555"))
        canv.drawString(12 * mm, 8 * mm, f"Generado: {now_cdmx_naive().strftime('%d/%m/%Y %H:%M')}")
        canv.drawRightString(A4[0] - 12 * mm, 8 * mm, f"Pagina {doc_.page}")
        canv.restoreState()

    filtros_texto = []
    if desde:
        filtros_texto.append(f"Desde: {desde}")
    if hasta:
        filtros_texto.append(f"Hasta: {hasta}")
    if estatus:
        filtros_texto.append(f"Estatus: {estatus}")
    if cliente:
        filtros_texto.append(f"Cliente/Empresa: {cliente}")
    if especialidad:
        filtros_texto.append(f"Especialidad: {especialidad}")
    if not filtros_texto:
        filtros_texto.append("Sin filtros")

    resumen_data = [
        [Paragraph(f"<b>Total cotizaciones:</b> {len(cotizaciones)}", styles["FollowupBody"]),
         Paragraph(f"<b>Con seguimiento:</b> {total_con_seguimiento}", styles["FollowupBody"])],
        [Paragraph(f"<b>Sin seguimiento:</b> {total_sin_seguimiento}", styles["FollowupBody"]),
         Paragraph(f"<b>Importe total:</b> {money(total_importe)}", styles["FollowupBody"])],
        [Paragraph(f"<b>Filtros:</b> {' | '.join(filtros_texto)}", styles["FollowupMeta"]),
         Paragraph(f"<b>Estatus:</b> {' | '.join(f'{k}: {v}' for k, v in estatus_counts.items())}", styles["FollowupMeta"])],
    ]
    resumen_tbl = Table(resumen_data, colWidths=[90 * mm, 90 * mm], hAlign="LEFT")
    resumen_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#f3f7fb")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cfd9e5")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9e2ec")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elems.append(Paragraph("Resumen", styles["FollowupHeading"]))
    elems.append(resumen_tbl)
    elems.append(Spacer(1, 8))

    if not cotizaciones:
        elems.append(Paragraph("No hay cotizaciones para los filtros seleccionados.", styles["FollowupBody"]))
    else:
        for idx, cot in enumerate(cotizaciones, start=1):
            cli = cot.cliente
            cliente_nombre = (cli.nombre_cliente if cli else "") or "Sin cliente"
            empresa_nombre = (cli.empresa if cli else "") or "-"
            fecha_cot = cot.fecha.strftime("%d/%m/%Y %H:%M") if cot.fecha else "-"

            block_items = [
                Paragraph(f"{idx}. {escape(cot.folio or '-')}", styles["FollowupHeading"]),
                Paragraph(
                    f"<b>Cliente:</b> {escape(cliente_nombre)} &nbsp;&nbsp;&nbsp; "
                    f"<b>Empresa:</b> {escape(empresa_nombre)} &nbsp;&nbsp;&nbsp; "
                    f"<b>Estatus:</b> {escape(cot.estatus or '-')}",
                    styles["FollowupBody"],
                ),
                Paragraph(
                    f"<b>Responsable:</b> {escape(cot.responsable or '-')} &nbsp;&nbsp;&nbsp; "
                    f"<b>Fecha:</b> {fecha_cot} &nbsp;&nbsp;&nbsp; "
                    f"<b>Total:</b> {money(cot.total)}",
                    styles["FollowupBody"],
                ),
                Spacer(1, 2),
            ]

            seguimientos = sorted(list(cot.seguimientos), key=lambda seg: seg.fecha_seguimiento or datetime.min)
            if seguimientos:
                for seg in seguimientos:
                    fecha_seg = seg.fecha_seguimiento.strftime("%d/%m/%Y %H:%M") if seg.fecha_seguimiento else "-"
                    editado = ""
                    if seg.actualizado_en and seg.fecha_seguimiento and seg.actualizado_en != seg.fecha_seguimiento:
                        editado = f" · Editado {seg.actualizado_en.strftime('%d/%m/%Y %H:%M')}"
                    block_items.append(Paragraph(f"<b>{fecha_seg}</b> · {escape(seg.autor or 'Sistema')}{editado}", styles["FollowupMeta"]))
                    comentario_html = escape(seg.comentario or "").replace("\n", "<br/>")
                    block_items.append(Paragraph(comentario_html, styles["FollowupComment"]))
            else:
                block_items.append(Paragraph("Sin seguimiento registrado.", styles["FollowupMeta"]))

            bloque = Table([[item] for item in block_items], colWidths=[180 * mm], hAlign="LEFT")
            bloque.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#edf4fb")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#c7d6e6")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            elems.append(KeepTogether([bloque, Spacer(1, 6)]))

    doc.build(elems, onFirstPage=_header_footer, onLaterPages=_header_footer)
    stamp = now_cdmx_naive().strftime("%Y%m%d_%H%M%S")
    return Response(
        buf.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="seguimientos_cotizaciones_{stamp}.pdf"'}
    )


@app.route("/api/dashboard/filter-summary")
@login_required
def api_dashboard_filter_summary():
    desde = (request.args.get("desde") or "").strip()
    hasta = (request.args.get("hasta") or "").strip()
    estatus = (request.args.get("estatus") or "").strip()
    cliente = (request.args.get("cliente") or "").strip()
    especialidad = (request.args.get("especialidad") or "").strip()

    try:
        q = _build_dashboard_cotizaciones_query(
            desde=desde,
            hasta=hasta,
            estatus=estatus,
            cliente=cliente,
            especialidad=especialidad,
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    cot_subq = q.with_entities(Cotizacion.id).subquery()
    cot_ids_select = db.select(cot_subq.c.id)

    total_importe = (
        db.session.query(db.func.coalesce(db.func.sum(Cotizacion.total), 0))
        .filter(Cotizacion.id.in_(cot_ids_select))
        .scalar()
        or 0
    )
    total_cotizaciones = (
        db.session.query(db.func.count())
        .select_from(cot_subq)
        .scalar()
        or 0
    )
    total_conceptos = (
        db.session.query(db.func.count(CotizacionDetalle.id))
        .filter(CotizacionDetalle.cotizacion_id.in_(cot_ids_select))
        .scalar()
        or 0
    )

    return jsonify({
        "total_importe": float(total_importe),
        "total_cotizaciones": int(total_cotizaciones),
        "total_conceptos": int(total_conceptos),
    })

# PDF - Diseño corporativo
# - Quitar RFC
# - "Condiciones comerciales"
# - RESPONSABLE: poner valor debajo del label (rellena el “espacio en blanco”)
# ---------------------------------------------------------
def draw_watermark(canvas, app):
    try:
        import os
        watermark_path = os.path.join(app.static_folder, "watermark.png")
        if os.path.exists(watermark_path):
            canvas.saveState()
            canvas.setFillAlpha(0.08)
            img = ImageReader(watermark_path)
            page_width, page_height = canvas._pagesize
            width = 300
            height = 300
            x = (page_width - width) / 2
            y = (page_height / 2) - 150
            canvas.drawImage(img, x, y, width=width, height=height, mask='auto')
            canvas.restoreState()
    except Exception:
        pass


def _build_cotizacion_pdf_response(c: Cotizacion):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=24*mm, bottomMargin=38*mm
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Encabezado", fontName="Helvetica", fontSize=9, leading=12, spaceAfter=4, splitLongWords=False))
    styles.add(ParagraphStyle(name="EncabezadoCorreo", parent=styles["Encabezado"], splitLongWords=True, wordWrap="CJK"))
    styles.add(ParagraphStyle(name="NormalCell", fontName="Helvetica", fontSize=8, leading=10, splitLongWords=False))
    styles.add(ParagraphStyle(name="NormalRight", fontName="Helvetica", fontSize=8, leading=10, alignment=2, splitLongWords=False))
    styles.add(ParagraphStyle(name="NormalCenter", fontName="Helvetica", fontSize=8, leading=10, alignment=1, splitLongWords=False))
    styles.add(ParagraphStyle(
        name="UnitCell",
        fontName="Helvetica",
        fontSize=6.5,
        leading=7,
        alignment=1,
        splitLongWords=False,
    ))

    elems = []

    def encabezado(canv, doc_):
        canv.saveState()
        canv.setFillColor(colors.HexColor(MAR_BLUE))
        canv.rect(0, A4[1]-40, A4[0], 40, stroke=0, fill=1)

        logo_path = os.path.join(app.static_folder or "static", "logo.png")
        if os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                max_w = 22.5 * mm
                scale = max_w / iw
                w = max_w
                h = ih * scale
                x_logo = 12
                y_logo = A4[1] - h - 8
                canv.drawImage(img, x_logo, y_logo, width=w, height=h, mask="auto")
            except Exception:
                pass

        canv.setFont("Helvetica-Bold", 14)
        canv.setFillColor(colors.white)
        canv.drawRightString(A4[0]-12, A4[1]-18, "COTIZACIÓN POLIUTECH")
        canv.setFont("Helvetica", 10)
        canv.drawRightString(A4[0]-12, A4[1]-31, "Recubrimientos Especializados")
        canv.restoreState()

    def footer(canv, doc_):
        canv.saveState()
        y_firma = 80
        canv.setFont("Helvetica", 9)
        canv.setFillColor(colors.black)
        canv.drawCentredString(A4[0]/2, y_firma + 18, "Atte.")
        canv.setFont("Helvetica-Bold", 9)
        canv.drawCentredString(A4[0]/2, y_firma + 6, "Ing. César Antonio Garza Guerrero")
        canv.setFont("Helvetica", 9)
        canv.drawCentredString(A4[0]/2, y_firma - 6, "DIRECTOR GENERAL")

        division_path = os.path.join(app.static_folder or "static", "division.png")
        if os.path.exists(division_path):
            try:
                canv.drawImage(division_path, (A4[0]-155*mm)/2, 45, width=155*mm, height=3*mm, mask="auto")
            except Exception:
                pass

        canv.setFont("Helvetica-Bold", 9)
        canv.setFillColor(colors.HexColor(MAR_BLUE))
        canv.drawCentredString(A4[0]/2, 35, "POLIUTECH – Recubrimientos Especializados")

        canv.setFont("Helvetica", 8)
        canv.setFillColor(colors.HexColor("#333333"))
        line1 = "Campos Elíseos 223 Oficina 602 · Col. Polanco V Sección · Miguel Hidalgo, CDMX 11560"
        line2 = "Tel: 55 5938 6530 / 55 5938 0536 · info@poliutech.com · www.poliutech.com"
        canv.drawCentredString(A4[0]/2, 25, line1)
        canv.drawCentredString(A4[0]/2, 15, line2)

        try:
            canv.setTitle(c.folio or "Cotizacion")
        except Exception:
            pass

        canv.restoreState()

    # === DATOS PRINCIPALES ===
    cli = c.cliente
    cliente_nombre = cli.nombre_cliente if cli else ""
    cliente_empresa = cli.empresa if cli else ""
    cliente_correo = cli.correo if cli else ""
    cliente_telefono = cli.telefono if cli else ""
    ciudad_trabajo = (getattr(c, "ciudad_trabajo", "") or "").strip()
    moneda = normalize_moneda(getattr(c, "moneda", None))
    try:
        correo_lineas = _parse_email_list(cliente_correo)
    except ValueError:
        correo_lineas = [part.strip() for part in str(cliente_correo or "").split(",") if part.strip()]
    correo_pdf = "<br/>".join(escape(correo) for correo in correo_lineas) if correo_lineas else ""

    meta_data = [
        [
            Paragraph(f"<b>Folio:</b> {c.folio}", styles["Encabezado"]),
            Paragraph(f"<b>Fecha:</b> {c.fecha.strftime('%d/%m/%Y %H:%M')}", styles["Encabezado"]),
        ],
        [
            Paragraph(f"<b>Responsable:</b> {c.responsable or ''}", styles["Encabezado"]),
            Paragraph(f"<b>Cliente:</b> {cliente_nombre}", styles["Encabezado"]),
        ],
        [
            Paragraph(f"<b>Empresa:</b> {cliente_empresa}", styles["Encabezado"]),
            Paragraph(f"<b>Correo:</b> {correo_pdf}", styles["EncabezadoCorreo"]),
        ],
        [
            Paragraph(f"<b>Teléfono:</b> {cliente_telefono}", styles["Encabezado"]),
            Paragraph(f"<b>Ciudad:</b> {escape(ciudad_trabajo)}", styles["Encabezado"]),
        ],
        [
            Paragraph(f"<b>Moneda:</b> {moneda_label(moneda)}", styles["Encabezado"]),
            Paragraph("", styles["Encabezado"]),
        ],
    ]
    meta_tbl = Table(meta_data, colWidths=[95*mm, 95*mm], hAlign="LEFT")
    meta_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))
    elems.append(meta_tbl)
    elems.append(Spacer(1, 4))

    # === TABLA DE CONCEPTOS ===
    data = [["Ptda", "Concepto", "Uni.", "Cant.", "Sistema", "Precio Unitario", "Subtotal"]]
    for d in c.detalles:
        data.append([
            Paragraph(_truncate_pdf_text(getattr(d, "capitulo", "") or "-", 28), styles["NormalCenter"]),
            Paragraph((d.nombre_concepto or "-").strip(), styles["NormalCell"]),
            Paragraph(" ".join(str(d.unidad or "-").strip().splitlines()), styles["UnitCell"]),
            Paragraph(f"{(d.cantidad or 0):.2f}", styles["NormalCenter"]),
            Paragraph((d.sistema or "-").strip(), styles["NormalCell"]),
            Paragraph(money_currency(d.precio_unitario, moneda), styles["NormalRight"]),
            Paragraph(money_currency(d.subtotal, moneda), styles["NormalRight"]),
        ])

    tbl = Table(
        data,
        colWidths=[12*mm, 78*mm, 16*mm, 16*mm, 24*mm, 22*mm, 22*mm],
        repeatRows=1,
        hAlign="CENTER"
    )
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(MAR_BLUE)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 1), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),
        ("ALIGN", (4, 1), (4, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("WORDWRAP", (0, 0), (-1, -1), True),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))

    elems.append(tbl)
    elems.append(Spacer(1, 6))

    # === CANTIDAD EN LETRA ===
    resumen_elems = []
    try:
        cantidad_letra = cantidad_en_letra(float(c.total or 0), moneda).replace("Cantidad en letra: ", "", 1)
        resumen_elems.append(Paragraph(f"<b>Cantidad en letra:</b> {cantidad_letra}", styles["Encabezado"]))
        resumen_elems.append(Spacer(1, 4))
    except Exception as e:
        print(f"[PDF] num2words error: {e}", file=sys.stderr)

    # === TOTALES ===
    # === TOTALES (con descuento si aplica) ===
    subtotal = float(c.subtotal or 0)
    descuento = float(c.descuento_total or 0)
    subtotal_desc = subtotal - descuento
    descuento_porc_pdf = (descuento / subtotal * 100.0) if subtotal > 0 else 0.0

    tot_data = [["Subtotal:", money_currency(subtotal, moneda)]]
    if descuento and descuento > 0.0001:
        tot_data.append([f"Descuento ({descuento_porc_pdf:g}%):", "-" + money_currency(descuento, moneda)])
        tot_data.append(["Subtotal c/ desc.:", money_currency(subtotal_desc, moneda)])
    tot_data.extend([
        [f"IVA ({c.iva_porc:.2f}%):", money_currency(c.iva_monto, moneda)],
        ["Total:", money_currency(c.total, moneda)],
    ])
    t2 = Table(tot_data, colWidths=[45*mm, 35*mm], hAlign="RIGHT")
    t2.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("LINEBELOW", (0, -1), (-1, -1), 0.5, colors.black),
    ]))
    resumen_elems.append(t2)
    elems.append(KeepTogether(resumen_elems))
    elems.append(Spacer(1, 6))

    # === CONDICIONES COMERCIALES ===
    condiciones = _condiciones_comerciales_finales(c.notas or "")
    if condiciones:
        elems.append(Paragraph("<b>Condiciones Comerciales:</b>", styles["Encabezado"]))
        nota_style = ParagraphStyle(
            "NotasJustify",
            parent=styles["Normal"],
            alignment=TA_JUSTIFY,
            leading=11,
            fontSize=9,
            leftIndent=8,
        )
        # Una condicion vacia representa un renglon en blanco intencional.
        # Asi, dos Enter en el textarea separan visualmente dos puntos.
        def _linea_condicion_pdf(value: object) -> str:
            texto = str(value).strip()
            if not texto:
                return ""
            texto_normalizado = _normalize_text_for_match(texto).rstrip(":").strip()
            if texto_normalizado == "clausulas":
                return escape(texto)
            return f"• {escape(texto)}"

        bullets = "<br/>".join(_linea_condicion_pdf(x) for x in condiciones)
        elems.append(Paragraph(bullets, nota_style))
        elems.append(Spacer(1, 8))

    doc.build(
        elems,
        onFirstPage=lambda canv, d: (draw_watermark(canv, app), encabezado(canv, d), footer(canv, d)),
        onLaterPages=lambda canv, d: (draw_watermark(canv, app), encabezado(canv, d), footer(canv, d))
    )

    buf.seek(0)
    response = Response(
        buf.getvalue(),
        mimetype="application/pdf",
        headers={'Content-Disposition': f'inline; filename="{c.folio}.pdf"'}
    )
    response.direct_passthrough = False
    return response


@app.route("/cotizaciones/<int:cot_id>/export.pdf")
def export_cotizacion_pdf(cot_id: int):
    c = _cotizacion_activa_or_404(cot_id)
    mobile_user = _mobile_user_from_token()
    if mobile_user:
        if not _mobile_user_can_access_quote(mobile_user, c):
            abort(403)
    elif current_user.is_authenticated:
        require_owner_or_admin(c)
    else:
        return login_manager.unauthorized()
    return _build_cotizacion_pdf_response(c)

@app.route("/cotizaciones/<int:cot_id>/pdf")
@login_required
def export_cotizacion_pdf_alias(cot_id: int):
    return export_cotizacion_pdf(cot_id)

# ---------------------------------------------------------
# PDF por FOLIO (compatibilidad)
# Soporta URLs tipo: /cotizaciones/PTCH-0002/export.pdf
# ---------------------------------------------------------
@app.route("/cotizaciones/<string:folio>/export.pdf")
@login_required
def export_cotizacion_pdf_by_folio(folio: str):
    folio = (folio or "").strip()
    if not folio:
        abort(404)
    c = _cotizaciones_activas_query().filter(Cotizacion.folio == folio).first_or_404()
    require_owner_or_admin(c)
    return export_cotizacion_pdf(c.id)

# ---------------------------------------------------------
# API Dashboard (series / kpis / breakdown) — FILTRADO por responsable
# ---------------------------------------------------------
@app.route("/api/cotizaciones/search")
@login_required
def api_cotizaciones_search():
    q = Cotizacion.query.join(Cliente, isouter=True)
    q = q.filter(Cotizacion.eliminada_en.is_(None))

    if not is_admin():
        q = q.filter(Cotizacion.responsable == responsable_actual())

    estatus = (request.args.get("estatus") or "").strip()
    fi = (request.args.get("fi") or "").strip()
    ff = (request.args.get("ff") or "").strip()
    mmin = (request.args.get("mmin") or "").strip()
    mmax = (request.args.get("mmax") or "").strip()

    if estatus:
        q = q.filter(Cotizacion.estatus == estatus)
    if fi:
        try: q = q.filter(Cotizacion.fecha >= datetime.fromisoformat(fi))
        except Exception: pass
    if ff:
        try: q = q.filter(Cotizacion.fecha <= datetime.fromisoformat(ff))
        except Exception: pass
    if mmin:
        try: q = q.filter(Cotizacion.total >= float(mmin))
        except Exception: pass
    if mmax:
        try: q = q.filter(Cotizacion.total <= float(mmax))
        except Exception: pass

    q = q.order_by(Cotizacion.fecha.desc()).limit(500)
    data = []
    for c in q.all():
        data.append({
            "id": c.id,
            "folio": c.folio,
            "cliente": c.cliente.nombre_cliente if c.cliente else "",
            "empresa": c.cliente.empresa if c.cliente else "",
            "fecha": c.fecha.strftime("%Y-%m-%d %H:%M"),
            "estatus": c.estatus,
            "estatus_aprobacion": c.estatus_aprobacion or "EN REVISIÓN",
            "especialidad": c.especialidad or "",
            "proyecto": c.proyecto or "",
            "total": round(c.total or 0, 2),
            "export_csv": url_for("export_cotizacion_csv", cot_id=c.id),
            "export_pdf": url_for("export_cotizacion_pdf", cot_id=c.id),
            "export_xlsx": url_for("export_cotizacion_xlsx", cot_id=c.id),
        })
    return jsonify(data)

@app.route("/api/dashboard/metrics")
@login_required
def api_dashboard_metrics():
    desde = (request.args.get("desde") or "").strip()
    hasta = (request.args.get("hasta") or "").strip()
    estatus = (request.args.get("estatus") or "").strip()
    cliente = (request.args.get("cliente") or "").strip()
    especialidad = (request.args.get("especialidad") or "").strip()

    try:
        q = _build_dashboard_cotizaciones_query(
            desde=desde,
            hasta=hasta,
            estatus=estatus,
            cliente=cliente,
            especialidad=especialidad,
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    rows = (
        q.with_entities(
            db.func.strftime("%Y-%m", Cotizacion.fecha).label("ym"),
            db.func.count(Cotizacion.id),
            db.func.coalesce(db.func.sum(Cotizacion.total), 0),
        )
        .filter(Cotizacion.fecha.isnot(None))
        .group_by("ym")
        .order_by("ym")
        .all()
    )
    series = [{"mes": ym, "cotizaciones": int(c), "total": float(t)} for ym, c, t in rows if ym]

    kpis = {
        "total_cotizaciones": q.count(),
        "total_importe": float(
            q.with_entities(db.func.coalesce(db.func.sum(Cotizacion.total), 0)).scalar() or 0
        ),
        "total_catalogo": Concepto.query.count(),
    }

    return jsonify({"series": series, "kpis": kpis})

@app.route("/api/dashboard/status_breakdown")
@login_required
def api_dashboard_status_breakdown():
    desde = (request.args.get("desde") or "").strip()
    hasta = (request.args.get("hasta") or "").strip()
    estatus = (request.args.get("estatus") or "").strip()
    cliente = (request.args.get("cliente") or "").strip()
    especialidad = (request.args.get("especialidad") or "").strip()

    try:
        q = _build_dashboard_cotizaciones_query(
            desde=desde,
            hasta=hasta,
            estatus=estatus,
            cliente=cliente,
            especialidad=especialidad,
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    rows = (
        q.with_entities(Cotizacion.estatus, db.func.count(Cotizacion.id))
        .group_by(Cotizacion.estatus)
        .all()
    )
    categorias = VALID_ESTATUS
    conteos_map = {(estatus or "").strip().upper(): cnt for estatus, cnt in rows}
    conteos = [int(conteos_map.get(cat, 0)) for cat in categorias]
    total = sum(conteos)
    porcentajes = [round((c * 100.0 / total), 2) if total > 0 else 0 for c in conteos]
    return jsonify({"labels": categorias, "counts": conteos, "percentages": porcentajes, "total": total})

# ---------------------------------------------------------
# Salud / Debug / Recordatorios
# ---------------------------------------------------------
@app.route("/health")
def health():
    return jsonify({"status": "ok", "now_cdmx": now_cdmx_naive().isoformat()}), 200

@app.route("/debug/send_test")
@login_required
def debug_send_test():
    if not is_admin():
        abort(403)
    msg = "✅ Mensaje de prueba - Sistema Poliutech (debug_send_test)."
    send_whatsapp_multi(ADMIN_LIST, msg)
    return jsonify({"sent": True, "to": ADMIN_LIST})

@app.route("/debug/mobile_push_hansel")
@login_required
def debug_mobile_push_hansel():
    if not is_admin():
        abort(403)
    reviewer_ids = _mobile_push_user_ids_for_approval_reviewer()
    hansel_ids = [18]
    tokens = _mobile_push_tokens_for_users(hansel_ids)
    used_fallback = False
    if not tokens:
        tokens = _mobile_push_tokens_for_users(reviewer_ids)
    if not tokens:
        tokens = _mobile_all_active_push_tokens()
        used_fallback = bool(tokens)
    result = _send_push_notification_debug(
        tokens,
        title="Prueba de notificación Hansel",
        body="Si ves esto, el token móvil y Firebase están funcionando.",
        data={
            "type": "quote_pending_approval",
            "target_user": "Hansel",
            "target_user_name": "Hansel",
            "recipient_user_name": "Hansel",
            "approval_reviewer": "Hansel",
        },
    )
    return jsonify({
        "ok": True,
        "hansel_ids": hansel_ids,
        "reviewer_ids": reviewer_ids,
        "tokens": len(tokens),
        "used_fallback": used_fallback,
        "sent": result.get("sent", 0),
        "failed": result.get("failed", 0),
        "errors": result.get("errors", []),
    })

@app.route("/debug/mobile_devices_hansel")
@login_required
def debug_mobile_devices_hansel():
    if not is_admin():
        abort(403)
    rows = (
        MobileDevice.query
        .filter(MobileDevice.usuario_id.in_([18]))
        .order_by(MobileDevice.updated_at.desc())
        .all()
    )
    return jsonify({
        "ok": True,
        "user_id": 18,
        "devices": [
            {
                "id": row.id,
                "active": bool(row.is_active),
                "platform": row.plataforma,
                "device_name": row.device_name,
                "app_version": row.app_version,
                "token_prefix": (row.token or "")[:18],
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "updated_at": row.updated_at.isoformat() if row.updated_at else None,
                "last_seen_at": row.last_seen_at.isoformat() if row.last_seen_at else None,
            }
            for row in rows
        ],
    })

@app.route("/debug/force_reminders")
@login_required
def debug_force_reminders():
    if not is_admin():
        abort(403)
    try:
        enviar_notificaciones_pendientes()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

def enviar_notificaciones_pendientes():
    with app.app_context():
        ahora = now_cdmx_naive()
        inicio_hoy = ahora.replace(hour=0, minute=0, second=0, microsecond=0)

        cotizaciones = (
            Cotizacion.query
            .filter(Cotizacion.eliminada_en.is_(None), db.func.upper(Cotizacion.estatus) != "FINALIZADA")
            .all()
        )

        for cot in cotizaciones:
            if cot.last_whatsapp_at is not None and cot.last_whatsapp_at >= inicio_hoy:
                continue
            try:
                _send_daily_status_reminder(cot, ahora)
            except Exception as e:
                print(f"[Scheduler] ERROR recordatorio ({cot.folio}): {e}", file=sys.stderr)

scheduler: Optional[BackgroundScheduler] = None
try:
    if os.environ.get("WERKZEUG_RUN_MAIN") == "true" or not app.debug:
        scheduler = BackgroundScheduler(timezone=TZ_CDMX, daemon=True)
        scheduler.add_job(
            enviar_notificaciones_pendientes,
            "cron",
            hour=10,
            minute=0,
            id="daily_quotes_status_reminder",
            replace_existing=True
        )
        scheduler.start()
        print("[Scheduler] Iniciado (10:00 AM CDMX).")
except Exception as e:
    print(f"[Scheduler] No pudo iniciar: {e}", file=sys.stderr)

@app.route("/admin/bitacora")
@login_required
def admin_bitacora():
    if not is_admin_account():
        abort(403)

    page = int(request.args.get("page", 1) or 1)
    per_page = int(request.args.get("per", 100) or 100)
    per_page = max(20, min(per_page, 300))

    q = (request.args.get("q") or "").strip()
    usuario_f = (request.args.get("usuario") or "").strip()
    metodo_f = (request.args.get("metodo") or "").strip().upper()
    status_f = (request.args.get("status") or "").strip()

    query = ActivityLog.query

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            ActivityLog.usuario.ilike(like),
            ActivityLog.ruta.ilike(like),
            ActivityLog.accion.ilike(like),
            ActivityLog.endpoint.ilike(like),
        ))
    if usuario_f:
        query = query.filter(ActivityLog.usuario == usuario_f)
    if metodo_f:
        query = query.filter(ActivityLog.metodo == metodo_f)
    if status_f.isdigit():
        query = query.filter(ActivityLog.status_code == int(status_f))

    total = query.count()
    logs = (query.order_by(ActivityLog.fecha.desc())
                .offset((page - 1) * per_page)
                .limit(per_page)
                .all())

    # usuarios distintos para dropdown
    usuarios = [u[0] for u in db.session.query(ActivityLog.usuario).distinct().order_by(ActivityLog.usuario).all()]

    return render_template(
        "admin_bitacora.html",
        logs=logs,
        page=page,
        per_page=per_page,
        total=total,
        q=q,
        usuario_f=usuario_f,
        metodo_f=metodo_f,
        status_f=status_f,
        usuarios=usuarios,
    )

# ---------------------------------------------------------
@app.route("/admin/usuarios", methods=["GET", "POST"])
@login_required
def admin_usuarios():
    if not is_admin_account():
        abort(403)

    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        nombre_visible = (request.form.get("nombre_visible") or "").strip()
        correo = (request.form.get("correo") or "").strip()
        password = (request.form.get("password") or "").strip()
        rol = normalize_user_role(request.form.get("rol"))

        if not nombre:
            flash("El usuario es obligatorio.", "danger")
            return redirect(url_for("admin_usuarios"))
        if not nombre_visible:
            flash("El nombre es obligatorio.", "danger")
            return redirect(url_for("admin_usuarios"))
        if not correo:
            flash("El correo del usuario es obligatorio.", "danger")
            return redirect(url_for("admin_usuarios"))
        try:
            correos_usuario = _parse_email_list(correo)
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("admin_usuarios"))
        if len(correos_usuario) != 1:
            flash("Captura un solo correo para el usuario.", "danger")
            return redirect(url_for("admin_usuarios"))
        correo = correos_usuario[0]
        if not password:
            flash("La contrasena es obligatoria para crear un usuario.", "danger")
            return redirect(url_for("admin_usuarios"))

        exists = Usuario.query.filter(db.func.lower(Usuario.nombre) == nombre.lower()).first()
        if exists:
            flash("Ya existe un usuario con ese usuario.", "danger")
            return redirect(url_for("admin_usuarios"))

        nuevo = Usuario(nombre=nombre, nombre_visible=nombre_visible, correo=correo, rol=rol)
        nuevo.set_password(password)
        db.session.add(nuevo)
        db.session.commit()
        try:
            _send_user_created_email(nuevo, current_user, password)
            flash(f"Usuario '{nombre}' creado correctamente y notificado a sistemas.", "success")
        except Exception as exc:
            try:
                logger.exception("No se pudo enviar correo de alta de usuario %s", nuevo.id)
            except Exception:
                pass
            flash(f"Usuario '{nombre}' creado correctamente, pero no se pudo enviar el correo a sistemas: {exc}", "warning")
        return redirect(url_for("admin_usuarios"))

    q = (request.args.get("q") or "").strip()
    usuarios_query = admin_users_base_query()
    if q:
        usuarios_query = usuarios_query.filter(
            or_(
                Usuario.nombre.ilike(f"%{q}%"),
                Usuario.nombre_visible.ilike(f"%{q}%"),
            )
        )

    usuarios = usuarios_query.all()
    total_admins = Usuario.query.filter(db.func.upper(Usuario.rol) == "ADMIN").count()
    return render_template(
        "admin_usuarios.html",
        usuarios=usuarios,
        q=q,
        total=len(usuarios),
        total_admins=total_admins,
    )

@app.route("/admin/usuarios/<int:user_id>/editar", methods=["POST"])
@login_required
def admin_usuario_editar(user_id: int):
    if not is_admin_account():
        abort(403)

    usuario = Usuario.query.get_or_404(user_id)
    nombre = (request.form.get("nombre") or "").strip()
    nombre_visible = (request.form.get("nombre_visible") or "").strip()
    correo = (request.form.get("correo") or "").strip()
    password = (request.form.get("password") or "").strip()
    rol = normalize_user_role(request.form.get("rol"))
    previous_nombre = usuario.nombre or ""
    previous_nombre_visible = _usuario_nombre_representante(usuario)
    previous_correo = usuario.correo or ""
    previous_rol = usuario.rol or "USER"

    if not nombre:
        flash("El usuario es obligatorio.", "danger")
        return redirect(url_for("admin_usuarios"))
    if not nombre_visible:
        flash("El nombre es obligatorio.", "danger")
        return redirect(url_for("admin_usuarios"))
    if not correo:
        flash("El correo del usuario es obligatorio.", "danger")
        return redirect(url_for("admin_usuarios"))
    try:
        correos_usuario = _parse_email_list(correo)
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("admin_usuarios"))
    if len(correos_usuario) != 1:
        flash("Captura un solo correo para el usuario.", "danger")
        return redirect(url_for("admin_usuarios"))
    correo = correos_usuario[0]

    duplicado = Usuario.query.filter(
        db.func.lower(Usuario.nombre) == nombre.lower(),
        Usuario.id != usuario.id,
    ).first()
    if duplicado:
        flash("Ya existe otro usuario con ese nombre.", "danger")
        return redirect(url_for("admin_usuarios"))

    if usuario.id == current_user.id and rol != "ADMIN":
        admins_restantes = Usuario.query.filter(
            db.func.upper(Usuario.rol) == "ADMIN",
            Usuario.id != usuario.id,
        ).count()
        if admins_restantes == 0:
            flash("No puedes quitar el rol ADMIN al unico administrador del sistema.", "danger")
            return redirect(url_for("admin_usuarios"))

    usuario.nombre = nombre
    usuario.nombre_visible = nombre_visible
    usuario.correo = correo
    usuario.rol = rol
    if password:
        usuario.set_password(password)

    db.session.commit()
    try:
        _send_user_updated_email(usuario, current_user, previous_nombre, previous_nombre_visible, previous_rol, previous_correo, password)
        flash(f"Usuario '{nombre}' actualizado correctamente y notificado a sistemas.", "success")
    except Exception as exc:
        try:
            logger.exception("No se pudo enviar correo de cambio de usuario %s", usuario.id)
        except Exception:
            pass
        flash(f"Usuario '{nombre}' actualizado correctamente, pero no se pudo enviar el correo a sistemas: {exc}", "warning")
    return redirect(url_for("admin_usuarios"))

@app.route("/admin/usuarios/<int:user_id>/eliminar", methods=["POST"])
@login_required
def admin_usuario_eliminar(user_id: int):
    if not is_admin_account():
        abort(403)

    usuario = Usuario.query.get_or_404(user_id)

    if usuario.id == current_user.id:
        flash("No puedes eliminar tu propio usuario mientras tienes la sesion activa.", "danger")
        return redirect(url_for("admin_usuarios"))

    if (usuario.rol or "").upper() == "ADMIN":
        admins_restantes = Usuario.query.filter(
            db.func.upper(Usuario.rol) == "ADMIN",
            Usuario.id != usuario.id,
        ).count()
        if admins_restantes == 0:
            flash("No puedes eliminar al ultimo administrador del sistema.", "danger")
            return redirect(url_for("admin_usuarios"))

    nombre = usuario.nombre
    db.session.delete(usuario)
    db.session.commit()
    flash(f"Usuario '{nombre}' eliminado correctamente.", "success")
    return redirect(url_for("admin_usuarios"))


# ---------------------------------------------------------
# Ordenes de compra
# ---------------------------------------------------------
ORDEN_COMPRA_ESTATUS = (
    "BORRADOR",
    "ENVIADA",
    "PARCIALMENTE RECIBIDA",
    "RECIBIDA COMPLETA",
    "CANCELADA",
    "FACTURADA",
    "PAGADA",
)
ORDEN_COMPRA_UPLOAD_EXTS = {"pdf", "png", "jpg", "jpeg", "webp"}
SOLICITUD_RECURSO_ESTATUS = ("SOLICITADA", "AUTORIZADA", "RECHAZADA", "ENTREGADA", "CANCELADA")
SOLICITUD_RECURSO_EMAILS = ("sistemas@poliutech.com", "hjaramillo@poliutech.com")
REPORTE_DIARIO_TO_EMAIL = "hjaramillo@poliutech.com"
REPORTE_DIARIO_BCC_EMAIL = "sistemas@poliutech.com"
REPORTE_DIARIO_CUMPLIMIENTO = ("100%", "80-99%", "60-79%", "MENOR A 60%")
REPORTE_DIARIO_ACTIVIDAD_ESTATUS = ("TERMINADA", "EN PROCESO", "PENDIENTE")
REPORTE_DIARIO_SEMAFORO = ("SIN INCIDENCIAS", "RIESGOS IDENTIFICADOS", "REQUIERE INTERVENCION INMEDIATA")


def _parse_date_or_none(raw: str):
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, "%Y-%m-%d")
    except ValueError:
        return None


def _solicitud_recurso_next_folio() -> str:
    year = now_cdmx_naive().year
    prefix = f"SR-{year}-"
    latest = (
        SolicitudRecurso.query
        .filter(SolicitudRecurso.folio.like(f"{prefix}%"))
        .order_by(SolicitudRecurso.id.desc())
        .first()
    )
    if latest and latest.folio:
        try:
            seq = int(latest.folio.rsplit("-", 1)[-1]) + 1
        except Exception:
            seq = latest.id + 1
    else:
        seq = 1
    return f"{prefix}{seq:04d}"


def _reporte_diario_next_folio() -> str:
    year = now_cdmx_naive().year
    prefix = f"RD-{year}-"
    latest = (
        ReporteDiario.query
        .filter(ReporteDiario.folio.like(f"{prefix}%"))
        .order_by(ReporteDiario.id.desc())
        .first()
    )
    if latest and latest.folio:
        try:
            seq = int(latest.folio.rsplit("-", 1)[-1]) + 1
        except Exception:
            seq = latest.id + 1
    else:
        seq = 1
    return f"{prefix}{seq:04d}"


def _json_dumps(value) -> str:
    return json.dumps(value or [], ensure_ascii=False)


def _json_loads_list(value: str | None) -> list:
    if not value:
        return []
    try:
        data = json.loads(value)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _reporte_diario_payload(reporte: ReporteDiario) -> dict:
    return {
        "actividades": _json_loads_list(reporte.actividades_json),
        "puntos": _json_loads_list(reporte.puntos_importantes_json),
        "prioridades": _json_loads_list(reporte.prioridades_siguientes_json),
        "tiempos": _json_loads_list(reporte.tiempos_json),
        "riesgos": _json_loads_list(reporte.problemas_riesgos_json),
    }


def _clean_parallel_rows(*columns: list[str]) -> list[tuple[str, ...]]:
    total = max((len(col) for col in columns), default=0)
    rows = []
    for idx in range(total):
        row = tuple((col[idx] if idx < len(col) else "").strip() for col in columns)
        if any(row):
            rows.append(row)
    return rows


def _reporte_diario_from_form(f) -> ReporteDiario:
    fecha = _parse_date_or_none(f.get("fecha")) or now_cdmx_naive()
    colaborador = (f.get("colaborador") or responsable_actual() or "").strip()
    puesto = (f.get("puesto") or "").strip()
    cumplimiento = (f.get("cumplimiento") or "").strip().upper()
    if cumplimiento not in REPORTE_DIARIO_CUMPLIMIENTO:
        cumplimiento = ""
    semaforo = (f.get("semaforo") or "SIN INCIDENCIAS").strip().upper()
    if semaforo not in REPORTE_DIARIO_SEMAFORO:
        semaforo = "SIN INCIDENCIAS"

    actividades = []
    for idx, (actividad, estatus, avance) in enumerate(_clean_parallel_rows(
        f.getlist("actividad[]"),
        f.getlist("actividad_estatus[]"),
        f.getlist("actividad_avance[]"),
    ), start=1):
        estatus = estatus.upper()
        actividades.append({
            "no": idx,
            "actividad": actividad,
            "estatus": estatus if estatus in REPORTE_DIARIO_ACTIVIDAD_ESTATUS else "PENDIENTE",
            "avance": avance,
        })

    puntos = []
    for idx, (prioridad, resultado, impacto) in enumerate(_clean_parallel_rows(
        f.getlist("punto_prioridad[]"),
        f.getlist("punto_resultado[]"),
        f.getlist("punto_impacto[]"),
    ), start=1):
        puntos.append({"no": idx, "prioridad": prioridad, "resultado": resultado, "impacto": impacto})

    prioridades = []
    for idx, (actividad, objetivo) in enumerate(_clean_parallel_rows(
        f.getlist("prioridad_actividad[]"),
        f.getlist("prioridad_objetivo[]"),
    ), start=1):
        prioridades.append({"no": idx, "actividad": actividad, "objetivo": objetivo})

    tiempos = []
    for tipo, horas in _clean_parallel_rows(f.getlist("tiempo_tipo[]"), f.getlist("tiempo_horas[]")):
        tiempos.append({"tipo": tipo, "horas": horas})

    riesgos = []
    for situacion, impacto, apoyo in _clean_parallel_rows(
        f.getlist("riesgo_situacion[]"),
        f.getlist("riesgo_impacto[]"),
        f.getlist("riesgo_apoyo[]"),
    ):
        riesgos.append({"situacion": situacion, "impacto": impacto, "apoyo": apoyo})

    return ReporteDiario(
        folio=_reporte_diario_next_folio(),
        colaborador=colaborador,
        puesto=puesto or None,
        fecha=fecha,
        hora_envio=now_cdmx_naive(),
        estatus="ENVIADO",
        cumplimiento=cumplimiento or None,
        semaforo=semaforo,
        actividades_json=_json_dumps(actividades),
        puntos_importantes_json=_json_dumps(puntos),
        prioridades_siguientes_json=_json_dumps(prioridades),
        tiempos_json=_json_dumps(tiempos),
        problemas_riesgos_json=_json_dumps(riesgos),
        apoyo_direccion=(f.get("apoyo_direccion") or "").strip() or None,
        observaciones=(f.get("observaciones") or "").strip() or None,
        usuario_id=getattr(current_user, "id", None),
    )


def _reporte_diario_mail_html(reporte: ReporteDiario, detail_url: str) -> str:
    payload = _reporte_diario_payload(reporte)

    def rows(items, cols):
        body = []
        for item in items:
            body.append("<tr>" + "".join(
                f"<td style='padding:9px;border-bottom:1px solid #e5e7eb;vertical-align:top;'>{escape(str(item.get(col, '') or ''))}</td>"
                for col in cols
            ) + "</tr>")
        return "".join(body) or f"<tr><td colspan='{len(cols)}' style='padding:9px;color:#64748b;'>Sin registros.</td></tr>"

    return f"""
    <div style="font-family:Arial,sans-serif;color:#0f172a;max-width:820px;margin:0 auto;">
      <h2 style="margin:0 0 8px;color:#0C3C78;">Reporte diario de actividades</h2>
      <p style="margin:0 0 18px;color:#475569;"><b>{escape(reporte.folio or str(reporte.id))}</b> enviado por {escape(reporte.colaborador or '')}.</p>
      <table style="width:100%;border-collapse:collapse;margin-bottom:16px;">
        <tr><td style="padding:7px;color:#64748b;font-weight:700;">Colaborador</td><td style="padding:7px;">{escape(reporte.colaborador or '-')}</td></tr>
        <tr><td style="padding:7px;color:#64748b;font-weight:700;">Puesto</td><td style="padding:7px;">{escape(reporte.puesto or '-')}</td></tr>
        <tr><td style="padding:7px;color:#64748b;font-weight:700;">Fecha</td><td style="padding:7px;">{reporte.fecha.strftime('%d/%m/%Y') if reporte.fecha else ''}</td></tr>
        <tr><td style="padding:7px;color:#64748b;font-weight:700;">Cumplimiento</td><td style="padding:7px;">{escape(reporte.cumplimiento or '-')}</td></tr>
        <tr><td style="padding:7px;color:#64748b;font-weight:700;">Semaforo</td><td style="padding:7px;font-weight:700;">{escape(reporte.semaforo or '-')}</td></tr>
      </table>
      <h3 style="font-size:16px;color:#0C3C78;">Actividades realizadas</h3>
      <table style="width:100%;border-collapse:collapse;border:1px solid #e5e7eb;margin-bottom:16px;">{rows(payload['actividades'], ['no', 'actividad', 'estatus', 'avance'])}</table>
      <h3 style="font-size:16px;color:#0C3C78;">Puntos importantes</h3>
      <table style="width:100%;border-collapse:collapse;border:1px solid #e5e7eb;margin-bottom:16px;">{rows(payload['puntos'], ['no', 'prioridad', 'resultado', 'impacto'])}</table>
      <h3 style="font-size:16px;color:#0C3C78;">Prioridades siguiente dia</h3>
      <table style="width:100%;border-collapse:collapse;border:1px solid #e5e7eb;margin-bottom:16px;">{rows(payload['prioridades'], ['no', 'actividad', 'objetivo'])}</table>
      <p><b>Apoyo requerido:</b> {escape(reporte.apoyo_direccion or '-')}</p>
      <p><b>Observaciones:</b> {escape(reporte.observaciones or '-')}</p>
      <p style="margin:18px 0;"><a href="{detail_url}" style="background:#0C3C78;color:#fff;text-decoration:none;padding:10px 14px;border-radius:6px;display:inline-block;">Ver reporte</a></p>
    </div>
    """


def _send_reporte_diario_email(reporte: ReporteDiario) -> None:
    recipients = _parse_email_list(REPORTE_DIARIO_TO_EMAIL)
    bcc = _parse_email_list(REPORTE_DIARIO_BCC_EMAIL)
    detail_url = url_for("reporte_diario_detalle", reporte_id=reporte.id, _external=True)
    msg = EmailMessage()
    msg["Subject"] = f"Reporte diario {reporte.folio or reporte.id} - {reporte.colaborador}"
    msg["From"] = f"SISTEMA MAR <{SMTP_FROM or SMTP_USERNAME}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content(
        f"Reporte diario {reporte.folio or reporte.id}\n"
        f"Colaborador: {reporte.colaborador}\n"
        f"Fecha: {reporte.fecha.strftime('%d/%m/%Y') if reporte.fecha else ''}\n"
        f"Cumplimiento: {reporte.cumplimiento or '-'}\n"
        f"Semaforo: {reporte.semaforo or '-'}\n"
        f"Ver: {detail_url}\n"
    )
    msg.add_alternative(_reporte_diario_mail_html(reporte, detail_url), subtype="html")
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg, to_addrs=[*recipients, *bcc])


def _send_reporte_diario_push_hansel(reporte: ReporteDiario) -> dict[str, int]:
    tokens = _mobile_push_tokens_for_users(_mobile_push_user_ids_for_hansel_only())
    if not tokens:
        logger.warning("Push reporte diario %s: Hjaramillo no tiene token movil activo.", reporte.folio or reporte.id)
    return _send_push_notification(
        tokens,
        title="Nuevo reporte diario",
        body=f"{reporte.colaborador} - {reporte.semaforo}",
        data={
            "type": "reporte_diario",
            "reporte_id": str(reporte.id),
            "folio": reporte.folio or "",
            "url": url_for("reporte_diario_detalle", reporte_id=reporte.id, _external=True),
        },
    )


def _notify_reporte_diario_created(reporte: ReporteDiario) -> None:
    try:
        _send_reporte_diario_email(reporte)
    except Exception as exc:
        logger.warning("Correo de reporte diario %s fallo: %s", reporte.folio or reporte.id, exc)
    try:
        _send_reporte_diario_push_hansel(reporte)
    except Exception as exc:
        logger.warning("Push de reporte diario %s fallo: %s", reporte.folio or reporte.id, exc)


def _solicitud_recurso_recalcular(solicitud: SolicitudRecurso) -> None:
    total = 0.0
    for partida in solicitud.partidas:
        partida.cantidad = fmt(partida.cantidad or 0)
        partida.importe = fmt(partida.importe or 0)
        row_total = getattr(partida, "total", None)
        if row_total is None or float(row_total or 0) <= 0:
            row_total = partida.cantidad * partida.importe
        partida.total = fmt(row_total)
        total += partida.total
    solicitud.total = fmt(total)
    solicitud.actualizado_en = now_cdmx_naive()


def _mobile_push_user_ids_for_hansel_only() -> list[int]:
    hansel_aliases = {"hansel", "hansel alejandro", "hansel angel", "hansel ángel"}
    hansel_emails = {"hjaramillo@poliutech.com"}
    fixed_hansel_ids = {18}
    user_ids: set[int] = set()
    for user in Usuario.query.all():
        user_name = (getattr(user, "nombre", "") or "").strip().lower()
        visible_name = (_mobile_user_responsable(user) or "").strip().lower()
        raw_visible_name = (getattr(user, "nombre_visible", "") or "").strip().lower()
        user_email = (getattr(user, "correo", "") or "").strip().lower()
        identity_parts = {user_name, visible_name, raw_visible_name}
        if (
            user.id in fixed_hansel_ids
            or user_email in hansel_emails
            or any(part in hansel_aliases or part.startswith("hansel ") for part in identity_parts if part)
        ):
            if user.id:
                user_ids.add(user.id)
    if not user_ids:
        logger.warning("Push solicitud de recursos: no se encontro usuario Hansel.")
    return list(user_ids)


def _solicitud_recurso_mail_html(solicitud: SolicitudRecurso, detail_url: str) -> str:
    rows = []
    for idx, partida in enumerate(solicitud.partidas, start=1):
        row_total = getattr(partida, "total", None)
        if row_total is None:
            row_total = float(partida.cantidad or 0) * float(partida.importe or 0)
        rows.append(
            "<tr>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;text-align:right;'>{idx}</td>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;text-align:right;'>{float(partida.cantidad or 0):,.2f}</td>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;'>{escape(partida.concepto or '')}</td>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;text-align:right;'>${float(partida.importe or 0):,.2f}</td>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;text-align:right;font-weight:700;'>${float(row_total or 0):,.2f}</td>"
            "</tr>"
        )
    partidas_html = "".join(rows) or "<tr><td colspan='5' style='padding:10px;'>Sin partidas.</td></tr>"
    return f"""
    <div style="font-family:Arial,sans-serif;color:#0f172a;max-width:760px;margin:0 auto;">
      <h2 style="margin:0 0 10px;color:#0C3C78;">Nueva solicitud de recursos</h2>
      <p style="margin:0 0 18px;color:#475569;">Se registro la solicitud <b>{escape(solicitud.folio or str(solicitud.id))}</b>.</p>
      <table style="width:100%;border-collapse:collapse;margin-bottom:18px;">
        <tr><td style="padding:8px;color:#64748b;font-weight:700;">Solicitante</td><td style="padding:8px;">{escape(solicitud.solicitante or '-')}</td></tr>
        <tr><td style="padding:8px;color:#64748b;font-weight:700;">Proyecto / obra</td><td style="padding:8px;">{escape(solicitud.proyecto or '-')}</td></tr>
        <tr><td style="padding:8px;color:#64748b;font-weight:700;">Fecha</td><td style="padding:8px;">{solicitud.fecha.strftime('%d/%m/%Y %H:%M') if solicitud.fecha else ''}</td></tr>
        <tr><td style="padding:8px;color:#64748b;font-weight:700;">Total</td><td style="padding:8px;font-weight:700;">${float(solicitud.total or 0):,.2f}</td></tr>
      </table>
      <table style="width:100%;border-collapse:collapse;border:1px solid #e5e7eb;">
        <thead>
          <tr style="background:#f8fafc;">
            <th style="padding:10px;text-align:right;">Partida</th>
            <th style="padding:10px;text-align:right;">Cantidad</th>
            <th style="padding:10px;text-align:left;">Descripcion</th>
            <th style="padding:10px;text-align:right;">Importe</th>
            <th style="padding:10px;text-align:right;">Total</th>
          </tr>
        </thead>
        <tbody>{partidas_html}</tbody>
      </table>
      <p style="margin:18px 0;"> <a href="{detail_url}" style="background:#0C3C78;color:#fff;text-decoration:none;padding:10px 14px;border-radius:6px;display:inline-block;">Ver solicitud</a></p>
      <p style="margin:0;color:#64748b;">{escape(solicitud.notas or '')}</p>
    </div>
    """


def _send_solicitud_recurso_email(solicitud: SolicitudRecurso) -> None:
    recipients = list(SOLICITUD_RECURSO_EMAILS)
    detail_url = url_for("solicitud_recurso_detalle", solicitud_id=solicitud.id, _external=True)
    msg = EmailMessage()
    msg["Subject"] = f"Nueva solicitud de recursos {solicitud.folio or solicitud.id}"
    msg["From"] = f"SISTEMA MAR <{SMTP_FROM or SMTP_USERNAME}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content(
        f"Nueva solicitud de recursos {solicitud.folio or solicitud.id}\n"
        f"Solicitante: {solicitud.solicitante or '-'}\n"
        f"Proyecto / obra: {solicitud.proyecto or '-'}\n"
        f"Total: ${float(solicitud.total or 0):,.2f}\n"
        f"Ver: {detail_url}\n"
    )
    msg.add_alternative(_solicitud_recurso_mail_html(solicitud, detail_url), subtype="html")
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg, to_addrs=recipients)


def _send_solicitud_recurso_push_hansel(solicitud: SolicitudRecurso) -> dict[str, int]:
    user_ids = _mobile_push_user_ids_for_hansel_only()
    tokens = _mobile_push_tokens_for_users(user_ids)
    if not tokens:
        logger.warning(
            "Push solicitud de recursos %s: Hansel no tiene token movil activo.",
            solicitud.folio or solicitud.id,
        )
    return _send_push_notification(
        tokens,
        title="Nueva solicitud de recursos",
        body=f"{solicitud.folio or solicitud.id} - ${float(solicitud.total or 0):,.2f}",
        data={
            "type": "solicitud_recurso",
            "solicitud_id": str(solicitud.id),
            "folio": solicitud.folio or "",
            "url": url_for("solicitud_recurso_detalle", solicitud_id=solicitud.id, _external=True),
        },
    )


def _notify_solicitud_recurso_created(solicitud: SolicitudRecurso) -> None:
    try:
        _send_solicitud_recurso_email(solicitud)
    except Exception as exc:
        logger.warning("Correo de solicitud de recursos %s fallo: %s", solicitud.folio or solicitud.id, exc)

    try:
        _send_solicitud_recurso_push_hansel(solicitud)
    except Exception as exc:
        logger.warning("Push de solicitud de recursos %s fallo: %s", solicitud.folio or solicitud.id, exc)


def _solicitud_recurso_solicitante_user(solicitud: SolicitudRecurso) -> Usuario | None:
    user = getattr(solicitud, "usuario", None)
    if user:
        return user
    user_id = getattr(solicitud, "usuario_id", None)
    if not user_id:
        return None
    return Usuario.query.get(user_id)


def _solicitud_recurso_resultado_mail_html(solicitud: SolicitudRecurso, detail_url: str) -> str:
    estatus = (solicitud.estatus or "").strip().upper()
    aprobada = estatus == "AUTORIZADA"
    titulo = "Solicitud de recursos autorizada" if aprobada else "Solicitud de recursos rechazada"
    color = "#15803d" if aprobada else "#b91c1c"
    mensaje = (
        "Tu solicitud fue autorizada y quedo registrada para seguimiento."
        if aprobada
        else "Tu solicitud fue rechazada. Revisa el detalle para dar seguimiento."
    )
    gasto_html = ""
    if aprobada and getattr(solicitud, "gasto_generado", None):
        gasto_html = (
            f"<tr><td style='padding:8px;color:#64748b;font-weight:700;'>Gasto generado</td>"
            f"<td style='padding:8px;'>{escape(solicitud.gasto_generado.folio or str(solicitud.gasto_generado.id))}</td></tr>"
        )
    return f"""
    <div style="font-family:Arial,sans-serif;color:#0f172a;max-width:680px;margin:0 auto;">
      <h2 style="margin:0 0 10px;color:{color};">{titulo}</h2>
      <p style="margin:0 0 18px;color:#475569;">{mensaje}</p>
      <table style="width:100%;border-collapse:collapse;margin-bottom:18px;">
        <tr><td style="padding:8px;color:#64748b;font-weight:700;">Folio</td><td style="padding:8px;">{escape(solicitud.folio or str(solicitud.id))}</td></tr>
        <tr><td style="padding:8px;color:#64748b;font-weight:700;">Estatus</td><td style="padding:8px;font-weight:700;color:{color};">{escape(estatus)}</td></tr>
        <tr><td style="padding:8px;color:#64748b;font-weight:700;">Proyecto / obra</td><td style="padding:8px;">{escape(solicitud.proyecto or '-')}</td></tr>
        <tr><td style="padding:8px;color:#64748b;font-weight:700;">Total</td><td style="padding:8px;font-weight:700;">${float(solicitud.total or 0):,.2f}</td></tr>
        {gasto_html}
      </table>
      <p style="margin:18px 0;"> <a href="{detail_url}" style="background:#0C3C78;color:#fff;text-decoration:none;padding:10px 14px;border-radius:6px;display:inline-block;">Ver solicitud</a></p>
    </div>
    """


def _send_solicitud_recurso_resultado_email(solicitud: SolicitudRecurso) -> None:
    solicitante = _solicitud_recurso_solicitante_user(solicitud)
    recipients = _parse_email_list(getattr(solicitante, "correo", None))
    if not recipients:
        raise ValueError("La solicitud no tiene correo del usuario solicitante.")

    estatus = (solicitud.estatus or "").strip().upper()
    accion = "autorizada" if estatus == "AUTORIZADA" else "rechazada"
    detail_url = url_for("solicitud_recurso_detalle", solicitud_id=solicitud.id, _external=True)
    msg = EmailMessage()
    msg["Subject"] = f"Solicitud de recursos {accion} {solicitud.folio or solicitud.id}"
    msg["From"] = f"SISTEMA MAR <{SMTP_FROM or SMTP_USERNAME}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content(
        f"Tu solicitud de recursos {solicitud.folio or solicitud.id} fue {accion}.\n"
        f"Proyecto / obra: {solicitud.proyecto or '-'}\n"
        f"Total: ${float(solicitud.total or 0):,.2f}\n"
        f"Ver: {detail_url}\n"
    )
    msg.add_alternative(_solicitud_recurso_resultado_mail_html(solicitud, detail_url), subtype="html")
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg, to_addrs=recipients)


def _send_solicitud_recurso_resultado_push(solicitud: SolicitudRecurso) -> dict[str, int]:
    solicitante = _solicitud_recurso_solicitante_user(solicitud)
    user_ids = [solicitante.id] if solicitante and solicitante.id else []
    tokens = _mobile_push_tokens_for_users(user_ids)
    if not tokens:
        logger.warning(
            "Push resultado solicitud de recursos %s: solicitante %s no tiene token movil activo.",
            solicitud.folio or solicitud.id,
            user_ids,
        )
    estatus = (solicitud.estatus or "").strip().upper()
    title = "Solicitud de recursos autorizada" if estatus == "AUTORIZADA" else "Solicitud de recursos rechazada"
    body = f"{solicitud.folio or solicitud.id} - ${float(solicitud.total or 0):,.2f}"
    return _send_push_notification(
        tokens,
        title=title,
        body=body,
        data={
            "type": "solicitud_recurso_resultado",
            "solicitud_id": str(solicitud.id),
            "folio": solicitud.folio or "",
            "estatus": estatus,
            "url": url_for("solicitud_recurso_detalle", solicitud_id=solicitud.id, _external=True),
            "target_user_id": str(user_ids[0]) if len(user_ids) == 1 else "",
        },
    )


def _notify_solicitud_recurso_resultado(solicitud: SolicitudRecurso) -> None:
    try:
        _send_solicitud_recurso_resultado_email(solicitud)
    except Exception as exc:
        logger.warning("Correo resultado solicitud de recursos %s fallo: %s", solicitud.folio or solicitud.id, exc)

    try:
        _send_solicitud_recurso_resultado_push(solicitud)
    except Exception as exc:
        logger.warning("Push resultado solicitud de recursos %s fallo: %s", solicitud.folio or solicitud.id, exc)


def _send_solicitud_recurso_autorizada_finanzas_email(solicitud: SolicitudRecurso) -> None:
    recipients = _finanzas_auth_notify_recipients()
    if not recipients:
        raise ValueError("No hay correos configurados para autorizaciones de finanzas.")

    detail_url = url_for("solicitud_recurso_detalle", solicitud_id=solicitud.id, _external=True)
    gasto_line = ""
    if getattr(solicitud, "gasto_generado", None):
        gasto_line = f"Gasto generado: {solicitud.gasto_generado.folio or solicitud.gasto_generado.id}\n"
    msg = EmailMessage()
    msg["Subject"] = f"Solicitud de recursos autorizada {solicitud.folio or solicitud.id}"
    msg["From"] = f"SISTEMA MAR <{SMTP_FROM or SMTP_USERNAME}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content(
        f"Solicitud de recursos autorizada {solicitud.folio or solicitud.id}\n"
        f"Solicitante: {solicitud.solicitante or '-'}\n"
        f"Proyecto / obra: {solicitud.proyecto or '-'}\n"
        f"Total: ${float(solicitud.total or 0):,.2f}\n"
        f"{gasto_line}"
        f"Ver: {detail_url}\n"
    )
    msg.add_alternative(_solicitud_recurso_resultado_mail_html(solicitud, detail_url), subtype="html")
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg, to_addrs=recipients)


def _send_solicitud_recurso_autorizada_finanzas_push(solicitud: SolicitudRecurso) -> dict[str, int]:
    user_ids = _mobile_push_user_ids_for_finanzas_auth_notify()
    tokens = _mobile_push_tokens_for_users(user_ids)
    if not tokens:
        logger.warning(
            "Push solicitud de recursos autorizada %s: Mescalera/Miguel sin token movil activo.",
            solicitud.folio or solicitud.id,
        )
    return _send_push_notification(
        tokens,
        title="Solicitud de recursos autorizada",
        body=f"{solicitud.folio or solicitud.id} - ${float(solicitud.total or 0):,.2f}",
        data={
            "type": "solicitud_recurso_autorizada_finanzas",
            "solicitud_id": str(solicitud.id),
            "folio": solicitud.folio or "",
            "estatus": solicitud.estatus or "",
            "url": url_for("solicitud_recurso_detalle", solicitud_id=solicitud.id, _external=True),
            "target_user_ids": ",".join(str(user_id) for user_id in user_ids),
        },
    )


def _notify_solicitud_recurso_autorizada_finanzas(solicitud: SolicitudRecurso) -> None:
    try:
        _send_solicitud_recurso_autorizada_finanzas_email(solicitud)
    except Exception as exc:
        logger.warning("Correo finanzas solicitud de recursos %s fallo: %s", solicitud.folio or solicitud.id, exc)

    try:
        _send_solicitud_recurso_autorizada_finanzas_push(solicitud)
    except Exception as exc:
        logger.warning("Push finanzas solicitud de recursos %s fallo: %s", solicitud.folio or solicitud.id, exc)


FINANZAS_CATEGORIA_CREDITO = "CREDITO_RECIBIDO"
FINANZAS_ESTATUS = ("PENDIENTE", "PARCIAL", "PAGADO", "VENCIDO", "CANCELADO")
GASTOS_ESTATUS = ("PENDIENTE", "EN REVISION", "APROBADO", "RECHAZADO", "REEMBOLSADO")
GASTOS_TIPOS = ("GASTO", "VIATICO", "RECURSO")
GASTOS_AGRUPACIONES = ("PROYECTO", "EVENTO")
GASTOS_UPLOAD_EXTS = {"pdf", "png", "jpg", "jpeg", "webp"}


def _finanzas_next_folio() -> str:
    year = now_cdmx_naive().year
    prefix = f"CRED-{year}-"
    latest = (
        MovimientoFinanciero.query
        .filter(MovimientoFinanciero.folio.like(f"{prefix}%"))
        .order_by(MovimientoFinanciero.id.desc())
        .first()
    )
    if latest and latest.folio:
        try:
            seq = int(latest.folio.rsplit("-", 1)[-1]) + 1
        except Exception:
            seq = latest.id + 1
    else:
        seq = 1
    return f"{prefix}{seq:04d}"


def _finanzas_estatus_real(mov: MovimientoFinanciero) -> str:
    estatus = (mov.estatus or "PENDIENTE").upper()
    if estatus in {"PAGADO", "CANCELADO"}:
        return estatus
    if float(mov.saldo or 0) <= 0:
        return "PAGADO"
    if mov.fecha_vencimiento and mov.fecha_vencimiento.date() < now_cdmx_naive().date():
        return "VENCIDO"
    if float(mov.saldo or 0) < float(mov.monto or 0):
        return "PARCIAL"
    return estatus


def _finanzas_badge_class(estatus: str) -> str:
    return {
        "PAGADO": "success",
        "PARCIAL": "info",
        "VENCIDO": "danger",
        "CANCELADO": "secondary",
    }.get((estatus or "").upper(), "warning")


def _finanzas_category_label(categoria: str) -> str:
    return "Credito recibido" if (categoria or "").upper() == FINANZAS_CATEGORIA_CREDITO else (categoria or "")


def _finanzas_fecha_input(fecha) -> str:
    return fecha.strftime("%Y-%m-%d") if fecha else ""


def _finanzas_dias_restantes(mov: MovimientoFinanciero):
    if not mov.fecha_vencimiento:
        return None
    return (mov.fecha_vencimiento.date() - now_cdmx_naive().date()).days


def _finanzas_pagado(mov: MovimientoFinanciero) -> float:
    return max(0.0, float(mov.monto or 0) - float(mov.saldo or 0))


def _finanzas_porcentaje_pagado(mov: MovimientoFinanciero) -> float:
    monto = float(mov.monto or 0)
    if monto <= 0:
        return 0.0
    return min(100.0, max(0.0, (_finanzas_pagado(mov) / monto) * 100.0))


def _finanzas_porcentaje_tiempo(mov: MovimientoFinanciero) -> float:
    if not mov.fecha or not mov.fecha_vencimiento:
        return 0.0
    inicio = mov.fecha.date()
    fin = mov.fecha_vencimiento.date()
    total = max(1, (fin - inicio).days)
    transcurridos = (now_cdmx_naive().date() - inicio).days
    return min(100.0, max(0.0, (transcurridos / total) * 100.0))


def _finanzas_tiempo_estado(mov: MovimientoFinanciero) -> dict:
    if _finanzas_estatus_real(mov) == "PAGADO":
        return {"texto": "Liquidado", "clase": "primary", "detalle": "Credito pagado"}
    dias = _finanzas_dias_restantes(mov)
    if dias is None:
        return {"texto": "Sin vencimiento", "clase": "secondary", "detalle": ""}
    if dias < 0:
        return {"texto": "Vencido", "clase": "danger", "detalle": f"Hace {abs(dias)} dias"}
    if dias <= 7:
        return {"texto": "Urgente", "clase": "danger", "detalle": f"{dias} dias restantes"}
    if dias <= 30:
        return {"texto": "Por vencer", "clase": "warning", "detalle": f"{dias} dias restantes"}
    return {"texto": "A tiempo", "clase": "success", "detalle": f"{dias} dias restantes"}


def _gastos_next_folio() -> str:
    year = now_cdmx_naive().year
    prefix = f"GAS-{year}-"
    latest = (
        ComprobacionGasto.query
        .filter(ComprobacionGasto.folio.like(f"{prefix}%"))
        .order_by(ComprobacionGasto.id.desc())
        .first()
    )
    if latest and latest.folio:
        try:
            seq = int(latest.folio.rsplit("-", 1)[-1]) + 1
        except Exception:
            seq = latest.id + 1
    else:
        seq = 1
    return f"{prefix}{seq:04d}"


def _solicitud_recurso_gasto_referencia(solicitud: SolicitudRecurso) -> str:
    return f"SR:{solicitud.folio or solicitud.id}"


def _solicitud_recurso_concepto_gasto(solicitud: SolicitudRecurso) -> str:
    conceptos = []
    for idx, partida in enumerate(solicitud.partidas, start=1):
        cantidad = float(partida.cantidad or 0)
        conceptos.append(f"{idx}. {cantidad:,.2f} x {partida.concepto}")
    return "Solicitud de recursos " + (solicitud.folio or f"#{solicitud.id}") + (": " + "; ".join(conceptos) if conceptos else "")


def _solicitud_recurso_registrar_gasto(solicitud: SolicitudRecurso) -> ComprobacionGasto:
    _solicitud_recurso_recalcular(solicitud)
    referencia = _solicitud_recurso_gasto_referencia(solicitud)
    gasto = None
    if getattr(solicitud, "gasto_generado_id", None):
        gasto = ComprobacionGasto.query.get(solicitud.gasto_generado_id)
    if gasto is None:
        gasto = ComprobacionGasto.query.filter_by(referencia=referencia).first()

    now = now_cdmx_naive()
    if gasto is None:
        gasto = ComprobacionGasto(
            folio=_gastos_next_folio(),
            fecha_registro=now,
            creado_en=now,
        )
        db.session.add(gasto)

    gasto.tipo_agrupacion = "PROYECTO"
    gasto.proyecto = (solicitud.proyecto or "").strip() or None
    gasto.evento = None
    gasto.tipo_gasto = "RECURSO"
    gasto.estatus = "APROBADO"
    gasto.proveedor = "Solicitud de dinero aprobada"
    gasto.concepto = _solicitud_recurso_concepto_gasto(solicitud)[:260]
    gasto.referencia = referencia
    gasto.solicitud_recurso_id = solicitud.id
    gasto.fecha_comprobante = solicitud.fecha or now
    gasto.subtotal = fmt(solicitud.total or 0)
    gasto.iva = 0.0
    gasto.total = fmt(solicitud.total or 0)
    gasto.moneda = "MXN"
    gasto.metodo_pago = None
    gasto.notas = (f"Generado automaticamente al autorizar {solicitud.folio or solicitud.id}. " + (solicitud.notas or "")).strip()
    gasto.ai_confianza = 0
    gasto.ai_resultado = None
    gasto.responsable = (solicitud.solicitante or "").strip() or responsable_actual() or None
    gasto.usuario_id = solicitud.usuario_id or getattr(current_user, "id", None)
    gasto.actualizado_en = now
    db.session.flush()

    solicitud.gasto_generado_id = gasto.id
    solicitud.gasto_generado_en = solicitud.gasto_generado_en or now
    return gasto


def _gastos_badge_class(estatus: str) -> str:
    return {
        "APROBADO": "success",
        "REEMBOLSADO": "primary",
        "RECHAZADO": "danger",
        "EN REVISION": "info",
    }.get((estatus or "").upper(), "warning")


def _gastos_status_row_class(estatus: str) -> str:
    return {
        "APROBADO": "gasto-row-aprobado",
        "RECHAZADO": "gasto-row-rechazado",
        "EN REVISION": "gasto-row-revision",
        "PENDIENTE": "gasto-row-revision",
        "REEMBOLSADO": "gasto-row-reembolsado",
    }.get((estatus or "").upper(), "")


def _gastos_es_recurso(gasto: "ComprobacionGasto") -> bool:
    tipo = (getattr(gasto, "tipo_gasto", "") or "").strip().upper()
    return tipo == "RECURSO"


def _gastos_monto_saldo(gasto: "ComprobacionGasto") -> float:
    if (gasto.estatus or "").upper() == "RECHAZADO":
        return 0.0
    total = float(gasto.total or 0)
    return total if _gastos_es_recurso(gasto) else -total


def _estado_cuenta_user_label(user: Usuario | None, fallback: str = "") -> str:
    if user:
        label = _usuario_nombre_representante(user)
        if label:
            return label
        if getattr(user, "correo", None):
            return user.correo
        return f"Usuario {user.id}"
    return (fallback or "").strip() or "Sin usuario"


def _estado_cuenta_user_key(user: Usuario | None, fallback: str = "") -> str:
    if user and user.id:
        return f"u-{user.id}"
    value = re.sub(r"[^a-z0-9]+", "-", (fallback or "sin-usuario").strip().lower()).strip("-")
    return f"r-{value or 'sin-usuario'}"


def _estado_cuenta_user_from_responsable(nombre: str) -> Usuario | None:
    nombre_l = (nombre or "").strip().lower()
    if not nombre_l:
        return None
    return Usuario.query.filter(
        or_(
            db.func.lower(Usuario.nombre) == nombre_l,
            db.func.lower(db.func.coalesce(Usuario.nombre_visible, "")) == nombre_l,
        )
    ).first()


def _estado_cuenta_user_for_gasto(gasto: "ComprobacionGasto") -> tuple[Usuario | None, str]:
    solicitud = getattr(gasto, "solicitud_recurso", None)
    if solicitud and getattr(solicitud, "usuario", None):
        return solicitud.usuario, solicitud.solicitante or gasto.responsable or ""
    if getattr(gasto, "usuario", None):
        return gasto.usuario, gasto.responsable or ""
    user = _estado_cuenta_user_from_responsable(gasto.responsable or "")
    return user, gasto.responsable or ""


def _estado_cuenta_new_bucket(user: Usuario | None, fallback: str = "") -> dict:
    label = _estado_cuenta_user_label(user, fallback)
    return {
        "key": _estado_cuenta_user_key(user, fallback or label),
        "usuario": user,
        "nombre": label,
        "correo": getattr(user, "correo", "") if user else "",
        "enviado": 0.0,
        "comprobado": 0.0,
        "saldo": 0.0,
        "recursos_count": 0,
        "comprobaciones_count": 0,
        "ultimo": None,
        "movimientos": [],
    }


def _estado_cuenta_recursos_data(user_key: str | None = None) -> list[dict]:
    buckets: dict[str, dict] = {}

    solicitudes = (
        SolicitudRecurso.query
        .filter(SolicitudRecurso.estatus == "AUTORIZADA")
        .order_by(SolicitudRecurso.fecha.asc(), SolicitudRecurso.id.asc())
        .all()
    )
    for solicitud in solicitudes:
        user = getattr(solicitud, "usuario", None)
        fallback = solicitud.solicitante or ""
        key = _estado_cuenta_user_key(user, fallback)
        bucket = buckets.setdefault(key, _estado_cuenta_new_bucket(user, fallback))
        monto = float(solicitud.total or 0)
        fecha = solicitud.gasto_generado_en or solicitud.actualizado_en or solicitud.fecha
        bucket["enviado"] += monto
        bucket["recursos_count"] += 1
        bucket["ultimo"] = max([d for d in [bucket["ultimo"], fecha] if d], default=None)
        bucket["movimientos"].append({
            "fecha": fecha,
            "tipo": "RECURSO AUTORIZADO",
            "folio": solicitud.folio or f"#{solicitud.id}",
            "concepto": solicitud.proyecto or "Solicitud de recursos",
            "referencia": solicitud.gasto_generado.folio if getattr(solicitud, "gasto_generado", None) else "",
            "estatus": solicitud.estatus,
            "monto_enviado": monto,
            "monto_comprobado": 0.0,
            "saldo_delta": monto,
            "url": url_for("solicitud_recurso_detalle", solicitud_id=solicitud.id),
        })

    comprobaciones = (
        ComprobacionGasto.query
        .filter(
            ComprobacionGasto.estatus.in_(("APROBADO", "REEMBOLSADO")),
            ComprobacionGasto.tipo_gasto != "RECURSO",
        )
        .order_by(ComprobacionGasto.fecha_comprobante.asc(), ComprobacionGasto.id.asc())
        .all()
    )
    for gasto in comprobaciones:
        user, fallback = _estado_cuenta_user_for_gasto(gasto)
        key = _estado_cuenta_user_key(user, fallback)
        bucket = buckets.setdefault(key, _estado_cuenta_new_bucket(user, fallback))
        monto = float(gasto.total or 0)
        fecha = gasto.fecha_comprobante or gasto.actualizado_en or gasto.fecha_registro
        bucket["comprobado"] += monto
        bucket["comprobaciones_count"] += 1
        bucket["ultimo"] = max([d for d in [bucket["ultimo"], fecha] if d], default=None)
        bucket["movimientos"].append({
            "fecha": fecha,
            "tipo": f"{gasto.tipo_gasto or 'GASTO'} COMPROBADO",
            "folio": gasto.folio or f"#{gasto.id}",
            "concepto": gasto.concepto or "",
            "referencia": gasto.proyecto or gasto.evento or gasto.referencia or "",
            "estatus": gasto.estatus,
            "monto_enviado": 0.0,
            "monto_comprobado": monto,
            "saldo_delta": -monto,
            "url": url_for("gastos_viaticos_detalle", gasto_id=gasto.id),
        })

    for bucket in buckets.values():
        bucket["enviado"] = fmt(bucket["enviado"])
        bucket["comprobado"] = fmt(bucket["comprobado"])
        bucket["saldo"] = fmt(bucket["enviado"] - bucket["comprobado"])
        bucket["movimientos"].sort(key=lambda item: (item["fecha"] or datetime.min, item["folio"]))
        saldo = 0.0
        for mov in bucket["movimientos"]:
            saldo = fmt(saldo + float(mov["saldo_delta"] or 0))
            mov["saldo"] = saldo

    rows = sorted(buckets.values(), key=lambda item: (float(item["saldo"] or 0), item["nombre"]), reverse=True)
    if user_key:
        rows = [item for item in rows if item["key"] == user_key]
    return rows


def _solicitudes_recurso_saldos(comprobaciones: list["ComprobacionGasto"]) -> list[dict]:
    by_id: dict[int, dict] = {}
    for solicitud in (
        SolicitudRecurso.query
        .filter(SolicitudRecurso.estatus == "AUTORIZADA")
        .order_by(SolicitudRecurso.fecha.desc(), SolicitudRecurso.id.desc())
        .all()
    ):
        if not solicitud.id:
            continue
        by_id[solicitud.id] = {
            "id": solicitud.id,
            "folio": solicitud.folio or f"#{solicitud.id}",
            "proyecto": (solicitud.proyecto or "").strip() or "Sin proyecto",
            "solicitante": solicitud.solicitante or "",
            "aprobado": float(solicitud.total or 0),
            "comprobado": 0.0,
            "pendiente_revision": 0.0,
            "saldo": float(solicitud.total or 0),
            "movimientos": 0,
        }

    for gasto in comprobaciones:
        solicitud_id = getattr(gasto, "solicitud_recurso_id", None)
        if not solicitud_id and (gasto.referencia or "").startswith("SR:"):
            folio = (gasto.referencia or "")[3:]
            solicitud = SolicitudRecurso.query.filter_by(folio=folio).first()
            solicitud_id = solicitud.id if solicitud else None
        if not solicitud_id or solicitud_id not in by_id or _gastos_es_recurso(gasto):
            continue
        by_id[solicitud_id]["movimientos"] += 1
        if (gasto.estatus or "") == "RECHAZADO":
            continue
        by_id[solicitud_id]["comprobado"] += float(gasto.total or 0)
        by_id[solicitud_id]["saldo"] -= float(gasto.total or 0)
        if (gasto.estatus or "") in {"PENDIENTE", "EN REVISION"}:
            by_id[solicitud_id]["pendiente_revision"] += float(gasto.total or 0)

    return list(by_id.values())


def _gastos_user_scope_filter():
    if is_admin():
        return None
    current_user_id = getattr(current_user, "id", None)
    responsable = responsable_actual()
    filters = []
    if current_user_id:
        filters.append(ComprobacionGasto.usuario_id == current_user_id)
    if responsable:
        filters.append(ComprobacionGasto.responsable == responsable)
    if not filters:
        return ComprobacionGasto.id == -1
    return or_(*filters)


def _gastos_apply_user_scope(query):
    scope = _gastos_user_scope_filter()
    if scope is not None:
        query = query.filter(scope)
    return query


def require_gasto_owner_or_admin(gasto: "ComprobacionGasto") -> None:
    if is_admin():
        return
    current_user_id = getattr(current_user, "id", None)
    if current_user_id and gasto.usuario_id == current_user_id:
        return
    responsable = responsable_actual()
    if responsable and (gasto.responsable or "") == responsable:
        return
    abort(403)


def _gastos_file_ext(filename: str) -> str:
    return (filename or "").rsplit(".", 1)[-1].lower() if "." in (filename or "") else ""


def _gastos_save_upload(uploaded, comprobacion_id: int) -> ComprobacionAdjunto | None:
    if not uploaded or not (uploaded.filename or "").strip():
        return None
    ext = _gastos_file_ext(uploaded.filename)
    if ext not in GASTOS_UPLOAD_EXTS:
        raise ValueError("Adjunta un PDF o imagen valida: pdf, png, jpg, jpeg o webp.")

    upload_dir = Path(app.static_folder or "static") / "uploads" / "gastos_viaticos" / str(comprobacion_id)
    upload_dir.mkdir(parents=True, exist_ok=True)
    original = secure_filename(uploaded.filename) or f"comprobante.{ext}"
    stem = Path(original).stem[:80] or "comprobante"
    filename = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{stem}.{ext}"
    disk_path = upload_dir / filename
    uploaded.save(disk_path)
    rel_path = f"uploads/gastos_viaticos/{comprobacion_id}/{filename}"
    return ComprobacionAdjunto(
        comprobacion_id=comprobacion_id,
        nombre_original=uploaded.filename,
        nombre_archivo=filename,
        ruta=rel_path,
        mime_type=uploaded.mimetype or mimetypes.guess_type(uploaded.filename)[0],
        tamano=disk_path.stat().st_size if disk_path.exists() else 0,
    )


def _gastos_pdf_text(file_bytes: bytes) -> str:
    try:
        import pdfplumber
        text_parts: list[str] = []
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages[:4]:
                text_parts.append(page.extract_text() or "")
        return "\n".join(text_parts).strip()
    except Exception:
        return ""


def _gastos_pdf_first_page_png(file_bytes: bytes) -> bytes:
    try:
        import pypdfium2 as pdfium
        pdf = pdfium.PdfDocument(file_bytes)
        if len(pdf) <= 0:
            return b""
        page = pdf[0]
        bitmap = page.render(scale=2.0)
        pil_image = bitmap.to_pil()
        out = io.BytesIO()
        pil_image.save(out, format="PNG")
        return out.getvalue()
    except Exception:
        return b""


def _gastos_parse_json(raw: str) -> dict:
    raw = (raw or "").strip()
    if not raw:
        return {}
    try:
        return json.loads(raw)
    except Exception:
        match = re.search(r"\{.*\}", raw, re.S)
        if match:
            try:
                return json.loads(match.group(0))
            except Exception:
                return {}
    return {}


def _gastos_guess_concepts_from_text(text: str) -> list[str]:
    skip_words = (
        "total", "subtotal", "iva", "importe", "cambio", "efectivo", "tarjeta",
        "rfc", "uuid", "folio", "factura", "ticket", "fecha", "hora", "cajero",
        "cliente", "regimen", "lugar", "expedicion", "metodo", "pago", "moneda",
    )
    concepts: list[str] = []
    for raw_line in (text or "").splitlines():
        line = re.sub(r"\s+", " ", raw_line or "").strip(" -:\t")
        if not line or len(line) < 4 or len(line) > 120:
            continue
        norm = line.lower()
        if any(word in norm for word in skip_words):
            continue
        if not re.search(r"[a-zA-ZáéíóúÁÉÍÓÚñÑ]", line):
            continue
        if re.fullmatch(r"[\d\s,.$%/:-]+", line):
            continue
        if re.search(r"\b\d{2,4}[-/]\d{1,2}[-/]\d{1,2}\b", line):
            continue
        cleaned = re.sub(r"\s+\$?\d{1,3}(?:,\d{3})*(?:\.\d{2})\s*$", "", line).strip(" -:")
        if len(cleaned) < 4:
            continue
        if cleaned.lower() in {item.lower() for item in concepts}:
            continue
        concepts.append(cleaned[:100])
        if len(concepts) >= 8:
            break
    return concepts


def _gastos_heuristic_extract(text: str) -> dict:
    amounts = []
    for match in re.finditer(r"(?:\$|MXN|M\.N\.)?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{2})|[0-9]+(?:\.[0-9]{2}))", text or "", re.I):
        amounts.append(parse_float(match.group(1), 0))
    amounts = [a for a in amounts if a > 0]

    date_value = ""
    date_match = re.search(r"\b(\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4})\b", text or "")
    if date_match:
        raw = date_match.group(1).replace("/", "-")
        for fmt_date in ("%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y"):
            try:
                date_value = datetime.strptime(raw, fmt_date).strftime("%Y-%m-%d")
                break
            except ValueError:
                continue

    lines = [line.strip() for line in (text or "").splitlines() if line.strip()]
    total = max(amounts) if amounts else 0.0
    iva = round(total * 0.16 / 1.16, 2) if total else 0.0
    subtotal = round(total - iva, 2) if total else 0.0
    concepts = _gastos_guess_concepts_from_text(text)
    concept_text = "; ".join(concepts) if concepts else "Comprobante de gasto"
    return {
        "proveedor": lines[0][:160] if lines else "",
        "concepto": concept_text[:260],
        "conceptos": concepts or ["Comprobante de gasto"],
        "referencia": "",
        "fecha_comprobante": date_value,
        "subtotal": subtotal,
        "iva": iva,
        "total": total,
        "moneda": "MXN",
        "metodo_pago": "",
        "tipo_gasto": "GASTO",
        "confianza": 0.45 if total or date_value else 0.0,
        "observaciones": "Extraccion automatica basica; revisa y ajusta los campos.",
    }


def _gastos_ai_extract(file_bytes: bytes, filename: str, mime_type: str, text: str = "") -> dict:
    prompt = (
        "Extrae datos de un comprobante de gastos o viaticos en Mexico. "
        "Responde solo JSON valido con estas llaves: proveedor, concepto, conceptos, referencia, "
        "fecha_comprobante en formato YYYY-MM-DD, subtotal, iva, total, moneda, metodo_pago, "
        "tipo_gasto como GASTO o VIATICO, confianza de 0 a 1, observaciones. "
        "conceptos debe ser una lista con cada concepto comprado o consumido detectado en el ticket/PDF, "
        "por ejemplo alimentos, gasolina, hospedaje o cada partida visible. No uses solo una categoria generica "
        "si hay articulos o partidas visibles. "
        "Si un dato no existe usa cadena vacia o 0."
    )
    if not OPENAI_API_KEY:
        return _gastos_heuristic_extract(text)

    headers = {"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"}
    content = [{"type": "text", "text": prompt}]
    if text:
        content.append({"type": "text", "text": f"Texto extraido del PDF:\n{text[:12000]}"})
    elif (mime_type or "").startswith("image/"):
        b64 = base64.b64encode(file_bytes).decode("ascii")
        content.append({"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{b64}"}})
    elif (mime_type or "").lower() == "application/pdf":
        png_bytes = _gastos_pdf_first_page_png(file_bytes)
        if not png_bytes:
            return _gastos_heuristic_extract(text)
        b64 = base64.b64encode(png_bytes).decode("ascii")
        content.append({"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}})
    else:
        return _gastos_heuristic_extract(text)

    payload = {
        "model": os.getenv("OPENAI_RECEIPT_MODEL", "gpt-4o-mini"),
        "messages": [{"role": "user", "content": content}],
        "temperature": 0,
        "response_format": {"type": "json_object"},
    }
    resp = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, json=payload, timeout=60)
    if resp.status_code >= 400:
        raise RuntimeError(resp.text[:500])
    message = ((resp.json().get("choices") or [{}])[0].get("message") or {}).get("content") or "{}"
    data = _gastos_parse_json(message)
    fallback = _gastos_heuristic_extract(text)
    fallback.update({k: v for k, v in data.items() if v not in (None, "", [], {})})
    return fallback


def _gastos_normalize_ai(data: dict) -> dict:
    data = data or {}
    conceptos_raw = data.get("conceptos")
    conceptos = []
    if isinstance(conceptos_raw, list):
        conceptos = [str(item).strip() for item in conceptos_raw if str(item or "").strip()]
    elif conceptos_raw:
        conceptos = [str(conceptos_raw).strip()]
    concepto_text = "; ".join(conceptos) or str(data.get("concepto") or "Comprobante de gasto")
    normalized = {
        "proveedor": str(data.get("proveedor") or "")[:180],
        "concepto": concepto_text[:260],
        "conceptos": conceptos[:20],
        "referencia": str(data.get("referencia") or "")[:120],
        "fecha_comprobante": str(data.get("fecha_comprobante") or "")[:10],
        "subtotal": fmt(parse_float(data.get("subtotal"), 0)),
        "iva": fmt(parse_float(data.get("iva"), 0)),
        "total": fmt(parse_float(data.get("total"), 0)),
        "moneda": (str(data.get("moneda") or "MXN").upper()[:10] or "MXN"),
        "metodo_pago": str(data.get("metodo_pago") or "")[:80],
        "tipo_gasto": (str(data.get("tipo_gasto") or "GASTO").upper()[:20] or "GASTO"),
        "confianza": min(1.0, max(0.0, parse_float(data.get("confianza"), 0))),
        "observaciones": str(data.get("observaciones") or "")[:500],
    }
    if normalized["tipo_gasto"] not in GASTOS_TIPOS:
        normalized["tipo_gasto"] = "GASTO"
    if not normalized["subtotal"] and normalized["total"] and normalized["iva"]:
        normalized["subtotal"] = fmt(normalized["total"] - normalized["iva"])
    return normalized


def _gastos_redirect():
    return redirect(request.referrer or url_for("gastos_viaticos_index"))


def _gastos_review_serializer() -> URLSafeTimedSerializer:
    return URLSafeTimedSerializer(app.secret_key, salt="gastos-viaticos-review")


def _gastos_review_token(gasto: "ComprobacionGasto", action: str = "view") -> str:
    return _gastos_review_serializer().dumps({"gasto_id": gasto.id, "action": action})


def _gastos_group_review_token(gastos: list["ComprobacionGasto"], action: str = "view") -> str:
    ids = [int(gasto.id) for gasto in gastos if getattr(gasto, "id", None)]
    return _gastos_review_serializer().dumps({"gasto_ids": ids, "action": action})


def _gastos_load_from_token(gasto_id: int, token: str, action: str = "view") -> "ComprobacionGasto":
    try:
        payload = _gastos_review_serializer().loads(token or "", max_age=60 * 60 * 24 * 45)
    except (BadSignature, SignatureExpired):
        abort(403)
    if int(payload.get("gasto_id") or 0) != int(gasto_id):
        abort(403)
    token_action = (payload.get("action") or "").strip()
    if action == "approve":
        if token_action != "approve":
            abort(403)
    elif token_action not in {action, "approve"}:
        abort(403)
    return ComprobacionGasto.query.get_or_404(gasto_id)


def _gastos_load_group_from_token(token: str, action: str = "view") -> list["ComprobacionGasto"]:
    try:
        payload = _gastos_review_serializer().loads(token or "", max_age=60 * 60 * 24 * 45)
    except (BadSignature, SignatureExpired):
        abort(403)
    ids = [int(item) for item in (payload.get("gasto_ids") or []) if str(item).isdigit()]
    if not ids:
        abort(403)
    token_action = (payload.get("action") or "").strip()
    if action == "approve":
        if token_action != "approve":
            abort(403)
    elif token_action not in {action, "approve"}:
        abort(403)
    gastos = (
        ComprobacionGasto.query
        .filter(ComprobacionGasto.id.in_(ids))
        .order_by(ComprobacionGasto.fecha_comprobante.asc(), ComprobacionGasto.id.asc())
        .all()
    )
    by_id = {gasto.id: gasto for gasto in gastos}
    return [by_id[item_id] for item_id in ids if item_id in by_id]


def _gastos_fecha_base(gasto: "ComprobacionGasto") -> datetime:
    return gasto.fecha_comprobante or gasto.fecha_registro or gasto.creado_en or now_cdmx_naive()


def _gastos_fecha_key(gasto: "ComprobacionGasto") -> str:
    return _gastos_fecha_base(gasto).strftime("%Y-%m-%d")


def _gastos_group_name(gasto: "ComprobacionGasto") -> str:
    if (gasto.tipo_agrupacion or "").upper() == "PROYECTO":
        return (gasto.proyecto or "").strip() or "Sin proyecto"
    return (gasto.evento or "").strip() or "Sin evento"


def _gastos_group_query(tipo_agrupacion: str, grupo: str, fecha: str, responsable: str):
    fecha_dt = _parse_date_or_none(fecha)
    if not fecha_dt:
        abort(400)
    tipo_agrupacion = (tipo_agrupacion or "").strip().upper()
    if tipo_agrupacion not in GASTOS_AGRUPACIONES:
        abort(400)
    field = ComprobacionGasto.proyecto if tipo_agrupacion == "PROYECTO" else ComprobacionGasto.evento
    query = _gastos_apply_user_scope(ComprobacionGasto.query).filter(
        ComprobacionGasto.tipo_agrupacion == tipo_agrupacion,
        ComprobacionGasto.estatus.in_(("PENDIENTE", "EN REVISION")),
        ComprobacionGasto.tipo_gasto != "RECURSO",
        field == (grupo or "").strip(),
    )
    if (responsable or "").strip():
        query = query.filter(ComprobacionGasto.responsable == responsable.strip())
    else:
        query = query.filter(or_(ComprobacionGasto.responsable.is_(None), ComprobacionGasto.responsable == ""))
    next_day = fecha_dt + timedelta(days=1)
    return query.filter(or_(
        and_(ComprobacionGasto.fecha_comprobante >= fecha_dt, ComprobacionGasto.fecha_comprobante < next_day),
        and_(ComprobacionGasto.fecha_comprobante.is_(None), ComprobacionGasto.fecha_registro >= fecha_dt, ComprobacionGasto.fecha_registro < next_day),
    ))


def _gastos_group_all_query(tipo_agrupacion: str, grupo: str, fecha: str, responsable: str):
    fecha_dt = _parse_date_or_none(fecha)
    if not fecha_dt:
        abort(400)
    tipo_agrupacion = (tipo_agrupacion or "").strip().upper()
    if tipo_agrupacion not in GASTOS_AGRUPACIONES:
        abort(400)
    field = ComprobacionGasto.proyecto if tipo_agrupacion == "PROYECTO" else ComprobacionGasto.evento
    query = _gastos_apply_user_scope(ComprobacionGasto.query).filter(
        ComprobacionGasto.tipo_agrupacion == tipo_agrupacion,
        field == (grupo or "").strip(),
    )
    if (responsable or "").strip():
        query = query.filter(ComprobacionGasto.responsable == responsable.strip())
    else:
        query = query.filter(or_(ComprobacionGasto.responsable.is_(None), ComprobacionGasto.responsable == ""))
    next_day = fecha_dt + timedelta(days=1)
    return query.filter(or_(
        and_(ComprobacionGasto.fecha_comprobante >= fecha_dt, ComprobacionGasto.fecha_comprobante < next_day),
        and_(ComprobacionGasto.fecha_comprobante.is_(None), ComprobacionGasto.fecha_registro >= fecha_dt, ComprobacionGasto.fecha_registro < next_day),
    ))


def _gastos_mail_html(gasto: "ComprobacionGasto", view_url: str, approve_url: str) -> str:
    concepto = escape(gasto.concepto or "")
    proveedor = escape(gasto.proveedor or "Sin proveedor")
    folio = escape(gasto.folio or f"#{gasto.id}")
    responsable = escape(gasto.responsable or "Sin responsable")
    total = f"${float(gasto.total or 0):,.2f} {escape(gasto.moneda or 'MXN')}"
    grupo = escape((gasto.proyecto if gasto.tipo_agrupacion == "PROYECTO" else gasto.evento) or "Sin grupo")
    fecha = escape(gasto.fecha_comprobante.strftime("%d/%m/%Y") if gasto.fecha_comprobante else "Sin fecha")
    referencia = escape(gasto.referencia or "Sin referencia")
    button_base = (
        "display:inline-block;min-width:156px;text-align:center;padding:15px 24px;"
        "border-radius:8px;text-decoration:none;font-weight:700;font-size:16px;"
        "letter-spacing:.2px;margin:0 8px 10px 0;"
    )
    return f"""
    <html>
      <body style="margin:0;padding:0;background:#eef2f7;font-family:Arial,Helvetica,sans-serif;color:#1f2937;">
        <div style="max-width:760px;margin:0 auto;padding:30px 16px;">
          <div style="background:#ffffff;border:1px solid #d9e2ec;border-radius:10px;overflow:hidden;box-shadow:0 8px 24px rgba(15,45,80,.08);">
            <div style="background:#0C3C78;color:#ffffff;padding:22px 26px;">
              <div style="font-size:12px;font-weight:700;letter-spacing:.9px;text-transform:uppercase;opacity:.9;">MAR · Poliutech</div>
              <div style="font-size:23px;font-weight:800;margin-top:5px;">Comprobante pendiente de revision</div>
              <div style="font-size:14px;opacity:.92;margin-top:6px;">Gastos y viaticos</div>
            </div>
            <div style="padding:26px;">
              <p style="margin:0 0 20px 0;font-size:15px;color:#475569;">Se registro un comprobante y requiere validacion administrativa.</p>

              <div style="border:1px solid #dbe4ef;border-radius:10px;overflow:hidden;margin-bottom:22px;">
                <div style="background:#f8fafc;padding:14px 18px;border-bottom:1px solid #dbe4ef;">
                  <div style="font-size:12px;text-transform:uppercase;letter-spacing:.7px;color:#64748b;font-weight:700;">Folio</div>
                  <div style="font-size:20px;font-weight:800;color:#0C3C78;margin-top:2px;">{folio}</div>
                </div>
                <table style="border-collapse:collapse;width:100%;background:#ffffff;">
                  <tr>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;width:34%;color:#64748b;font-weight:700;">Proveedor</td>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#111827;font-weight:600;">{proveedor}</td>
                  </tr>
                  <tr>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:700;">Concepto(s)</td>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#111827;">{concepto}</td>
                  </tr>
                  <tr>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:700;">Grupo</td>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#111827;">{grupo}</td>
                  </tr>
                  <tr>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:700;">Quien hizo la compra</td>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#111827;">{responsable}</td>
                  </tr>
                  <tr>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#64748b;font-weight:700;">Fecha</td>
                    <td style="padding:13px 16px;border-bottom:1px solid #edf2f7;color:#111827;">{fecha}</td>
                  </tr>
                  <tr>
                    <td style="padding:13px 16px;color:#64748b;font-weight:700;">Referencia</td>
                    <td style="padding:13px 16px;color:#111827;">{referencia}</td>
                  </tr>
                </table>
              </div>

              <div style="background:#f0f7ff;border:1px solid #cfe3ff;border-radius:10px;padding:16px 18px;margin-bottom:24px;">
                <div style="font-size:12px;text-transform:uppercase;letter-spacing:.7px;color:#0C3C78;font-weight:800;">Total del comprobante</div>
                <div style="font-size:30px;font-weight:900;color:#0C3C78;margin-top:3px;">{total}</div>
              </div>

              <div style="margin-top:4px;">
                <a href="{view_url}" style="{button_base}background:#0C3C78;color:#ffffff;border:1px solid #0C3C78;">Ver Detalle</a>
                <a href="{approve_url}" style="{button_base}background:#16854f;color:#ffffff;border:1px solid #16854f;">Aprobar</a>
              </div>

              <div style="margin-top:18px;padding-top:16px;border-top:1px solid #e5e7eb;color:#64748b;font-size:12px;">
                Este mensaje fue generado automaticamente por MAR. Si los botones no abren, usa la vista HTML del correo.
              </div>
            </div>
          </div>
        </div>
      </body>
    </html>
    """.strip()


def _gastos_group_mail_html(gastos: list["ComprobacionGasto"], view_url: str, approve_url: str) -> str:
    first = gastos[0]
    grupo = escape(_gastos_group_name(first))
    tipo = escape(first.tipo_agrupacion or "")
    responsable = escape(first.responsable or "Sin responsable")
    fecha = escape(_gastos_fecha_base(first).strftime("%d/%m/%Y"))
    total = sum(float(gasto.total or 0) for gasto in gastos)
    rows = []
    for gasto in gastos:
        rows.append(
            "<tr>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;font-weight:700;color:#0C3C78;'>{escape(gasto.folio or f'#{gasto.id}')}</td>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;'>{escape(gasto.proveedor or 'Sin proveedor')}</td>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;'>{escape(gasto.concepto or '')}</td>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;text-align:right;font-weight:700;'>${float(gasto.total or 0):,.2f} {escape(gasto.moneda or 'MXN')}</td>"
            "</tr>"
        )
    button_base = (
        "display:inline-block;min-width:156px;text-align:center;padding:15px 24px;"
        "border-radius:8px;text-decoration:none;font-weight:700;font-size:16px;"
        "letter-spacing:.2px;margin:0 8px 10px 0;"
    )
    return f"""
    <html>
      <body style="margin:0;padding:0;background:#eef2f7;font-family:Arial,Helvetica,sans-serif;color:#1f2937;">
        <div style="max-width:860px;margin:0 auto;padding:30px 16px;">
          <div style="background:#ffffff;border:1px solid #d9e2ec;border-radius:10px;overflow:hidden;box-shadow:0 8px 24px rgba(15,45,80,.08);">
            <div style="background:#0C3C78;color:#ffffff;padding:22px 26px;">
              <div style="font-size:12px;font-weight:700;letter-spacing:.9px;text-transform:uppercase;opacity:.9;">MAR · Poliutech</div>
              <div style="font-size:23px;font-weight:800;margin-top:5px;">Salida agrupada pendiente de revision</div>
              <div style="font-size:14px;opacity:.92;margin-top:6px;">Gastos y viaticos</div>
            </div>
            <div style="padding:26px;">
              <p style="margin:0 0 20px 0;font-size:15px;color:#475569;">Se envio una salida agrupada con {len(gastos)} comprobante(s) para validacion administrativa.</p>
              <div style="background:#f8fafc;border:1px solid #dbe4ef;border-radius:10px;padding:16px 18px;margin-bottom:20px;">
                <div style="font-size:20px;font-weight:900;color:#0C3C78;">{grupo}</div>
                <div style="margin-top:6px;color:#475569;">{tipo} · {fecha} · {responsable}</div>
              </div>
              <table style="border-collapse:collapse;width:100%;background:#ffffff;border:1px solid #dbe4ef;margin-bottom:22px;">
                <thead>
                  <tr style="background:#f1f5f9;color:#334155;">
                    <th style="padding:10px;text-align:left;">Folio</th>
                    <th style="padding:10px;text-align:left;">Proveedor</th>
                    <th style="padding:10px;text-align:left;">Concepto(s)</th>
                    <th style="padding:10px;text-align:right;">Total</th>
                  </tr>
                </thead>
                <tbody>{''.join(rows)}</tbody>
              </table>
              <div style="background:#f0f7ff;border:1px solid #cfe3ff;border-radius:10px;padding:16px 18px;margin-bottom:24px;">
                <div style="font-size:12px;text-transform:uppercase;letter-spacing:.7px;color:#0C3C78;font-weight:800;">Total de la salida</div>
                <div style="font-size:30px;font-weight:900;color:#0C3C78;margin-top:3px;">${total:,.2f}</div>
              </div>
              <div style="margin-top:4px;">
                <a href="{view_url}" style="{button_base}background:#0C3C78;color:#ffffff;border:1px solid #0C3C78;">Ver Salida</a>
                <a href="{approve_url}" style="{button_base}background:#16854f;color:#ffffff;border:1px solid #16854f;">Aprobar Salida</a>
              </div>
            </div>
          </div>
        </div>
      </body>
    </html>
    """.strip()


def _gastos_add_email_attachments(msg: EmailMessage, gasto: "ComprobacionGasto") -> None:
    for adj in getattr(gasto, "adjuntos", []) or []:
        rel_path = (getattr(adj, "ruta", "") or "").replace("\\", "/").lstrip("/")
        if not rel_path:
            continue
        file_path = Path(app.static_folder or "static") / rel_path
        if not file_path.exists() or not file_path.is_file():
            logger.warning("Adjunto de gasto no encontrado: %s", file_path)
            continue
        mime_type = getattr(adj, "mime_type", None) or mimetypes.guess_type(str(file_path))[0] or "application/octet-stream"
        maintype, subtype = (mime_type.split("/", 1) + ["octet-stream"])[:2] if "/" in mime_type else ("application", "octet-stream")
        filename = getattr(adj, "nombre_original", None) or file_path.name
        msg.add_attachment(file_path.read_bytes(), maintype=maintype, subtype=subtype, filename=filename)


def _send_gastos_review_push_hansel(gasto: "ComprobacionGasto") -> dict[str, int]:
    tokens = _mobile_push_tokens_for_users(_mobile_push_user_ids_for_hansel_only())
    if not tokens:
        logger.warning("Push gasto %s: Hansel no tiene token movil activo.", gasto.folio or gasto.id)
    return _send_push_notification(
        tokens,
        title="Nuevo comprobante de gasto",
        body=f"{gasto.folio or gasto.id} - ${float(gasto.total or 0):,.2f} {gasto.moneda or 'MXN'}",
        data={
            "type": "gasto_viatico",
            "gasto_id": str(gasto.id),
            "folio": gasto.folio or "",
            "url": url_for("gastos_viaticos_detalle", gasto_id=gasto.id, _external=True),
        },
    )


def _notify_gasto_created_for_review(gasto: "ComprobacionGasto") -> None:
    try:
        _send_gastos_review_email(gasto)
    except Exception as exc:
        logger.exception("No se pudo enviar correo de revision de gasto %s", gasto.folio or gasto.id)
        raise exc
    try:
        _send_gastos_review_push_hansel(gasto)
    except Exception as exc:
        logger.warning("Push de gasto %s fallo: %s", gasto.folio or gasto.id, exc)


def _send_gastos_review_email(gasto: "ComprobacionGasto") -> None:
    recipients = _parse_email_list(GASTOS_REVIEW_EMAIL)
    bcc = _parse_email_list(GASTOS_REVIEW_BCC_EMAIL)
    if not recipients:
        raise ValueError("No hay correo configurado para revision de gastos.")
    view_url = url_for("gastos_viaticos_revision", gasto_id=gasto.id, token=_gastos_review_token(gasto, "view"), _external=True)
    approve_url = url_for("gastos_viaticos_revision_aprobar", gasto_id=gasto.id, token=_gastos_review_token(gasto, "approve"), _external=True)

    msg = EmailMessage()
    msg["Subject"] = f"Revision de comprobante {gasto.folio or gasto.id}"
    msg["From"] = f"REGISTRO DE GASTOS Y/O VIATICOS <{SMTP_FROM or SMTP_USERNAME}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content(
        f"Nuevo comprobante {gasto.folio or gasto.id}\n"
        f"Proveedor: {gasto.proveedor or 'Sin proveedor'}\n"
        f"Concepto(s): {gasto.concepto or ''}\n"
        f"Total: ${float(gasto.total or 0):,.2f} {gasto.moneda or 'MXN'}\n\n"
        "Abre este correo en vista HTML para usar los botones Ver y Aprobar.\n"
        f"Ver: {view_url}\n"
        f"Aprobar: {approve_url}\n"
    )
    msg.add_alternative(_gastos_mail_html(gasto, view_url, approve_url), subtype="html")
    _gastos_add_email_attachments(msg, gasto)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg, to_addrs=[*recipients, *bcc])


def _send_gastos_group_review_email(gastos: list["ComprobacionGasto"]) -> None:
    recipients = _parse_email_list(GASTOS_REVIEW_EMAIL)
    bcc = _parse_email_list(GASTOS_REVIEW_BCC_EMAIL)
    if not recipients:
        raise ValueError("No hay correo configurado para revision de gastos.")
    if not gastos:
        raise ValueError("No hay gastos para enviar a revision.")

    view_url = url_for("gastos_viaticos_revision_grupo", token=_gastos_group_review_token(gastos, "view"), _external=True)
    approve_url = url_for("gastos_viaticos_revision_grupo_aprobar", token=_gastos_group_review_token(gastos, "approve"), _external=True)
    first = gastos[0]
    grupo = _gastos_group_name(first)
    fecha = _gastos_fecha_base(first).strftime("%d/%m/%Y")
    total = sum(float(gasto.total or 0) for gasto in gastos)
    lines = [
        f"Salida agrupada: {grupo}",
        f"Fecha: {fecha}",
        f"Comprobantes: {len(gastos)}",
        f"Total: ${total:,.2f}",
        "",
    ]
    for gasto in gastos:
        lines.append(f"- {gasto.folio or gasto.id}: {gasto.proveedor or 'Sin proveedor'} | {gasto.concepto or ''} | ${float(gasto.total or 0):,.2f} {gasto.moneda or 'MXN'}")
    lines.extend(["", "Abre este correo en vista HTML para usar los botones Ver y Aprobar.", f"Ver: {view_url}", f"Aprobar: {approve_url}"])

    msg = EmailMessage()
    msg["Subject"] = f"Revision de salida {grupo} - {fecha} ({len(gastos)} comprobantes)"
    msg["From"] = f"REGISTRO DE GASTOS Y/O VIATICOS <{SMTP_FROM or SMTP_USERNAME}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content("\n".join(lines))
    msg.add_alternative(_gastos_group_mail_html(gastos, view_url, approve_url), subtype="html")

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg, to_addrs=[*recipients, *bcc])


def _gastos_authorized_mail_html(gastos: list["ComprobacionGasto"], detail_url: str) -> str:
    total = sum(float(gasto.total or 0) for gasto in gastos)
    rows = []
    for gasto in gastos:
        grupo = gasto.proyecto or gasto.evento or "-"
        rows.append(
            "<tr>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;font-weight:700;color:#0C3C78;'>{escape(gasto.folio or f'#{gasto.id}')}</td>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;'>{escape(gasto.tipo_gasto or 'GASTO')}</td>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;'>{escape(grupo)}</td>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;'>{escape(gasto.responsable or '-')}</td>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;'>{escape(gasto.concepto or '')}</td>"
            f"<td style='padding:10px;border-bottom:1px solid #e5e7eb;text-align:right;font-weight:700;'>${float(gasto.total or 0):,.2f} {escape(gasto.moneda or 'MXN')}</td>"
            "</tr>"
        )
    title = "Comprobante autorizado" if len(gastos) == 1 else "Comprobantes autorizados"
    return f"""
    <div style="font-family:Arial,sans-serif;color:#0f172a;max-width:860px;margin:0 auto;">
      <h2 style="margin:0 0 10px;color:#15803d;">{title}</h2>
      <p style="margin:0 0 18px;color:#475569;">Se autorizo {len(gastos)} comprobante(s) de gasto/viatico.</p>
      <table style="width:100%;border-collapse:collapse;border:1px solid #e5e7eb;">
        <thead>
          <tr style="background:#f8fafc;color:#334155;">
            <th style="padding:10px;text-align:left;">Folio</th>
            <th style="padding:10px;text-align:left;">Tipo</th>
            <th style="padding:10px;text-align:left;">Proyecto / evento</th>
            <th style="padding:10px;text-align:left;">Responsable</th>
            <th style="padding:10px;text-align:left;">Concepto</th>
            <th style="padding:10px;text-align:right;">Total</th>
          </tr>
        </thead>
        <tbody>{''.join(rows)}</tbody>
      </table>
      <div style="margin:18px 0;padding:14px 16px;border:1px solid #bbf7d0;background:#f0fdf4;border-radius:8px;">
        <div style="font-size:12px;text-transform:uppercase;color:#166534;font-weight:800;">Total autorizado</div>
        <div style="font-size:24px;font-weight:900;color:#166534;">${total:,.2f}</div>
      </div>
      <p style="margin:18px 0;"> <a href="{detail_url}" style="background:#0C3C78;color:#fff;text-decoration:none;padding:10px 14px;border-radius:6px;display:inline-block;">Ver detalle</a></p>
    </div>
    """


def _send_gastos_authorized_finanzas_email(gastos: list["ComprobacionGasto"]) -> None:
    gastos = [gasto for gasto in gastos if not _gastos_es_recurso(gasto)]
    if not gastos:
        return
    recipients = _finanzas_auth_notify_recipients()
    if not recipients:
        raise ValueError("No hay correos configurados para autorizaciones de finanzas.")

    first = gastos[0]
    detail_url = url_for("gastos_viaticos_detalle", gasto_id=first.id, _external=True)
    total = sum(float(gasto.total or 0) for gasto in gastos)
    subject = (
        f"Comprobante autorizado {first.folio or first.id}"
        if len(gastos) == 1
        else f"{len(gastos)} comprobantes autorizados - ${total:,.2f}"
    )
    lines = [
        subject,
        f"Total autorizado: ${total:,.2f}",
        "",
    ]
    for gasto in gastos:
        lines.append(
            f"- {gasto.folio or gasto.id}: {gasto.tipo_gasto or 'GASTO'} | "
            f"{gasto.proyecto or gasto.evento or '-'} | {gasto.responsable or '-'} | "
            f"${float(gasto.total or 0):,.2f} {gasto.moneda or 'MXN'}"
        )
    lines.extend(["", f"Ver detalle: {detail_url}"])

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = f"REGISTRO DE GASTOS Y/O VIATICOS <{SMTP_FROM or SMTP_USERNAME}>"
    msg["To"] = ", ".join(recipients)
    msg.set_content("\n".join(lines))
    msg.add_alternative(_gastos_authorized_mail_html(gastos, detail_url), subtype="html")

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.ehlo()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(msg, to_addrs=recipients)


def _send_gastos_authorized_finanzas_push(gastos: list["ComprobacionGasto"]) -> dict[str, int]:
    gastos = [gasto for gasto in gastos if not _gastos_es_recurso(gasto)]
    if not gastos:
        return {"sent": 0, "failed": 0}
    user_ids = _mobile_push_user_ids_for_finanzas_auth_notify()
    tokens = _mobile_push_tokens_for_users(user_ids)
    if not tokens:
        logger.warning("Push gastos autorizados: Mescalera/Miguel sin token movil activo.")
    total = sum(float(gasto.total or 0) for gasto in gastos)
    first = gastos[0]
    title = "Comprobante autorizado" if len(gastos) == 1 else "Comprobantes autorizados"
    body = (
        f"{first.folio or first.id} - ${float(first.total or 0):,.2f} {first.moneda or 'MXN'}"
        if len(gastos) == 1
        else f"{len(gastos)} comprobantes - ${total:,.2f}"
    )
    return _send_push_notification(
        tokens,
        title=title,
        body=body,
        data={
            "type": "gastos_viaticos_autorizados_finanzas",
            "gasto_id": str(first.id),
            "folio": first.folio or "",
            "count": str(len(gastos)),
            "total": f"{total:.2f}",
            "url": url_for("gastos_viaticos_detalle", gasto_id=first.id, _external=True),
            "target_user_ids": ",".join(str(user_id) for user_id in user_ids),
        },
    )


def _notify_gastos_authorized_finanzas(gastos: list["ComprobacionGasto"]) -> None:
    gastos = [gasto for gasto in gastos if not _gastos_es_recurso(gasto)]
    if not gastos:
        return
    try:
        _send_gastos_authorized_finanzas_email(gastos)
    except Exception as exc:
        logger.warning("Correo finanzas gastos autorizados fallo: %s", exc)

    try:
        _send_gastos_authorized_finanzas_push(gastos)
    except Exception as exc:
        logger.warning("Push finanzas gastos autorizados fallo: %s", exc)


def _gastos_query_from_request():
    q = (request.args.get("q") or "").strip()
    agrupacion = (request.args.get("agrupacion") or "").strip().upper()
    estatus = (request.args.get("estatus") or "").strip().upper()

    query = _gastos_apply_user_scope(ComprobacionGasto.query)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            ComprobacionGasto.folio.ilike(like),
            ComprobacionGasto.proveedor.ilike(like),
            ComprobacionGasto.concepto.ilike(like),
            ComprobacionGasto.proyecto.ilike(like),
            ComprobacionGasto.evento.ilike(like),
            ComprobacionGasto.referencia.ilike(like),
            ComprobacionGasto.responsable.ilike(like),
        ))
    if agrupacion in GASTOS_AGRUPACIONES:
        query = query.filter(ComprobacionGasto.tipo_agrupacion == agrupacion)
    if estatus in GASTOS_ESTATUS:
        query = query.filter(ComprobacionGasto.estatus == estatus)
    return query, q, agrupacion, estatus


@app.route("/gastos-viaticos")
@login_required
def gastos_viaticos_index():
    query, q, agrupacion, estatus = _gastos_query_from_request()
    comprobaciones = query.order_by(ComprobacionGasto.fecha_registro.desc(), ComprobacionGasto.id.desc()).all()
    total_recursos = sum(float(g.total or 0) for g in comprobaciones if _gastos_es_recurso(g) and (g.estatus or "") != "RECHAZADO")
    total_gastos = sum(float(g.total or 0) for g in comprobaciones if not _gastos_es_recurso(g) and (g.estatus or "") != "RECHAZADO")
    saldo_total = sum(_gastos_monto_saldo(g) for g in comprobaciones)
    total_general = total_gastos
    total_pendiente = sum(float(g.total or 0) for g in comprobaciones if not _gastos_es_recurso(g) and (g.estatus or "") in {"PENDIENTE", "EN REVISION"})
    total_aprobado = sum(float(g.total or 0) for g in comprobaciones if not _gastos_es_recurso(g) and (g.estatus or "") in {"APROBADO", "REEMBOLSADO"})
    solicitudes_saldo = _solicitudes_recurso_saldos(comprobaciones)

    proyectos_saldo: dict[str, dict] = {}
    grupos: dict[tuple[str, str, str, str], dict] = {}
    for gasto in comprobaciones:
        if (gasto.tipo_agrupacion or "").upper() == "PROYECTO":
            proyecto_nombre = (gasto.proyecto or "").strip() or "Sin proyecto"
            proyecto_item = proyectos_saldo.setdefault(proyecto_nombre.lower(), {
                "nombre": proyecto_nombre,
                "recursos": 0.0,
                "gastos": 0.0,
                "saldo": 0.0,
                "movimientos": 0,
                "pendientes": 0.0,
            })
            proyecto_item["movimientos"] += 1
            proyecto_item["saldo"] += _gastos_monto_saldo(gasto)
            if _gastos_es_recurso(gasto) and (gasto.estatus or "") != "RECHAZADO":
                proyecto_item["recursos"] += float(gasto.total or 0)
            elif not _gastos_es_recurso(gasto) and (gasto.estatus or "") != "RECHAZADO":
                proyecto_item["gastos"] += float(gasto.total or 0)
                if (gasto.estatus or "") in {"PENDIENTE", "EN REVISION"}:
                    proyecto_item["pendientes"] += float(gasto.total or 0)

        nombre = _gastos_group_name(gasto)
        fecha = _gastos_fecha_key(gasto)
        responsable = (gasto.responsable or "").strip()
        key = (gasto.tipo_agrupacion or "PROYECTO", nombre, fecha, responsable)
        item = grupos.setdefault(key, {
            "nombre": nombre,
            "conteo": 0,
            "pendientes": 0,
            "total": 0.0,
            "recursos": 0.0,
            "gastos": 0.0,
            "saldo": 0.0,
            "tipo": gasto.tipo_agrupacion,
            "fecha": fecha,
            "responsable": responsable,
        })
        item["conteo"] += 1
        item["total"] += float(gasto.total or 0)
        item["saldo"] += _gastos_monto_saldo(gasto)
        if _gastos_es_recurso(gasto) and (gasto.estatus or "") != "RECHAZADO":
            item["recursos"] += float(gasto.total or 0)
        elif not _gastos_es_recurso(gasto) and (gasto.estatus or "") != "RECHAZADO":
            item["gastos"] += float(gasto.total or 0)
        if not _gastos_es_recurso(gasto) and (gasto.estatus or "") in {"PENDIENTE", "EN REVISION"}:
            item["pendientes"] += 1

    return render_template(
        "gastos_viaticos.html",
        title="Gastos y viaticos",
        comprobaciones=comprobaciones,
        grupos=sorted(grupos.values(), key=lambda item: (item["fecha"], item["total"]), reverse=True),
        total_general=total_general,
        total_pendiente=total_pendiente,
        total_aprobado=total_aprobado,
        total_recursos=total_recursos,
        total_gastos=total_gastos,
        saldo_total=saldo_total,
        proyectos_saldo=sorted(proyectos_saldo.values(), key=lambda item: item["nombre"].lower()),
        solicitudes_saldo=solicitudes_saldo,
        solicitud_options=solicitudes_saldo,
        q=q,
        agrupacion=agrupacion,
        estatus=estatus,
        estatus_options=GASTOS_ESTATUS,
        gasto_tipos=GASTOS_TIPOS,
        agrupaciones=GASTOS_AGRUPACIONES,
        gastos_badge_class=_gastos_badge_class,
        gastos_row_class=_gastos_status_row_class,
        gastos_es_recurso=_gastos_es_recurso,
        gastos_monto_saldo=_gastos_monto_saldo,
        fecha_input=_finanzas_fecha_input,
        responsable_default=responsable_actual() or "",
        gastos_can_view_all=is_admin(),
        project_options=_known_project_names(),
    )


def _gastos_admin_query():
    if not _gastos_admin_can_view():
        abort(403)
    return ComprobacionGasto.query.filter(ComprobacionGasto.tipo_gasto != "RECURSO")


def _gastos_empleado_nombre(gasto: "ComprobacionGasto") -> str:
    responsable = (getattr(gasto, "responsable", "") or "").strip()
    if responsable:
        return responsable
    usuario = getattr(gasto, "usuario", None)
    if usuario:
        return (getattr(usuario, "nombre_visible", "") or getattr(usuario, "nombre", "") or "").strip() or "Sin empleado"
    return "Sin empleado"


@app.route("/gastos-viaticos/admin")
@login_required
def gastos_admin_panel():
    gastos = _gastos_admin_query().order_by(ComprobacionGasto.fecha_registro.desc(), ComprobacionGasto.id.desc()).all()
    empleados: dict[str, dict] = {}
    for gasto in gastos:
        nombre = _gastos_empleado_nombre(gasto)
        item = empleados.setdefault(nombre.lower(), {
            "nombre": nombre,
            "total": 0.0,
            "pendiente": 0.0,
            "aprobado": 0.0,
            "rechazado": 0.0,
            "conteo": 0,
            "adjuntos": 0,
            "ultimo": None,
        })
        total = float(gasto.total or 0)
        estatus_gasto = (gasto.estatus or "").upper()
        item["conteo"] += 1
        item["adjuntos"] += len(gasto.adjuntos or [])
        if estatus_gasto == "RECHAZADO":
            item["rechazado"] += total
        else:
            item["total"] += total
            if estatus_gasto in {"PENDIENTE", "EN REVISION"}:
                item["pendiente"] += total
            elif estatus_gasto in {"APROBADO", "REEMBOLSADO"}:
                item["aprobado"] += total
        fecha_base = gasto.fecha_comprobante or gasto.fecha_registro or gasto.creado_en
        if fecha_base and (item["ultimo"] is None or fecha_base > item["ultimo"]):
            item["ultimo"] = fecha_base

    empleados_list = sorted(empleados.values(), key=lambda item: (item["pendiente"], item["total"], item["nombre"]), reverse=True)
    return render_template(
        "gastos_admin_panel.html",
        title="Panel admin de gastos",
        empleados=empleados_list,
        total_general=sum(item["total"] for item in empleados_list),
        total_pendiente=sum(item["pendiente"] for item in empleados_list),
        total_aprobado=sum(item["aprobado"] for item in empleados_list),
        total_comprobantes=sum(item["conteo"] for item in empleados_list),
    )


@app.route("/gastos-viaticos/admin/empleado")
@login_required
def gastos_admin_empleado():
    empleado = (request.args.get("empleado") or "").strip()
    if not empleado:
        abort(404)
    gastos = _gastos_admin_query().order_by(ComprobacionGasto.fecha_registro.desc(), ComprobacionGasto.id.desc()).all()
    gastos = [gasto for gasto in gastos if _gastos_empleado_nombre(gasto).strip().lower() == empleado.lower()]
    if not gastos:
        abort(404)

    total = sum(float(gasto.total or 0) for gasto in gastos if (gasto.estatus or "").upper() != "RECHAZADO")
    pendiente = sum(float(gasto.total or 0) for gasto in gastos if (gasto.estatus or "").upper() in {"PENDIENTE", "EN REVISION"})
    aprobado = sum(float(gasto.total or 0) for gasto in gastos if (gasto.estatus or "").upper() in {"APROBADO", "REEMBOLSADO"})
    rechazado = sum(float(gasto.total or 0) for gasto in gastos if (gasto.estatus or "").upper() == "RECHAZADO")
    return render_template(
        "gastos_admin_empleado.html",
        title=f"Gastos de {empleado}",
        empleado=empleado,
        gastos=gastos,
        total=total,
        pendiente=pendiente,
        aprobado=aprobado,
        rechazado=rechazado,
        gastos_badge_class=_gastos_badge_class,
        fecha_input=_finanzas_fecha_input,
    )


@app.route("/gastos-viaticos/estado-cuenta")
@login_required
def estado_cuenta_recursos_panel():
    if not _estado_cuenta_recursos_can_view():
        abort(403)
    usuarios = _estado_cuenta_recursos_data()
    total_enviado = sum(float(item["enviado"] or 0) for item in usuarios)
    total_comprobado = sum(float(item["comprobado"] or 0) for item in usuarios)
    total_saldo = sum(float(item["saldo"] or 0) for item in usuarios)
    return render_template(
        "estado_cuenta_recursos.html",
        title="Estado de cuenta de recursos",
        usuarios=usuarios,
        total_enviado=total_enviado,
        total_comprobado=total_comprobado,
        total_saldo=total_saldo,
    )


@app.route("/gastos-viaticos/estado-cuenta/<path:user_key>")
@login_required
def estado_cuenta_recursos_detalle(user_key: str):
    if not _estado_cuenta_recursos_can_view():
        abort(403)
    rows = _estado_cuenta_recursos_data(user_key)
    if not rows:
        abort(404)
    usuario = rows[0]
    return render_template(
        "estado_cuenta_recursos_detalle.html",
        title=f"Estado de cuenta {usuario['nombre']}",
        usuario=usuario,
    )


@app.route("/gastos-viaticos/export.xlsx")
@login_required
def gastos_viaticos_export_xlsx():
    if Workbook is None:
        abort(501, description="openpyxl no instalado en el servidor.")

    query, _q, _agrupacion, _estatus = _gastos_query_from_request()
    comprobaciones = query.order_by(ComprobacionGasto.fecha_registro.desc(), ComprobacionGasto.id.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos y viaticos"
    headers = [
        "Folio", "Agrupacion", "Proyecto", "Evento", "Tipo", "Estatus",
        "Quien hizo la compra", "Proveedor", "Concepto(s)", "Referencia",
        "Fecha comprobante", "Subtotal", "IVA", "Total", "Movimiento saldo", "Moneda",
        "Metodo pago", "Notas", "Confianza IA", "Adjuntos",
    ]
    ws.append(headers)
    for gasto in comprobaciones:
        ws.append([
            gasto.folio or "",
            gasto.tipo_agrupacion or "",
            gasto.proyecto or "",
            gasto.evento or "",
            gasto.tipo_gasto or "",
            gasto.estatus or "",
            gasto.responsable or "",
            gasto.proveedor or "",
            gasto.concepto or "",
            gasto.referencia or "",
            gasto.fecha_comprobante.strftime("%Y-%m-%d") if gasto.fecha_comprobante else "",
            float(gasto.subtotal or 0),
            float(gasto.iva or 0),
            float(gasto.total or 0),
            float(_gastos_monto_saldo(gasto)),
            gasto.moneda or "",
            gasto.metodo_pago or "",
            gasto.notas or "",
            float(gasto.ai_confianza or 0),
            len(gasto.adjuntos or []),
        ])

    proyectos_saldo: dict[str, dict] = {}
    for gasto in comprobaciones:
        if (gasto.tipo_agrupacion or "").upper() != "PROYECTO":
            continue
        proyecto_nombre = (gasto.proyecto or "").strip() or "Sin proyecto"
        item = proyectos_saldo.setdefault(proyecto_nombre.lower(), {
            "nombre": proyecto_nombre,
            "recursos": 0.0,
            "gastos": 0.0,
            "pendientes": 0.0,
            "saldo": 0.0,
            "movimientos": 0,
        })
        item["movimientos"] += 1
        item["saldo"] += _gastos_monto_saldo(gasto)
        if _gastos_es_recurso(gasto) and (gasto.estatus or "") != "RECHAZADO":
            item["recursos"] += float(gasto.total or 0)
        elif not _gastos_es_recurso(gasto) and (gasto.estatus or "") != "RECHAZADO":
            item["gastos"] += float(gasto.total or 0)
            if (gasto.estatus or "") in {"PENDIENTE", "EN REVISION"}:
                item["pendientes"] += float(gasto.total or 0)

    ws_proyectos = wb.create_sheet("Saldos por proyecto")
    ws_proyectos.append(["Proyecto", "Dinero aprobado", "Gastos registrados", "Pendiente / revision", "Saldo", "Movimientos"])
    for item in sorted(proyectos_saldo.values(), key=lambda row: row["nombre"].lower()):
        ws_proyectos.append([
            item["nombre"],
            float(item["recursos"] or 0),
            float(item["gastos"] or 0),
            float(item["pendientes"] or 0),
            float(item["saldo"] or 0),
            int(item["movimientos"] or 0),
        ])

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0D6EFD")
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(header) + 3, 14), 28)
    for idx, header in enumerate(["Proyecto", "Dinero aprobado", "Gastos registrados", "Pendiente / revision", "Saldo", "Movimientos"], start=1):
        cell = ws_proyectos.cell(row=1, column=idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0D6EFD")
        ws_proyectos.column_dimensions[get_column_letter(idx)].width = min(max(len(header) + 4, 16), 34)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="gastos_viaticos_{stamp}.xlsx"'},
    )


@app.route("/gastos-viaticos/analizar", methods=["POST"])
@login_required
def gastos_viaticos_analizar():
    uploaded = request.files.get("comprobante")
    if not uploaded or not (uploaded.filename or "").strip():
        return jsonify({"ok": False, "error": "Adjunta una foto, imagen o PDF del comprobante."}), 400
    ext = _gastos_file_ext(uploaded.filename)
    if ext not in GASTOS_UPLOAD_EXTS:
        return jsonify({"ok": False, "error": "Solo se aceptan PDF o imagenes: pdf, png, jpg, jpeg, webp."}), 400

    file_bytes = uploaded.read()
    mime_type = uploaded.mimetype or mimetypes.guess_type(uploaded.filename)[0] or "application/octet-stream"
    text = _gastos_pdf_text(file_bytes) if ext == "pdf" else ""
    try:
        data = _gastos_normalize_ai(_gastos_ai_extract(file_bytes, uploaded.filename, mime_type, text))
    except Exception as exc:
        try:
            logger.exception("Error analizando comprobante")
        except Exception:
            pass
        return jsonify({"ok": False, "error": f"No se pudo analizar el comprobante: {exc}"}), 500
    return jsonify({"ok": True, "data": data, "text_detected": bool(text)})


@app.route("/gastos-viaticos/crear", methods=["POST"])
@login_required
def gastos_viaticos_crear():
    f = request.form
    tipo_agrupacion = (f.get("tipo_agrupacion") or "PROYECTO").strip().upper()
    if tipo_agrupacion not in GASTOS_AGRUPACIONES:
        tipo_agrupacion = "PROYECTO"
    estatus = (f.get("estatus") or "PENDIENTE").strip().upper()
    if estatus not in GASTOS_ESTATUS:
        estatus = "PENDIENTE"

    if tipo_agrupacion == "PROYECTO" and not (f.get("proyecto") or "").strip():
        flash("Captura el proyecto para agrupar la comprobacion.", "warning")
        return _gastos_redirect()
    if tipo_agrupacion == "EVENTO" and not (f.get("evento") or "").strip():
        flash("Captura el evento para agrupar la comprobacion.", "warning")
        return _gastos_redirect()

    conceptos = f.getlist("concepto[]") or [f.get("concepto")]
    proveedores = f.getlist("proveedor[]") or [f.get("proveedor")]
    fechas = f.getlist("fecha_comprobante[]") or [f.get("fecha_comprobante")]
    subtotales = f.getlist("subtotal[]") or [f.get("subtotal")]
    ivas = f.getlist("iva[]") or [f.get("iva")]
    iva_capturado = "iva[]" in f or "iva" in f
    totales = f.getlist("total[]") or [f.get("total")]
    monedas = f.getlist("moneda[]") or [f.get("moneda")]
    metodos_pago = f.getlist("metodo_pago[]") or [f.get("metodo_pago")]
    referencias = f.getlist("referencia[]") or [f.get("referencia")]
    notas_rows = f.getlist("notas[]") or [f.get("notas")]
    tipos_gasto = f.getlist("tipo_gasto[]") or [f.get("tipo_gasto")]
    comprobantes = request.files.getlist("comprobante[]")
    fecha_salida = _parse_date_or_none(f.get("fecha_salida")) or now_cdmx_naive().replace(hour=0, minute=0, second=0, microsecond=0)
    responsable_salida = ((f.get("responsable") or "").strip() if is_admin() else "") or responsable_actual() or None
    proyecto = (f.get("proyecto") or "").strip() or None
    evento = (f.get("evento") or "").strip() or None
    solicitud_recurso_id = int(parse_float(f.get("solicitud_recurso_id"), 0) or 0)
    solicitud_recurso = None
    if solicitud_recurso_id:
        solicitud_recurso = SolicitudRecurso.query.filter_by(id=solicitud_recurso_id, estatus="AUTORIZADA").first()
        if not solicitud_recurso:
            flash("Selecciona una solicitud de recurso autorizada para comprobar.", "warning")
            return _gastos_redirect()
        tipo_agrupacion = "PROYECTO"
        proyecto = (solicitud_recurso.proyecto or "").strip() or proyecto
        estatus = "EN REVISION"

    gastos_creados: list[ComprobacionGasto] = []
    max_rows = max(len(conceptos), len(totales), 1)
    now = now_cdmx_naive()
    for idx in range(max_rows):
        concepto = (conceptos[idx] if idx < len(conceptos) else "").strip()
        subtotal = fmt(parse_float(subtotales[idx] if idx < len(subtotales) else 0, 0))
        iva = fmt(parse_float(ivas[idx] if idx < len(ivas) else 0, 0))
        # Conserva compatibilidad con clientes antiguos que no envian IVA. Si el
        # campo si fue enviado, vacio o cero significan expresamente "sin IVA".
        if subtotal > 0 and not iva_capturado:
            iva = fmt(subtotal * 0.16)
        total = fmt(parse_float(totales[idx] if idx < len(totales) else 0, 0))
        if total <= 0 and (subtotal > 0 or iva > 0):
            total = fmt(subtotal + iva)
        if not concepto and total <= 0:
            continue
        if not concepto or total <= 0:
            db.session.rollback()
            flash(f"Revisa el renglon {idx + 1}: captura concepto y total mayor a cero.", "warning")
            return _gastos_redirect()
        tipo_gasto = ((tipos_gasto[idx] if idx < len(tipos_gasto) else "") or "GASTO").strip().upper()
        if tipo_gasto not in GASTOS_TIPOS:
            tipo_gasto = "GASTO"
        gasto = ComprobacionGasto(
            folio=_gastos_next_folio(),
            tipo_agrupacion=tipo_agrupacion,
            proyecto=proyecto,
            evento=evento,
            tipo_gasto=tipo_gasto,
            estatus=estatus,
            proveedor=((proveedores[idx] if idx < len(proveedores) else "") or "").strip() or None,
            concepto=concepto,
            referencia=((referencias[idx] if idx < len(referencias) else "") or "").strip() or (_solicitud_recurso_gasto_referencia(solicitud_recurso) if solicitud_recurso else None),
            solicitud_recurso_id=solicitud_recurso.id if solicitud_recurso else None,
            fecha_comprobante=_parse_date_or_none(fechas[idx] if idx < len(fechas) else "") or fecha_salida,
            fecha_registro=now,
            subtotal=subtotal,
            iva=iva,
            total=total,
            moneda=(((monedas[idx] if idx < len(monedas) else "") or "MXN").strip().upper()[:10] or "MXN"),
            metodo_pago=((metodos_pago[idx] if idx < len(metodos_pago) else "") or "").strip() or None,
            notas=((notas_rows[idx] if idx < len(notas_rows) else "") or "").strip() or None,
            ai_confianza=0,
            ai_resultado=None,
            responsable=responsable_salida,
            usuario_id=getattr(current_user, "id", None),
        )
        db.session.add(gasto)
        db.session.flush()
        try:
            uploaded = comprobantes[idx] if idx < len(comprobantes) else None
            adjunto = _gastos_save_upload(uploaded, gasto.id)
            if adjunto:
                db.session.add(adjunto)
        except ValueError as exc:
            db.session.rollback()
            flash(f"Renglon {idx + 1}: {exc}", "warning")
            return _gastos_redirect()
        gastos_creados.append(gasto)

    if not gastos_creados:
        flash("Captura al menos un renglon con concepto y total mayor a cero.", "warning")
        return _gastos_redirect()

    db.session.commit()
    notified = 0
    notify_errors: list[str] = []
    for gasto in gastos_creados:
        if _gastos_es_recurso(gasto):
            continue
        try:
            _notify_gasto_created_for_review(gasto)
            notified += 1
        except Exception as exc:
            notify_errors.append(str(exc))
    grupo = proyecto if tipo_agrupacion == "PROYECTO" else evento
    if notify_errors:
        flash(f"Salida '{grupo}' registrada, pero no se pudo notificar revision: {notify_errors[0]}", "warning")
    elif notified:
        flash(f"Comprobacion registrada y notificada por correo y movil.", "success")
    else:
        flash(f"Salida '{grupo}' registrada con {len(gastos_creados)} gasto(s). Quedo lista para enviarse a revision.", "success")
    return _gastos_redirect()


@app.route("/gastos-viaticos/enviar-grupo", methods=["POST"])
@login_required
def gastos_viaticos_enviar_grupo():
    tipo_agrupacion = (request.form.get("tipo_agrupacion") or "").strip().upper()
    grupo = (request.form.get("grupo") or "").strip()
    fecha = (request.form.get("fecha") or "").strip()
    responsable = (request.form.get("responsable") or "").strip()
    gastos = (
        _gastos_group_query(tipo_agrupacion, grupo, fecha, responsable)
        .order_by(ComprobacionGasto.fecha_comprobante.asc(), ComprobacionGasto.id.asc())
        .all()
    )
    if not gastos:
        flash("No hay gastos pendientes en esa salida para enviar a revision.", "info")
        return _gastos_redirect()
    for gasto in gastos:
        gasto.estatus = "EN REVISION"
        gasto.actualizado_en = now_cdmx_naive()
    db.session.commit()
    try:
        _send_gastos_group_review_email(gastos)
        flash(f"Salida '{grupo}' enviada a revision con {len(gastos)} comprobante(s).", "success")
    except Exception as exc:
        try:
            logger.exception("No se pudo enviar correo de revision agrupada de gastos %s", grupo)
        except Exception:
            pass
        flash(f"La salida quedo en revision, pero no se pudo enviar el correo: {exc}", "warning")
    return _gastos_redirect()


@app.route("/gastos-viaticos/grupo")
@login_required
def gastos_viaticos_grupo_detalle():
    tipo_agrupacion = (request.args.get("tipo_agrupacion") or "").strip().upper()
    grupo = (request.args.get("grupo") or "").strip()
    fecha = (request.args.get("fecha") or "").strip()
    responsable = (request.args.get("responsable") or "").strip()
    gastos = (
        _gastos_group_all_query(tipo_agrupacion, grupo, fecha, responsable)
        .order_by(ComprobacionGasto.tipo_gasto.desc(), ComprobacionGasto.fecha_comprobante.asc(), ComprobacionGasto.id.asc())
        .all()
    )
    if not gastos:
        abort(404)

    solicitud = None
    for gasto in gastos:
        if getattr(gasto, "solicitud_recurso", None):
            solicitud = gasto.solicitud_recurso
            break
    if solicitud is None:
        for gasto in gastos:
            if _gastos_es_recurso(gasto) and (gasto.referencia or "").startswith("SR:"):
                solicitud = SolicitudRecurso.query.filter_by(folio=(gasto.referencia or "")[3:]).first()
                break
    if solicitud:
        gastos = (
            _gastos_apply_user_scope(ComprobacionGasto.query)
            .filter(ComprobacionGasto.solicitud_recurso_id == solicitud.id)
            .order_by(ComprobacionGasto.tipo_gasto.desc(), ComprobacionGasto.fecha_comprobante.asc(), ComprobacionGasto.id.asc())
            .all()
        )

    aprobado = float(solicitud.total or 0) if solicitud else sum(float(g.total or 0) for g in gastos if _gastos_es_recurso(g) and (g.estatus or "") != "RECHAZADO")
    comprobado = sum(float(g.total or 0) for g in gastos if not _gastos_es_recurso(g) and (g.estatus or "") != "RECHAZADO")
    pendiente_revision = sum(float(g.total or 0) for g in gastos if not _gastos_es_recurso(g) and (g.estatus or "") in {"PENDIENTE", "EN REVISION"})
    saldo = aprobado - comprobado

    return render_template(
        "gastos_viaticos_grupo.html",
        title=f"Salida {grupo}",
        gastos=gastos,
        grupo=grupo,
        tipo_agrupacion=tipo_agrupacion,
        fecha=fecha,
        responsable=responsable,
        solicitud=solicitud,
        aprobado=aprobado,
        comprobado=comprobado,
        pendiente_revision=pendiente_revision,
        saldo=saldo,
        estatus_options=GASTOS_ESTATUS,
        gasto_tipos=tuple(item for item in GASTOS_TIPOS if item != "RECURSO"),
        gastos_badge_class=_gastos_badge_class,
        gastos_es_recurso=_gastos_es_recurso,
        gastos_monto_saldo=_gastos_monto_saldo,
        fecha_input=_finanzas_fecha_input,
        gastos_can_view_all=is_admin(),
    )


@app.route("/gastos-viaticos/<int:gasto_id>/detalle")
@login_required
def gastos_viaticos_detalle(gasto_id: int):
    gasto = ComprobacionGasto.query.get_or_404(gasto_id)
    require_gasto_owner_or_admin(gasto)
    return render_template(
        "gastos_viaticos_detalle.html",
        title=f"Comprobante {gasto.folio or gasto.id}",
        gasto=gasto,
        public_view=False,
        approve_url=url_for("gastos_viaticos_marcar_aprobado", gasto_id=gasto.id),
        gastos_badge_class=_gastos_badge_class,
    )


@app.route("/gastos-viaticos/<int:gasto_id>/aprobar", methods=["POST"])
@login_required
def gastos_viaticos_marcar_aprobado(gasto_id: int):
    gasto = ComprobacionGasto.query.get_or_404(gasto_id)
    require_gasto_owner_or_admin(gasto)
    anterior = (gasto.estatus or "").strip().upper()
    gasto.estatus = "APROBADO"
    gasto.actualizado_en = now_cdmx_naive()
    db.session.commit()
    if anterior != "APROBADO":
        _notify_gastos_authorized_finanzas([gasto])
    flash(f"Comprobacion {gasto.folio} aprobada.", "success")
    return redirect(url_for("gastos_viaticos_detalle", gasto_id=gasto.id))


@app.route("/gastos-viaticos/revision/<int:gasto_id>")
def gastos_viaticos_revision(gasto_id: int):
    gasto = _gastos_load_from_token(gasto_id, request.args.get("token"), "view")
    token = request.args.get("token") or ""
    return render_template(
        "gastos_viaticos_detalle.html",
        title=f"Comprobante {gasto.folio or gasto.id}",
        gasto=gasto,
        public_view=True,
        approve_url=url_for("gastos_viaticos_revision_aprobar", gasto_id=gasto.id, token=token),
        gastos_badge_class=_gastos_badge_class,
    )


@app.route("/gastos-viaticos/revision/<int:gasto_id>/aprobar")
def gastos_viaticos_revision_aprobar(gasto_id: int):
    gasto = _gastos_load_from_token(gasto_id, request.args.get("token"), "approve")
    anterior = (gasto.estatus or "").strip().upper()
    gasto.estatus = "APROBADO"
    gasto.actualizado_en = now_cdmx_naive()
    db.session.commit()
    if anterior != "APROBADO":
        _notify_gastos_authorized_finanzas([gasto])
    return render_template(
        "gastos_viaticos_detalle.html",
        title=f"Comprobante {gasto.folio or gasto.id}",
        gasto=gasto,
        public_view=True,
        approved_now=True,
        approve_url=url_for("gastos_viaticos_revision_aprobar", gasto_id=gasto.id, token=request.args.get("token") or ""),
        gastos_badge_class=_gastos_badge_class,
    )


@app.route("/gastos-viaticos/revision-grupo")
def gastos_viaticos_revision_grupo():
    token = request.args.get("token") or ""
    gastos = _gastos_load_group_from_token(token, "view")
    if not gastos:
        abort(404)
    return render_template(
        "gastos_viaticos_revision_grupo.html",
        title="Salida de gastos",
        gastos=gastos,
        public_view=True,
        approved_now=False,
        grupo=_gastos_group_name(gastos[0]),
        fecha=_gastos_fecha_base(gastos[0]),
        total_salida=sum(float(gasto.total or 0) for gasto in gastos),
        approve_url=url_for("gastos_viaticos_revision_grupo_aprobar", token=token),
        gastos_badge_class=_gastos_badge_class,
    )


@app.route("/gastos-viaticos/revision-grupo/aprobar")
def gastos_viaticos_revision_grupo_aprobar():
    token = request.args.get("token") or ""
    gastos = _gastos_load_group_from_token(token, "approve")
    if not gastos:
        abort(404)
    now = now_cdmx_naive()
    newly_approved = []
    for gasto in gastos:
        anterior = (gasto.estatus or "").strip().upper()
        gasto.estatus = "APROBADO"
        gasto.actualizado_en = now
        if anterior != "APROBADO":
            newly_approved.append(gasto)
    db.session.commit()
    _notify_gastos_authorized_finanzas(newly_approved)
    return render_template(
        "gastos_viaticos_revision_grupo.html",
        title="Salida de gastos",
        gastos=gastos,
        public_view=True,
        approved_now=True,
        grupo=_gastos_group_name(gastos[0]),
        fecha=_gastos_fecha_base(gastos[0]),
        total_salida=sum(float(gasto.total or 0) for gasto in gastos),
        approve_url=url_for("gastos_viaticos_revision_grupo_aprobar", token=token),
        gastos_badge_class=_gastos_badge_class,
    )


@app.route("/gastos-viaticos/<int:gasto_id>/actualizar", methods=["POST"])
@login_required
def gastos_viaticos_actualizar(gasto_id: int):
    gasto = ComprobacionGasto.query.get_or_404(gasto_id)
    require_gasto_owner_or_admin(gasto)
    f = request.form
    estatus = (f.get("estatus") or gasto.estatus or "PENDIENTE").strip().upper()
    if estatus not in GASTOS_ESTATUS:
        flash("Selecciona un estatus valido.", "warning")
        return _gastos_redirect()

    anterior = (gasto.estatus or "").strip().upper()
    gasto.estatus = estatus
    gasto.proveedor = (f.get("proveedor") or "").strip() or None
    gasto.concepto = (f.get("concepto") or "").strip() or gasto.concepto
    gasto.proyecto = (f.get("proyecto") or "").strip() or None
    gasto.evento = (f.get("evento") or "").strip() or None
    gasto.referencia = (f.get("referencia") or "").strip() or None
    gasto.fecha_comprobante = _parse_date_or_none(f.get("fecha_comprobante"))
    gasto.subtotal = fmt(parse_float(f.get("subtotal"), 0))
    gasto.iva = fmt(parse_float(f.get("iva"), 0))
    gasto.total = fmt(parse_float(f.get("total"), 0))
    gasto.moneda = (f.get("moneda") or "MXN").strip().upper()[:10] or "MXN"
    gasto.metodo_pago = (f.get("metodo_pago") or "").strip() or None
    gasto.notas = (f.get("notas") or "").strip() or None
    if is_admin():
        gasto.responsable = (f.get("responsable") or "").strip() or gasto.responsable
    else:
        gasto.responsable = responsable_actual() or gasto.responsable
    gasto.actualizado_en = now_cdmx_naive()
    db.session.commit()
    if estatus == "APROBADO" and anterior != "APROBADO":
        _notify_gastos_authorized_finanzas([gasto])
    flash(f"Comprobacion {gasto.folio} actualizada.", "success")
    return _gastos_redirect()


@app.route("/gastos-viaticos/<int:gasto_id>/eliminar", methods=["POST"])
@login_required
def gastos_viaticos_eliminar(gasto_id: int):
    gasto = ComprobacionGasto.query.get_or_404(gasto_id)
    require_gasto_owner_or_admin(gasto)
    folio = gasto.folio or f"#{gasto.id}"
    db.session.delete(gasto)
    db.session.commit()
    flash(f"Comprobacion {folio} eliminada.", "success")
    return _gastos_redirect()


def _finanzas_redirect():
    return redirect(request.referrer or url_for("finanzas_index"))


@app.route("/finanzas")
@login_required
def finanzas_index():
    q = (request.args.get("q") or "").strip()
    estatus = (request.args.get("estatus") or "").strip().upper()

    query = MovimientoFinanciero.query.filter(MovimientoFinanciero.categoria == FINANZAS_CATEGORIA_CREDITO)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            MovimientoFinanciero.folio.ilike(like),
            MovimientoFinanciero.contraparte.ilike(like),
            MovimientoFinanciero.concepto.ilike(like),
            MovimientoFinanciero.proyecto.ilike(like),
            MovimientoFinanciero.referencia.ilike(like),
        ))
    if estatus:
        if estatus == "VENCIDO":
            query = query.filter(
                MovimientoFinanciero.saldo > 0,
                MovimientoFinanciero.estatus.notin_(["PAGADO", "CANCELADO"]),
                MovimientoFinanciero.fecha_vencimiento < now_cdmx_naive(),
            )
        else:
            query = query.filter(MovimientoFinanciero.estatus == estatus)

    movimientos = query.order_by(
        MovimientoFinanciero.fecha_vencimiento.is_(None),
        MovimientoFinanciero.fecha_vencimiento.asc(),
        MovimientoFinanciero.fecha.desc(),
        MovimientoFinanciero.id.desc(),
    ).all()

    activos = [
        m for m in MovimientoFinanciero.query.filter_by(categoria=FINANZAS_CATEGORIA_CREDITO).all()
        if _finanzas_estatus_real(m) not in {"PAGADO", "CANCELADO"}
    ]
    saldo_total = sum(float(m.saldo or 0) for m in activos)
    monto_total = sum(float(m.monto or 0) for m in MovimientoFinanciero.query.filter_by(categoria=FINANZAS_CATEGORIA_CREDITO).all())
    pagado_total = sum(_finanzas_pagado(m) for m in MovimientoFinanciero.query.filter_by(categoria=FINANZAS_CATEGORIA_CREDITO).all())
    vencido = sum(float(m.saldo or 0) for m in activos if _finanzas_estatus_real(m) == "VENCIDO")
    por_vencer_30 = sum(
        float(m.saldo or 0)
        for m in activos
        if m.fecha_vencimiento and 0 <= _finanzas_dias_restantes(m) <= 30
    )
    pagos_recientes = (
        MovimientoFinancieroPago.query
        .join(MovimientoFinanciero)
        .filter(MovimientoFinanciero.categoria == FINANZAS_CATEGORIA_CREDITO)
        .order_by(MovimientoFinancieroPago.fecha.desc(), MovimientoFinancieroPago.id.desc())
        .limit(12)
        .all()
    )
    pagos_credito = (
        MovimientoFinancieroPago.query
        .filter(MovimientoFinancieroPago.movimiento_id.in_([m.id for m in movimientos] or [0]))
        .order_by(MovimientoFinancieroPago.fecha.desc(), MovimientoFinancieroPago.id.desc())
        .all()
    )
    pagos_por_credito: dict[int, list[MovimientoFinancieroPago]] = {}
    for pago in pagos_credito:
        pagos_por_credito.setdefault(pago.movimiento_id, []).append(pago)

    return render_template(
        "finanzas.html",
        title="Creditos",
        creditos=movimientos,
        pagos_recientes=pagos_recientes,
        pagos_por_credito=pagos_por_credito,
        estatus_options=FINANZAS_ESTATUS,
        q=q,
        estatus=estatus,
        saldo_total=saldo_total,
        monto_total=monto_total,
        pagado_total=pagado_total,
        vencido=vencido,
        por_vencer_30=por_vencer_30,
        finanzas_estatus_real=_finanzas_estatus_real,
        finanzas_badge_class=_finanzas_badge_class,
        finanzas_category_label=_finanzas_category_label,
        finanzas_fecha_input=_finanzas_fecha_input,
        finanzas_dias_restantes=_finanzas_dias_restantes,
        finanzas_pagado=_finanzas_pagado,
        finanzas_porcentaje_pagado=_finanzas_porcentaje_pagado,
        finanzas_porcentaje_tiempo=_finanzas_porcentaje_tiempo,
        finanzas_tiempo_estado=_finanzas_tiempo_estado,
    )


@app.route("/finanzas/crear", methods=["POST"])
@login_required
def finanzas_crear():
    f = request.form
    contraparte = (f.get("contraparte") or "").strip()
    concepto = (f.get("concepto") or "").strip()
    monto = fmt(parse_float(f.get("monto"), 0))
    if not contraparte or not concepto or monto <= 0:
        flash("Captura quien otorgo el credito, concepto y monto mayor a cero.", "warning")
        return _finanzas_redirect()

    fecha = _parse_date_or_none(f.get("fecha")) or now_cdmx_naive()
    dias_credito = max(0, int(parse_float(f.get("dias_credito"), 0)))
    fecha_vencimiento = _parse_date_or_none(f.get("fecha_vencimiento"))
    if not fecha_vencimiento and dias_credito:
        fecha_vencimiento = fecha + timedelta(days=dias_credito)

    mov = MovimientoFinanciero(
        folio=_finanzas_next_folio(),
        categoria=FINANZAS_CATEGORIA_CREDITO,
        estatus="PENDIENTE",
        contraparte=contraparte,
        concepto=concepto,
        proyecto=(f.get("proyecto") or "").strip() or None,
        referencia=(f.get("referencia") or "").strip() or None,
        fecha=fecha,
        fecha_vencimiento=fecha_vencimiento,
        dias_credito=dias_credito,
        monto=monto,
        saldo=monto,
        moneda=(f.get("moneda") or "MXN").strip().upper()[:10] or "MXN",
        notas=(f.get("notas") or "").strip() or None,
        responsable=responsable_actual() or None,
        usuario_id=getattr(current_user, "id", None),
    )
    db.session.add(mov)
    db.session.commit()
    flash(f"Credito {mov.folio} registrado.", "success")
    return _finanzas_redirect()


@app.route("/finanzas/<int:mov_id>/actualizar", methods=["POST"])
@login_required
def finanzas_actualizar(mov_id: int):
    mov = MovimientoFinanciero.query.get_or_404(mov_id)
    f = request.form
    contraparte = (f.get("contraparte") or "").strip()
    concepto = (f.get("concepto") or "").strip()
    estatus = (f.get("estatus") or "").strip().upper()
    monto = fmt(parse_float(f.get("monto"), 0))
    saldo = fmt(parse_float(f.get("saldo"), 0))

    if estatus not in FINANZAS_ESTATUS:
        flash("Selecciona un estatus valido.", "warning")
        return _finanzas_redirect()
    if not contraparte or not concepto or monto <= 0:
        flash("Captura quien otorgo el credito, concepto y monto mayor a cero.", "warning")
        return _finanzas_redirect()

    fecha = _parse_date_or_none(f.get("fecha")) or mov.fecha or now_cdmx_naive()
    dias_credito = max(0, int(parse_float(f.get("dias_credito"), 0)))
    fecha_vencimiento = _parse_date_or_none(f.get("fecha_vencimiento"))
    if not fecha_vencimiento and dias_credito:
        fecha_vencimiento = fecha + timedelta(days=dias_credito)

    if estatus in {"PAGADO", "CANCELADO"}:
        saldo = 0.0
    else:
        saldo = min(max(0.0, saldo), monto)
        if saldo <= 0:
            estatus = "PAGADO"
        elif saldo < monto and estatus == "PENDIENTE":
            estatus = "PARCIAL"

    mov.categoria = FINANZAS_CATEGORIA_CREDITO
    mov.estatus = estatus
    mov.contraparte = contraparte
    mov.concepto = concepto
    mov.proyecto = (f.get("proyecto") or "").strip() or None
    mov.referencia = (f.get("referencia") or "").strip() or None
    mov.fecha = fecha
    mov.fecha_vencimiento = fecha_vencimiento
    mov.dias_credito = dias_credito
    mov.monto = monto
    mov.saldo = saldo
    mov.moneda = (f.get("moneda") or "MXN").strip().upper()[:10] or "MXN"
    mov.notas = (f.get("notas") or "").strip() or None
    mov.actualizado_en = now_cdmx_naive()
    db.session.commit()
    flash(f"Credito {mov.folio} actualizado.", "success")
    return _finanzas_redirect()


@app.route("/finanzas/<int:mov_id>/eliminar", methods=["POST"])
@login_required
def finanzas_eliminar(mov_id: int):
    mov = MovimientoFinanciero.query.get_or_404(mov_id)
    folio = mov.folio or f"#{mov.id}"
    db.session.delete(mov)
    db.session.commit()
    flash(f"Credito {folio} eliminado.", "success")
    return _finanzas_redirect()


@app.route("/finanzas/<int:mov_id>/abono", methods=["POST"])
@login_required
def finanzas_abono(mov_id: int):
    mov = MovimientoFinanciero.query.get_or_404(mov_id)
    abono = fmt(parse_float(request.form.get("abono"), 0))
    if abono <= 0:
        flash("Captura un abono mayor a cero.", "warning")
        return _finanzas_redirect()
    abono = min(abono, float(mov.saldo or 0))
    fecha_pago = _parse_date_or_none(request.form.get("fecha_pago")) or now_cdmx_naive()
    referencia = (request.form.get("referencia_pago") or "").strip() or None
    nota = (request.form.get("nota_abono") or "").strip() or None
    pago = MovimientoFinancieroPago(
        movimiento=mov,
        fecha=fecha_pago,
        monto=abono,
        referencia=referencia,
        notas=nota,
        responsable=responsable_actual() or None,
        usuario_id=getattr(current_user, "id", None),
    )
    db.session.add(pago)
    mov.saldo = fmt(max(0.0, float(mov.saldo or 0) - abono))
    mov.estatus = "PAGADO" if mov.saldo <= 0 else "PARCIAL"
    mov.actualizado_en = now_cdmx_naive()
    db.session.commit()
    flash("Pago registrado.", "success")
    return _finanzas_redirect()


@app.route("/finanzas/<int:mov_id>/estatus", methods=["POST"])
@login_required
def finanzas_estatus(mov_id: int):
    mov = MovimientoFinanciero.query.get_or_404(mov_id)
    estatus = (request.form.get("estatus") or "").strip().upper()
    if estatus not in FINANZAS_ESTATUS:
        flash("Estatus invalido.", "warning")
        return _finanzas_redirect()
    mov.estatus = estatus
    if estatus == "PAGADO":
        mov.saldo = 0.0
    elif estatus == "CANCELADO":
        mov.saldo = 0.0
    mov.actualizado_en = now_cdmx_naive()
    db.session.commit()
    flash("Estatus del credito actualizado.", "success")
    return _finanzas_redirect()


@app.route("/finanzas/export.xlsx")
@login_required
def finanzas_export_xlsx():
    if Workbook is None:
        abort(501, description="openpyxl no instalado en el servidor.")

    movimientos = (
        MovimientoFinanciero.query
        .filter_by(categoria=FINANZAS_CATEGORIA_CREDITO)
        .order_by(MovimientoFinanciero.fecha.desc(), MovimientoFinanciero.id.desc())
        .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Creditos"
    ws.append([
        "Folio", "Otorgante", "Concepto", "Estatus", "Fecha credito", "Vencimiento",
        "Dias credito", "Dias restantes", "Monto", "Pagado", "Saldo", "Moneda",
        "Referencia", "Responsable", "Notas",
    ])
    for mov in movimientos:
        ws.append([
            mov.folio or "",
            mov.contraparte or "",
            mov.concepto or "",
            _finanzas_estatus_real(mov),
            mov.fecha.strftime("%d/%m/%Y") if mov.fecha else "",
            mov.fecha_vencimiento.strftime("%d/%m/%Y") if mov.fecha_vencimiento else "",
            int(mov.dias_credito or 0),
            _finanzas_dias_restantes(mov),
            float(mov.monto or 0),
            _finanzas_pagado(mov),
            float(mov.saldo or 0),
            mov.moneda or "MXN",
            mov.referencia or "",
            mov.responsable or "",
            mov.notas or "",
        ])
    ws_pagos = wb.create_sheet("Pagos")
    ws_pagos.append(["Credito", "Otorgante", "Fecha pago", "Monto", "Referencia", "Responsable", "Notas"])
    pagos = (
        MovimientoFinancieroPago.query
        .join(MovimientoFinanciero)
        .filter(MovimientoFinanciero.categoria == FINANZAS_CATEGORIA_CREDITO)
        .order_by(MovimientoFinancieroPago.fecha.desc(), MovimientoFinancieroPago.id.desc())
        .all()
    )
    for pago in pagos:
        ws_pagos.append([
            pago.movimiento.folio if pago.movimiento else "",
            pago.movimiento.contraparte if pago.movimiento else "",
            pago.fecha.strftime("%d/%m/%Y") if pago.fecha else "",
            float(pago.monto or 0),
            pago.referencia or "",
            pago.responsable or "",
            pago.notas or "",
        ])
    for sheet in (ws, ws_pagos):
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=MAR_BLUE_XLSX)
            cell.alignment = Alignment(horizontal="center")
    for col in ("I", "J", "K"):
        for cell in ws[col][1:]:
            cell.number_format = '"$"#,##0.00'
    for cell in ws_pagos["D"][1:]:
        cell.number_format = '"$"#,##0.00'
    for idx, width in enumerate([18, 30, 36, 16, 14, 14, 14, 14, 16, 16, 16, 10, 20, 18, 44], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for idx, width in enumerate([18, 30, 14, 16, 24, 18, 44], start=1):
        ws_pagos.column_dimensions[get_column_letter(idx)].width = width

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    stamp = now_cdmx_naive().strftime("%Y%m%d_%H%M%S")
    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="creditos_recibidos_{stamp}.xlsx"'},
    )


def _orden_compra_next_folio() -> str:
    year = now_cdmx_naive().year
    prefix = f"OC-{year}-"
    latest = (
        OrdenCompra.query
        .filter(OrdenCompra.folio.like(f"{prefix}%"))
        .order_by(OrdenCompra.id.desc())
        .first()
    )
    if latest and latest.folio:
        try:
            seq = int(latest.folio.rsplit("-", 1)[-1]) + 1
        except Exception:
            seq = latest.id + 1
    else:
        seq = 1
    return f"{prefix}{seq:04d}"


def _orden_compra_recalcular(orden: OrdenCompra) -> None:
    subtotal = 0.0
    for partida in orden.partidas:
        partida.cantidad = fmt(partida.cantidad or 0)
        partida.precio_unitario = fmt(partida.precio_unitario or 0)
        partida.subtotal = fmt(partida.cantidad * partida.precio_unitario)
        subtotal += partida.subtotal
    orden.subtotal = fmt(subtotal)
    descuento_porc = min(100.0, max(0.0, parse_float(getattr(orden, "descuento_total", 0), 0)))
    orden.descuento_total = fmt(descuento_porc)
    descuento_monto = fmt(orden.subtotal * (descuento_porc / 100.0))
    base_iva = max(0.0, orden.subtotal - descuento_monto)
    orden.iva_porc = fmt(orden.iva_porc if orden.iva_porc is not None else 16.0)
    orden.iva_monto = fmt(base_iva * (orden.iva_porc / 100.0))
    orden.total = fmt(base_iva + orden.iva_monto)
    orden.actualizado_en = now_cdmx_naive()


def _orden_compra_actualizar_estatus_recepcion(orden: OrdenCompra) -> None:
    if orden.estatus in {"CANCELADA", "FACTURADA", "PAGADA"}:
        return
    total_pedido = sum(float(p.cantidad or 0) for p in orden.partidas)
    total_recibido = sum(float(p.cantidad_recibida or 0) for p in orden.partidas)
    if total_pedido <= 0 or total_recibido <= 0:
        return
    orden.estatus = "RECIBIDA COMPLETA" if total_recibido + 0.0001 >= total_pedido else "PARCIALMENTE RECIBIDA"


def _provider_options_for_orders() -> list[str]:
    values = []
    for row in _load_provider_numbers():
        empresa = (row.get("empresa") or "").strip()
        razon = (row.get("razon_social_poliutech") or "").strip()
        if empresa:
            values.append(empresa)
        elif razon:
            values.append(razon)
    return sorted(set(values), key=str.lower)


def _orden_compra_guardar_archivo(uploaded, orden: OrdenCompra, prefijo: str) -> str | None:
    if not uploaded or not (uploaded.filename or "").strip():
        return None
    filename = secure_filename(uploaded.filename or "")
    if not filename or "." not in filename:
        raise ValueError("El archivo debe ser PDF o imagen.")
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext not in ORDEN_COMPRA_UPLOAD_EXTS:
        raise ValueError("Solo se permiten archivos PDF o imagen.")

    upload_dir = Path(app.static_folder or "static") / "uploads" / "ordenes_compra"
    upload_dir.mkdir(parents=True, exist_ok=True)
    stamp = now_cdmx_naive().strftime("%Y%m%d%H%M%S")
    folio = secure_filename(orden.folio or f"oc_{orden.id}")
    saved_name = f"{folio}_{prefijo}_{stamp}.{ext}"
    uploaded.save(upload_dir / saved_name)
    return f"uploads/ordenes_compra/{saved_name}"


def _reportes_diarios_can_view_all() -> bool:
    email = (getattr(current_user, "correo", "") or "").strip().lower()
    nombre = (getattr(current_user, "nombre", "") or "").strip().lower()
    visible = (getattr(current_user, "nombre_visible", "") or "").strip().lower()
    rol = (getattr(current_user, "rol", "") or "").strip().upper()
    return (
        rol == "ADMIN"
        or nombre == "admin"
        or email == "hjaramillo@poliutech.com"
        or nombre in {"hjaramillo", "hansel"}
        or nombre.startswith("hansel")
        or visible.startswith("hansel")
    )


def _evaluacion_departamental_can_view() -> bool:
    """Restringe la evaluacion departamental exclusivamente a Admin y Hansel."""
    if not getattr(current_user, "is_authenticated", False):
        return False

    user_id = getattr(current_user, "id", None)
    email = (getattr(current_user, "correo", "") or "").strip().lower()
    nombre = (getattr(current_user, "nombre", "") or "").strip().lower()
    rol = (getattr(current_user, "rol", "") or "").strip().upper()
    return (
        rol == "ADMIN"
        or nombre == "admin"
        or user_id == 18
        or email == "hjaramillo@poliutech.com"
        or nombre in {"hansel", "hjaramillo"}
    )


def _reportes_diarios_query():
    query = ReporteDiario.query
    if not _reportes_diarios_can_view_all():
        query = query.filter(ReporteDiario.usuario_id == getattr(current_user, "id", None))
    return query


def _reporte_diario_score(reporte: ReporteDiario) -> tuple[int, list[str], list[str]]:
    payload = _reporte_diario_payload(reporte)
    cumplimiento_score = {
        "100%": 100,
        "80-99%": 90,
        "60-79%": 70,
        "MENOR A 60%": 45,
    }.get((reporte.cumplimiento or "").upper(), 0)
    semaforo_score = {
        "SIN INCIDENCIAS": 100,
        "RIESGOS IDENTIFICADOS": 70,
        "REQUIERE INTERVENCION INMEDIATA": 35,
    }.get((reporte.semaforo or "").upper(), 0)

    actividades = payload["actividades"]
    avances = []
    terminadas = 0
    pendientes = 0
    fortalezas = []
    debilidades = []
    for item in actividades:
        estatus = (item.get("estatus") or "").upper()
        if estatus == "TERMINADA":
            terminadas += 1
        if estatus == "PENDIENTE":
            pendientes += 1
            actividad = (item.get("actividad") or "").strip()
            if actividad:
                debilidades.append(f"Pendiente: {actividad}")
        try:
            avance = float(item.get("avance") or 0)
            if avance > 0:
                avances.append(max(0, min(100, avance)))
        except Exception:
            pass

    actividad_score = round(sum(avances) / len(avances)) if avances else (100 if actividades and pendientes == 0 else 0)
    score = round((cumplimiento_score * 0.35) + (actividad_score * 0.4) + (semaforo_score * 0.25))

    if terminadas:
        fortalezas.append(f"{terminadas} actividades terminadas")
    for punto in payload["puntos"]:
        impacto = (punto.get("impacto") or "").strip()
        if impacto:
            fortalezas.append(impacto)
    for riesgo in payload["riesgos"]:
        situacion = (riesgo.get("situacion") or "").strip()
        if situacion:
            debilidades.append(situacion)
    if reporte.apoyo_direccion:
        debilidades.append(f"Apoyo requerido: {reporte.apoyo_direccion.strip()}")
    if (reporte.semaforo or "").upper() == "REQUIERE INTERVENCION INMEDIATA":
        debilidades.append("Requiere intervencion inmediata")

    return score, fortalezas[:5], debilidades[:5]


def _departamento_reporte_diario(reporte: ReporteDiario) -> str:
    puesto = (reporte.puesto or "").strip()
    if puesto:
        return puesto
    return "Sin departamento"


def _evaluacion_nivel(score: int, alertas: int = 0) -> str:
    if alertas >= 2 and score < 80:
        return "Atencion prioritaria"
    if score >= 90:
        return "Sobresaliente"
    if score >= 80:
        return "Fuerte"
    if score >= 70:
        return "Estable con seguimiento"
    if score >= 60:
        return "En riesgo"
    return "Critico"


def _evaluacion_reportes_filtrados():
    fecha_ini_raw = (request.args.get("fecha_ini") or "").strip()
    fecha_fin_raw = (request.args.get("fecha_fin") or "").strip()
    departamento = (request.args.get("departamento") or "").strip()
    colaborador = (request.args.get("colaborador") or "").strip()
    fecha_fin = _parse_date_or_none(fecha_fin_raw) or now_cdmx_naive()
    fecha_ini = _parse_date_or_none(fecha_ini_raw) or (fecha_fin - timedelta(days=30))
    start = fecha_ini.replace(hour=0, minute=0, second=0, microsecond=0)
    end = fecha_fin.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)

    query = ReporteDiario.query.filter(ReporteDiario.fecha >= start, ReporteDiario.fecha < end)
    if departamento:
        query = query.filter(ReporteDiario.puesto.ilike(f"%{departamento}%"))
    if colaborador:
        query = query.filter(ReporteDiario.colaborador.ilike(f"%{colaborador}%"))

    reportes = query.order_by(ReporteDiario.fecha.desc(), ReporteDiario.id.desc()).all()
    evaluacion = _evaluacion_departamentos(reportes)
    empleado_seleccionado = None
    if colaborador:
        colaborador_l = colaborador.lower()
        exactos = [item for item in evaluacion["colaboradores"] if item["nombre"].lower() == colaborador_l]
        empleado_seleccionado = exactos[0] if exactos else (evaluacion["colaboradores"][0] if len(evaluacion["colaboradores"]) == 1 else None)

    return {
        "fecha_ini": start,
        "fecha_fin": end - timedelta(days=1),
        "departamento": departamento,
        "colaborador": colaborador,
        "reportes": reportes,
        "evaluacion": evaluacion,
        "empleado_seleccionado": empleado_seleccionado,
    }


def _evaluacion_departamentos(reportes: list[ReporteDiario]) -> dict:
    departamentos: dict[str, dict] = {}
    colaboradores: dict[str, dict] = {}
    semaforo_counts = {status: 0 for status in REPORTE_DIARIO_SEMAFORO}
    cumplimiento_counts = {status: 0 for status in REPORTE_DIARIO_CUMPLIMIENTO}
    timeline: dict[str, dict[str, int]] = {}

    for reporte in reportes:
        score, fortalezas, debilidades = _reporte_diario_score(reporte)
        departamento = _departamento_reporte_diario(reporte)
        dep = departamentos.setdefault(departamento, {
            "nombre": departamento,
            "reportes": 0,
            "score_total": 0,
            "riesgos": 0,
            "intervenciones": 0,
            "fortalezas": {},
            "debilidades": {},
            "colaboradores": set(),
        })
        dep["reportes"] += 1
        dep["score_total"] += score
        dep["colaboradores"].add(reporte.colaborador or "Sin colaborador")
        if (reporte.semaforo or "").upper() == "RIESGOS IDENTIFICADOS":
            dep["riesgos"] += 1
        if (reporte.semaforo or "").upper() == "REQUIERE INTERVENCION INMEDIATA":
            dep["intervenciones"] += 1
        for item in fortalezas:
            dep["fortalezas"][item] = dep["fortalezas"].get(item, 0) + 1
        for item in debilidades:
            dep["debilidades"][item] = dep["debilidades"].get(item, 0) + 1

        col_key = (reporte.colaborador or "Sin colaborador").strip()
        col = colaboradores.setdefault(col_key, {
            "nombre": col_key,
            "departamento": departamento,
            "reportes": 0,
            "score_total": 0,
            "ultimo_reporte": reporte,
            "alertas": 0,
            "fortalezas": {},
            "debilidades": {},
            "semaforo_counts": {status: 0 for status in REPORTE_DIARIO_SEMAFORO},
            "cumplimiento_counts": {status: 0 for status in REPORTE_DIARIO_CUMPLIMIENTO},
            "reportes_rows": [],
        })
        col["reportes"] += 1
        col["score_total"] += score
        if reporte.fecha and (not col["ultimo_reporte"].fecha or reporte.fecha > col["ultimo_reporte"].fecha):
            col["ultimo_reporte"] = reporte
        if (reporte.semaforo or "").upper() != "SIN INCIDENCIAS":
            col["alertas"] += 1
        for item in fortalezas:
            col["fortalezas"][item] = col["fortalezas"].get(item, 0) + 1
        for item in debilidades:
            col["debilidades"][item] = col["debilidades"].get(item, 0) + 1
        col["reportes_rows"].append({
            "id": reporte.id,
            "folio": reporte.folio,
            "fecha": reporte.fecha,
            "score": score,
            "cumplimiento": reporte.cumplimiento,
            "semaforo": reporte.semaforo,
        })

        semaforo = (reporte.semaforo or "SIN INCIDENCIAS").upper()
        if semaforo in semaforo_counts:
            semaforo_counts[semaforo] += 1
        if semaforo in col["semaforo_counts"]:
            col["semaforo_counts"][semaforo] += 1
        cumplimiento = (reporte.cumplimiento or "").upper()
        if cumplimiento in cumplimiento_counts:
            cumplimiento_counts[cumplimiento] += 1
        if cumplimiento in col["cumplimiento_counts"]:
            col["cumplimiento_counts"][cumplimiento] += 1
        if reporte.fecha:
            day = reporte.fecha.strftime("%Y-%m-%d")
            timeline.setdefault(day, {status: 0 for status in REPORTE_DIARIO_SEMAFORO})
            if semaforo in timeline[day]:
                timeline[day][semaforo] += 1

    for dep in departamentos.values():
        dep["score"] = round(dep["score_total"] / dep["reportes"]) if dep["reportes"] else 0
        dep["nivel"] = _evaluacion_nivel(dep["score"], dep["riesgos"] + dep["intervenciones"])
        dep["colaboradores_count"] = len(dep["colaboradores"])
        dep["fortalezas_top"] = sorted(dep["fortalezas"].items(), key=lambda item: item[1], reverse=True)[:4]
        dep["debilidades_top"] = sorted(dep["debilidades"].items(), key=lambda item: item[1], reverse=True)[:4]

    for col in colaboradores.values():
        col["score"] = round(col["score_total"] / col["reportes"]) if col["reportes"] else 0
        col["nivel"] = _evaluacion_nivel(col["score"], col["alertas"])
        col["fortalezas_top"] = sorted(col["fortalezas"].items(), key=lambda item: item[1], reverse=True)[:5]
        col["debilidades_top"] = sorted(col["debilidades"].items(), key=lambda item: item[1], reverse=True)[:5]
        col["reportes_rows"] = sorted(col["reportes_rows"], key=lambda item: item["fecha"] or datetime.min)
        col["chart_labels"] = [
            (item["fecha"].strftime("%d/%m") if item["fecha"] else item["folio"] or "")
            for item in col["reportes_rows"]
        ]
        col["chart_scores"] = [item["score"] for item in col["reportes_rows"]]

    dept_list = sorted(departamentos.values(), key=lambda item: item["score"], reverse=True)
    col_list = sorted(colaboradores.values(), key=lambda item: (item["alertas"], -item["score"]), reverse=True)
    return {
        "departamentos": dept_list,
        "colaboradores": col_list,
        "semaforo_counts": semaforo_counts,
        "cumplimiento_counts": cumplimiento_counts,
        "timeline": [{"fecha": day, **values} for day, values in sorted(timeline.items())],
    }


@app.route("/reportes-diarios/evaluacion")
@login_required
def reportes_diarios_evaluacion():
    if not _evaluacion_departamental_can_view():
        abort(403)

    contexto = _evaluacion_reportes_filtrados()
    reportes = contexto["reportes"]
    evaluacion = contexto["evaluacion"]
    empleado_seleccionado = contexto["empleado_seleccionado"]
    departamento = contexto["departamento"]
    colaborador = contexto["colaborador"]
    departamentos_options = [
        row[0] or "Sin departamento"
        for row in db.session.query(ReporteDiario.puesto).distinct().order_by(ReporteDiario.puesto.asc()).all()
    ]
    colaboradores_options = [
        row[0]
        for row in db.session.query(ReporteDiario.colaborador).distinct().order_by(ReporteDiario.colaborador.asc()).all()
        if row[0]
    ]

    return render_template(
        "reportes_diarios_evaluacion.html",
        title="Evaluacion departamental",
        reportes=reportes,
        evaluacion=evaluacion,
        fecha_ini=contexto["fecha_ini"].strftime("%Y-%m-%d"),
        fecha_fin=contexto["fecha_fin"].strftime("%Y-%m-%d"),
        departamento=departamento,
        colaborador=colaborador,
        empleado_seleccionado=empleado_seleccionado,
        evaluacion_nivel=_evaluacion_nivel,
        departamentos_options=departamentos_options,
        colaboradores_options=colaboradores_options,
        semaforo_options=REPORTE_DIARIO_SEMAFORO,
        cumplimiento_options=REPORTE_DIARIO_CUMPLIMIENTO,
    )


@app.route("/reportes-diarios/evaluacion/export.xlsx")
@login_required
def reportes_diarios_evaluacion_export_xlsx():
    if not _evaluacion_departamental_can_view():
        abort(403)
    if Workbook is None:
        abort(501, description="openpyxl no instalado en el servidor.")

    contexto = _evaluacion_reportes_filtrados()
    reportes = contexto["reportes"]
    evaluacion = contexto["evaluacion"]
    empleado = contexto["empleado_seleccionado"]
    fecha_ini = contexto["fecha_ini"].strftime("%d/%m/%Y")
    fecha_fin = contexto["fecha_fin"].strftime("%d/%m/%Y")
    vista = empleado["nombre"] if empleado else "Global"
    total_reportes = len(reportes)
    total_departamentos = len(evaluacion["departamentos"])
    promedio = round(sum(dep["score"] for dep in evaluacion["departamentos"]) / total_departamentos) if total_departamentos else 0
    score_vista = empleado["score"] if empleado else promedio
    alertas = evaluacion["semaforo_counts"].get("RIESGOS IDENTIFICADOS", 0) + evaluacion["semaforo_counts"].get("REQUIERE INTERVENCION INMEDIATA", 0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    header_fill = PatternFill("solid", fgColor=MAR_BLUE_XLSX)
    light_fill = PatternFill("solid", fgColor="EAF1FB")
    danger_fill = PatternFill("solid", fgColor="FCE4E4")
    border = Border(
        left=Side(style="thin", color="D9E2EF"),
        right=Side(style="thin", color="D9E2EF"),
        top=Side(style="thin", color="D9E2EF"),
        bottom=Side(style="thin", color="D9E2EF"),
    )

    def style_header(row_idx: int):
        for cell in ws[row_idx]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    def append_kv(sheet, key: str, value: object):
        sheet.append([key, value])
        sheet.cell(sheet.max_row, 1).font = Font(bold=True)
        sheet.cell(sheet.max_row, 1).fill = light_fill
        for cell in sheet[sheet.max_row]:
            cell.border = border

    ws.append(["Reporte de evaluacion por reportes diarios"])
    ws["A1"].font = Font(bold=True, size=16, color=MAR_BLUE_XLSX)
    ws.merge_cells("A1:D1")
    ws.append([])
    append_kv(ws, "Periodo", f"{fecha_ini} - {fecha_fin}")
    append_kv(ws, "Vista", vista)
    append_kv(ws, "Departamento / puesto", contexto["departamento"] or "Todos")
    append_kv(ws, "Reportes analizados", total_reportes)
    append_kv(ws, "Score", f"{score_vista}%")
    append_kv(ws, "Nivel", _evaluacion_nivel(score_vista, alertas if not empleado else empleado["alertas"]))
    append_kv(ws, "Alertas", alertas if not empleado else empleado["alertas"])
    ws.append([])
    ws.append(["Criterio", "Peso", "Como se interpreta"])
    style_header(ws.max_row)
    criterios = [
        ["Cumplimiento del dia", "35%", "Indicador declarado en el reporte: 100%, 80-99%, 60-79% o menor a 60%."],
        ["Avance de actividades", "40%", "Promedio de avances capturados y actividades terminadas/pendientes."],
        ["Riesgo operativo", "25%", "Semaforo del reporte: sin incidencias, riesgos identificados o intervencion inmediata."],
        ["Fortalezas", "Cualitativo", "Impactos positivos, resultados obtenidos y actividades terminadas recurrentes."],
        ["Debilidades", "Cualitativo", "Pendientes, riesgos, apoyos requeridos y alertas recurrentes."],
    ]
    for row in criterios:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_dep = wb.create_sheet("Departamentos")
    ws_dep.append(["Departamento", "Reportes", "Colaboradores", "Score", "Nivel", "Riesgos", "Intervenciones", "Fortalezas", "Debilidades"])
    for cell in ws_dep[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
    for dep in evaluacion["departamentos"]:
        ws_dep.append([
            dep["nombre"],
            dep["reportes"],
            dep["colaboradores_count"],
            dep["score"],
            dep["nivel"],
            dep["riesgos"],
            dep["intervenciones"],
            "; ".join(item for item, count in dep["fortalezas_top"]),
            "; ".join(item for item, count in dep["debilidades_top"]),
        ])
        for cell in ws_dep[ws_dep.max_row]:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_emp = wb.create_sheet("Empleados")
    ws_emp.append(["Empleado", "Departamento", "Reportes", "Score", "Nivel", "Alertas", "Fortalezas", "Debilidades", "Ultimo reporte"])
    for cell in ws_emp[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
    for col in evaluacion["colaboradores"]:
        ultimo = col["ultimo_reporte"]
        ws_emp.append([
            col["nombre"],
            col["departamento"],
            col["reportes"],
            col["score"],
            col["nivel"],
            col["alertas"],
            "; ".join(item for item, count in col["fortalezas_top"]),
            "; ".join(item for item, count in col["debilidades_top"]),
            f"{ultimo.folio or ultimo.id} - {ultimo.fecha.strftime('%d/%m/%Y') if ultimo.fecha else ''}",
        ])
        for cell in ws_emp[ws_emp.max_row]:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if col["alertas"]:
            ws_emp.cell(ws_emp.max_row, 6).fill = danger_fill

    ws_rep = wb.create_sheet("Reportes")
    ws_rep.append(["Fecha", "Folio", "Empleado", "Departamento", "Score", "Nivel", "Cumplimiento", "Semaforo", "Fortalezas", "Debilidades"])
    for cell in ws_rep[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
    for reporte in reportes:
        score, fortalezas, debilidades = _reporte_diario_score(reporte)
        alerta_reporte = 1 if (reporte.semaforo or "").upper() != "SIN INCIDENCIAS" else 0
        ws_rep.append([
            reporte.fecha.strftime("%d/%m/%Y") if reporte.fecha else "",
            reporte.folio or reporte.id,
            reporte.colaborador or "",
            _departamento_reporte_diario(reporte),
            score,
            _evaluacion_nivel(score, alerta_reporte),
            reporte.cumplimiento or "",
            reporte.semaforo or "",
            "; ".join(fortalezas),
            "; ".join(debilidades),
        ])
        for cell in ws_rep[ws_rep.max_row]:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if alerta_reporte:
            ws_rep.cell(ws_rep.max_row, 8).fill = danger_fill

    for sheet in wb.worksheets:
        widths = {
            1: 24,
            2: 22,
            3: 16,
            4: 18,
            5: 16,
            6: 16,
            7: 34,
            8: 34,
            9: 34,
            10: 34,
        }
        for col_idx, width in widths.items():
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        sheet.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    stamp = now_cdmx_naive().strftime("%Y%m%d_%H%M%S")
    filename = f"evaluacion_reportes_diarios_{stamp}.xlsx"
    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.route("/reportes-diarios")
@login_required
def reportes_diarios_index():
    q = (request.args.get("q") or "").strip()
    semaforo = (request.args.get("semaforo") or "").strip().upper()
    fecha_raw = (request.args.get("fecha") or "").strip()
    fecha = _parse_date_or_none(fecha_raw)

    query = _reportes_diarios_query()
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            ReporteDiario.folio.ilike(like),
            ReporteDiario.colaborador.ilike(like),
            ReporteDiario.puesto.ilike(like),
            ReporteDiario.actividades_json.ilike(like),
            ReporteDiario.observaciones.ilike(like),
        ))
    if semaforo in REPORTE_DIARIO_SEMAFORO:
        query = query.filter(ReporteDiario.semaforo == semaforo)
    if fecha:
        start = fecha.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=1)
        query = query.filter(ReporteDiario.fecha >= start, ReporteDiario.fecha < end)

    reportes = query.order_by(ReporteDiario.fecha.desc(), ReporteDiario.hora_envio.desc(), ReporteDiario.id.desc()).all()
    borradores = [reporte for reporte in reportes if reporte.estatus == "BORRADOR"]
    reportes_enviados = [reporte for reporte in reportes if reporte.estatus != "BORRADOR"]
    kanban = {status: [] for status in REPORTE_DIARIO_SEMAFORO}
    for reporte in reportes_enviados:
        kanban.setdefault(reporte.semaforo or "SIN INCIDENCIAS", []).append(reporte)

    editar_id = request.args.get("editar", type=int)
    borrador_edicion = None
    borrador_payload = {"actividades": [], "puntos": [], "prioridades": [], "tiempos": [], "riesgos": []}
    if editar_id:
        borrador_edicion = ReporteDiario.query.filter(
            ReporteDiario.id == editar_id,
            ReporteDiario.usuario_id == getattr(current_user, "id", None),
            ReporteDiario.estatus == "BORRADOR",
        ).first_or_404()
        borrador_payload = _reporte_diario_payload(borrador_edicion)

    hoy = now_cdmx_naive().replace(hour=0, minute=0, second=0, microsecond=0)
    ya_envio_hoy = ReporteDiario.query.filter(
        ReporteDiario.usuario_id == getattr(current_user, "id", None),
        ReporteDiario.estatus == "ENVIADO",
        ReporteDiario.fecha >= hoy,
        ReporteDiario.fecha < hoy + timedelta(days=1),
    ).first()

    return render_template(
        "reportes_diarios.html",
        title="Reportes diarios",
        reportes=reportes,
        borradores=borradores,
        kanban=kanban,
        q=q,
        semaforo=semaforo,
        fecha=fecha_raw,
        semaforo_options=REPORTE_DIARIO_SEMAFORO,
        cumplimiento_options=REPORTE_DIARIO_CUMPLIMIENTO,
        actividad_estatus_options=REPORTE_DIARIO_ACTIVIDAD_ESTATUS,
        fecha_hoy=now_cdmx_naive().strftime("%Y-%m-%d"),
        responsable_default=responsable_actual() or "",
        ya_envio_hoy=ya_envio_hoy,
        borrador_edicion=borrador_edicion,
        borrador_payload=borrador_payload,
        can_view_all=_reportes_diarios_can_view_all(),
    )


@app.route("/reportes-diarios/crear", methods=["POST"])
@login_required
def reporte_diario_crear():
    accion = (request.form.get("accion") or "enviar").strip().lower()
    guardar_borrador = accion == "guardar"
    reporte_id = request.form.get("reporte_id", type=int)
    reporte = _reporte_diario_from_form(request.form)
    if not reporte.colaborador:
        message = "Captura el colaborador del reporte."
        if request.accept_mimetypes.best == "application/json":
            return jsonify({"ok": False, "message": message}), 400
        flash(message, "warning")
        return redirect(url_for("reportes_diarios_index"))
    if not guardar_borrador and not _json_loads_list(reporte.actividades_json):
        message = "Agrega al menos una actividad realizada."
        if request.accept_mimetypes.best == "application/json":
            return jsonify({"ok": False, "message": message}), 400
        flash(message, "warning")
        return redirect(url_for("reportes_diarios_index"))

    start = reporte.fecha.replace(hour=0, minute=0, second=0, microsecond=0)
    if reporte_id:
        existing = ReporteDiario.query.filter(
            ReporteDiario.id == reporte_id,
            ReporteDiario.usuario_id == getattr(current_user, "id", None),
            ReporteDiario.estatus == "BORRADOR",
        ).first_or_404()
    else:
        existing = ReporteDiario.query.filter(
            ReporteDiario.usuario_id == getattr(current_user, "id", None),
            ReporteDiario.fecha >= start,
            ReporteDiario.fecha < start + timedelta(days=1),
        ).first()
    if existing:
        if existing.estatus == "ENVIADO":
            message = f"El reporte {existing.folio} de esa fecha ya fue enviado."
            if request.accept_mimetypes.best == "application/json":
                return jsonify({"ok": False, "message": message}), 409
            flash(message, "warning")
            return redirect(url_for("reporte_diario_detalle", reporte_id=existing.id))
        for field in (
            "colaborador", "puesto", "fecha", "cumplimiento", "semaforo",
            "actividades_json", "puntos_importantes_json", "prioridades_siguientes_json",
            "tiempos_json", "problemas_riesgos_json", "apoyo_direccion", "observaciones",
        ):
            setattr(existing, field, getattr(reporte, field))
        reporte = existing
    else:
        db.session.add(reporte)

    reporte.estatus = "BORRADOR" if guardar_borrador else "ENVIADO"
    if not guardar_borrador:
        reporte.hora_envio = now_cdmx_naive()
    db.session.commit()

    if guardar_borrador:
        message = f"Borrador {reporte.folio} guardado correctamente."
        if request.accept_mimetypes.best == "application/json":
            return jsonify({"ok": True, "message": message, "reporte_id": reporte.id})
        flash(message, "success")
        return redirect(url_for("reportes_diarios_index"))

    _notify_reporte_diario_created(reporte)
    flash(f"Reporte {reporte.folio} enviado correctamente.", "success")
    return redirect(url_for("reporte_diario_detalle", reporte_id=reporte.id))


@app.route("/reportes-diarios/<int:reporte_id>")
@login_required
def reporte_diario_detalle(reporte_id: int):
    reporte = _reportes_diarios_query().filter(ReporteDiario.id == reporte_id).first_or_404()
    return render_template(
        "reporte_diario_detalle.html",
        title=f"Reporte {reporte.folio}",
        reporte=reporte,
        payload=_reporte_diario_payload(reporte),
    )


@app.route("/reportes-diarios/<int:reporte_id>/eliminar", methods=["POST"])
@login_required
def reporte_diario_eliminar(reporte_id: int):
    reporte = _reportes_diarios_query().filter(ReporteDiario.id == reporte_id).first_or_404()
    if not (_reportes_diarios_can_view_all() or reporte.usuario_id == getattr(current_user, "id", None)):
        abort(403)
    folio = reporte.folio or f"#{reporte.id}"
    db.session.delete(reporte)
    db.session.commit()
    flash(f"Reporte {folio} eliminado.", "success")
    return redirect(url_for("reportes_diarios_index"))


@app.route("/solicitudes-recursos")
@login_required
def solicitudes_recursos_index():
    q = (request.args.get("q") or "").strip()
    estatus = (request.args.get("estatus") or "").strip().upper()

    query = SolicitudRecurso.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            SolicitudRecurso.folio.ilike(like),
            SolicitudRecurso.solicitante.ilike(like),
            SolicitudRecurso.proyecto.ilike(like),
            SolicitudRecurso.notas.ilike(like),
            SolicitudRecurso.partidas.any(SolicitudRecursoPartida.concepto.ilike(like)),
        ))
    if estatus:
        query = query.filter(SolicitudRecurso.estatus == estatus)

    solicitudes = query.order_by(SolicitudRecurso.fecha.desc(), SolicitudRecurso.id.desc()).all()
    status_counts = {status: SolicitudRecurso.query.filter_by(estatus=status).count() for status in SOLICITUD_RECURSO_ESTATUS}
    total_visible = fmt(sum(float(s.total or 0) for s in solicitudes))

    return render_template(
        "solicitudes_recursos.html",
        title="Solicitudes de recursos",
        solicitudes=solicitudes,
        estatus_options=SOLICITUD_RECURSO_ESTATUS,
        status_counts=status_counts,
        total_visible=total_visible,
        q=q,
        estatus=estatus,
        project_options=_known_project_names(),
    )


@app.route("/solicitudes-recursos/crear", methods=["POST"])
@login_required
def solicitud_recurso_crear():
    f = request.form
    proyecto = (f.get("proyecto") or "").strip()
    if not proyecto:
        flash("Selecciona o captura el proyecto para agrupar la solicitud de recursos.", "warning")
        return redirect(url_for("solicitudes_recursos_index"))

    solicitud = SolicitudRecurso(
        folio=_solicitud_recurso_next_folio(),
        fecha=now_cdmx_naive(),
        solicitante=(f.get("solicitante") or responsable_actual() or "").strip() or None,
        proyecto=proyecto,
        estatus="SOLICITADA",
        notas=(f.get("notas") or "").strip() or None,
        usuario_id=getattr(current_user, "id", None),
    )
    db.session.add(solicitud)

    cantidades = f.getlist("cantidad[]")
    conceptos = f.getlist("descripcion[]") or f.getlist("concepto[]")
    importes = f.getlist("importe[]")
    total_rows = max(len(cantidades), len(conceptos), len(importes))
    for idx in range(total_rows):
        concepto = (conceptos[idx] if idx < len(conceptos) else "").strip()
        cantidad = parse_float(cantidades[idx] if idx < len(cantidades) else 0, 0)
        importe = parse_float(importes[idx] if idx < len(importes) else 0, 0)
        if not concepto or cantidad <= 0:
            continue
        total_partida = fmt(cantidad * importe)
        solicitud.partidas.append(SolicitudRecursoPartida(
            cantidad=fmt(cantidad),
            concepto=concepto,
            importe=fmt(importe),
            total=total_partida,
        ))

    if not solicitud.partidas:
        db.session.rollback()
        flash("Agrega al menos un renglon con cantidad y descripcion.", "warning")
        return redirect(url_for("solicitudes_recursos_index"))

    _solicitud_recurso_recalcular(solicitud)
    db.session.commit()
    _notify_solicitud_recurso_created(solicitud)
    flash(f"Solicitud {solicitud.folio} registrada.", "success")
    return redirect(url_for("solicitud_recurso_detalle", solicitud_id=solicitud.id))


@app.route("/solicitudes-recursos/<int:solicitud_id>")
@login_required
def solicitud_recurso_detalle(solicitud_id: int):
    solicitud = SolicitudRecurso.query.get_or_404(solicitud_id)
    return render_template(
        "solicitud_recurso_detalle.html",
        title=f"Solicitud {solicitud.folio}",
        solicitud=solicitud,
        cliente_info=_cliente_seguimiento_payload(
            nombre=solicitud.solicitante,
            responsable=solicitud.solicitante,
            titulo="Datos del cliente / solicitud",
            extras=[{"label": "Proyecto / obra", "value": solicitud.proyecto}],
        ),
        estatus_options=SOLICITUD_RECURSO_ESTATUS,
    )


@app.route("/solicitudes-recursos/<int:solicitud_id>/eliminar", methods=["POST"])
@login_required
def solicitud_recurso_eliminar(solicitud_id: int):
    solicitud = SolicitudRecurso.query.get_or_404(solicitud_id)
    folio = solicitud.folio or str(solicitud.id)
    db.session.delete(solicitud)
    db.session.commit()
    flash(f"Solicitud {folio} eliminada.", "success")
    return redirect(url_for("solicitudes_recursos_index"))


@app.route("/solicitudes-recursos/<int:solicitud_id>/estatus", methods=["POST"])
@login_required
def solicitud_recurso_estatus(solicitud_id: int):
    solicitud = SolicitudRecurso.query.get_or_404(solicitud_id)
    nuevo = (request.form.get("estatus") or "").strip().upper()
    if nuevo not in SOLICITUD_RECURSO_ESTATUS:
        flash("Estatus no valido.", "warning")
        return redirect(url_for("solicitud_recurso_detalle", solicitud_id=solicitud.id))
    if nuevo == "AUTORIZADA" and not (solicitud.proyecto or "").strip():
        flash("La solicitud necesita proyecto para registrarse automaticamente en gastos.", "warning")
        return redirect(url_for("solicitud_recurso_detalle", solicitud_id=solicitud.id))
    anterior = (solicitud.estatus or "").strip().upper()
    solicitud.estatus = nuevo
    solicitud.actualizado_en = now_cdmx_naive()
    gasto = _solicitud_recurso_registrar_gasto(solicitud) if nuevo == "AUTORIZADA" else None
    db.session.commit()
    if nuevo in {"AUTORIZADA", "RECHAZADA"} and anterior != nuevo:
        _notify_solicitud_recurso_resultado(solicitud)
    if nuevo == "AUTORIZADA" and anterior != nuevo:
        _notify_solicitud_recurso_autorizada_finanzas(solicitud)
    if gasto:
        flash(f"Estatus actualizado. Se registro en gastos como {gasto.folio}.", "success")
    else:
        flash("Estatus actualizado.", "success")
    return redirect(url_for("solicitud_recurso_detalle", solicitud_id=solicitud.id))


@app.route("/ordenes-compra")
@login_required
def ordenes_compra_index():
    q = (request.args.get("q") or "").strip()
    estatus = (request.args.get("estatus") or "").strip().upper()
    proveedor = (request.args.get("proveedor") or "").strip()

    query = OrdenCompra.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            OrdenCompra.folio.ilike(like),
            OrdenCompra.proveedor.ilike(like),
            OrdenCompra.notas.ilike(like),
        ))
    if estatus:
        query = query.filter(OrdenCompra.estatus == estatus)
    if proveedor:
        query = query.filter(OrdenCompra.proveedor == proveedor)

    ordenes = query.order_by(OrdenCompra.fecha.desc(), OrdenCompra.id.desc()).all()
    productos = InventarioProducto.query.filter_by(activo=True).order_by(InventarioProducto.nombre.asc()).all()
    proveedor_options = _provider_options_for_orders()
    status_counts = {status: OrdenCompra.query.filter_by(estatus=status).count() for status in ORDEN_COMPRA_ESTATUS}

    return render_template(
        "ordenes_compra.html",
        title="Ordenes de compra",
        ordenes=ordenes,
        productos=productos,
        proveedor_options=proveedor_options,
        estatus_options=ORDEN_COMPRA_ESTATUS,
        status_counts=status_counts,
        q=q,
        estatus=estatus,
        proveedor=proveedor,
    )


@app.route("/ordenes-compra/crear", methods=["POST"])
@login_required
def orden_compra_crear():
    f = request.form
    proveedor = (f.get("proveedor") or "").strip()
    if not proveedor:
        flash("Captura el proveedor de la orden.", "warning")
        return redirect(url_for("ordenes_compra_index"))

    orden = OrdenCompra(
        folio=_orden_compra_next_folio(),
        proveedor=proveedor,
        numero_cliente_proveedor=(f.get("numero_cliente_proveedor") or "").strip() or None,
        contacto=(f.get("contacto") or "").strip() or None,
        telefono=(f.get("telefono") or "").strip() or None,
        correo=(f.get("correo") or "").strip() or None,
        fecha=now_cdmx_naive(),
        fecha_entrega=_parse_date_or_none(f.get("fecha_entrega")),
        forma_pago=(f.get("forma_pago") or "CONTADO").strip().upper(),
        estatus="BORRADOR",
        descuento_total=fmt(parse_float(f.get("descuento_total"), 0)),
        iva_porc=fmt(parse_float(f.get("iva_porc"), 16.0)),
        condiciones=(f.get("condiciones") or "").strip() or None,
        notas=(f.get("notas") or "").strip() or None,
        responsable=responsable_actual() or None,
        usuario_id=getattr(current_user, "id", None),
    )
    db.session.add(orden)

    producto_ids = f.getlist("producto_id[]")
    descripciones = f.getlist("descripcion[]")
    unidades = f.getlist("unidad[]")
    cantidades = f.getlist("cantidad[]")
    precios = f.getlist("precio_unitario[]")
    observaciones = f.getlist("observaciones[]")

    total_rows = max(len(producto_ids), len(descripciones), len(cantidades), len(precios))
    for idx in range(total_rows):
        producto = None
        producto_id = producto_ids[idx] if idx < len(producto_ids) else ""
        if producto_id:
            producto = InventarioProducto.query.get(int(producto_id))

        descripcion = (descripciones[idx] if idx < len(descripciones) else "").strip()
        unidad = (unidades[idx] if idx < len(unidades) else "").strip()
        cantidad = parse_float(cantidades[idx] if idx < len(cantidades) else 0, 0)
        precio = parse_float(precios[idx] if idx < len(precios) else 0, 0)
        if producto:
            descripcion = descripcion or producto.nombre
            unidad = unidad or producto.unidad or "pieza"
            if precio <= 0:
                precio = float(producto.costo_unitario or 0)
        if not descripcion or cantidad <= 0:
            continue

        orden.partidas.append(OrdenCompraPartida(
            inventario_producto_id=producto.id if producto else None,
            codigo=(producto.codigo if producto else None),
            descripcion=descripcion,
            unidad=unidad or "pieza",
            cantidad=fmt(cantidad),
            cantidad_recibida=0.0,
            precio_unitario=fmt(precio),
            observaciones=(observaciones[idx] if idx < len(observaciones) else "").strip() or None,
        ))

    if not orden.partidas:
        db.session.rollback()
        flash("Agrega al menos una partida con cantidad mayor a cero.", "warning")
        return redirect(url_for("ordenes_compra_index"))

    _orden_compra_recalcular(orden)
    db.session.commit()
    flash(f"Orden {orden.folio} creada.", "success")
    return redirect(url_for("orden_compra_detalle", orden_id=orden.id))


@app.route("/ordenes-compra/<int:orden_id>")
@login_required
def orden_compra_detalle(orden_id: int):
    orden = OrdenCompra.query.get_or_404(orden_id)
    productos = InventarioProducto.query.filter_by(activo=True).order_by(InventarioProducto.nombre.asc()).all()
    return render_template(
        "orden_compra_detalle.html",
        title=f"Orden {orden.folio}",
        orden=orden,
        productos=productos,
        estatus_options=ORDEN_COMPRA_ESTATUS,
    )


@app.route("/ordenes-compra/<int:orden_id>/actualizar", methods=["POST"])
@login_required
def orden_compra_actualizar(orden_id: int):
    orden = OrdenCompra.query.get_or_404(orden_id)
    f = request.form
    proveedor = (f.get("proveedor") or "").strip()
    if not proveedor:
        flash("Captura el proveedor de la orden.", "warning")
        return redirect(url_for("orden_compra_detalle", orden_id=orden.id))

    orden.proveedor = proveedor
    orden.numero_cliente_proveedor = (f.get("numero_cliente_proveedor") or "").strip() or None
    orden.contacto = (f.get("contacto") or "").strip() or None
    orden.telefono = (f.get("telefono") or "").strip() or None
    orden.correo = (f.get("correo") or "").strip() or None
    orden.fecha_entrega = _parse_date_or_none(f.get("fecha_entrega"))
    orden.forma_pago = (f.get("forma_pago") or "CONTADO").strip().upper()
    orden.descuento_total = fmt(parse_float(f.get("descuento_total"), 0))
    orden.iva_porc = fmt(parse_float(f.get("iva_porc"), 16.0))
    orden.condiciones = (f.get("condiciones") or "").strip() or None
    orden.notas = (f.get("notas") or "").strip() or None

    producto_ids = f.getlist("producto_id[]")
    partida_ids = f.getlist("partida_id[]")
    descripciones = f.getlist("descripcion[]")
    unidades = f.getlist("unidad[]")
    cantidades = f.getlist("cantidad[]")
    precios = f.getlist("precio_unitario[]")
    observaciones = f.getlist("observaciones[]")

    total_rows = max(len(producto_ids), len(descripciones), len(cantidades), len(precios), len(partida_ids))
    partidas_por_id = {str(p.id): p for p in orden.partidas}
    partidas_validas: set[int] = set()

    for idx in range(total_rows):
        partida_id = partida_ids[idx] if idx < len(partida_ids) else ""
        partida = partidas_por_id.get(partida_id)
        producto = None
        producto_id = producto_ids[idx] if idx < len(producto_ids) else ""
        if producto_id:
            producto = InventarioProducto.query.get(int(producto_id))

        descripcion = (descripciones[idx] if idx < len(descripciones) else "").strip()
        unidad = (unidades[idx] if idx < len(unidades) else "").strip()
        cantidad = parse_float(cantidades[idx] if idx < len(cantidades) else 0, 0)
        precio = parse_float(precios[idx] if idx < len(precios) else 0, 0)
        observacion = (observaciones[idx] if idx < len(observaciones) else "").strip()

        if producto:
            descripcion = descripcion or producto.nombre
            unidad = unidad or producto.unidad or "pieza"
            if precio <= 0:
                precio = float(producto.costo_unitario or 0)
        if not descripcion or cantidad <= 0:
            continue

        if partida and float(partida.cantidad_recibida or 0) > cantidad:
            cantidad = float(partida.cantidad_recibida or 0)

        if not partida:
            partida = OrdenCompraPartida(orden=orden, cantidad_recibida=0.0)
            db.session.add(partida)
            db.session.flush()

        partida.inventario_producto_id = producto.id if producto else None
        partida.codigo = producto.codigo if producto else None
        partida.descripcion = descripcion
        partida.unidad = unidad or "pieza"
        partida.cantidad = fmt(cantidad)
        partida.precio_unitario = fmt(precio)
        partida.observaciones = observacion or None
        if partida.id:
            partidas_validas.add(partida.id)

    for partida in list(orden.partidas):
        if partida.id not in partidas_validas and float(partida.cantidad_recibida or 0) <= 0:
            db.session.delete(partida)

    if not partidas_validas:
        flash("La orden debe conservar al menos una partida.", "warning")
        return redirect(url_for("orden_compra_detalle", orden_id=orden.id))

    _orden_compra_recalcular(orden)
    _orden_compra_actualizar_estatus_recepcion(orden)
    db.session.commit()
    flash("Orden de compra actualizada.", "success")
    return redirect(url_for("orden_compra_detalle", orden_id=orden.id))


@app.route("/ordenes-compra/<int:orden_id>/estatus", methods=["POST"])
@login_required
def orden_compra_estatus(orden_id: int):
    orden = OrdenCompra.query.get_or_404(orden_id)
    nuevo = (request.form.get("estatus") or "").strip().upper()
    if nuevo not in ORDEN_COMPRA_ESTATUS:
        flash("Estatus invalido.", "warning")
        return redirect(url_for("orden_compra_detalle", orden_id=orden.id))
    orden.estatus = nuevo
    orden.factura_folio = (request.form.get("factura_folio") or orden.factura_folio or "").strip() or None
    orden.factura_monto = fmt(parse_float(request.form.get("factura_monto"), orden.factura_monto or 0))
    orden.pago_referencia = (request.form.get("pago_referencia") or orden.pago_referencia or "").strip() or None
    orden.pago_monto = fmt(parse_float(request.form.get("pago_monto"), orden.pago_monto or 0))
    try:
        factura_archivo = _orden_compra_guardar_archivo(request.files.get("factura_archivo"), orden, "factura")
        pago_archivo = _orden_compra_guardar_archivo(request.files.get("pago_archivo"), orden, "pago")
    except ValueError as exc:
        flash(str(exc), "warning")
        return redirect(url_for("orden_compra_detalle", orden_id=orden.id))
    if factura_archivo:
        orden.factura_archivo = factura_archivo
    if pago_archivo:
        orden.pago_archivo = pago_archivo
    orden.actualizado_en = now_cdmx_naive()
    db.session.commit()
    flash("Estatus actualizado.", "success")
    return redirect(url_for("orden_compra_detalle", orden_id=orden.id))


@app.route("/ordenes-compra/<int:orden_id>/recibir", methods=["POST"])
@login_required
def orden_compra_recibir(orden_id: int):
    orden = OrdenCompra.query.get_or_404(orden_id)
    if orden.estatus == "CANCELADA":
        flash("No se puede recibir una orden cancelada.", "warning")
        return redirect(url_for("orden_compra_detalle", orden_id=orden.id))

    recibidas = 0
    for partida in orden.partidas:
        cantidad = parse_float(request.form.get(f"recibir_{partida.id}"), 0)
        if cantidad <= 0:
            continue
        pendiente = max(0.0, float(partida.cantidad or 0) - float(partida.cantidad_recibida or 0))
        cantidad = min(cantidad, pendiente)
        if cantidad <= 0:
            continue
        partida.cantidad_recibida = fmt(float(partida.cantidad_recibida or 0) + cantidad)
        recibidas += 1
        if partida.producto:
            _inventario_registrar_movimiento(
                producto=partida.producto,
                tipo="ENTRADA",
                motivo="COMPRA",
                cantidad=cantidad,
                costo_unitario=partida.precio_unitario or partida.producto.costo_unitario or 0,
                proveedor=orden.proveedor,
                referencia=orden.folio or "",
                observaciones=f"Recepcion de orden de compra {orden.folio}.",
            )

    if not recibidas:
        flash("Captura una cantidad a recibir.", "warning")
        return redirect(url_for("orden_compra_detalle", orden_id=orden.id))

    _orden_compra_actualizar_estatus_recepcion(orden)
    db.session.commit()
    flash("Recepcion registrada y Kardex actualizado.", "success")
    return redirect(url_for("orden_compra_detalle", orden_id=orden.id))


@app.route("/ordenes-compra/export.xlsx")
@login_required
def ordenes_compra_export_xlsx():
    if Workbook is None:
        abort(501, description="openpyxl no instalado en el servidor.")

    ordenes = OrdenCompra.query.order_by(OrdenCompra.fecha.desc(), OrdenCompra.id.desc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Ordenes de compra"
    ws.append([
        "Folio", "Fecha", "Proveedor", "Numero cliente proveedor", "Forma pago",
        "Estatus", "Subtotal", "Descuento %", "IVA", "Total", "Fecha entrega",
        "Factura", "Pago", "Responsable",
    ])
    for orden in ordenes:
        ws.append([
            orden.folio or "",
            orden.fecha.strftime("%d/%m/%Y") if orden.fecha else "",
            orden.proveedor or "",
            orden.numero_cliente_proveedor or "",
            orden.forma_pago or "",
            orden.estatus or "",
            float(orden.subtotal or 0),
            float(orden.descuento_total or 0),
            float(orden.iva_monto or 0),
            float(orden.total or 0),
            orden.fecha_entrega.strftime("%d/%m/%Y") if orden.fecha_entrega else "",
            orden.factura_folio or "",
            orden.pago_referencia or "",
            orden.responsable or "",
        ])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=MAR_BLUE_XLSX)
        cell.alignment = Alignment(horizontal="center")
    for col in ("G", "I", "J"):
        for cell in ws[col][1:]:
            cell.number_format = '"$"#,##0.00'
    for cell in ws["H"][1:]:
        cell.number_format = '0.00'
    for idx, width in enumerate([18, 14, 34, 24, 14, 24, 16, 16, 16, 16, 16, 18, 20, 18], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    stamp = now_cdmx_naive().strftime("%Y%m%d_%H%M%S")
    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="ordenes_compra_{stamp}.xlsx"'},
    )


@app.route("/ordenes-compra/<int:orden_id>/pdf")
@login_required
def orden_compra_pdf(orden_id: int):
    orden = OrdenCompra.query.get_or_404(orden_id)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=26 * mm,
        bottomMargin=35 * mm,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="OCCell", fontName="Helvetica", fontSize=8, leading=10, splitLongWords=False))
    styles.add(ParagraphStyle(name="OCSmall", fontName="Helvetica", fontSize=8.5, leading=11, splitLongWords=False))

    def encabezado(canv, doc_):
        canv.saveState()
        canv.setFillColor(colors.HexColor(MAR_BLUE))
        canv.rect(0, A4[1] - 40, A4[0], 40, stroke=0, fill=1)
        logo_path = os.path.join(app.static_folder or "static", "logo.png")
        if os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                max_w = 22 * mm
                scale = max_w / iw
                canv.drawImage(img, 12, A4[1] - (ih * scale) - 8, width=max_w, height=ih * scale, mask="auto")
            except Exception:
                pass
        canv.setFont("Helvetica-Bold", 14)
        canv.setFillColor(colors.white)
        canv.drawRightString(A4[0] - 12, A4[1] - 18, "ORDEN DE COMPRA")
        canv.setFont("Helvetica", 10)
        canv.drawRightString(A4[0] - 12, A4[1] - 31, orden.folio or "")
        canv.restoreState()

    def footer(canv, doc_):
        canv.saveState()
        canv.setFont("Helvetica", 8)
        canv.setFillColor(colors.HexColor("#333333"))
        canv.drawCentredString(A4[0] / 2, 24, "POLIUTECH - Recubrimientos Especializados")
        canv.drawCentredString(A4[0] / 2, 14, "Tel: 55 5938 6530 / 55 5938 0536 - info@poliutech.com - www.poliutech.com")
        canv.restoreState()

    elems = []
    meta = [
        [
            Paragraph(f"<b>Proveedor:</b> {escape(orden.proveedor or '')}", styles["OCSmall"]),
            Paragraph(f"<b>Fecha:</b> {orden.fecha.strftime('%d/%m/%Y') if orden.fecha else ''}", styles["OCSmall"]),
        ],
        [
            Paragraph(f"<b>Numero cliente con proveedor:</b> {escape(orden.numero_cliente_proveedor or '-')}", styles["OCSmall"]),
            Paragraph(f"<b>Forma de pago:</b> {escape(orden.forma_pago or 'CONTADO')}", styles["OCSmall"]),
        ],
        [
            Paragraph(f"<b>Contacto:</b> {escape(orden.contacto or '-')}", styles["OCSmall"]),
            Paragraph(f"<b>Entrega estimada:</b> {orden.fecha_entrega.strftime('%d/%m/%Y') if orden.fecha_entrega else '-'}", styles["OCSmall"]),
        ],
        [
            Paragraph(f"<b>Correo:</b> {escape(orden.correo or '-')}", styles["OCSmall"]),
            Paragraph(f"<b>Estatus:</b> {escape(orden.estatus or '')}", styles["OCSmall"]),
        ],
    ]
    meta_tbl = Table(meta, colWidths=[118 * mm, 62 * mm])
    meta_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elems.append(meta_tbl)
    elems.append(Spacer(1, 8))

    data = [["Codigo", "Descripcion", "Unidad", "Cantidad", "Recibida", "P. Unit.", "Importe"]]
    for partida in orden.partidas:
        data.append([
            Paragraph(escape(partida.codigo or ""), styles["OCCell"]),
            Paragraph(escape(partida.descripcion or ""), styles["OCCell"]),
            partida.unidad or "",
            f"{float(partida.cantidad or 0):,.2f}",
            f"{float(partida.cantidad_recibida or 0):,.2f}",
            f"${float(partida.precio_unitario or 0):,.2f}",
            f"${float(partida.subtotal or 0):,.2f}",
        ])
    tbl = Table(data, colWidths=[20 * mm, 66 * mm, 18 * mm, 20 * mm, 20 * mm, 24 * mm, 24 * mm], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(MAR_BLUE)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    elems.append(tbl)
    elems.append(Spacer(1, 8))
    totals = [
        ["Subtotal", f"${float(orden.subtotal or 0):,.2f}"],
        [f"Descuento {float(orden.descuento_total or 0):g}%", f"-${float(orden.subtotal or 0) * float(orden.descuento_total or 0) / 100.0:,.2f}"],
        [f"IVA {float(orden.iva_porc or 0):g}%", f"${float(orden.iva_monto or 0):,.2f}"],
        ["Total", f"${float(orden.total or 0):,.2f}"],
    ]
    totals_tbl = Table(totals, colWidths=[32 * mm, 32 * mm], hAlign="RIGHT")
    totals_tbl.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 0.5, colors.HexColor(MAR_BLUE)),
    ]))
    elems.append(totals_tbl)
    if orden.condiciones or orden.notas:
        elems.append(Spacer(1, 8))
        elems.append(Paragraph(f"<b>Condiciones:</b> {escape(orden.condiciones or '-')}", styles["OCSmall"]))
        elems.append(Paragraph(f"<b>Notas:</b> {escape(orden.notas or '-')}", styles["OCSmall"]))

    doc.build(
        elems,
        onFirstPage=lambda canv, d: (draw_watermark(canv, app), encabezado(canv, d), footer(canv, d)),
        onLaterPages=lambda canv, d: (draw_watermark(canv, app), encabezado(canv, d), footer(canv, d)),
    )
    buf.seek(0)
    filename = orden.folio or "orden_compra"
    return Response(
        buf.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}.pdf"'},
    )


# ---------------------------------------------------------
# Inventario
# ---------------------------------------------------------
INVENTARIO_TIPOS = ("ENTRADA", "SALIDA", "AJUSTE")
INVENTARIO_MOTIVOS = {
    "ENTRADA": ("COMPRA", "DEVOLUCION", "AJUSTE POSITIVO"),
    "SALIDA": ("OBRA", "CONSUMO INTERNO", "VENTA", "AJUSTE NEGATIVO"),
    "AJUSTE": ("AJUSTE POSITIVO", "AJUSTE NEGATIVO", "CONTEO FISICO"),
}


def _inventario_stock_bajo_query():
    return InventarioProducto.query.filter(
        InventarioProducto.activo.is_(True),
        or_(
            and_(
                InventarioProducto.stock_maximo > 0,
                InventarioProducto.stock_actual < (InventarioProducto.stock_maximo * 0.15),
            ),
            and_(
                or_(InventarioProducto.stock_maximo.is_(None), InventarioProducto.stock_maximo <= 0),
                InventarioProducto.stock_actual <= InventarioProducto.stock_minimo,
            ),
        ),
    )


def _inventario_porcentaje(producto: InventarioProducto) -> float:
    stock = float(producto.stock_actual or 0)
    maximo = float(getattr(producto, "stock_maximo", 0) or 0)
    if maximo <= 0:
        maximo = max(stock, float(producto.stock_minimo or 0), 1.0)
    return max(0.0, min(100.0, (stock / maximo) * 100.0))


def _inventario_estado(producto: InventarioProducto) -> dict:
    pct = _inventario_porcentaje(producto)
    if pct >= 75:
        return {"porcentaje": pct, "color": "success", "texto": "OK"}
    if pct >= 25:
        return {"porcentaje": pct, "color": "warning", "texto": "MEDIO"}
    if pct >= 15:
        return {"porcentaje": pct, "color": "orange", "texto": "REORDEN"}
    return {"porcentaje": pct, "color": "danger", "texto": "BAJO"}


def _inventario_registrar_movimiento(
    producto: InventarioProducto,
    tipo: str,
    motivo: str,
    cantidad: float,
    costo_unitario: float = 0.0,
    proveedor: str = "",
    obra: str = "",
    referencia: str = "",
    observaciones: str = "",
) -> InventarioMovimiento:
    tipo = (tipo or "").strip().upper()
    motivo = (motivo or "").strip().upper()
    cantidad = fmt(abs(cantidad))
    costo_unitario = fmt(costo_unitario)

    if tipo not in INVENTARIO_TIPOS:
        raise ValueError("Tipo de movimiento invalido.")
    if cantidad <= 0:
        raise ValueError("La cantidad debe ser mayor a cero.")
    if motivo not in INVENTARIO_MOTIVOS.get(tipo, ()):
        raise ValueError("Motivo de movimiento invalido.")
    if tipo == "AJUSTE" and not is_admin():
        raise PermissionError("Solo el administrador puede registrar ajustes.")

    stock_antes = float(producto.stock_actual or 0)
    if tipo == "ENTRADA":
        delta = cantidad
    elif tipo == "SALIDA":
        delta = -cantidad
    else:
        delta = cantidad if "POSITIVO" in motivo else -cantidad

    stock_despues = stock_antes + delta
    if stock_despues < -0.0001:
        raise ValueError(f"Stock insuficiente. Disponible: {stock_antes:g} {producto.unidad or ''}.")

    mov = InventarioMovimiento(
        producto_id=producto.id,
        fecha=now_cdmx_naive(),
        tipo=tipo,
        motivo=motivo,
        cantidad=cantidad,
        costo_unitario=costo_unitario or float(producto.costo_unitario or 0),
        stock_antes=fmt(stock_antes),
        stock_despues=fmt(stock_despues),
        proveedor=(proveedor or "").strip() or None,
        obra=(obra or "").strip() or None,
        referencia=(referencia or "").strip() or None,
        responsable=responsable_actual() or None,
        observaciones=(observaciones or "").strip() or None,
        usuario_id=getattr(current_user, "id", None),
    )
    producto.stock_actual = fmt(stock_despues)
    if float(getattr(producto, "stock_maximo", 0) or 0) <= 0 and stock_despues > 0:
        producto.stock_maximo = fmt(max(stock_despues, float(producto.stock_minimo or 0)))
    if tipo == "ENTRADA" and costo_unitario > 0:
        producto.costo_unitario = costo_unitario
    producto.actualizado_en = now_cdmx_naive()
    db.session.add(mov)
    return mov


@app.route("/inventario")
@login_required
def inventario_index():
    q = (request.args.get("q") or "").strip()
    categoria = (request.args.get("categoria") or "").strip()
    estado = (request.args.get("estado") or "").strip()

    productos_q = InventarioProducto.query
    if q:
        like = f"%{q}%"
        productos_q = productos_q.filter(or_(
            InventarioProducto.codigo.ilike(like),
            InventarioProducto.nombre.ilike(like),
            InventarioProducto.proveedor.ilike(like),
            InventarioProducto.ubicacion.ilike(like),
        ))
    if categoria:
        productos_q = productos_q.filter(InventarioProducto.categoria == categoria)
    if estado == "bajo":
        productos_q = productos_q.filter(
            InventarioProducto.activo.is_(True),
            or_(
                and_(
                    InventarioProducto.stock_maximo > 0,
                    InventarioProducto.stock_actual < (InventarioProducto.stock_maximo * 0.15),
                ),
                and_(
                    or_(InventarioProducto.stock_maximo.is_(None), InventarioProducto.stock_maximo <= 0),
                    InventarioProducto.stock_actual <= InventarioProducto.stock_minimo,
                ),
            )
        )
    elif estado == "reorden":
        productos_q = productos_q.filter(
            InventarioProducto.activo.is_(True),
            InventarioProducto.stock_actual >= (InventarioProducto.stock_maximo * 0.15),
            InventarioProducto.stock_actual < (InventarioProducto.stock_maximo * 0.25),
        )
    elif estado == "inactivo":
        productos_q = productos_q.filter(InventarioProducto.activo.is_(False))
    else:
        productos_q = productos_q.filter(InventarioProducto.activo.is_(True))

    productos = productos_q.order_by(InventarioProducto.nombre.asc()).all()
    recientes = (
        InventarioMovimiento.query
        .join(InventarioProducto)
        .order_by(InventarioMovimiento.fecha.desc(), InventarioMovimiento.id.desc())
        .limit(12)
        .all()
    )
    categorias = [
        row[0] for row in db.session.query(InventarioProducto.categoria)
        .filter(InventarioProducto.categoria.isnot(None), InventarioProducto.categoria != "")
        .distinct()
        .order_by(InventarioProducto.categoria.asc())
        .all()
    ]
    obras = RegistroObra.query.order_by(RegistroObra.obra.asc()).limit(300).all()
    total_valor = sum(float(p.stock_actual or 0) * float(p.costo_unitario or 0) for p in productos)

    return render_template(
        "inventario.html",
        title="Inventario",
        productos=productos,
        recientes=recientes,
        categorias=categorias,
        obras=obras,
        q=q,
        categoria=categoria,
        estado=estado,
        total_valor=total_valor,
        stock_bajo_count=_inventario_stock_bajo_query().count(),
        inventario_motivos=INVENTARIO_MOTIVOS,
        inventario_estado=_inventario_estado,
    )


@app.route("/inventario/productos/crear", methods=["POST"])
@login_required
def inventario_producto_crear():
    if not is_admin():
        abort(403)

    f = request.form
    codigo = (f.get("codigo") or "").strip().upper() or None
    nombre = (f.get("nombre") or "").strip()
    if not nombre:
        flash("Captura el nombre del material.", "warning")
        return redirect(url_for("inventario_index"))
    if codigo and InventarioProducto.query.filter_by(codigo=codigo).first():
        flash("Ya existe un material con ese codigo.", "danger")
        return redirect(url_for("inventario_index"))

    producto = InventarioProducto(
        codigo=codigo,
        nombre=nombre,
        categoria=(f.get("categoria") or "").strip() or None,
        unidad=(f.get("unidad") or "").strip() or "pieza",
        stock_minimo=fmt(parse_float(f.get("stock_minimo"), 0)),
        stock_maximo=fmt(parse_float(f.get("stock_maximo"), parse_float(f.get("stock_inicial"), 0))),
        costo_unitario=fmt(parse_float(f.get("costo_unitario"), 0)),
        proveedor=(f.get("proveedor") or "").strip() or None,
        ubicacion=(f.get("ubicacion") or "").strip() or None,
        activo=True,
    )
    db.session.add(producto)
    db.session.flush()

    stock_inicial = parse_float(f.get("stock_inicial"), 0)
    if stock_inicial > 0:
        _inventario_registrar_movimiento(
            producto,
            "ENTRADA",
            "AJUSTE POSITIVO",
            stock_inicial,
            producto.costo_unitario or 0,
            proveedor=producto.proveedor or "",
            referencia="STOCK INICIAL",
            observaciones="Alta inicial de inventario.",
        )

    db.session.commit()
    flash("Material agregado al inventario.", "success")
    return redirect(url_for("inventario_index"))


@app.route("/inventario/productos/<int:producto_id>/actualizar", methods=["POST"])
@login_required
def inventario_producto_actualizar(producto_id: int):
    if not is_admin():
        abort(403)

    producto = InventarioProducto.query.get_or_404(producto_id)
    f = request.form
    codigo = (f.get("codigo") or "").strip().upper() or None
    if codigo and codigo != producto.codigo and InventarioProducto.query.filter_by(codigo=codigo).first():
        flash("Ya existe otro material con ese codigo.", "danger")
        return redirect(url_for("inventario_kardex", producto_id=producto.id))

    producto.codigo = codigo
    producto.nombre = (f.get("nombre") or producto.nombre).strip()
    producto.categoria = (f.get("categoria") or "").strip() or None
    producto.unidad = (f.get("unidad") or producto.unidad or "pieza").strip()
    producto.stock_minimo = fmt(parse_float(f.get("stock_minimo"), producto.stock_minimo or 0))
    producto.stock_maximo = fmt(parse_float(f.get("stock_maximo"), producto.stock_maximo or 0))
    producto.costo_unitario = fmt(parse_float(f.get("costo_unitario"), producto.costo_unitario or 0))
    producto.proveedor = (f.get("proveedor") or "").strip() or None
    producto.ubicacion = (f.get("ubicacion") or "").strip() or None
    producto.activo = (f.get("activo") or "1") == "1"
    producto.actualizado_en = now_cdmx_naive()
    db.session.commit()
    flash("Material actualizado.", "success")
    return redirect(url_for("inventario_kardex", producto_id=producto.id))


@app.route("/inventario/movimientos/crear", methods=["POST"])
@login_required
def inventario_movimiento_crear():
    producto = InventarioProducto.query.get_or_404(int(request.form.get("producto_id") or 0))
    try:
        _inventario_registrar_movimiento(
            producto=producto,
            tipo=request.form.get("tipo"),
            motivo=request.form.get("motivo"),
            cantidad=parse_float(request.form.get("cantidad"), 0),
            costo_unitario=parse_float(request.form.get("costo_unitario"), producto.costo_unitario or 0),
            proveedor=request.form.get("proveedor") or producto.proveedor or "",
            obra=request.form.get("obra") or "",
            referencia=request.form.get("referencia") or "",
            observaciones=request.form.get("observaciones") or "",
        )
        db.session.commit()
        flash("Movimiento registrado correctamente.", "success")
    except PermissionError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "warning")
    return redirect(request.referrer or url_for("inventario_index"))


@app.route("/inventario/productos/<int:producto_id>/kardex")
@login_required
def inventario_kardex(producto_id: int):
    producto = InventarioProducto.query.get_or_404(producto_id)
    movimientos = (
        InventarioMovimiento.query
        .filter_by(producto_id=producto.id)
        .order_by(InventarioMovimiento.fecha.desc(), InventarioMovimiento.id.desc())
        .all()
    )
    obras = RegistroObra.query.order_by(RegistroObra.obra.asc()).limit(300).all()
    return render_template(
        "inventario_kardex.html",
        title=f"Kardex {producto.nombre}",
        producto=producto,
        movimientos=movimientos,
        obras=obras,
        inventario_motivos=INVENTARIO_MOTIVOS,
        inventario_estado=_inventario_estado,
    )


@app.route("/inventario/export.xlsx")
@login_required
def inventario_export_xlsx():
    if Workbook is None:
        abort(501, description="openpyxl no instalado en el servidor.")

    productos = InventarioProducto.query.order_by(InventarioProducto.nombre.asc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    headers = [
        "Codigo", "Material", "Categoria", "Unidad", "Stock actual", "Stock minimo",
        "Stock maximo", "% inventario", "Estado", "Costo unitario", "Valor", "Proveedor", "Ubicacion", "Activo",
    ]
    ws.append(headers)
    fills_estado = {
        "OK": PatternFill("solid", fgColor="D1E7DD"),
        "MEDIO": PatternFill("solid", fgColor="FFF3CD"),
        "REORDEN": PatternFill("solid", fgColor="FED7AA"),
        "BAJO": PatternFill("solid", fgColor="F8D7DA"),
    }
    for p in productos:
        stock = float(p.stock_actual or 0)
        costo = float(p.costo_unitario or 0)
        estado_inv = _inventario_estado(p)
        ws.append([
            p.codigo or "",
            p.nombre or "",
            p.categoria or "",
            p.unidad or "",
            stock,
            float(p.stock_minimo or 0),
            float(p.stock_maximo or 0),
            estado_inv["porcentaje"] / 100.0,
            estado_inv["texto"],
            costo,
            stock * costo,
            p.proveedor or "",
            p.ubicacion or "",
            "SI" if p.activo else "NO",
        ])
        fill = fills_estado.get(estado_inv["texto"])
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=MAR_BLUE_XLSX)
        cell.alignment = Alignment(horizontal="center")
    for col in ("E", "F", "G"):
        for cell in ws[col][1:]:
            cell.number_format = '0.00'
    for cell in ws["H"][1:]:
        cell.number_format = '0.00%'
    for col in ("J", "K"):
        for cell in ws[col][1:]:
            cell.number_format = '"$"#,##0.00'
    widths = [16, 34, 18, 12, 14, 14, 14, 14, 14, 16, 16, 24, 20, 10]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    stamp = now_cdmx_naive().strftime("%Y%m%d_%H%M%S")
    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="inventario_{stamp}.xlsx"'}
    )


# Blueprints (Catálogos) — si existen en tu repo
# ---------------------------------------------------------
try:
    from catalogos_routes import bp as catalogos_bp
    app.register_blueprint(catalogos_bp, url_prefix="/catalogos")
except Exception as e:
    print(f"[WARN] No se pudo cargar blueprint catalogos_routes: {e}", file=sys.stderr)

try:
    from pu_routes import pu_bp
    app.register_blueprint(pu_bp, url_prefix="/pu")
except Exception as e:
    print(f"[WARN] No se pudo cargar blueprint pu_routes: {e}", file=sys.stderr)

try:
    from facturacion_routes import facturacion_bp
    app.register_blueprint(facturacion_bp)
except Exception as e:
    print(f"[WARN] No se pudo cargar blueprint facturacion_routes: {e}", file=sys.stderr)

# ---------------------------------------------------------
# Main
# ---------------------------------------------------------
if __name__ == "__main__":
    try:
        os.makedirs(app.static_folder or "static", exist_ok=True)
    except Exception:
        pass
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=True)


