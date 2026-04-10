from __future__ import annotations

from pathlib import Path
import shutil
from threading import Lock
from typing import Any

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import login_required
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


pu_bp = Blueprint("pu", __name__, template_folder="templates")

PROJECT_ROOT = Path(__file__).resolve().parent
PU_TEMPLATE_PATH = PROJECT_ROOT / "APU General (1).xlsm"
PU_RUNTIME_PATH = PROJECT_ROOT / "apu_runtime.xlsm"
WORKBOOK_LOCK = Lock()
DEFAULT_ROWS = 35
DEFAULT_COLS = 16
DEFAULT_SCALE = 0.86


def _pesos(cantidad: Any) -> str:
    try:
        valor = float(cantidad or 0)
    except Exception:
        valor = 0.0

    decimales = round(100 * (valor - int(valor)))
    cantidad_txt = f"{valor:.2f}".rstrip("0").rstrip(".")
    partes = cantidad_txt.split(".")
    entero = partes[0]

    places = {
        2: " MIL ",
        3: " MILLONES ",
        4: " BILLONES ",
        5: " TRILLONES ",
    }

    def get_digit(digit: str) -> str:
        return {
            "1": "UN",
            "2": "DOS",
            "3": "TRES",
            "4": "CUATRO",
            "5": "CINCO",
            "6": "SEIS",
            "7": "SIETE",
            "8": "OCHO",
            "9": "NUEVE",
        }.get(digit, "")

    def get_tens(text: str) -> str:
        if len(text) == 1:
            return get_digit(text)
        if text[0] == "1":
            return {
                "10": "DIEZ",
                "11": "ONCE",
                "12": "DOCE",
                "13": "TRECE",
                "14": "CATORCE",
                "15": "QUINCE",
                "16": "DIECISEIS",
                "17": "DIECISIETE",
                "18": "DIECIOCHO",
                "19": "DIECINUEVE",
            }.get(text, "")
        if text[1] == "0":
            return {
                "2": "VEINTE",
                "3": "TREINTA",
                "4": "CUARENTA",
                "5": "CINCUENTA",
                "6": "SESENTA",
                "7": "SETENTA",
                "8": "OCHENTA",
                "9": "NOVENTA",
            }.get(text[0], "")
        prefijo = {
            "2": "VEINTI",
            "3": "TREINTA Y ",
            "4": "CUARENTA Y ",
            "5": "CINCUENTA Y ",
            "6": "SESENTA Y ",
            "7": "SETENTA Y ",
            "8": "OCHENTA Y ",
            "9": "NOVENTA Y ",
        }.get(text[0], "")
        return prefijo + get_digit(text[1])

    def get_hundreds(text: str) -> str:
        if int(text) == 0:
            return ""
        text = text.zfill(3)
        result = ""
        if text[0] != "0":
            if text == "100":
                result = "CIEN "
            else:
                result = {
                    "1": "CIENTO ",
                    "2": "DOSCIENTOS ",
                    "3": "TRESCIENTOS ",
                    "4": "CUATROCIENTOS ",
                    "5": "QUINIENTOS ",
                    "6": "SEISCIENTOS ",
                    "7": "SETECIENTOS ",
                    "8": "OCHOCIENTOS ",
                    "9": "NOVECIENTOS ",
                }.get(text[0], "")
        if text[1] != "0":
            result += get_tens(text[1:])
        else:
            result += get_digit(text[2])
        return result

    count = 1
    pesotes = ""
    original_len = len(entero)
    original_num = int(entero or 0)
    while entero:
        temp = get_hundreds(entero[-3:])
        if temp:
            if temp == "UN" and count > 2:
                pesotes = temp + places[count][:-3] + " " + pesotes
            else:
                pesotes = temp + places.get(count, "") + pesotes
        entero = entero[:-3]
        count += 1

    if not pesotes:
        pesotes = " (CERO PESOS"
    elif pesotes == "UN":
        pesotes = " (UN PESO"
    elif original_len > 6 and (original_num % 1000000) == 0:
        pesotes = f" ({pesotes} DE PESOS"
    else:
        pesotes = f" ({pesotes} PESOS"

    centavos = f" {int(decimales):02d}/100 M.N.)"
    return (pesotes + centavos).upper().strip()


def _ensure_runtime_workbook() -> None:
    if PU_RUNTIME_PATH.exists():
        return
    if not PU_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"No se encontró la plantilla base: {PU_TEMPLATE_PATH.name}")
    shutil.copy2(PU_TEMPLATE_PATH, PU_RUNTIME_PATH)


def _open_workbook(data_only: bool = False):
    _ensure_runtime_workbook()
    return load_workbook(PU_RUNTIME_PATH, keep_vba=True, data_only=data_only)


def _save_workbook(workbook) -> None:
    workbook.save(PU_RUNTIME_PATH)


def _sheet_bounds(ws) -> tuple[int, int, int, int]:
    min_row = None
    min_col = None
    max_row = None
    max_col = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            min_row = cell.row if min_row is None else min(min_row, cell.row)
            min_col = cell.column if min_col is None else min(min_col, cell.column)
            max_row = cell.row if max_row is None else max(max_row, cell.row)
            max_col = cell.column if max_col is None else max(max_col, cell.column)
    if min_row is None:
        return 1, 1, 1, 1
    return min_row, min_col, max_row, max_col


def _display_value(formula_cell, data_cell=None) -> str:
    effective = data_cell if data_cell is not None else formula_cell
    if effective.value is None:
        return ""
    return str(effective.value)


def _excel_color(color) -> str | None:
    rgb = getattr(color, "rgb", None)
    if rgb and len(rgb) == 8:
        return f"#{rgb[-6:]}"
    if rgb and len(rgb) == 6:
        return f"#{rgb}"
    return None


def _border_css(side) -> str:
    style = getattr(side, "style", None)
    color = _excel_color(getattr(side, "color", None)) or "#d9e2ec"
    width = {
        "thin": "1px",
        "hair": "1px",
        "dotted": "1px",
        "dashed": "1px",
        "double": "3px",
        "medium": "2px",
        "mediumDashed": "2px",
        "mediumDashDot": "2px",
        "mediumDashDotDot": "2px",
        "slantDashDot": "2px",
        "thick": "3px",
    }.get(style, "1px")
    kind = "solid"
    if style in {"dotted"}:
        kind = "dotted"
    elif style in {"dashed", "mediumDashed"}:
        kind = "dashed"
    elif style == "double":
        kind = "double"
    return f"{width} {kind} {color}" if style else "1px solid transparent"


def _cell_style(cell) -> str:
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    pieces = []
    fill_color = _excel_color(getattr(fill, "fgColor", None))
    if fill_color and fill.patternType == "solid":
        pieces.append(f"background:{fill_color}")
    font_color = _excel_color(getattr(font, "color", None))
    if font_color:
        pieces.append(f"color:{font_color}")
    if font.bold:
        pieces.append("font-weight:700")
    if font.italic:
        pieces.append("font-style:italic")
    if font.sz:
        pieces.append(f"font-size:{max(9, round(font.sz * DEFAULT_SCALE, 2))}pt")
    if font.name:
        pieces.append(f"font-family:'{font.name}',sans-serif")
    if alignment and alignment.horizontal:
        pieces.append(f"text-align:{alignment.horizontal}")
    if alignment and alignment.vertical:
        pieces.append(f"vertical-align:{alignment.vertical}")
    if alignment and alignment.wrapText:
        pieces.append("white-space:pre-wrap")
    else:
        pieces.append("white-space:nowrap")
    pieces.append(f"border-top:{_border_css(border.top)}")
    pieces.append(f"border-right:{_border_css(border.right)}")
    pieces.append(f"border-bottom:{_border_css(border.bottom)}")
    pieces.append(f"border-left:{_border_css(border.left)}")
    return ";".join(pieces)


def _column_pixel_width(ws, col_idx: int) -> int:
    dim = ws.column_dimensions.get(get_column_letter(col_idx))
    width = dim.width if dim and dim.width else 8.43
    return max(48, min(420, int(width * 7 + 12)))


def _row_pixel_height(ws, row_idx: int) -> int:
    dim = ws.row_dimensions.get(row_idx)
    height = dim.height if dim and dim.height else 15
    return max(18, min(240, int(height * 1.35)))


def _is_hidden_row(ws, row_idx: int) -> bool:
    dim = ws.row_dimensions.get(row_idx)
    return bool(dim.hidden) if dim else False


def _is_hidden_col(ws, col_idx: int) -> bool:
    dim = ws.column_dimensions.get(get_column_letter(col_idx))
    return bool(dim.hidden) if dim else False


def _build_merge_maps(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    merge_root = {}
    merge_skip = set()
    for merged in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = merged.bounds
        if max_r < start_row or min_r > end_row or max_c < start_col or min_c > end_col:
            continue
        top_left = (min_r, min_c)
        merge_root[top_left] = {
            "rowspan": max_r - min_r + 1,
            "colspan": max_c - min_c + 1,
        }
        for row_idx in range(min_r, max_r + 1):
            for col_idx in range(min_c, max_c + 1):
                if (row_idx, col_idx) != top_left:
                    merge_skip.add((row_idx, col_idx))
    return merge_root, merge_skip


def _sheet_window(ws, ws_data, start_row: int, start_col: int, rows: int, cols: int):
    row_indices = []
    row_cursor = start_row
    while len(row_indices) < rows and row_cursor <= ws.max_row:
        if not _is_hidden_row(ws, row_cursor):
            row_indices.append(row_cursor)
        row_cursor += 1

    col_indices = []
    col_cursor = start_col
    while len(col_indices) < cols and col_cursor <= ws.max_column:
        if not _is_hidden_col(ws, col_cursor):
            col_indices.append(col_cursor)
        col_cursor += 1

    if not row_indices:
        row_indices = [start_row]
    if not col_indices:
        col_indices = [start_col]

    end_row = row_indices[-1]
    end_col = col_indices[-1]
    merge_root, merge_skip = _build_merge_maps(ws, start_row, end_row, start_col, end_col)

    grid_rows = []
    for row_idx in row_indices:
        row_cells = []
        for col_idx in col_indices:
            if (row_idx, col_idx) in merge_skip:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            data_cell = ws_data.cell(row=row_idx, column=col_idx) if ws_data is not None else None
            merge = merge_root.get((row_idx, col_idx), {"rowspan": 1, "colspan": 1})
            row_cells.append({
                "coord": cell.coordinate,
                "value": _display_value(cell, data_cell),
                "is_formula": cell.data_type == "f",
                "formula": f"={cell.value}" if cell.data_type == "f" and cell.value else "",
                "rowspan": merge["rowspan"],
                "colspan": merge["colspan"],
                "style": _cell_style(cell),
                "width": _column_pixel_width(ws, col_idx),
                "height": _row_pixel_height(ws, row_idx),
            })
        grid_rows.append({
            "row_number": row_idx,
            "height": _row_pixel_height(ws, row_idx),
            "cells": row_cells,
        })

    header_cols = [{
        "label": get_column_letter(col_idx),
        "width": _column_pixel_width(ws, col_idx),
    } for col_idx in col_indices]

    return header_cols, grid_rows, row_indices, col_indices


def _parse_input_value(raw: str):
    text = (raw or "").strip()
    if text == "":
        return None
    if text.startswith("="):
        return text
    lowered = text.lower()
    if lowered in {"true", "false"}:
        return lowered == "true"
    try:
        if "." in text:
            return float(text)
        return int(text)
    except Exception:
        return text


def _copy_range_values(ws, source_start: str, source_end: str, target_start: str) -> None:
    start_col = column_index_from_string("".join(ch for ch in source_start if ch.isalpha()))
    start_row = int("".join(ch for ch in source_start if ch.isdigit()))
    end_col = column_index_from_string("".join(ch for ch in source_end if ch.isalpha()))
    end_row = int("".join(ch for ch in source_end if ch.isdigit()))
    target_col = column_index_from_string("".join(ch for ch in target_start if ch.isalpha()))
    target_row = int("".join(ch for ch in target_start if ch.isdigit()))

    for r_offset, row in enumerate(range(start_row, end_row + 1)):
        for c_offset, col in enumerate(range(start_col, end_col + 1)):
            target_cell = ws.cell(row=target_row + r_offset, column=target_col + c_offset)
            target_cell.value = ws.cell(row=row, column=col).value


def _clear_range(ws, start: str, end: str) -> None:
    start_col = column_index_from_string("".join(ch for ch in start if ch.isalpha()))
    start_row = int("".join(ch for ch in start if ch.isdigit()))
    end_col = column_index_from_string("".join(ch for ch in end if ch.isalpha()))
    end_row = int("".join(ch for ch in end if ch.isdigit()))
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).value = None


def _sort_rows(ws, header_row: int, start_row: int, end_row: int, start_col: str, end_col: str, key_col: str) -> None:
    start_col_idx = column_index_from_string(start_col)
    end_col_idx = column_index_from_string(end_col)
    key_col_idx = column_index_from_string(key_col)

    rows = []
    for row_idx in range(start_row, end_row + 1):
        row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(start_col_idx, end_col_idx + 1)]
        rows.append(row_values)

    key_offset = key_col_idx - start_col_idx

    def sort_key(row_values):
        value = row_values[key_offset]
        if value in (None, ""):
            return (1, "")
        return (0, str(value))

    rows.sort(key=sort_key)

    for row_offset, row_values in enumerate(rows):
        for col_offset, value in enumerate(row_values):
            ws.cell(row=start_row + row_offset, column=start_col_idx + col_offset).value = value


def _macro_buttons(sheet_name: str) -> list[dict[str, str]]:
    buttons = []
    if sheet_name == "VOLUMETRIA":
        buttons.extend([
            {"action": "volumen", "label": "Volumen"},
            {"action": "borrar", "label": "Borrar"},
        ])
    if sheet_name == "RESUMEN":
        buttons.extend([
            {"action": "resumen", "label": "Resumen"},
            {"action": "borra_res", "label": "Borra_Res"},
        ])
    return buttons


@pu_bp.route("/")
@login_required
def obras_index():
    return redirect(url_for("pu.sheet_view", sheet_name="PORTADA"))


@pu_bp.route("/sheet/<sheet_name>")
@login_required
def sheet_view(sheet_name: str):
    with WORKBOOK_LOCK:
        workbook = _open_workbook()
        workbook_data = _open_workbook(data_only=True)
        if sheet_name not in workbook.sheetnames:
            flash("La hoja solicitada no existe en el libro de APU.", "warning")
            return redirect(url_for("pu.obras_index"))

        ws = workbook[sheet_name]
        ws_data = workbook_data[sheet_name]
        min_row, min_col, max_row, max_col = _sheet_bounds(ws)
        start_row = max(1, int(request.args.get("row", min_row)))
        start_col = max(1, int(request.args.get("col", min_col)))
        rows = max(5, min(80, int(request.args.get("rows", DEFAULT_ROWS))))
        cols = max(5, min(30, int(request.args.get("cols", DEFAULT_COLS))))
        header_cols, grid_rows, row_indices, col_indices = _sheet_window(ws, ws_data, start_row, start_col, rows, cols)
        end_row = row_indices[-1]
        end_col = col_indices[-1]

        pesos_result = ""
        pesos_input = request.args.get("pesos", "").strip()
        if pesos_input:
            pesos_result = _pesos(pesos_input)

        return render_template(
            "pu_sheet.html",
            title=f"Precios Unitarios - {sheet_name}",
            workbook_name=PU_RUNTIME_PATH.name,
            sheets=workbook.sheetnames,
            hidden_sheets={"Apoyo"},
            freeze_panes=ws.freeze_panes,
            sheet_name=sheet_name,
            header_cols=header_cols,
            grid_rows=grid_rows,
            min_row=min_row,
            min_col=min_col,
            max_row=max_row,
            max_col=max_col,
            start_row=start_row,
            start_col=start_col,
            end_row=end_row,
            end_col=end_col,
            rows=rows,
            cols=cols,
            macro_buttons=_macro_buttons(sheet_name),
            pesos_input=pesos_input,
            pesos_result=pesos_result,
        )


@pu_bp.route("/action/<action_name>", methods=["POST"])
@login_required
def run_action(action_name: str):
    sheet_name = (request.form.get("sheet_name") or "PORTADA").strip() or "PORTADA"
    with WORKBOOK_LOCK:
        workbook = _open_workbook()
        if action_name == "volumen":
            ws = workbook["VOLUMETRIA"]
            _copy_range_values(ws, "Z21", "AP1020", "BK21")
            _sort_rows(ws, header_row=20, start_row=21, end_row=1020, start_col="BK", end_col="CA", key_col="BK")
            flash("Macro Volumen aplicada.", "success")
        elif action_name == "borrar":
            ws = workbook["VOLUMETRIA"]
            _clear_range(ws, "AA21", "AP1020")
            _clear_range(ws, "BK21", "CA1020")
            flash("Macro Borrar aplicada.", "success")
        elif action_name == "resumen":
            ws = workbook["RESUMEN"]
            _copy_range_values(ws, "AA21", "AH1020", "AQ21")
            _sort_rows(ws, header_row=20, start_row=21, end_row=1020, start_col="AQ", end_col="AX", key_col="AX")
            flash("Macro Resumen aplicada.", "success")
        elif action_name == "borra_res":
            ws = workbook["RESUMEN"]
            _clear_range(ws, "AQ21", "AX1020")
            flash("Macro Borra_Res aplicada.", "success")
        elif action_name == "reset":
            shutil.copy2(PU_TEMPLATE_PATH, PU_RUNTIME_PATH)
            flash("Se restauró el libro desde la plantilla base.", "success")
            return redirect(url_for("pu.sheet_view", sheet_name=sheet_name))
        else:
            flash("Acción no reconocida para el libro de APU.", "warning")
            return redirect(url_for("pu.sheet_view", sheet_name=sheet_name))
        _save_workbook(workbook)
    return redirect(url_for("pu.sheet_view", sheet_name=sheet_name))


@pu_bp.route("/update-cell", methods=["POST"])
@login_required
def update_cell():
    sheet_name = (request.form.get("sheet_name") or "").strip()
    cell_ref = (request.form.get("cell_ref") or "").strip().upper()
    cell_value = request.form.get("cell_value", "")
    if not sheet_name or not cell_ref:
        flash("Indica hoja y celda para actualizar.", "warning")
        return redirect(url_for("pu.obras_index"))

    with WORKBOOK_LOCK:
        workbook = _open_workbook()
        if sheet_name not in workbook.sheetnames:
            flash("La hoja indicada no existe.", "warning")
            return redirect(url_for("pu.obras_index"))
        workbook[sheet_name][cell_ref] = _parse_input_value(cell_value)
        _save_workbook(workbook)

    flash(f"Celda {cell_ref} actualizada.", "success")
    return redirect(url_for("pu.sheet_view", sheet_name=sheet_name))


@pu_bp.route("/download")
@login_required
def download_runtime():
    _ensure_runtime_workbook()
    return send_file(PU_RUNTIME_PATH, as_attachment=True, download_name=PU_RUNTIME_PATH.name)
