from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSM = ROOT / "APU General (1).xlsm"
OUTPUT_JSON = ROOT / "apu_template.json"


def display_value(value) -> str:
    if value is None:
        return ""
    return str(value)


def export_workbook() -> dict:
    wb_formula = load_workbook(SOURCE_XLSM, data_only=False, keep_vba=True)
    wb_data = load_workbook(SOURCE_XLSM, data_only=True, keep_vba=True)

    result = {
        "source": SOURCE_XLSM.name,
        "sheets": [],
    }

    for sheet_name in wb_formula.sheetnames:
        ws_formula = wb_formula[sheet_name]
        ws_data = wb_data[sheet_name]
        sheet_payload = {
            "name": sheet_name,
            "hidden": ws_formula.sheet_state != "visible",
            "max_row": ws_formula.max_row,
            "max_col": ws_formula.max_column,
            "freeze_panes": str(ws_formula.freeze_panes or ""),
            "merged_ranges": [str(rng) for rng in ws_formula.merged_cells.ranges],
            "rows": {},
            "cols": {},
            "cells": {},
        }

        for idx, dim in ws_formula.row_dimensions.items():
            if dim.height or dim.hidden:
                sheet_payload["rows"][str(idx)] = {
                    "height": dim.height,
                    "hidden": bool(dim.hidden),
                }

        for key, dim in ws_formula.column_dimensions.items():
            if dim.width or dim.hidden:
                sheet_payload["cols"][key] = {
                    "width": dim.width,
                    "hidden": bool(dim.hidden),
                }

        for row in ws_formula.iter_rows():
            for cell in row:
                if cell.value in (None, ""):
                    continue
                cached_value = ws_data[cell.coordinate].value
                sheet_payload["cells"][cell.coordinate] = {
                    "value": display_value(cached_value),
                    "formula": f"={cell.value}" if cell.data_type == "f" else "",
                    "raw": display_value(cell.value),
                    "row": cell.row,
                    "col": cell.column,
                    "col_letter": get_column_letter(cell.column),
                }

        result["sheets"].append(sheet_payload)

    return result


def main() -> None:
    payload = export_workbook()
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(f"Plantilla exportada a {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
