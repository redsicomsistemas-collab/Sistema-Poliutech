import json
import io
import os
import re
from datetime import datetime

from flask import Blueprint, Response, flash, jsonify, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import or_, text

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except Exception:
    Workbook = None

from models import Cliente, Concepto, Cotizacion, CotizacionDetalle, db
from .calc import explosion_insumos_obra, programa_obra, programa_recursos_obra, recalcular_apu, recalcular_obra
from .models import APU, APUDetalle, ManoObra, Maquinaria, Material, Obra, ObraCargo, ObraPartida

apu_bp = Blueprint("apu", __name__, url_prefix="/apu", template_folder="templates")

BASE_DIR = os.path.dirname(__file__)
PROJECT_ROOT = os.path.dirname(BASE_DIR)
PLANTILLAS_FILE = os.path.join(BASE_DIR, "plantillas.json")
TIPOS_INSUMO = [
    ("material", "Material"),
    ("mano_obra", "Mano de obra"),
    ("maquinaria", "Maquinaria"),
    ("auxiliar", "Auxiliar"),
]
TYPE_LABELS = {key: label for key, label in TIPOS_INSUMO}
SUGERENCIAS_CARGOS = [
    ("Agua", "servicios", "indirecto", "mes"),
    ("Baños", "servicios", "indirecto", "mes"),
    ("Energía eléctrica", "servicios", "indirecto", "mes"),
    ("Seguristas", "seguridad", "indirecto", "jor"),
    ("Rescatistas", "seguridad", "indirecto", "jor"),
    ("Inspectores", "supervision", "indirecto", "jor"),
    ("Fianzas", "legal", "cargo_adicional", "lote"),
    ("SIROC", "legal", "indirecto", "tramite"),
    ("Cuotas sindicales", "legal", "retencion", "lote"),
    ("Equipo de seguridad", "seguridad", "directo_global", "lote"),
    ("Equipo especializado", "equipo", "directo_global", "lote"),
    ("Certificados DC3", "legal", "indirecto", "tramite"),
    ("Certificados DC5", "legal", "indirecto", "tramite"),
]


def _f(value, default=0.0):
    try:
        if value is None or value == "":
            return float(default)
        return float(str(value).replace(",", "").replace("$", "").strip())
    except Exception:
        return float(default)


def _s(value):
    return (value or "").strip() or None


def _load_plantillas():
    try:
        with open(PLANTILLAS_FILE, "r", encoding="utf-8") as file_handle:
            return json.load(file_handle)
    except Exception:
        return []


def _responsable_actual():
    try:
        nombre = (getattr(current_user, "nombre", "") or "").strip()
        if not nombre:
            return None
        first = nombre.split()[0].strip()
        return first[:1].upper() + first[1:].lower() if first else None
    except Exception:
        return None


def _generar_folio():
    prefix = "PTCH-"
    maxn = 0
    rows = db.session.execute(text("SELECT folio FROM cotizacion WHERE folio LIKE 'PTCH-%'")).fetchall()
    for (folio,) in rows:
        match = re.match(r"PTCH-(\d{4})$", str(folio))
        if match:
            maxn = max(maxn, int(match.group(1)))
    return f"{prefix}{maxn + 1:04d}"


def _safe_detalle_kwargs(**kwargs):
    valid_columns = set(CotizacionDetalle.__table__.columns.keys())
    return {key: value for key, value in kwargs.items() if key in valid_columns}


def _fmt_money(value):
    return "${:,.2f}".format(float(value or 0))


def _truncate_text(value, limit=120):
    text_value = (value or "").strip()
    if len(text_value) <= limit:
        return text_value
    return text_value[: limit - 1].rstrip() + "…"


def _resource_for_tipo(tipo_insumo):
    return {
        "material": Material,
        "mano_obra": ManoObra,
        "maquinaria": Maquinaria,
    }.get(tipo_insumo)


def _buscar_recurso(tipo_insumo, referencia_id):
    if tipo_insumo == "auxiliar":
        return APU.query.get(referencia_id)
    model = _resource_for_tipo(tipo_insumo)
    if not model:
        return None
    return model.query.get(referencia_id)


def _resource_label(item):
    if isinstance(item, APU):
        clave = f"{item.clave} · " if getattr(item, "clave", None) else ""
        categoria = f" [{item.categoria}]" if getattr(item, "categoria", None) else ""
        return f"{clave}{item.concepto}{categoria} · {item.unidad} · ${item.precio_unitario or 0:,.2f}"
    clave = f"{item.clave} · " if getattr(item, "clave", None) else ""
    categoria = f" [{item.categoria}]" if getattr(item, "categoria", None) else ""
    return f"{clave}{item.nombre}{categoria} · {item.unidad} · ${item.precio_unitario or 0:,.2f}"


def _set_apu_defaults(apu):
    if apu.indirecto_pct is None:
        apu.indirecto_pct = 12.0
    if apu.utilidad_pct is None:
        apu.utilidad_pct = 10.0
    if apu.financiamiento_pct is None:
        apu.financiamiento_pct = 2.5
    if apu.cargos_adicionales_pct is None:
        apu.cargos_adicionales_pct = 0.0
    if apu.jornada_horas is None:
        apu.jornada_horas = 8.0
    if apu.cantidad_objetivo is None:
        apu.cantidad_objetivo = 1.0
    if apu.rendimiento_base is None:
        apu.rendimiento_base = 1.0
    if apu.herramienta_menor_pct is None:
        apu.herramienta_menor_pct = 3.0
    if getattr(apu, "es_auxiliar", None) is None:
        apu.es_auxiliar = False


def _guardar_apu_desde_form(apu):
    apu.clave = _s(request.form.get("clave"))
    apu.concepto = (request.form.get("concepto") or "").strip()
    apu.descripcion = _s(request.form.get("descripcion"))
    apu.categoria = _s(request.form.get("categoria"))
    apu.capitulo = _s(request.form.get("capitulo"))
    apu.subcapitulo = _s(request.form.get("subcapitulo"))
    apu.alcance = _s(request.form.get("alcance"))
    apu.unidad = (request.form.get("unidad") or "m2").strip() or "m2"
    apu.es_auxiliar = request.form.get("es_auxiliar") == "1"
    apu.cantidad_objetivo = _f(request.form.get("cantidad_objetivo"), 1.0)
    apu.rendimiento_base = _f(request.form.get("rendimiento_base"), 1.0)
    apu.jornada_horas = _f(request.form.get("jornada_horas"), 8.0)
    apu.desperdicio_general_pct = _f(request.form.get("desperdicio_general_pct"))
    apu.herramienta_menor_pct = _f(request.form.get("herramienta_menor_pct"))
    apu.indirecto_pct = _f(request.form.get("indirecto_pct"))
    apu.utilidad_pct = _f(request.form.get("utilidad_pct"))
    apu.financiamiento_pct = _f(request.form.get("financiamiento_pct"))
    apu.cargos_adicionales_pct = _f(request.form.get("cargos_adicionales_pct"))
    apu.notas = _s(request.form.get("notas"))
    recalcular_apu(apu)


def _build_catalog_context():
    materiales = Material.query.order_by(Material.nombre.asc()).all()
    mano_obra = ManoObra.query.order_by(ManoObra.nombre.asc()).all()
    maquinarias = Maquinaria.query.order_by(Maquinaria.nombre.asc()).all()
    return {
        "materiales": materiales,
        "mano_obra": mano_obra,
        "maquinarias": maquinarias,
        "catalogo_por_tipo": {
            "material": materiales,
            "mano_obra": mano_obra,
            "maquinaria": maquinarias,
            "auxiliar": APU.query.filter(APU.es_auxiliar.is_(True)).order_by(APU.concepto.asc()).all(),
        },
    }


def _decorate_apu(apu):
    recalcular_apu(apu)
    detalles = sorted(apu.detalles, key=lambda item: (item.tipo_insumo or "", item.id or 0))
    tipo_totales = []
    for tipo, label in TIPOS_INSUMO:
        total = sum((d.subtotal or 0.0) for d in detalles if d.tipo_insumo == tipo)
        tipo_totales.append(
            {
                "key": tipo,
                "label": label,
                "total": total,
                "pct_directo": (total / apu.costo_directo * 100.0) if apu.costo_directo else 0.0,
            }
        )

    apu.detalles_ordenados = detalles
    apu.tipo_totales = tipo_totales
    apu.tarjetas_resumen = [
        {"label": "Materiales", "value": apu.costo_materiales, "tone": "sand"},
        {"label": "Mano de obra", "value": apu.costo_mano_obra, "tone": "mint"},
        {"label": "Maquinaria", "value": apu.costo_maquinaria, "tone": "steel"},
        {"label": "Herr. menor", "value": apu.costo_herramienta, "tone": "amber"},
        {"label": "Costo directo", "value": apu.costo_directo, "tone": "navy"},
        {"label": "P.U. final", "value": apu.precio_unitario, "tone": "ink"},
    ]
    apu.sobrecostos = [
        {"label": "Indirectos", "pct": apu.indirecto_pct or 0.0, "amount": apu.indirecto_monto or 0.0},
        {"label": "Financiamiento", "pct": apu.financiamiento_pct or 0.0, "amount": apu.financiamiento_monto or 0.0},
        {"label": "Utilidad", "pct": apu.utilidad_pct or 0.0, "amount": apu.utilidad_monto or 0.0},
        {"label": "Cargos adicionales", "pct": apu.cargos_adicionales_pct or 0.0, "amount": apu.cargos_adicionales_monto or 0.0},
    ]
    apu.resumen_programa = {
        "cantidad_objetivo": apu.cantidad_objetivo or 0.0,
        "rendimiento_base": apu.rendimiento_base or 0.0,
        "jornada_horas": apu.jornada_horas or 0.0,
        "jornadas_estimadas": getattr(apu, "jornadas_estimadas", 0.0),
        "importe_partida": apu.importe_partida or 0.0,
        "factor_sobrecosto": getattr(apu, "factor_sobrecosto", 0.0),
    }
    return apu


def _dashboard_data():
    apus = APU.query.order_by(APU.actualizado_en.desc(), APU.id.desc()).all()
    obras = Obra.query.order_by(Obra.actualizado_en.desc(), Obra.id.desc()).limit(12).all()
    total_directo = 0.0
    total_venta = 0.0
    total_materiales = Material.query.count()
    total_mano_obra = ManoObra.query.count()
    total_maquinaria = Maquinaria.query.count()

    for apu in apus:
        recalcular_apu(apu)
        total_directo += float(apu.costo_directo or 0)
        total_venta += float(apu.precio_unitario or 0)

    for obra in obras:
        recalcular_obra(obra)

    return {
        "totales": [
            {"label": "APU activos", "value": len(apus)},
            {"label": "Auxiliares", "value": APU.query.filter(APU.es_auxiliar.is_(True)).count()},
            {"label": "Obras", "value": Obra.query.count()},
            {"label": "Materiales", "value": total_materiales},
            {"label": "Mano de obra", "value": total_mano_obra},
            {"label": "Maquinaria", "value": total_maquinaria},
        ],
        "indicadores": {
            "promedio_directo": (total_directo / len(apus)) if apus else 0.0,
            "promedio_venta": (total_venta / len(apus)) if apus else 0.0,
            "relacion_venta_directo": (total_venta / total_directo) if total_directo else 0.0,
        },
        "apus": apus[:20],
        "obras": obras,
    }


def _set_obra_defaults(obra):
    if obra.indirecto_pct is None:
        obra.indirecto_pct = 12.0
    if obra.financiamiento_pct is None:
        obra.financiamiento_pct = 2.5
    if obra.utilidad_pct is None:
        obra.utilidad_pct = 10.0
    if obra.cargos_adicionales_pct is None:
        obra.cargos_adicionales_pct = 0.0
    if obra.unidad_venta is None:
        obra.unidad_venta = "obra"
    if getattr(obra, "programa_intervalo_dias", None) is None:
        obra.programa_intervalo_dias = 7
    if getattr(obra, "frentes", None) is None:
        obra.frentes = 1.0
    if getattr(obra, "indirecto_campo_pct", None) is None:
        obra.indirecto_campo_pct = 0.0
    if getattr(obra, "indirecto_oficina_pct", None) is None:
        obra.indirecto_oficina_pct = 0.0


def _guardar_obra_desde_form(obra):
    obra.clave = _s(request.form.get("clave"))
    obra.nombre = (request.form.get("nombre") or "").strip()
    obra.cliente = _s(request.form.get("cliente"))
    obra.descripcion = _s(request.form.get("descripcion"))
    obra.ubicacion = _s(request.form.get("ubicacion"))
    obra.unidad_venta = (request.form.get("unidad_venta") or "obra").strip() or "obra"
    obra.indirecto_pct = _f(request.form.get("indirecto_pct"), obra.indirecto_pct or 0.0)
    obra.indirecto_campo_pct = _f(request.form.get("indirecto_campo_pct"), getattr(obra, "indirecto_campo_pct", 0.0) or 0.0)
    obra.indirecto_oficina_pct = _f(request.form.get("indirecto_oficina_pct"), getattr(obra, "indirecto_oficina_pct", 0.0) or 0.0)
    obra.financiamiento_pct = _f(request.form.get("financiamiento_pct"), obra.financiamiento_pct or 0.0)
    obra.utilidad_pct = _f(request.form.get("utilidad_pct"), obra.utilidad_pct or 0.0)
    obra.cargos_adicionales_pct = _f(request.form.get("cargos_adicionales_pct"), obra.cargos_adicionales_pct or 0.0)
    obra.plazo_dias = request.form.get("plazo_dias", type=int) or 0
    obra.programa_intervalo_dias = request.form.get("programa_intervalo_dias", type=int) or 7
    obra.frentes = _f(request.form.get("frentes"), getattr(obra, "frentes", 1.0) or 1.0)
    fecha_inicio = _s(request.form.get("fecha_inicio"))
    fecha_fin = _s(request.form.get("fecha_fin"))
    obra.fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d") if fecha_inicio else None
    obra.fecha_fin = datetime.strptime(fecha_fin, "%Y-%m-%d") if fecha_fin else None
    recalcular_obra(obra)


def _decorate_obra(obra):
    recalcular_obra(obra)
    explosion = explosion_insumos_obra(obra)
    programa = programa_obra(obra)
    partidas = sorted(obra.partidas, key=lambda item: (item.orden or 0, item.id or 0))
    capitulos = {}
    for partida in partidas:
        capitulo = partida.capitulo or "Sin capitulo"
        bucket = capitulos.setdefault(capitulo, {"label": capitulo, "importe": 0.0, "partidas": 0})
        bucket["importe"] += float(partida.importe_venta or 0)
        bucket["partidas"] += 1

    obra.partidas_ordenadas = partidas
    obra.resumen_capitulos = sorted(capitulos.values(), key=lambda item: item["label"])
    obra.explosion_insumos = explosion[:60]
    obra.explosion_total = sum(float(item["importe"] or 0) for item in explosion)
    obra.programa_resumen = programa
    obra.programa_recursos = programa_recursos_obra(obra)
    obra.kpis = [
        {"label": "Partidas", "value": len(partidas)},
        {"label": "Directo", "value": obra.subtotal_directo or 0.0},
        {"label": "Venta", "value": obra.total_venta or 0.0},
        {"label": "Plazo", "value": f"{programa['plazo_dias'] or 0} dias"},
    ]
    obra.sobrecostos = [
        {"label": "Indirectos", "pct": obra.indirecto_pct or 0.0, "amount": obra.indirecto_monto or 0.0},
        {"label": "Financiamiento", "pct": obra.financiamiento_pct or 0.0, "amount": obra.financiamiento_monto or 0.0},
        {"label": "Utilidad", "pct": obra.utilidad_pct or 0.0, "amount": obra.utilidad_monto or 0.0},
        {"label": "Cargos adicionales", "pct": obra.cargos_adicionales_pct or 0.0, "amount": obra.cargos_adicionales_monto or 0.0},
    ]
    obra.indirectos_desglosados = [
        {"label": "Indirecto campo", "pct": getattr(obra, "indirecto_campo_pct", 0.0) or 0.0},
        {"label": "Indirecto oficina", "pct": getattr(obra, "indirecto_oficina_pct", 0.0) or 0.0},
    ]
    obra.cargos_resumen = {
        "directo_global": sum(float(c.importe or 0) for c in obra.cargos if c.incidencia == "directo_global"),
        "indirecto": sum(float(c.importe or 0) for c in obra.cargos if c.incidencia == "indirecto"),
        "cargo_adicional": sum(float(c.importe or 0) for c in obra.cargos if c.incidencia == "cargo_adicional"),
        "retencion": sum(float(c.importe or 0) for c in obra.cargos if c.incidencia == "retencion"),
    }
    return obra


def _save_resource(item, include_provider=False):
    item.nombre = (request.form.get("nombre") or "").strip()
    item.clave = _s(request.form.get("clave"))
    item.categoria = _s(request.form.get("categoria"))
    item.unidad = (request.form.get("unidad") or item.unidad or "pza").strip() or "pza"
    item.precio_unitario = _f(request.form.get("precio_unitario"))
    if isinstance(item, Material):
        item.unidad_compra = _s(request.form.get("unidad_compra"))
        item.cantidad_presentacion = _f(request.form.get("cantidad_presentacion"))
        item.precio_presentacion = _f(request.form.get("precio_presentacion"))
        item.compra_minima = _f(request.form.get("compra_minima"))
    if include_provider:
        item.proveedor = _s(request.form.get("proveedor"))


def _resource_list(model, template_name, title):
    q = (request.args.get("q") or "").strip()
    query = model.query
    if q:
        pattern = f"%{q}%"
        clauses = [model.nombre.ilike(pattern)]
        if hasattr(model, "clave"):
            clauses.append(model.clave.ilike(pattern))
        if hasattr(model, "categoria"):
            clauses.append(model.categoria.ilike(pattern))
        query = query.filter(or_(*clauses))
    items = query.order_by(model.nombre.asc()).all()
    return render_template(template_name, items=items, title=title, q=q)


def _render_apu_form(apu=None, title="APU"):
    if apu:
        _decorate_apu(apu)
        db.session.flush()
    context = _build_catalog_context()
    return render_template(
        "neodata/apu_form.html",
        title=title,
        apu=apu,
        tipos_insumo=TIPOS_INSUMO,
        type_labels=TYPE_LABELS,
        resource_label=_resource_label,
        **context,
    )


@apu_bp.route("/")
@login_required
def index():
    dashboard = _dashboard_data()
    return render_template("neodata/apu_index.html", dashboard=dashboard, title="MAR DATA")


@apu_bp.route("/lista")
@login_required
def apu_list():
    q = (request.args.get("q") or "").strip()
    query = APU.query
    if q:
        pattern = f"%{q}%"
        query = query.filter(or_(APU.concepto.ilike(pattern), APU.clave.ilike(pattern), APU.categoria.ilike(pattern)))
    items = query.order_by(APU.actualizado_en.desc(), APU.id.desc()).all()
    for item in items:
        recalcular_apu(item)
    return render_template("neodata/apu_list.html", items=items, title="Lista de APU", q=q)


@apu_bp.route("/nuevo", methods=["GET", "POST"])
@login_required
def apu_new():
    apu = APU()
    _set_apu_defaults(apu)
    if request.method == "POST":
        _guardar_apu_desde_form(apu)
        db.session.add(apu)
        db.session.commit()
        flash("APU creado correctamente.", "success")
        return redirect(url_for("apu.apu_edit", apu_id=apu.id))
    return _render_apu_form(apu=apu, title="Nuevo APU")


@apu_bp.route("/<int:apu_id>/editar", methods=["GET", "POST"])
@login_required
def apu_edit(apu_id):
    apu = APU.query.get_or_404(apu_id)
    _set_apu_defaults(apu)
    if request.method == "POST":
        _guardar_apu_desde_form(apu)
        db.session.commit()
        flash("APU actualizado.", "success")
        return redirect(url_for("apu.apu_edit", apu_id=apu.id))
    return _render_apu_form(apu=apu, title="Editar APU")


@apu_bp.route("/<int:apu_id>/eliminar", methods=["POST"])
@login_required
def apu_delete(apu_id):
    apu = APU.query.get_or_404(apu_id)
    db.session.delete(apu)
    db.session.commit()
    flash("APU eliminado.", "success")
    return redirect(url_for("apu.apu_list"))


@apu_bp.route("/<int:apu_id>/duplicar", methods=["POST"])
@login_required
def apu_duplica(apu_id):
    apu = APU.query.get_or_404(apu_id)
    nuevo = APU(
        clave=f"{apu.clave}-COPIA" if apu.clave else None,
        concepto=f"{apu.concepto} (Copia)",
        descripcion=apu.descripcion,
        categoria=apu.categoria,
        capitulo=apu.capitulo,
        subcapitulo=apu.subcapitulo,
        alcance=apu.alcance,
        es_auxiliar=apu.es_auxiliar,
        unidad=apu.unidad,
        cantidad_objetivo=apu.cantidad_objetivo,
        rendimiento_base=apu.rendimiento_base,
        jornada_horas=apu.jornada_horas,
        desperdicio_general_pct=apu.desperdicio_general_pct,
        herramienta_menor_pct=apu.herramienta_menor_pct,
        indirecto_pct=apu.indirecto_pct,
        utilidad_pct=apu.utilidad_pct,
        financiamiento_pct=apu.financiamiento_pct,
        cargos_adicionales_pct=apu.cargos_adicionales_pct,
        notas=apu.notas,
    )
    db.session.add(nuevo)
    db.session.flush()

    for detalle in apu.detalles:
        db.session.add(
            APUDetalle(
                apu_id=nuevo.id,
                tipo_insumo=detalle.tipo_insumo,
                referencia_id=detalle.referencia_id,
                descripcion=detalle.descripcion,
                codigo=detalle.codigo,
                categoria=detalle.categoria,
                unidad=detalle.unidad,
                cantidad=detalle.cantidad,
                factor=detalle.factor,
                cuadrilla=detalle.cuadrilla,
                rendimiento=detalle.rendimiento,
                desperdicio_pct=detalle.desperdicio_pct,
                comentario=detalle.comentario,
                precio_unitario=detalle.precio_unitario,
                subtotal=detalle.subtotal,
                auxiliar_apu_id=detalle.auxiliar_apu_id,
            )
        )

    recalcular_apu(nuevo)
    db.session.commit()
    flash("APU duplicado.", "success")
    return redirect(url_for("apu.apu_edit", apu_id=nuevo.id))


@apu_bp.route("/<int:apu_id>/a-catalogo", methods=["POST"])
@login_required
def apu_to_catalogo(apu_id):
    apu = APU.query.get_or_404(apu_id)
    recalcular_apu(apu)

    filtro = db.func.lower(Concepto.nombre_concepto) == apu.concepto.lower()
    if apu.clave and hasattr(Concepto, "clave"):
        filtro = or_(filtro, db.func.lower(Concepto.clave) == apu.clave.lower())

    concepto = Concepto.query.filter(filtro).first()
    descripcion = apu.descripcion or f"Generado desde MAR DATA {apu.clave or apu.id}"

    if concepto:
        concepto.nombre_concepto = apu.concepto
        concepto.unidad = apu.unidad
        concepto.precio_unitario = apu.precio_unitario
        concepto.sistema = "MAR DATA"
        concepto.descripcion = descripcion
        if apu.clave and hasattr(concepto, "clave"):
            concepto.clave = apu.clave
    else:
        kwargs = dict(
            nombre_concepto=apu.concepto,
            unidad=apu.unidad,
            precio_unitario=apu.precio_unitario,
            sistema="MAR DATA",
            descripcion=descripcion,
        )
        if apu.clave and hasattr(Concepto, "clave"):
            kwargs["clave"] = apu.clave
        concepto = Concepto(**kwargs)
        db.session.add(concepto)

    db.session.commit()
    flash("APU enviado al catalogo de conceptos.", "success")
    return redirect(url_for("apu.apu_edit", apu_id=apu.id))


@apu_bp.route("/<int:apu_id>/detalle/agregar", methods=["POST"])
@login_required
def apu_detalle_add(apu_id):
    apu = APU.query.get_or_404(apu_id)
    tipo_insumo = (request.form.get("tipo_insumo") or "").strip()
    referencia_id = request.form.get(f"referencia_id_{tipo_insumo}", type=int)
    recurso = _buscar_recurso(tipo_insumo, referencia_id)

    if not recurso:
        flash("No se encontro el insumo seleccionado.", "danger")
        return redirect(url_for("apu.apu_edit", apu_id=apu.id))

    detalle = APUDetalle(
        apu_id=apu.id,
        tipo_insumo=tipo_insumo,
        referencia_id=recurso.id,
        descripcion=getattr(recurso, "nombre", None) or getattr(recurso, "concepto", ""),
        codigo=getattr(recurso, "clave", None),
        categoria=getattr(recurso, "categoria", None),
        unidad=getattr(recurso, "unidad", "pza"),
        cantidad=_f(request.form.get("cantidad"), 1.0),
        factor=_f(request.form.get("factor"), 1.0),
        cuadrilla=_f(request.form.get("cuadrilla"), 1.0),
        rendimiento=_f(request.form.get("rendimiento"), 0.0),
        desperdicio_pct=_f(request.form.get("desperdicio_pct"), apu.desperdicio_general_pct or 0.0),
        comentario=_s(request.form.get("comentario")),
        precio_unitario=getattr(recurso, "precio_unitario", 0.0),
        unidad_compra=getattr(recurso, "unidad_compra", None),
        cantidad_presentacion=getattr(recurso, "cantidad_presentacion", 0.0),
        precio_presentacion=getattr(recurso, "precio_presentacion", 0.0),
        compra_minima=getattr(recurso, "compra_minima", 0.0),
        auxiliar_apu_id=recurso.id if tipo_insumo == "auxiliar" else None,
    )
    db.session.add(detalle)
    recalcular_apu(apu)
    db.session.commit()
    flash("Insumo agregado al APU.", "success")
    return redirect(url_for("apu.apu_edit", apu_id=apu.id))


@apu_bp.route("/<int:apu_id>/detalle/manual", methods=["POST"])
@login_required
def apu_detalle_manual(apu_id):
    apu = APU.query.get_or_404(apu_id)
    tipo_insumo = (request.form.get("tipo_insumo_manual") or "material").strip()
    descripcion = (request.form.get("descripcion_manual") or "").strip()
    if not descripcion:
        flash("La descripcion del renglon manual es obligatoria.", "danger")
        return redirect(url_for("apu.apu_edit", apu_id=apu.id))

    detalle = APUDetalle(
        apu_id=apu.id,
        tipo_insumo=tipo_insumo,
        descripcion=descripcion,
        codigo=_s(request.form.get("codigo_manual")),
        categoria=_s(request.form.get("categoria_manual")),
        unidad=(request.form.get("unidad_manual") or "pza").strip() or "pza",
        cantidad=_f(request.form.get("cantidad_manual"), 1.0),
        factor=_f(request.form.get("factor_manual"), 1.0),
        cuadrilla=_f(request.form.get("cuadrilla_manual"), 1.0),
        rendimiento=_f(request.form.get("rendimiento_manual"), 0.0),
        desperdicio_pct=_f(request.form.get("desperdicio_manual"), apu.desperdicio_general_pct or 0.0),
        comentario=_s(request.form.get("comentario_manual")),
        precio_unitario=_f(request.form.get("precio_unitario_manual")),
    )
    db.session.add(detalle)
    recalcular_apu(apu)
    db.session.commit()
    flash("Renglon manual agregado.", "success")
    return redirect(url_for("apu.apu_edit", apu_id=apu.id))


@apu_bp.route("/detalle/<int:detalle_id>/actualizar", methods=["POST"])
@login_required
def apu_detalle_update(detalle_id):
    detalle = APUDetalle.query.get_or_404(detalle_id)
    detalle.descripcion = (request.form.get("descripcion") or detalle.descripcion).strip()
    detalle.codigo = _s(request.form.get("codigo"))
    detalle.categoria = _s(request.form.get("categoria"))
    detalle.unidad = (request.form.get("unidad") or detalle.unidad or "pza").strip() or "pza"
    detalle.cantidad = _f(request.form.get("cantidad"), detalle.cantidad)
    detalle.factor = _f(request.form.get("factor"), detalle.factor or 1.0)
    detalle.cuadrilla = _f(request.form.get("cuadrilla"), detalle.cuadrilla or 1.0)
    detalle.rendimiento = _f(request.form.get("rendimiento"), detalle.rendimiento or 0.0)
    detalle.desperdicio_pct = _f(request.form.get("desperdicio_pct"), detalle.desperdicio_pct or 0.0)
    detalle.precio_unitario = _f(request.form.get("precio_unitario"), detalle.precio_unitario)
    detalle.comentario = _s(request.form.get("comentario"))
    recalcular_apu(detalle.apu)
    db.session.commit()
    flash("Renglon actualizado.", "success")
    return redirect(url_for("apu.apu_edit", apu_id=detalle.apu_id))


@apu_bp.route("/detalle/<int:detalle_id>/eliminar", methods=["POST"])
@login_required
def apu_detalle_delete(detalle_id):
    detalle = APUDetalle.query.get_or_404(detalle_id)
    apu_id = detalle.apu_id
    apu = detalle.apu
    db.session.delete(detalle)
    recalcular_apu(apu)
    db.session.commit()
    flash("Renglon eliminado.", "success")
    return redirect(url_for("apu.apu_edit", apu_id=apu_id))


@apu_bp.route("/materiales")
@login_required
def materiales_list():
    return _resource_list(Material, "neodata/materiales_list.html", "Materiales")


@apu_bp.route("/materiales/nuevo", methods=["GET", "POST"])
@login_required
def materiales_new():
    item = Material(unidad="kg")
    if request.method == "POST":
        _save_resource(item, include_provider=True)
        db.session.add(item)
        db.session.commit()
        flash("Material creado.", "success")
        return redirect(url_for("apu.materiales_list"))
    return render_template("neodata/recurso_form.html", title="Nuevo material", item=item, back=url_for("apu.materiales_list"), include_provider=True, include_packaging=True)


@apu_bp.route("/materiales/<int:item_id>/editar", methods=["GET", "POST"])
@login_required
def materiales_edit(item_id):
    item = Material.query.get_or_404(item_id)
    if request.method == "POST":
        _save_resource(item, include_provider=True)
        db.session.commit()
        flash("Material actualizado.", "success")
        return redirect(url_for("apu.materiales_list"))
    return render_template("neodata/recurso_form.html", title="Editar material", item=item, back=url_for("apu.materiales_list"), include_provider=True, include_packaging=True)


@apu_bp.route("/materiales/<int:item_id>/eliminar", methods=["POST"])
@login_required
def materiales_delete(item_id):
    item = Material.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash("Material eliminado.", "success")
    return redirect(url_for("apu.materiales_list"))


@apu_bp.route("/mano-obra")
@login_required
def mano_obra_list():
    return _resource_list(ManoObra, "neodata/mano_obra_list.html", "Mano de obra")


@apu_bp.route("/mano-obra/nuevo", methods=["GET", "POST"])
@login_required
def mano_obra_new():
    item = ManoObra(unidad="jor")
    if request.method == "POST":
        _save_resource(item)
        db.session.add(item)
        db.session.commit()
        flash("Recurso de mano de obra creado.", "success")
        return redirect(url_for("apu.mano_obra_list"))
    return render_template("neodata/recurso_form.html", title="Nueva mano de obra", item=item, back=url_for("apu.mano_obra_list"), include_provider=False, include_packaging=False)


@apu_bp.route("/mano-obra/<int:item_id>/editar", methods=["GET", "POST"])
@login_required
def mano_obra_edit(item_id):
    item = ManoObra.query.get_or_404(item_id)
    if request.method == "POST":
        _save_resource(item)
        db.session.commit()
        flash("Recurso actualizado.", "success")
        return redirect(url_for("apu.mano_obra_list"))
    return render_template("neodata/recurso_form.html", title="Editar mano de obra", item=item, back=url_for("apu.mano_obra_list"), include_provider=False, include_packaging=False)


@apu_bp.route("/mano-obra/<int:item_id>/eliminar", methods=["POST"])
@login_required
def mano_obra_delete(item_id):
    item = ManoObra.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash("Recurso eliminado.", "success")
    return redirect(url_for("apu.mano_obra_list"))


@apu_bp.route("/maquinaria")
@login_required
def maquinaria_list():
    return _resource_list(Maquinaria, "neodata/maquinaria_list.html", "Maquinaria")


@apu_bp.route("/maquinaria/nuevo", methods=["GET", "POST"])
@login_required
def maquinaria_new():
    item = Maquinaria(unidad="hr")
    if request.method == "POST":
        _save_resource(item)
        db.session.add(item)
        db.session.commit()
        flash("Maquinaria creada.", "success")
        return redirect(url_for("apu.maquinaria_list"))
    return render_template("neodata/recurso_form.html", title="Nueva maquinaria", item=item, back=url_for("apu.maquinaria_list"), include_provider=False, include_packaging=False)


@apu_bp.route("/maquinaria/<int:item_id>/editar", methods=["GET", "POST"])
@login_required
def maquinaria_edit(item_id):
    item = Maquinaria.query.get_or_404(item_id)
    if request.method == "POST":
        _save_resource(item)
        db.session.commit()
        flash("Maquinaria actualizada.", "success")
        return redirect(url_for("apu.maquinaria_list"))
    return render_template("neodata/recurso_form.html", title="Editar maquinaria", item=item, back=url_for("apu.maquinaria_list"), include_provider=False, include_packaging=False)


@apu_bp.route("/maquinaria/<int:item_id>/eliminar", methods=["POST"])
@login_required
def maquinaria_delete(item_id):
    item = Maquinaria.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash("Maquinaria eliminada.", "success")
    return redirect(url_for("apu.maquinaria_list"))


@apu_bp.route("/plantillas")
@login_required
def plantillas():
    return render_template("neodata/plantillas.html", items=_load_plantillas(), title="Plantillas")


@apu_bp.route("/generador", methods=["GET", "POST"])
@login_required
def generador():
    plantillas_data = _load_plantillas()
    if request.method == "POST":
        plantilla_nombre = (request.form.get("plantilla") or "").strip()
        espesor_mm = _f(request.form.get("espesor_mm"), 1.0)
        rendimiento = _f(request.form.get("rendimiento"), 1.0) or 1.0
        plantilla = next((p for p in plantillas_data if p.get("nombre") == plantilla_nombre), None)

        if not plantilla:
            flash("No se encontro la plantilla seleccionada.", "danger")
            return redirect(url_for("apu.generador"))

        apu = APU(
            clave=plantilla.get("clave"),
            concepto=f"{plantilla.get('nombre')} {espesor_mm:g} mm",
            descripcion=plantilla.get("descripcion"),
            categoria=plantilla.get("categoria"),
            unidad=plantilla.get("unidad") or "m2",
            cantidad_objetivo=1.0,
            rendimiento_base=rendimiento,
            jornada_horas=8.0,
            herramienta_menor_pct=_f(plantilla.get("herramienta_menor"), 3.0),
            indirecto_pct=_f(plantilla.get("indirecto"), 12.0),
            utilidad_pct=_f(plantilla.get("utilidad"), 10.0),
            financiamiento_pct=_f(plantilla.get("financiamiento"), 2.5),
            cargos_adicionales_pct=_f(plantilla.get("cargos"), 0.0),
        )
        db.session.add(apu)
        db.session.flush()

        for componente in plantilla.get("componentes", []):
            tipo = componente.get("tipo")
            model = _resource_for_tipo(tipo)
            if not model:
                continue

            buscar = (componente.get("buscar") or "").strip()
            recurso = model.query.filter(model.nombre.ilike(f"%{buscar}%")).order_by(model.nombre.asc()).first()
            if not recurso:
                continue

            cantidad = _f(componente.get("consumo_fijo"))
            if not cantidad:
                cantidad = _f(componente.get("factor_por_mm")) * espesor_mm

            db.session.add(
                APUDetalle(
                    apu_id=apu.id,
                    tipo_insumo=tipo,
                    referencia_id=recurso.id,
                    descripcion=recurso.nombre,
                    codigo=getattr(recurso, "clave", None),
                    categoria=getattr(recurso, "categoria", None),
                    unidad=recurso.unidad,
                    cantidad=cantidad,
                    factor=1.0,
                    cuadrilla=_f(componente.get("cuadrilla"), 1.0),
                    rendimiento=rendimiento if tipo in {"mano_obra", "maquinaria"} else 0.0,
                    desperdicio_pct=_f(componente.get("desperdicio_pct"), 0.0),
                    comentario=_s(componente.get("comentario")),
                    precio_unitario=recurso.precio_unitario,
                )
            )

        recalcular_apu(apu)
        db.session.commit()
        flash("APU generado desde plantilla.", "success")
        return redirect(url_for("apu.apu_edit", apu_id=apu.id))

    return render_template("neodata/generador.html", plantillas=plantillas_data, title="Generador automatico")


@apu_bp.route("/obras")
@login_required
def obras_list():
    q = (request.args.get("q") or "").strip()
    query = Obra.query
    if q:
        pattern = f"%{q}%"
        query = query.filter(or_(Obra.nombre.ilike(pattern), Obra.clave.ilike(pattern), Obra.cliente.ilike(pattern)))
    items = query.order_by(Obra.actualizado_en.desc(), Obra.id.desc()).all()
    for item in items:
        recalcular_obra(item)
    return render_template("neodata/obra_list.html", items=items, title="Obras MAR Data", q=q)


@apu_bp.route("/obras/nueva", methods=["GET", "POST"])
@login_required
def obra_new():
    obra = Obra()
    _set_obra_defaults(obra)
    if request.method == "POST":
        if not (request.form.get("nombre") or "").strip():
            flash("El nombre de la obra es obligatorio.", "danger")
            return render_template("neodata/obra_form.html", obra=obra, apus=[], title="Nueva obra")
        _guardar_obra_desde_form(obra)
        db.session.add(obra)
        db.session.commit()
        flash("Obra creada.", "success")
        return redirect(url_for("apu.obra_edit", obra_id=obra.id))
    return render_template("neodata/obra_form.html", obra=obra, apus=APU.query.order_by(APU.concepto.asc()).all(), title="Nueva obra")


@apu_bp.route("/obras/<int:obra_id>/editar", methods=["GET", "POST"])
@login_required
def obra_edit(obra_id):
    obra = Obra.query.get_or_404(obra_id)
    _set_obra_defaults(obra)
    if request.method == "POST":
        _guardar_obra_desde_form(obra)
        db.session.commit()
        flash("Encabezado de obra actualizado.", "success")
        return redirect(url_for("apu.obra_edit", obra_id=obra.id))
    _decorate_obra(obra)
    apus = APU.query.order_by(APU.categoria.asc(), APU.concepto.asc()).all()
    return render_template("neodata/obra_form.html", obra=obra, apus=apus, title="Editar obra")


@apu_bp.route("/obras/<int:obra_id>/eliminar", methods=["POST"])
@login_required
def obra_delete(obra_id):
    obra = Obra.query.get_or_404(obra_id)
    db.session.delete(obra)
    db.session.commit()
    flash("Obra eliminada.", "success")
    return redirect(url_for("apu.obras_list"))


@apu_bp.route("/obras/<int:obra_id>/partidas/agregar", methods=["POST"])
@login_required
def obra_partida_add(obra_id):
    obra = Obra.query.get_or_404(obra_id)
    apu_id = request.form.get("apu_id", type=int)
    apu = APU.query.get_or_404(apu_id)
    partida = ObraPartida(
        obra_id=obra.id,
        apu_id=apu.id,
        orden=len(obra.partidas) + 1,
        capitulo=apu.capitulo or apu.categoria,
        subcapitulo=apu.subcapitulo,
        clave=apu.clave,
        concepto=apu.concepto,
        unidad=apu.unidad,
        cantidad=_f(request.form.get("cantidad"), 1.0),
        rendimiento=_f(request.form.get("rendimiento"), apu.rendimiento_base or 0.0),
        comentario=_s(request.form.get("comentario")),
    )
    db.session.add(partida)
    recalcular_obra(obra)
    db.session.commit()
    flash("Partida agregada al presupuesto.", "success")
    return redirect(url_for("apu.obra_edit", obra_id=obra.id))


@apu_bp.route("/obras/<int:obra_id>/partidas/importar", methods=["POST"])
@login_required
def obra_partidas_import(obra_id):
    obra = Obra.query.get_or_404(obra_id)
    raw = (request.form.get("matriz_masiva") or "").strip()
    if not raw:
        flash("Pega primero las partidas a importar.", "danger")
        return redirect(url_for("apu.obra_edit", obra_id=obra.id))

    creadas = 0
    errores = []
    base_orden = len(obra.partidas)
    for idx, line in enumerate(raw.splitlines(), start=1):
        row = line.strip()
        if not row:
            continue
        cols = [c.strip() for c in re.split(r"\t|,", row)]
        if len(cols) < 2:
            errores.append(f"Linea {idx}: faltan columnas.")
            continue

        clave = cols[0] if len(cols) > 0 else ""
        cantidad = _f(cols[-1], 0.0)
        capitulo = cols[1] if len(cols) > 2 else None
        subcapitulo = cols[2] if len(cols) > 3 else None

        apu = None
        if clave:
            apu = APU.query.filter(db.func.lower(APU.clave) == clave.lower()).first()
        if not apu and len(cols) > 3:
            concepto_ref = cols[3]
            apu = APU.query.filter(db.func.lower(APU.concepto) == concepto_ref.lower()).first()
        if not apu:
            errores.append(f"Linea {idx}: no se encontro APU para '{clave or (cols[3] if len(cols) > 3 else '')}'.")
            continue

        partida = ObraPartida(
            obra_id=obra.id,
            apu_id=apu.id,
            orden=base_orden + creadas + 1,
            capitulo=capitulo or apu.capitulo or apu.categoria,
            subcapitulo=subcapitulo or apu.subcapitulo,
            clave=apu.clave,
            concepto=apu.concepto,
            unidad=apu.unidad,
            cantidad=cantidad or 1.0,
            rendimiento=apu.rendimiento_base or 0.0,
            comentario=None,
        )
        db.session.add(partida)
        creadas += 1

    recalcular_obra(obra)
    db.session.commit()
    if creadas:
        flash(f"Importacion masiva completada: {creadas} partidas agregadas.", "success")
    if errores:
        flash(" | ".join(errores[:4]), "warning")
    return redirect(url_for("apu.obra_edit", obra_id=obra.id))


@apu_bp.route("/obras/<int:obra_id>/partidas/actualizar", methods=["POST"])
@login_required
def obra_partidas_update(obra_id):
    obra = Obra.query.get_or_404(obra_id)
    ids = request.form.getlist("partida_id[]")
    ordenes = request.form.getlist("orden[]")
    capitulos = request.form.getlist("capitulo[]")
    subcapitulos = request.form.getlist("subcapitulo[]")
    claves = request.form.getlist("clave[]")
    conceptos = request.form.getlist("concepto[]")
    unidades = request.form.getlist("unidad[]")
    cantidades = request.form.getlist("cantidad[]")
    rendimientos = request.form.getlist("rendimiento[]")
    comentarios = request.form.getlist("comentario[]")

    for i, partida_id in enumerate(ids):
        partida = ObraPartida.query.get(int(partida_id))
        if not partida or partida.obra_id != obra.id:
            continue
        partida.orden = int(_f(ordenes[i], partida.orden)) if i < len(ordenes) else partida.orden
        partida.capitulo = (capitulos[i] if i < len(capitulos) else partida.capitulo) or None
        partida.subcapitulo = (subcapitulos[i] if i < len(subcapitulos) else partida.subcapitulo) or None
        partida.clave = (claves[i] if i < len(claves) else partida.clave) or None
        partida.concepto = (conceptos[i] if i < len(conceptos) else partida.concepto) or None
        partida.unidad = (unidades[i] if i < len(unidades) else partida.unidad) or partida.unidad
        partida.cantidad = _f(cantidades[i], partida.cantidad) if i < len(cantidades) else partida.cantidad
        partida.rendimiento = _f(rendimientos[i], partida.rendimiento) if i < len(rendimientos) else partida.rendimiento
        partida.comentario = (comentarios[i] if i < len(comentarios) else partida.comentario) or None

    recalcular_obra(obra)
    db.session.commit()
    flash("Presupuesto actualizado en bloque.", "success")
    return redirect(url_for("apu.obra_edit", obra_id=obra.id))


@apu_bp.route("/obras/partidas/<int:partida_id>/eliminar", methods=["POST"])
@login_required
def obra_partida_delete(partida_id):
    partida = ObraPartida.query.get_or_404(partida_id)
    obra_id = partida.obra_id
    db.session.delete(partida)
    obra = Obra.query.get_or_404(obra_id)
    recalcular_obra(obra)
    db.session.commit()
    flash("Partida eliminada.", "success")
    return redirect(url_for("apu.obra_edit", obra_id=obra_id))


@apu_bp.route("/obras/<int:obra_id>/cargos/agregar", methods=["POST"])
@login_required
def obra_cargo_add(obra_id):
    obra = Obra.query.get_or_404(obra_id)
    nombre = (request.form.get("nombre") or "").strip()
    if not nombre:
        flash("El nombre del cargo global es obligatorio.", "danger")
        return redirect(url_for("apu.obra_edit", obra_id=obra.id))
    cargo = ObraCargo(
        obra_id=obra.id,
        nombre=nombre,
        categoria=_s(request.form.get("categoria")),
        incidencia=(request.form.get("incidencia") or "indirecto").strip(),
        unidad=(request.form.get("unidad") or "mes").strip() or "mes",
        cantidad=_f(request.form.get("cantidad"), 1.0),
        precio_unitario=_f(request.form.get("precio_unitario")),
        comentario=_s(request.form.get("comentario")),
    )
    db.session.add(cargo)
    recalcular_obra(obra)
    db.session.commit()
    flash("Cargo global agregado.", "success")
    return redirect(url_for("apu.obra_edit", obra_id=obra.id))


@apu_bp.route("/obras/<int:obra_id>/cargos/sugeridos", methods=["POST"])
@login_required
def obra_cargos_seed(obra_id):
    obra = Obra.query.get_or_404(obra_id)
    existentes = {c.nombre.lower() for c in obra.cargos}
    creados = 0
    for nombre, categoria, incidencia, unidad in SUGERENCIAS_CARGOS:
        if nombre.lower() in existentes:
            continue
        db.session.add(
            ObraCargo(
                obra_id=obra.id,
                nombre=nombre,
                categoria=categoria,
                incidencia=incidencia,
                unidad=unidad,
                cantidad=1.0,
                precio_unitario=0.0,
            )
        )
        creados += 1
    recalcular_obra(obra)
    db.session.commit()
    flash(f"Se cargaron {creados} cargos sugeridos.", "success")
    return redirect(url_for("apu.obra_edit", obra_id=obra.id))


@apu_bp.route("/obras/cargos/<int:cargo_id>/eliminar", methods=["POST"])
@login_required
def obra_cargo_delete(cargo_id):
    cargo = ObraCargo.query.get_or_404(cargo_id)
    obra_id = cargo.obra_id
    db.session.delete(cargo)
    obra = Obra.query.get_or_404(obra_id)
    recalcular_obra(obra)
    db.session.commit()
    flash("Cargo global eliminado.", "success")
    return redirect(url_for("apu.obra_edit", obra_id=obra_id))


@apu_bp.route("/obras/<int:obra_id>/cargos/actualizar", methods=["POST"])
@login_required
def obra_cargos_update(obra_id):
    obra = Obra.query.get_or_404(obra_id)
    ids = request.form.getlist("cargo_id[]")
    nombres = request.form.getlist("cargo_nombre[]")
    categorias = request.form.getlist("cargo_categoria[]")
    incidencias = request.form.getlist("cargo_incidencia[]")
    unidades = request.form.getlist("cargo_unidad[]")
    cantidades = request.form.getlist("cargo_cantidad[]")
    precios = request.form.getlist("cargo_precio[]")
    comentarios = request.form.getlist("cargo_comentario[]")

    for i, cargo_id in enumerate(ids):
        cargo = ObraCargo.query.get(int(cargo_id))
        if not cargo or cargo.obra_id != obra.id:
            continue
        cargo.nombre = (nombres[i] if i < len(nombres) else cargo.nombre).strip() or cargo.nombre
        cargo.categoria = (categorias[i] if i < len(categorias) else cargo.categoria) or None
        cargo.incidencia = (incidencias[i] if i < len(incidencias) else cargo.incidencia) or cargo.incidencia
        cargo.unidad = (unidades[i] if i < len(unidades) else cargo.unidad) or cargo.unidad
        cargo.cantidad = _f(cantidades[i], cargo.cantidad) if i < len(cantidades) else cargo.cantidad
        cargo.precio_unitario = _f(precios[i], cargo.precio_unitario) if i < len(precios) else cargo.precio_unitario
        cargo.comentario = (comentarios[i] if i < len(comentarios) else cargo.comentario) or None

    recalcular_obra(obra)
    db.session.commit()
    flash("Cargos globales actualizados.", "success")
    return redirect(url_for("apu.obra_edit", obra_id=obra.id))


@apu_bp.route("/obras/<int:obra_id>/reporte")
@login_required
def obra_report(obra_id):
    obra = Obra.query.get_or_404(obra_id)
    _decorate_obra(obra)
    return render_template("neodata/obra_report.html", obra=obra, title=f"Reporte {obra.nombre}")


@apu_bp.route("/obras/<int:obra_id>/reporte-ejecutivo")
@login_required
def obra_report_ejecutivo(obra_id):
    obra = Obra.query.get_or_404(obra_id)
    _decorate_obra(obra)
    top_recursos = sorted(obra.programa_recursos, key=lambda item: float(item.get("importe_total") or 0), reverse=True)[:12]
    return render_template("neodata/obra_executive.html", obra=obra, top_recursos=top_recursos, title=f"Ejecutivo {obra.nombre}")


@apu_bp.route("/obras/<int:obra_id>/exportar-csv")
@login_required
def obra_export_csv(obra_id):
    obra = Obra.query.get_or_404(obra_id)
    _decorate_obra(obra)

    rows = [
        "SECCION,ORDEN,CAPITULO,SUBCAPITULO,CLAVE,CONCEPTO,UNIDAD,CANTIDAD,PRECIO_UNITARIO,IMPORTE_DIRECTO,IMPORTE_VENTA",
    ]
    for partida in obra.partidas_ordenadas:
        values = [
            "PARTIDA",
            str(partida.orden or ""),
            (partida.capitulo or "").replace(",", " "),
            (partida.subcapitulo or "").replace(",", " "),
            (partida.clave or "").replace(",", " "),
            (partida.concepto or "").replace(",", " "),
            (partida.unidad or "").replace(",", " "),
            f"{float(partida.cantidad or 0):.4f}",
            f"{float(partida.precio_unitario or 0):.4f}",
            f"{float(partida.importe_directo or 0):.4f}",
            f"{float(partida.importe_venta or 0):.4f}",
        ]
        rows.append(",".join(values))

    rows.append("")
    rows.append("SECCION,TIPO,CODIGO,DESCRIPCION,UNIDAD,CANTIDAD,PRECIO_UNITARIO,IMPORTE,PARTIDAS")
    for item in obra.explosion_insumos:
        values = [
            "EXPLOSION",
            (item.get("tipo") or "").replace(",", " "),
            (item.get("codigo") or "").replace(",", " "),
            (item.get("descripcion") or "").replace(",", " "),
            (item.get("unidad") or "").replace(",", " "),
            f"{float(item.get('cantidad') or 0):.4f}",
            f"{float(item.get('precio_unitario') or 0):.4f}",
            f"{float(item.get('importe') or 0):.4f}",
            " | ".join(item.get("partidas") or []).replace(",", " "),
        ]
        rows.append(",".join(values))

    filename = f"obra_{obra.clave or obra.id}.csv"
    return Response(
        "\n".join(rows),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@apu_bp.route("/obras/<int:obra_id>/exportar-xlsx")
@login_required
def obra_export_xlsx(obra_id):
    if Workbook is None:
        flash("openpyxl no esta disponible en este servidor.", "danger")
        return redirect(url_for("apu.obra_edit", obra_id=obra_id))

    obra = Obra.query.get_or_404(obra_id)
    _decorate_obra(obra)

    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"
    ws["A1"] = "PRESUPUESTO MAR DATA"
    ws["A2"] = obra.nombre
    ws["A3"] = f"Clave: {obra.clave or '-'}"
    ws["D2"] = f"Cliente: {obra.cliente or '-'}"
    ws["D3"] = f"Ubicación: {obra.ubicacion or '-'}"
    ws["A5"] = "Ord"
    ws["B5"] = "Capítulo"
    ws["C5"] = "Subcapítulo"
    ws["D5"] = "Clave"
    ws["E5"] = "Concepto"
    ws["F5"] = "Unidad"
    ws["G5"] = "Cantidad"
    ws["H5"] = "P.U."
    ws["I5"] = "Importe"
    fill = PatternFill("solid", fgColor="1F56A7")
    for cell in ws["5:5"]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    row = 6
    for partida in obra.partidas_ordenadas:
        ws.cell(row=row, column=1, value=partida.orden or 0)
        ws.cell(row=row, column=2, value=partida.capitulo or "")
        ws.cell(row=row, column=3, value=partida.subcapitulo or "")
        ws.cell(row=row, column=4, value=partida.clave or "")
        ws.cell(row=row, column=5, value=partida.concepto or "")
        ws.cell(row=row, column=6, value=partida.unidad or "")
        ws.cell(row=row, column=7, value=float(partida.cantidad or 0))
        ws.cell(row=row, column=8, value=float(partida.precio_unitario or 0))
        ws.cell(row=row, column=9, value=float(partida.importe_venta or 0))
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Explosión consolidada")
    row += 1
    headers = ["Tipo", "Código", "Descripción", "Unidad", "Cantidad", "P.U.", "Importe"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    row += 1
    for item in obra.explosion_insumos:
        ws.cell(row=row, column=1, value=item.get("tipo") or "")
        ws.cell(row=row, column=2, value=item.get("codigo") or "")
        ws.cell(row=row, column=3, value=item.get("descripcion") or "")
        ws.cell(row=row, column=4, value=item.get("unidad") or "")
        ws.cell(row=row, column=5, value=float(item.get("cantidad") or 0))
        ws.cell(row=row, column=6, value=float(item.get("precio_unitario") or 0))
        ws.cell(row=row, column=7, value=float(item.get("importe") or 0))
        row += 1

    ws2 = wb.create_sheet("Programa")
    ws2["A1"] = "PROGRAMA BASE POR PERIODOS"
    headers_programa = ["Periodo", "Orden", "Capítulo", "Concepto", "Inicio", "Fin", "Jornadas"]
    for col, header in enumerate(headers_programa, start=1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    row2 = 4
    for item in obra.programa_resumen.get("calendario", []):
        ws2.cell(row=row2, column=1, value=item.get("periodo") or "")
        ws2.cell(row=row2, column=2, value=item.get("orden") or 0)
        ws2.cell(row=row2, column=3, value=item.get("capitulo") or "")
        ws2.cell(row=row2, column=4, value=item.get("concepto") or "")
        ws2.cell(row=row2, column=5, value=item.get("inicio").strftime("%d/%m/%Y") if item.get("inicio") else "")
        ws2.cell(row=row2, column=6, value=item.get("fin").strftime("%d/%m/%Y") if item.get("fin") else "")
        ws2.cell(row=row2, column=7, value=float(item.get("jornadas") or 0))
        row2 += 1

    ws3 = wb.create_sheet("Recursos")
    ws3["A1"] = "PROGRAMA DE RECURSOS / EQUIPO"
    headers_rec = ["Tipo", "Código", "Descripción", "Unidad", "Cantidad Total", "Importe Total", "Periodos"]
    for col, header in enumerate(headers_rec, start=1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    row3 = 4
    for item in obra.programa_recursos:
        periodos = ", ".join(p.get("periodo") or "" for p in item.get("periodos", []))
        ws3.cell(row=row3, column=1, value=item.get("tipo") or "")
        ws3.cell(row=row3, column=2, value=item.get("codigo") or "")
        ws3.cell(row=row3, column=3, value=item.get("descripcion") or "")
        ws3.cell(row=row3, column=4, value=item.get("unidad") or "")
        ws3.cell(row=row3, column=5, value=float(item.get("cantidad_total") or 0))
        ws3.cell(row=row3, column=6, value=float(item.get("importe_total") or 0))
        ws3.cell(row=row3, column=7, value=periodos)
        row3 += 1

    ws4 = wb.create_sheet("Cargos")
    ws4["A1"] = "CARGOS GLOBALES DE OBRA"
    headers_cargos = ["Nombre", "Categoría", "Incidencia", "Unidad", "Cantidad", "P.U.", "Importe", "Comentario"]
    for col, header in enumerate(headers_cargos, start=1):
        cell = ws4.cell(row=3, column=col, value=header)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    row4 = 4
    for cargo in obra.cargos:
        ws4.cell(row=row4, column=1, value=cargo.nombre or "")
        ws4.cell(row=row4, column=2, value=cargo.categoria or "")
        ws4.cell(row=row4, column=3, value=cargo.incidencia or "")
        ws4.cell(row=row4, column=4, value=cargo.unidad or "")
        ws4.cell(row=row4, column=5, value=float(cargo.cantidad or 0))
        ws4.cell(row=row4, column=6, value=float(cargo.precio_unitario or 0))
        ws4.cell(row=row4, column=7, value=float(cargo.importe or 0))
        ws4.cell(row=row4, column=8, value=cargo.comentario or "")
        row4 += 1

    auto_widths = {"A": 10, "B": 18, "C": 18, "D": 14, "E": 48, "F": 10, "G": 14, "H": 14, "I": 14}
    for col, width in auto_widths.items():
        ws.column_dimensions[col].width = width
    for sheet in (ws2, ws3):
        for col, width in {"A": 14, "B": 10, "C": 18, "D": 42, "E": 14, "F": 14, "G": 18}.items():
            sheet.column_dimensions[col].width = width
    for col, width in {"A": 24, "B": 18, "C": 18, "D": 12, "E": 12, "F": 12, "G": 14, "H": 30}.items():
        ws4.column_dimensions[col].width = width

    out = io.BytesIO()
    wb.save(out)
    payload = out.getvalue()
    out.close()
    filename = f"obra_{obra.clave or obra.id}.xlsx"
    return Response(payload, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@apu_bp.route("/obras/<int:obra_id>/exportar-pdf")
@login_required
def obra_export_pdf(obra_id):
    obra = Obra.query.get_or_404(obra_id)
    _decorate_obra(obra)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=26 * mm,
        bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="MarSmall", fontName="Helvetica", fontSize=8, leading=10, textColor=colors.HexColor("#46576B"), splitLongWords=False))
    styles.add(ParagraphStyle(name="MarCell", fontName="Helvetica", fontSize=7.3, leading=9, splitLongWords=False))
    styles.add(ParagraphStyle(name="MarCellCenter", fontName="Helvetica", fontSize=7.3, leading=9, alignment=1, splitLongWords=False))
    styles.add(ParagraphStyle(name="MarCellRight", fontName="Helvetica", fontSize=7.3, leading=9, alignment=2, splitLongWords=False))

    logo_path = os.path.join(PROJECT_ROOT, "static", "logo.png")

    def _header_footer(canv, _doc):
        canv.saveState()
        canv.setFillColor(colors.HexColor("#1f56a7"))
        canv.rect(0, A4[1] - 32, A4[0], 32, stroke=0, fill=1)
        if os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                scale = min((16 * mm) / iw, (16 * mm) / ih)
                canv.drawImage(img, 12, A4[1] - (ih * scale) - 8, width=iw * scale, height=ih * scale, mask="auto")
            except Exception:
                pass
        canv.setFillColor(colors.white)
        canv.setFont("Helvetica-Bold", 15)
        canv.drawRightString(A4[0] - 12, A4[1] - 14, "PRESUPUESTO MAR DATA")
        canv.setFont("Helvetica", 9)
        canv.drawRightString(A4[0] - 12, A4[1] - 25, "Poliutech · Recubrimientos Especializados")
        canv.setFillColor(colors.HexColor("#6B7A8A"))
        canv.setFont("Helvetica", 8)
        canv.drawString(12, 10, f"Obra: {obra.clave or obra.id} · Generado {datetime.utcnow().strftime('%d/%m/%Y %H:%M')}")
        canv.restoreState()

    story = []
    story.append(Paragraph(f"<b>{obra.nombre}</b>", styles["Title"]))
    meta_left = [
        f"Clave: {obra.clave or '-'}",
        f"Cliente: {obra.cliente or '-'}",
        f"Ubicación: {_truncate_text(obra.ubicacion or '-', 70)}",
    ]
    meta_right = [
        f"Plazo: {obra.plazo_dias or 0} dias",
        f"Unidad de venta: {obra.unidad_venta or 'obra'}",
        f"Partidas: {len(obra.partidas_ordenadas)}",
    ]
    meta_table = Table(
        [
            [
                Paragraph("<br/>".join(meta_left), styles["MarSmall"]),
                Paragraph("<br/>".join(meta_right), styles["MarSmall"]),
            ]
        ],
        colWidths=[93 * mm, 92 * mm],
    )
    meta_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story.append(meta_table)
    story.append(Spacer(1, 4 * mm))

    summary_data = [["Directo", "Indirecto", "Financiamiento", "Utilidad", "Cargos", "Venta"]]
    summary_data.append(
        [
            Paragraph(_fmt_money(obra.subtotal_directo), styles["MarCellCenter"]),
            Paragraph(_fmt_money(obra.indirecto_monto), styles["MarCellCenter"]),
            Paragraph(_fmt_money(obra.financiamiento_monto), styles["MarCellCenter"]),
            Paragraph(_fmt_money(obra.utilidad_monto), styles["MarCellCenter"]),
            Paragraph(_fmt_money(obra.cargos_adicionales_monto), styles["MarCellCenter"]),
            Paragraph(_fmt_money(obra.total_venta), styles["MarCellCenter"]),
        ]
    )
    summary_table = Table(summary_data, colWidths=[30 * mm] * 6, repeatRows=1)
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#102542")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C8D4E0")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 5 * mm))

    if getattr(obra, "indirectos_desglosados", None):
        story.append(Paragraph("<b>Desglose de indirectos y globales</b>", styles["Heading3"]))
        indirect_rows = [["Concepto", "%", "Monto"]]
        for item in obra.sobrecostos:
            indirect_rows.append(
                [
                    Paragraph(item["label"], styles["MarCell"]),
                    Paragraph("{:,.2f}%".format(float(item["pct"] or 0)), styles["MarCellCenter"]),
                    Paragraph(_fmt_money(item["amount"]), styles["MarCellRight"]),
                ]
            )
        for item in obra.indirectos_desglosados:
            indirect_rows.append(
                [
                    Paragraph(item["label"], styles["MarCell"]),
                    Paragraph("{:,.2f}%".format(float(item["pct"] or 0)), styles["MarCellCenter"]),
                    Paragraph("-", styles["MarCellRight"]),
                ]
            )
        indirect_table = Table(indirect_rows, colWidths=[90 * mm, 28 * mm, 42 * mm], repeatRows=1)
        indirect_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#102542")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0D9E3")),
                    ("PADDING", (0, 0), (-1, -1), 4),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ]
            )
        )
        story.append(indirect_table)
        story.append(Spacer(1, 5 * mm))

    if getattr(obra, "cargos", None):
        story.append(Paragraph("<b>Cargos globales de obra</b>", styles["Heading3"]))
        cargos_rows = [["Nombre", "Incidencia", "Unidad", "Cantidad", "Importe"]]
        for cargo in obra.cargos[:28]:
            cargos_rows.append(
                [
                    Paragraph(_truncate_text(cargo.nombre or "", 34), styles["MarCell"]),
                    Paragraph(_truncate_text(cargo.incidencia or "", 18), styles["MarCellCenter"]),
                    Paragraph(_truncate_text(cargo.unidad or "", 12), styles["MarCellCenter"]),
                    Paragraph("{:,.4f}".format(float(cargo.cantidad or 0)), styles["MarCellRight"]),
                    Paragraph(_fmt_money(cargo.importe or 0), styles["MarCellRight"]),
                ]
            )
        cargos_table = Table(cargos_rows, colWidths=[80 * mm, 30 * mm, 20 * mm, 24 * mm, 32 * mm], repeatRows=1)
        cargos_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#102542")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0D9E3")),
                    ("PADDING", (0, 0), (-1, -1), 4),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ]
            )
        )
        story.append(cargos_table)
        story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("<b>Presupuesto por partidas</b>", styles["Heading3"]))
    part_rows = [[
        "Ord",
        "Capítulo",
        "Clave",
        "Concepto",
        "Uni.",
        "Cant.",
        "P.U.",
        "Importe",
    ]]
    for partida in obra.partidas_ordenadas:
        part_rows.append(
            [
                Paragraph(str(partida.orden or ""), styles["MarCellCenter"]),
                Paragraph(_truncate_text(partida.capitulo or "-", 26), styles["MarCell"]),
                Paragraph(_truncate_text(partida.clave or "-", 18), styles["MarCell"]),
                Paragraph(_truncate_text(partida.concepto or "", 88), styles["MarCell"]),
                Paragraph(_truncate_text(partida.unidad or "", 10), styles["MarCellCenter"]),
                Paragraph("{:,.4f}".format(float(partida.cantidad or 0)), styles["MarCellRight"]),
                Paragraph(_fmt_money(partida.precio_unitario), styles["MarCellRight"]),
                Paragraph(_fmt_money(partida.importe_venta), styles["MarCellRight"]),
            ]
        )
    part_table = Table(part_rows, colWidths=[10 * mm, 25 * mm, 18 * mm, 73 * mm, 12 * mm, 18 * mm, 20 * mm, 24 * mm], repeatRows=1)
    part_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f56a7")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0D9E3")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFD")]),
                ("PADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(part_table)
    story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("<b>Explosión consolidada de insumos</b>", styles["Heading3"]))
    exp_rows = [["Tipo", "Código", "Descripción", "Uni.", "Cantidad", "Importe"]]
    for item in obra.explosion_insumos[:45]:
        exp_rows.append(
            [
                Paragraph(_truncate_text(item.get("tipo") or "", 14), styles["MarCell"]),
                Paragraph(_truncate_text(item.get("codigo") or "-", 18), styles["MarCell"]),
                Paragraph(_truncate_text(item.get("descripcion") or "", 74), styles["MarCell"]),
                Paragraph(_truncate_text(item.get("unidad") or "", 10), styles["MarCellCenter"]),
                Paragraph("{:,.4f}".format(float(item.get("cantidad") or 0)), styles["MarCellRight"]),
                Paragraph(_fmt_money(item.get("importe") or 0), styles["MarCellRight"]),
            ]
        )
    exp_table = Table(exp_rows, colWidths=[18 * mm, 22 * mm, 89 * mm, 12 * mm, 22 * mm, 24 * mm], repeatRows=1)
    exp_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#102542")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0D9E3")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFCFE")]),
                ("PADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(exp_table)

    if obra.programa_resumen.get("calendario"):
        story.append(Spacer(1, 5 * mm))
        story.append(Paragraph("<b>Programa base por periodos</b>", styles["Heading3"]))
        cal_rows = [["Periodo", "Inicio", "Fin", "Concepto", "Jornadas"]]
        for item in obra.programa_resumen["calendario"][:35]:
            cal_rows.append(
                [
                    Paragraph(item.get("periodo") or "-", styles["MarCellCenter"]),
                    Paragraph(item["inicio"].strftime("%d/%m/%Y") if item.get("inicio") else "-", styles["MarCellCenter"]),
                    Paragraph(item["fin"].strftime("%d/%m/%Y") if item.get("fin") else "-", styles["MarCellCenter"]),
                    Paragraph(_truncate_text(item.get("concepto") or "", 68), styles["MarCell"]),
                    Paragraph("{:,.4f}".format(float(item.get("jornadas") or 0)), styles["MarCellRight"]),
                ]
            )
        cal_table = Table(cal_rows, colWidths=[18 * mm, 24 * mm, 24 * mm, 100 * mm, 22 * mm], repeatRows=1)
        cal_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f56a7")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0D9E3")),
                    ("PADDING", (0, 0), (-1, -1), 4),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ]
            )
        )
        story.append(cal_table)

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)
    pdf = buf.getvalue()
    buf.close()
    filename = f"obra_{obra.clave or obra.id}.pdf"
    return Response(pdf, mimetype="application/pdf", headers={"Content-Disposition": f'inline; filename="{filename}"'})


@apu_bp.route("/obras/<int:obra_id>/a-cotizacion", methods=["POST"])
@login_required
def obra_to_cotizacion(obra_id):
    obra = Obra.query.get_or_404(obra_id)
    _decorate_obra(obra)

    cliente = None
    nombre_cliente = (obra.cliente or "").strip()
    if nombre_cliente:
        cliente = Cliente.query.filter(db.func.lower(Cliente.nombre_cliente) == nombre_cliente.lower()).first()
        if not cliente:
            cliente = Cliente(
                nombre_cliente=nombre_cliente,
                empresa=None,
                responsable=_responsable_actual(),
                correo=None,
                telefono=None,
                direccion=obra.ubicacion,
            )
            db.session.add(cliente)
            db.session.flush()

    notas_base = [
        f"Generado desde MAR Data / Obra: {obra.nombre}",
        f"Clave obra: {obra.clave or obra.id}",
    ]
    if obra.descripcion:
        notas_base.append(obra.descripcion)

    cot = Cotizacion(
        folio=_generar_folio(),
        fecha=datetime.utcnow(),
        cliente_id=cliente.id if cliente else None,
        estatus="PENDIENTE",
        subtotal=round(float(obra.subtotal_directo or 0), 2),
        descuento_total=0.0,
        iva_porc=16.0,
        iva_monto=round(float(obra.total_venta or 0) * 0.16, 2),
        total=round(float(obra.total_venta or 0) * 1.16, 2),
        notas="\n".join(notas_base),
        responsable=_responsable_actual(),
    )
    db.session.add(cot)
    db.session.flush()

    subtotal_venta = 0.0
    for partida in obra.partidas_ordenadas:
        apu = partida.apu
        concepto = Concepto.query.filter_by(nombre_concepto=partida.concepto).first()
        if not concepto:
            concepto = Concepto(
                nombre_concepto=partida.concepto,
                unidad=partida.unidad,
                precio_unitario=partida.precio_unitario,
                sistema="MAR DATA",
                descripcion=getattr(apu, "descripcion", None) or f"Obra {obra.nombre}",
            )
            db.session.add(concepto)
            db.session.flush()

        line_subtotal = float(partida.importe_venta or 0)
        subtotal_venta += line_subtotal
        resumen = {
            "obra_id": obra.id,
            "obra_nombre": obra.nombre,
            "partida_id": partida.id,
            "apu_id": apu.id if apu else None,
            "apu_directo": float(getattr(apu, "costo_directo", 0.0) if apu else 0.0),
            "precio_unitario": float(partida.precio_unitario or 0.0),
        }
        det = CotizacionDetalle(
            **_safe_detalle_kwargs(
                cotizacion_id=cot.id,
                concepto_id=concepto.id,
                nombre_concepto=partida.concepto,
                unidad=partida.unidad,
                cantidad=float(partida.cantidad or 0),
                precio_unitario=float(partida.precio_unitario or 0),
                capitulo=partida.capitulo,
                sistema=f"MAR DATA / {obra.nombre}",
                descripcion=(getattr(apu, "descripcion", None) or partida.comentario or f"Partida de obra {obra.nombre}"),
                subtotal=round(line_subtotal, 2),
                origen="APU",
                apu_id=apu.id if apu else None,
                apu_clave=partida.clave,
                apu_directo=float(getattr(apu, "costo_directo", 0) if apu else 0),
                apu_resumen_json=json.dumps(resumen, ensure_ascii=False),
            )
        )
        db.session.add(det)

    cot.subtotal = round(subtotal_venta, 2)
    cot.iva_monto = round(subtotal_venta * 0.16, 2)
    cot.total = round(subtotal_venta + cot.iva_monto, 2)
    db.session.commit()
    flash("La obra se convirtió en cotización con todas sus partidas.", "success")
    return redirect(url_for("view_cotizacion", cot_id=cot.id))


@apu_bp.route("/api/suggest")
@login_required
def api_apu_suggest():
    q = (request.args.get("q") or "").strip()
    if len(q) < 1:
        return jsonify([])
    filtros = [APU.concepto.ilike(f"%{q}%"), APU.clave.ilike(f"%{q}%")]
    if q.isdigit():
        filtros.append(APU.id == int(q))

    rows = APU.query.filter(or_(*filtros)).order_by(APU.concepto.asc()).limit(15).all()
    return jsonify(
        [
            {
                "id": item.id,
                "concepto": item.concepto,
                "unidad": item.unidad,
                "precio_unitario": item.precio_unitario,
                "clave": item.clave or "",
                "categoria": item.categoria or "",
                "capitulo": item.capitulo or "",
                "es_auxiliar": bool(item.es_auxiliar),
            }
            for item in rows
        ]
    )


@apu_bp.route("/api/<int:apu_id>/resumen")
@login_required
def api_apu_resumen(apu_id):
    apu = APU.query.get_or_404(apu_id)
    recalcular_apu(apu)
    db.session.commit()
    return jsonify(
        {
            "id": apu.id,
            "clave": apu.clave,
            "concepto": apu.concepto,
            "categoria": apu.categoria,
            "capitulo": apu.capitulo,
            "subcapitulo": apu.subcapitulo,
            "unidad": apu.unidad,
            "precio_unitario": apu.precio_unitario,
            "costo_directo": apu.costo_directo,
            "indirecto_monto": apu.indirecto_monto,
            "financiamiento_monto": apu.financiamiento_monto,
            "utilidad_monto": apu.utilidad_monto,
            "cargos_adicionales_monto": apu.cargos_adicionales_monto,
            "importe_partida": apu.importe_partida,
            "descripcion": apu.descripcion,
            "es_auxiliar": bool(apu.es_auxiliar),
        }
    )


@apu_bp.route("/cotizador-rapido", methods=["GET", "POST"])
@login_required
def apu_cotizador_rapido():
    if request.method == "POST":
        apu_id = int(request.form.get("apu_id"))
        cantidad = _f(request.form.get("cantidad"), 1.0)
        nombre_cliente = (request.form.get("cliente") or "").strip()
        empresa = (request.form.get("empresa") or "").strip()
        correo = (request.form.get("correo") or "").strip() or None
        telefono = (request.form.get("telefono") or "").strip() or None
        direccion = (request.form.get("direccion") or "").strip() or None
        notas = (request.form.get("notas") or "").strip() or None

        apu = APU.query.get_or_404(apu_id)
        recalcular_apu(apu)
        db.session.commit()

        concepto = Concepto.query.filter_by(nombre_concepto=apu.concepto).first()
        if not concepto:
            concepto = Concepto(
                nombre_concepto=apu.concepto,
                unidad=apu.unidad,
                precio_unitario=apu.precio_unitario,
                sistema="MAR DATA",
                descripcion=apu.descripcion or f"Generado desde MAR DATA {apu.clave or apu.id}",
            )
            db.session.add(concepto)
            db.session.flush()

        cliente = None
        if nombre_cliente:
            cliente = Cliente.query.filter(db.func.lower(Cliente.nombre_cliente) == nombre_cliente.lower()).first()
            if not cliente:
                cliente = Cliente(
                    nombre_cliente=nombre_cliente,
                    empresa=empresa or None,
                    responsable=_responsable_actual(),
                    correo=correo,
                    telefono=telefono,
                    direccion=direccion,
                )
                db.session.add(cliente)
                db.session.flush()

        subtotal = float(cantidad) * float(apu.precio_unitario or 0)
        iva_porc = 16.0
        iva_monto = subtotal * 0.16
        total = subtotal + iva_monto

        cot = Cotizacion(
            folio=_generar_folio(),
            fecha=datetime.utcnow(),
            cliente_id=cliente.id if cliente else None,
            estatus="PENDIENTE",
            subtotal=round(subtotal, 2),
            descuento_total=0.0,
            iva_porc=iva_porc,
            iva_monto=round(iva_monto, 2),
            total=round(total, 2),
            notas=notas,
            responsable=_responsable_actual(),
        )
        db.session.add(cot)
        db.session.flush()

        det = CotizacionDetalle(
            cotizacion_id=cot.id,
            concepto_id=concepto.id,
            nombre_concepto=apu.concepto,
            unidad=apu.unidad,
            cantidad=float(cantidad),
            precio_unitario=float(apu.precio_unitario or 0),
            sistema="MAR DATA",
            descripcion=f"Creado desde APU {apu.clave or apu.id}",
            subtotal=round(subtotal, 2),
        )
        db.session.add(det)
        db.session.commit()

        flash("Cotizacion rapida generada desde MAR DATA.", "success")
        return redirect(url_for("view_cotizacion", cot_id=cot.id))

    return render_template("apu_cotizador_rapido.html", title="Cotizador rapido MAR DATA")
